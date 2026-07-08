"""
ActuarIA — Service partagé : Génération de rapports Excel
Direction Vie & EP-RE

Fondation commune pour tous les agents de la direction.
Produit des fichiers .xlsx auditables, traçables et conformes
aux exigences des commissaires aux comptes et de l'ACPR.

PALETTE :
  Navy   #0F2E52  — entêtes et fond
  Or     #C9A84C  — titres et valeurs clés
  Blanc  #F0F4F8  — texte sur fond foncé
  Gris   #8A9AB0  — libellés secondaires
  Vert   #2ECC71  — statuts VERT
  Ambre  #F39C12  — statuts AMBRE
  Rouge  #E74C3C  — statuts ROUGE

UTILISATION :
  from direction_vie_epre.services.rapport_excel import (
      creer_workbook_actuariel,
      ajouter_onglet_hypotheses,
      ajouter_onglet_resultats,
      ajouter_onglet_audit_trail,
      ajouter_onglet_validation,
      ajouter_onglet_sensibilites,
      finaliser_et_retourner,
  )
"""

import io
from datetime import datetime
from typing import Dict, List, Optional, Any

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.worksheet import Worksheet
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

# ── Palette couleurs ─────────────────────────────────────────────────────────
NAVY   = "0F2E52"
OR     = "C9A84C"
BLANC  = "F0F4F8"
GRIS   = "8A9AB0"
VERT_H = "2ECC71"
AMBRE  = "F39C12"
ROUGE  = "E74C3C"
NAVY_L = "1B3A5C"
GRIS_L = "EAF0F6"

# ── Styles réutilisables ─────────────────────────────────────────────────────
def _font(bold=False, color=BLANC, size=11, italic=False):
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _border_thin():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _statut_couleur(statut: str) -> str:
    return {"VERT": VERT_H, "AMBRE": AMBRE, "ROUGE": ROUGE}.get(
        statut.upper(), GRIS
    )

# ── Helpers de mise en forme ─────────────────────────────────────────────────
def _set_col_width(ws: "Worksheet", col: int, width: float):
    ws.column_dimensions[get_column_letter(col)].width = width

def _cellule(
    ws: "Worksheet", row: int, col: int, valeur,
    bold=False, color_font=BLANC, fill_color=None,
    align_h="left", fmt=None, border=True, wrap=False, italic=False,
):
    cell = ws.cell(row=row, column=col, value=valeur)
    cell.font      = _font(bold=bold, color=color_font, italic=italic)
    cell.alignment = _align(h=align_h, wrap=wrap)
    if fill_color:
        cell.fill = _fill(fill_color)
    if border:
        cell.border = _border_thin()
    if fmt:
        cell.number_format = fmt
    return cell

def _entete_colonne(ws, row, col, texte, width=18):
    _cellule(ws, row, col, texte, bold=True,
             color_font=BLANC, fill_color=NAVY, align_h="center")
    _set_col_width(ws, col, width)

def _ligne_section(ws, row, n_cols, titre):
    """Ligne de titre de section — fond Navy clair, texte Or."""
    for c in range(1, n_cols + 1):
        ws.merge_cells(
            start_row=row, start_column=1,
            end_row=row, end_column=n_cols
        ) if c == 1 else None
        break
    cell = ws.cell(row=row, column=1, value=f"  {titre}")
    cell.font      = _font(bold=True, color=OR, size=11)
    cell.fill      = _fill(NAVY_L)
    cell.alignment = _align(h="left")
    cell.border    = _border_thin()

def _bandeau_titre(ws, titre, sous_titre, agent, audit_id, date_rapport, n_cols=8):
    """
    Bandeau d'en-tête standardisé ActuarIA.
    Lignes 1-5 : logo textuel, titre, sous-titre, agent, date/audit_id.
    """
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 18
    ws.row_dimensions[4].height = 16
    ws.row_dimensions[5].height = 14

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    c1 = ws.cell(row=1, column=1, value="  ActuarIA — Direction Vie & EP-RE")
    c1.font      = _font(bold=True, color=OR, size=14)
    c1.fill      = _fill(NAVY)
    c1.alignment = _align(h="left", v="center")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    c2 = ws.cell(row=2, column=1, value=f"  {titre}")
    c2.font      = _font(bold=True, color=BLANC, size=12)
    c2.fill      = _fill(NAVY)
    c2.alignment = _align(h="left", v="center")

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=n_cols)
    c3 = ws.cell(row=3, column=1, value=f"  {sous_titre}")
    c3.font      = _font(color=GRIS, size=10, italic=True)
    c3.fill      = _fill(NAVY)
    c3.alignment = _align(h="left", v="center")

    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=n_cols)
    c4 = ws.cell(row=4, column=1, value=f"  Agent : {agent}")
    c4.font      = _font(color=BLANC, size=10)
    c4.fill      = _fill(NAVY_L)
    c4.alignment = _align(h="left", v="center")

    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=n_cols)
    c5 = ws.cell(
        row=5, column=1,
        value=f"  Date : {date_rapport}   |   Audit ID : {audit_id}"
    )
    c5.font      = _font(color=GRIS, size=9, italic=True)
    c5.fill      = _fill(NAVY_L)
    c5.alignment = _align(h="left", v="center")

    return 6  # première ligne disponible après le bandeau


# ══════════════════════════════════════════════════════════════════════════════
# FONCTIONS PUBLIQUES
# ══════════════════════════════════════════════════════════════════════════════

def creer_workbook_actuariel() -> "Workbook":
    """Crée un Workbook vierge avec les paramètres de base."""
    if not OPENPYXL_OK:
        raise ImportError("openpyxl requis : pip install openpyxl")
    wb = Workbook()
    # Supprimer l'onglet par défaut — les agents ajoutent leurs propres onglets
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    return wb


def ajouter_onglet_hypotheses(
    wb: "Workbook",
    agent: str,
    audit_id: str,
    date_rapport: str,
    hypotheses: List[Dict],
    sources: Optional[Dict] = None,
) -> "Worksheet":
    """
    Onglet "Hypothèses" : paramètres d'entrée avec leur source.

    hypotheses : liste de dicts avec clés :
        label     (str)   — libellé de l'hypothèse
        valeur    (Any)   — valeur utilisée
        unite     (str)   — unité (%, €, ans, ...)
        source    (str)   — origine (saisie manuelle / agent amont / table officielle)
        reference (str)   — référence réglementaire ou actuarielle
    """
    ws = wb.create_sheet("Hypothèses")
    n_cols = 6
    row = _bandeau_titre(
        ws,
        titre=f"Hypothèses actuarielles — {agent}",
        sous_titre="Paramètres d'entrée utilisés dans les calculs",
        agent=agent, audit_id=audit_id, date_rapport=date_rapport,
        n_cols=n_cols,
    )

    # En-têtes colonnes
    headers = ["Hypothèse", "Valeur", "Unité", "Source", "Référence réglementaire", "Statut"]
    widths  = [35, 15, 12, 30, 40, 14]
    for c, (h, w) in enumerate(zip(headers, widths), 1):
        _entete_colonne(ws, row, c, h, width=w)
    row += 1

    # Données
    for hyp in hypotheses:
        fill = GRIS_L if row % 2 == 0 else "FFFFFF"
        _cellule(ws, row, 1, hyp.get("label", ""),       fill_color=fill, color_font="0F2E52")
        _cellule(ws, row, 2, hyp.get("valeur", ""),      fill_color=fill, color_font="0F2E52",
                 align_h="right", fmt=hyp.get("fmt_excel"))
        _cellule(ws, row, 3, hyp.get("unite", ""),       fill_color=fill, color_font="0F2E52", align_h="center")
        _cellule(ws, row, 4, hyp.get("source", "saisie manuelle"), fill_color=fill, color_font="0F2E52")
        _cellule(ws, row, 5, hyp.get("reference", ""),   fill_color=fill, color_font="0F2E52")
        statut = hyp.get("statut", "")
        c_statut = _statut_couleur(statut) if statut else GRIS
        _cellule(ws, row, 6, statut, fill_color=c_statut if statut else fill,
                 color_font=BLANC if statut else "0F2E52", align_h="center", bold=bool(statut))
        row += 1

    # Sources globales (si fournies)
    if sources:
        row += 1
        _ligne_section(ws, row, n_cols, "Sources des données d'entrée")
        row += 1
        for k, v in sources.items():
            _cellule(ws, row, 1, k,    fill_color=GRIS_L, color_font="0F2E52")
            _cellule(ws, row, 2, v,    fill_color=GRIS_L, color_font="0F2E52")
            for c in range(3, n_cols + 1):
                _cellule(ws, row, c, "", fill_color=GRIS_L, color_font="0F2E52")
            row += 1

    ws.freeze_panes = "A7"
    return ws


def ajouter_onglet_resultats(
    wb: "Workbook",
    agent: str,
    audit_id: str,
    date_rapport: str,
    sections: List[Dict],
) -> "Worksheet":
    """
    Onglet "Résultats" : résultats structurés par section.

    sections : liste de dicts avec clés :
        titre   (str)        — titre de la section
        lignes  (List[Dict]) — chaque dict : {label, valeur, unite, statut, commentaire}
    """
    ws = wb.create_sheet("Résultats")
    n_cols = 5
    row = _bandeau_titre(
        ws,
        titre=f"Résultats actuariels — {agent}",
        sous_titre="Résultats calculés — auditables et traçables",
        agent=agent, audit_id=audit_id, date_rapport=date_rapport,
        n_cols=n_cols,
    )

    headers = ["Indicateur", "Valeur", "Unité", "Statut RAG", "Commentaire"]
    widths  = [40, 20, 14, 14, 55]
    for c, (h, w) in enumerate(zip(headers, widths), 1):
        _entete_colonne(ws, row, c, h, width=w)
    row += 1

    for section in sections:
        _ligne_section(ws, row, n_cols, section.get("titre", ""))
        row += 1
        for ligne in section.get("lignes", []):
            fill = GRIS_L if row % 2 == 0 else "FFFFFF"
            _cellule(ws, row, 1, ligne.get("label", ""),    fill_color=fill, color_font="0F2E52")
            _cellule(ws, row, 2, ligne.get("valeur", ""),   fill_color=fill, color_font="0F2E52",
                     align_h="right", fmt=ligne.get("fmt_excel"))
            _cellule(ws, row, 3, ligne.get("unite", ""),    fill_color=fill, color_font="0F2E52", align_h="center")
            statut = ligne.get("statut", "")
            c_statut = _statut_couleur(statut) if statut else fill
            _cellule(ws, row, 4, statut,
                     fill_color=c_statut if statut else fill,
                     color_font=BLANC if statut else "0F2E52",
                     align_h="center", bold=bool(statut))
            _cellule(ws, row, 5, ligne.get("commentaire", ""),
                     fill_color=fill, color_font="0F2E52", wrap=True)
            row += 1
        row += 1

    ws.freeze_panes = "A7"
    return ws


def ajouter_onglet_validation(
    wb: "Workbook",
    agent: str,
    audit_id: str,
    date_rapport: str,
    validation: Dict,
    cles_controles: List[str],
) -> "Worksheet":
    """
    Onglet "Validation" : scorecard des contrôles actuariels.

    validation    : dict retour de _valider_*()
    cles_controles: clés des contrôles dans le dict (ex: ["h1_taux", "h2_qx", "h3_comp"])
    """
    ws = wb.create_sheet("Validation")
    n_cols = 5
    row = _bandeau_titre(
        ws,
        titre=f"Scorecard de validation — {agent}",
        sous_titre="Contrôles actuariels — conformité réglementaire",
        agent=agent, audit_id=audit_id, date_rapport=date_rapport,
        n_cols=n_cols,
    )

    headers = ["Contrôle", "Statut", "Message", "Conseil", "Référence"]
    widths  = [35, 12, 55, 55, 35]
    for c, (h, w) in enumerate(zip(headers, widths), 1):
        _entete_colonne(ws, row, c, h, width=w)
    row += 1

    for cle in cles_controles:
        ctrl = validation.get(cle, {})
        if not ctrl:
            continue
        statut = ctrl.get("statut", "")
        fill   = GRIS_L if row % 2 == 0 else "FFFFFF"
        label  = (ctrl.get("titre_graphique", cle)
                  .replace("✅ ", "").replace("⚠️ ", "").replace("❌ ", ""))
        _cellule(ws, row, 1, label,                        fill_color=fill,    color_font="0F2E52")
        c_statut = _statut_couleur(statut)
        _cellule(ws, row, 2, statut,    fill_color=c_statut, color_font=BLANC,  align_h="center", bold=True)
        _cellule(ws, row, 3, ctrl.get("message", ""),      fill_color=fill,    color_font="0F2E52", wrap=True)
        _cellule(ws, row, 4, ctrl.get("conseil", ""),      fill_color=fill,    color_font="0F2E52", wrap=True)
        _cellule(ws, row, 5, ctrl.get("reference", ""),    fill_color=fill,    color_font="0F2E52", wrap=True)
        ws.row_dimensions[row].height = 32
        row += 1

    # Conclusion globale
    row += 1
    sg  = validation.get("statut_global", "")
    ccl = validation.get("conclusion", "")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    cell = ws.cell(row=row, column=1, value=f"  Conclusion : {ccl}")
    cell.font      = _font(bold=True, color=BLANC, size=11)
    cell.fill      = _fill(_statut_couleur(sg) if sg else NAVY)
    cell.alignment = _align(h="left", v="center")
    ws.row_dimensions[row].height = 24

    ws.freeze_panes = "A7"
    return ws


def ajouter_onglet_sensibilites(
    wb: "Workbook",
    agent: str,
    audit_id: str,
    date_rapport: str,
    sensibilites: List[Dict],
) -> "Worksheet":
    """
    Onglet "Sensibilités" : analyse de sensibilité des résultats.

    sensibilites : liste de dicts :
        parametre  (str)   — paramètre choqué
        choc       (str)   — description du choc (ex: "+100bp")
        valeur_ref (float) — valeur de référence
        valeur_cho (float) — valeur après choc
        impact_abs (float) — impact absolu
        impact_pct (float) — impact en %
        unite      (str)   — unité
        reference  (str)   — base réglementaire du choc
    """
    ws = wb.create_sheet("Sensibilités")
    n_cols = 8
    row = _bandeau_titre(
        ws,
        titre=f"Analyse de sensibilité — {agent}",
        sous_titre="Impact des chocs sur les résultats actuariels",
        agent=agent, audit_id=audit_id, date_rapport=date_rapport,
        n_cols=n_cols,
    )

    headers = ["Paramètre", "Choc", "Valeur réf.", "Valeur choquée",
               "Impact abs.", "Impact %", "Unité", "Référence"]
    widths  = [35, 16, 18, 18, 18, 12, 12, 35]
    for c, (h, w) in enumerate(zip(headers, widths), 1):
        _entete_colonne(ws, row, c, h, width=w)
    row += 1

    FMT_NB  = "#,##0.00"
    FMT_PCT = "0.00%"

    for s in sensibilites:
        fill = GRIS_L if row % 2 == 0 else "FFFFFF"
        _cellule(ws, row, 1, s.get("parametre", ""),  fill_color=fill, color_font="0F2E52")
        _cellule(ws, row, 2, s.get("choc", ""),       fill_color=fill, color_font="0F2E52", align_h="center")
        _cellule(ws, row, 3, s.get("valeur_ref", ""), fill_color=fill, color_font="0F2E52",
                 align_h="right", fmt=FMT_NB)
        _cellule(ws, row, 4, s.get("valeur_cho", ""), fill_color=fill, color_font="0F2E52",
                 align_h="right", fmt=FMT_NB)
        # Impact : coloration selon le signe
        imp = s.get("impact_abs", 0) or 0
        c_imp = "E74C3C" if imp < 0 else "2ECC71" if imp > 0 else fill
        _cellule(ws, row, 5, imp,                     fill_color=fill,  color_font=c_imp,
                 align_h="right", fmt=FMT_NB, bold=True)
        _cellule(ws, row, 6, (s.get("impact_pct") or 0) / 100,
                 fill_color=fill, color_font="0F2E52", align_h="right", fmt=FMT_PCT)
        _cellule(ws, row, 7, s.get("unite", ""),      fill_color=fill, color_font="0F2E52", align_h="center")
        _cellule(ws, row, 8, s.get("reference", ""),  fill_color=fill, color_font="0F2E52")
        row += 1

    ws.freeze_panes = "A7"
    return ws


def ajouter_onglet_audit_trail(
    wb: "Workbook",
    agent: str,
    audit_id: str,
    date_rapport: str,
    result: Dict,
    actuaire_resp: str = "À renseigner",
) -> "Worksheet":
    """
    Onglet "Audit Trail" : traçabilité complète et immuable.
    Requis par l'ACPR et les commissaires aux comptes.
    Contient : identifiants, paramètres, statuts, chaîne de sources, horodatage.
    """
    ws = wb.create_sheet("Audit Trail")
    n_cols = 3
    row = _bandeau_titre(
        ws,
        titre=f"Audit Trail — {agent}",
        sous_titre="Traçabilité complète — ne pas modifier",
        agent=agent, audit_id=audit_id, date_rapport=date_rapport,
        n_cols=n_cols,
    )

    def _ligne_at(label, valeur, fill=None):
        nonlocal row
        bg = fill or (GRIS_L if row % 2 == 0 else "FFFFFF")
        _cellule(ws, row, 1, label,  fill_color=NAVY_L, color_font=OR,    bold=True)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=n_cols)
        _cellule(ws, row, 2, str(valeur), fill_color=bg,    color_font="0F2E52", wrap=True)
        ws.row_dimensions[row].height = 18
        row += 1

    _set_col_width(ws, 1, 30)
    _set_col_width(ws, 2, 55)
    _set_col_width(ws, 3, 20)

    _ligne_section(ws, row, n_cols, "Identification")
    row += 1
    _ligne_at("Audit ID",          audit_id)
    _ligne_at("Agent",             agent)
    _ligne_at("Date / Heure",      date_rapport)
    _ligne_at("Actuaire responsable", actuaire_resp)
    _ligne_at("Statut global",     result.get("statut_rag", "N/A"))

    row += 1
    _ligne_section(ws, row, n_cols, "Sources des données")
    row += 1
    sources = result.get("sources", {"parametres": "saisie manuelle"})
    for k, v in sources.items():
        _ligne_at(k, v)

    row += 1
    _ligne_section(ws, row, n_cols, "Paramètres utilisés")
    row += 1
    # Sérialiser les paramètres scalaires du result
    for k, v in result.items():
        if isinstance(v, (int, float, str, bool)) and k not in (
            "success", "agent", "audit_id", "commentaire", "erreur",
            "statut_rag", "be_ifrs17_mode"
        ):
            _ligne_at(k, v)

    row += 1
    _ligne_section(ws, row, n_cols, "Commentaire actuariel")
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    cell = ws.cell(row=row, column=1, value=result.get("commentaire", ""))
    cell.font      = _font(color="0F2E52", size=10)
    cell.fill      = _fill("FFFFFF")
    cell.alignment = _align(h="left", v="top", wrap=True)
    cell.border    = _border_thin()
    ws.row_dimensions[row].height = 80
    row += 1

    row += 1
    _ligne_section(ws, row, n_cols, "Certification")
    row += 1
    _ligne_at("Certifié par",      actuaire_resp)
    _ligne_at("Certifié le",       date_rapport)
    _ligne_at("Outil",             "ActuarIA — Direction Vie & EP-RE")
    _ligne_at("Version",           "1.0 — Production")
    _ligne_at("Clause",
              "Ce rapport est produit par un outil de calcul actuariel. "
              "Il doit être validé et signé par un actuaire qualifié "
              "avant toute utilisation réglementaire (ACPR, CA, audit).")

    ws.freeze_panes = "A7"
    return ws


def finaliser_et_retourner(wb: "Workbook") -> bytes:
    """
    Finalise le workbook et retourne les bytes du fichier .xlsx.
    Les bytes peuvent être stockés dans excel_bytes du dict retour de l'agent.
    """
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
