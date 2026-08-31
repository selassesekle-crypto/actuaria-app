"""
ActuarIA — Service partagé : Rapport Excel Tarification Non-Vie
Direction Non-Vie · Équipe Tarification

Produit des rapports Excel multi-onglets pour A3 (GLM), A4 (ML) et A6 (Comparaison).
Palette Navy/Gold ActuarIA, formats €/%, audit trail intégré.

ONGLETS PAR AGENT :
  A3 GLM  : Synthèse · Relativités · Métriques GLM · Hypothèses · Audit
  A4 ML   : Synthèse · Classement · SHAP · Hypothèses · Monitoring · Audit
  A6      : Synthèse · Classement final · Backtesting A/E · Fiche décision · Audit
"""

import io
from core.conformite_reglementaire import (
    avertissement_controle_effet, avertissement_walk_forward,
    synthese_exclusions, synthese_alertes_experience,
    synthese_colonnes_plan_ecartees, synthese_modele_dl,
)
from core.qualite_donnees import synthese_qualite_donnees
from core.plan_tarifaire import synthese_colonnes_plan_manquantes
from core.mapping_client import synthese_mapping
import logging
from datetime import datetime
from typing import Dict, List, Optional, Any

logger = logging.getLogger('actuaria.tarif.excel')

try:
    from openpyxl import Workbook
except ImportError:
    pass

# ── Helpers Excel partagés — audit V4 point #12 ──────────────────────────────
# Palette, formats et fonctions factorisés dans excel_helpers.py (source
# unique, éliminait une duplication avec rapport_equipe_tarif.py).
try:
    from .excel_helpers import (
        OPENPYXL_OK, NAVY, OR, BLANC, GRIS, VERT_H, AMBRE, ROUGE,
        NAVY_L, GRIS_L, NOIR, FMT_EUR, FMT_PCT, FMT_DEC4, FMT_NB,
        _font, _fill, _border, _align, _statut_fill, _col_w,
        _cell, _header, _section, _kpi, _bandeau,
    )
except ImportError:
    from direction_non_vie.tarification.services.excel_helpers import (
        OPENPYXL_OK, NAVY, OR, BLANC, GRIS, VERT_H, AMBRE, ROUGE,
        NAVY_L, GRIS_L, NOIR, FMT_EUR, FMT_PCT, FMT_DEC4, FMT_NB,
        _font, _fill, _border, _align, _statut_fill, _col_w,
        _cell, _header, _section, _kpi, _bandeau,
    )

if not OPENPYXL_OK:
    logger.warning("openpyxl non disponible — export Excel désactivé")


# =============================================================================
#  EXPORT A3 — GLM
#  6 onglets : Synthèse · Relativités Poisson · Relativités Gamma
#              Hypothèses H1-H4 · Audit Trail · Crédibilité & Géo
#  ⚠️⚠️ Constat `services/C8` : l'en-tête et la docstring annonçaient 5. Le
#  code crée ws1..ws6 — « 6-Crédibilité & Géo » a été ajouté sans que le
#  compte suive. Mesuré : c'est le SEUL écart, a1/a2/a4/a5/a6 sont exacts.
# =============================================================================

def export_excel_a3(result_a3: Dict, audit_id: str = "", arrete: Optional[str] = None) -> bytes:
    """Génère le rapport Excel A3 GLM (6 onglets). Retourne bytes ou b''."""
    if not OPENPYXL_OK or not result_a3 or not result_a3.get('success'):
        return b''
    try:
        wb  = Workbook()
        now = datetime.now().strftime("%d/%m/%Y %H:%M")
        aid = audit_id or result_a3.get('audit_id', 'N/A')
        met = result_a3.get('metriques', {})
        val = result_a3.get('validation_glm', {})

        # ── Onglet 1 : Synthèse ───────────────────────────────────────────────
        ws1 = wb.active
        ws1.title = "1-Synthèse"
        _bandeau(ws1, "Rapport Tarification GLM", "Synthèse des résultats",
                 "A3 — GLM Poisson/Gamma/Tweedie", aid, arrete)
        r = 7
        _section(ws1, r, "▶ MÉTRIQUES GLM"); r += 1
        for modele, label in [('poisson','GLM Poisson'), ('gamma','GLM Gamma'), ('tweedie','GLM Tweedie')]:
            m = met.get(modele, {})
            if m:
                _kpi(ws1, r, f"{label} — Gini test",     round(m.get('gini', 0), 4), fmt=FMT_DEC4)
                r += 1
                _kpi(ws1, r, f"{label} — AIC",            round(m.get('aic', 0), 0),  fmt=FMT_NB)
                r += 1
                _kpi(ws1, r, f"{label} — Déviance nulle", round(m.get('deviance_nulle', 0), 2))
                r += 1
                _kpi(ws1, r, f"{label} — Pseudo-R²",      round(m.get('pseudo_r2', 0), 4))
                r += 1
                _kpi(ws1, r, f"{label} — Vars retenues",  m.get('nb_vars_retenues', 0))
                r += 2

        _section(ws1, r, "▶ STATUT GLOBAL"); r += 1
        statut = val.get('statut_global', 'N/A')
        _kpi(ws1, r, "Statut validation", statut, statut=statut); r += 1
        _kpi(ws1, r, "Conclusion", val.get('conclusion', ''), wrap=True); r += 1

        # ── Onglet 2 : Relativités Poisson ───────────────────────────────────
        ws2 = wb.create_sheet("2-Relativités Poisson")
        _bandeau(ws2, "Relativités Tarifaires", "GLM Poisson — exp(β) avec IC 95%",
                 "A3 — GLM Poisson", aid, arrete)
        rels_p = result_a3.get('relativites_poisson', {})
        r = 7
        _section(ws2, r, "▶ RELATIVITÉS GLM POISSON — exp(β)  [Source : Mildenhall 1999]"); r += 1
        for col, txt, w in [(1,"Variable",28),(2,"β",12),(3,"Relativité exp(β)",20),
                             (4,"IC 95% bas",16),(5,"IC 95% haut",16),
                             (6,"p-value",12),(7,"Significatif",14),(8,"Sens",14)]:
            _header(ws2, r, col, txt, w)
        r += 1
        for var, d in sorted(rels_p.items(), key=lambda x: -abs(x[1].get('beta', 0))):
            vals = [var, d.get('beta',0), d.get('relativite',0),
                    d.get('ic95_low',0), d.get('ic95_high',0),
                    d.get('pvalue',0), "Oui" if d.get('significatif') else "Non",
                    d.get('sens','')]
            fmts = [None, FMT_DEC4, FMT_DEC4, FMT_DEC4, FMT_DEC4, FMT_DEC4, None, None]
            bg = GRIS_L if list(rels_p.keys()).index(var) % 2 == 0 else None
            for j, (v, f) in enumerate(zip(vals, fmts), 1):
                _cell(ws2, r, j, v, cf=NOIR, fill=bg, fmt=f, ah="right" if isinstance(v,(int,float)) else "left")
            # Couleur sens
            col_sens = 8
            fill_s = "EAF3DE" if d.get('sens') == 'allegant' else "FCEBEB"
            ws2.cell(row=r, column=col_sens).fill = _fill(fill_s)
            r += 1

        # ── Onglet 3 : Relativités Gamma ─────────────────────────────────────
        ws3 = wb.create_sheet("3-Relativités Gamma")
        _bandeau(ws3, "Relativités Tarifaires", "GLM Gamma — exp(β) avec IC 95%",
                 "A3 — GLM Gamma", aid, arrete)
        rels_g = result_a3.get('relativites_gamma', {})
        r = 7
        _section(ws3, r, "▶ RELATIVITÉS GLM GAMMA — exp(β) (coût moyen)"); r += 1
        for col, txt, w in [(1,"Variable",28),(2,"β",12),(3,"Relativité exp(β)",20),
                             (4,"IC 95% bas",16),(5,"IC 95% haut",16),
                             (6,"p-value",12),(7,"Significatif",14),(8,"Sens",14)]:
            _header(ws3, r, col, txt, w)
        r += 1
        for var, d in sorted(rels_g.items(), key=lambda x: -abs(x[1].get('beta', 0))):
            vals = [var, d.get('beta',0), d.get('relativite',0),
                    d.get('ic95_low',0), d.get('ic95_high',0),
                    d.get('pvalue',0), "Oui" if d.get('significatif') else "Non",
                    d.get('sens','')]
            fmts = [None, FMT_DEC4, FMT_DEC4, FMT_DEC4, FMT_DEC4, FMT_DEC4, None, None]
            bg = GRIS_L if list(rels_g.keys()).index(var) % 2 == 0 else None
            for j, (v, f) in enumerate(zip(vals, fmts), 1):
                _cell(ws3, r, j, v, cf=NOIR, fill=bg, fmt=f, ah="right" if isinstance(v,(int,float)) else "left")
            r += 1

        # ── Onglet 4 : Hypothèses H1-H4 ──────────────────────────────────────
        ws4 = wb.create_sheet("4-Hypothèses H1-H4")
        _bandeau(ws4, "Validation Hypothèses GLM", "H1 Distribution · H2 Homoscédasticité · H3 Gini · H4 Stabilité",
                 "A3 — Validation réglementaire", aid, arrete)
        r = 7
        hyp_map = [
            ("h1_poisson",  "H1 — Sur-dispersion Poisson",         "ratio_disp"),
            ("h2_homosc",   "H2 — Homoscédasticité résidus",       "ratio_variance"),
            ("h3_ajustement","H3 — Qualité ajustement (Gini)",     "gini_max"),
            ("h4_stabilite","H4 — Stabilité relativités bootstrap","cv_max"),
        ]
        for hkey, hlabel, hval_key in hyp_map:
            h = val.get(hkey, {})
            if not h:
                continue
            st = h.get('statut', 'N/A')
            _section(ws4, r, f"▶ {hlabel}"); r += 1
            _kpi(ws4, r, "Statut",     st,                    statut=st); r += 1
            # ⚠️ MEME DEFAUT QU'AU BLOC A4, ET IL Y EN AVAIT DEUX : 0 pour
            # une valeur absente, et un plantage sur un None.
            _v = h.get(hval_key)
            _kpi(ws4, r, "Valeur",
                 '—' if _v is None else round(_v, 4)); r += 1
            _kpi(ws4, r, "Message",    h.get('message', ''),  wrap=True); r += 1
            _kpi(ws4, r, "Conseil",    h.get('conseil', ''),  wrap=True); r += 2

        # ── Onglet 5 : Audit Trail ────────────────────────────────────────────
        ws5 = wb.create_sheet("5-Audit Trail")
        _bandeau(ws5, "Audit Trail", "Traçabilité ACPR — Agent A3 GLM",
                 "A3 — Audit", aid, arrete)
        r = 7
        _section(ws5, r, "▶ INFORMATIONS AUDIT"); r += 1
        _kpi(ws5, r, "Audit ID",    aid);  r += 1
        _kpi(ws5, r, "Date",        now);  r += 1
        _kpi(ws5, r, "Agent",       "A3 GLM Poisson/Gamma/Tweedie"); r += 1
        _kpi(ws5, r, "Branche",     result_a3.get('branche', 'N/A')); r += 1
        _kpi(ws5, r, "Statut RAG",  result_a3.get('statut_rag', 'N/A'),
             statut=result_a3.get('statut_rag')); r += 1
        _kpi(ws5, r, "Nb obs train",result_a3.get('metriques',{}).get('poisson',{}).get('nb_obs_train',0),
             fmt=FMT_NB); r += 1
        _kpi(ws5, r, "Nb obs test", result_a3.get('metriques',{}).get('poisson',{}).get('nb_obs_test',0),
             fmt=FMT_NB); r += 1
        _kpi(ws5, r, "Vars retenues (Poisson)",
             ', '.join(result_a3.get('metriques',{}).get('poisson',{}).get('vars_retenues', [])),
             wrap=True); r += 1

        # ── Onglet 6 : Crédibilité Bühlmann-Straub & Lissage Géographique ────
        # Réf. : Bühlmann & Straub (1970) ASTIN Bulletin ; Gelfand et al. (2010)
        cred = result_a3.get('credibilite', {})
        geo  = result_a3.get('lissage_geo', {})
        ws6 = wb.create_sheet("6-Crédibilité & Géo")
        _bandeau(ws6, "Crédibilité & Lissage Géo", "Modules avancés P2 — Agent A3 GLM",
                 "A3 — Bühlmann-Straub / Krigeage", aid, arrete)
        r = 7
        _section(ws6, r, "▶ CRÉDIBILITÉ BÜHLMANN-STRAUB"); r += 1
        if cred.get('appliquee'):
            _kpi(ws6, r, "Statut", "✓ Appliquée", statut="VERT"); r += 1
            _kpi(ws6, r, "Colonne de groupe", cred.get('col_groupe', 'N/A')); r += 1
            _kpi(ws6, r, "Nombre de groupes", cred.get('n_groupes', 0), fmt=FMT_NB); r += 1
            _kpi(ws6, r, "Facteur k (σ²intra/σ²entre)", round(cred.get('k', 0), 4), fmt=FMT_DEC4); r += 1
            _kpi(ws6, r, "Z moyen", round(cred.get('z_moyen', 0), 4), fmt=FMT_DEC4); r += 1
            _kpi(ws6, r, "Z min — Z max",
                 f"{cred.get('z_min',0):.4f} — {cred.get('z_max',0):.4f}"); r += 1
            _kpi(ws6, r, "μ marché (taux global)", round(cred.get('mu_marche', 0), 6), fmt=FMT_DEC4); r += 1
            _kpi(ws6, r, "σ²intra", round(cred.get('sigma2_intra', 0), 8), fmt=FMT_DEC4); r += 1
            _kpi(ws6, r, "σ²entre", round(cred.get('sigma2_entre', 0), 8), fmt=FMT_DEC4); r += 1
            r += 1
            _section(ws6, r, "▶ PRIMES CRÉDIBILISÉES PAR GROUPE (top 20)"); r += 1
            headers = ['Groupe', 'Exposition', 'Taux observé', 'Z', 'Prime crédibilisée']
            for ci, h in enumerate(headers, 1):
                _header(ws6, r, ci, h, width=20)
            r += 1
            for ligne in cred.get('primes_par_groupe', [])[:20]:
                vals = list(ligne.values())
                for ci, v in enumerate(vals, 1):
                    _cell(ws6, r, ci, v, ah="center")
                r += 1
            r += 1
            _kpi(ws6, r, "Référence", cred.get('reference', ''), wrap=True); r += 1
        else:
            _kpi(ws6, r, "Statut", "○ Non applicable", statut="AMBRE"); r += 1
            _kpi(ws6, r, "Raison", cred.get('raison', 'N/A'), wrap=True); r += 1

        r += 2
        _section(ws6, r, "▶ LISSAGE GÉOGRAPHIQUE"); r += 1
        if geo.get('applique'):
            _kpi(ws6, r, "Statut", "✓ Appliqué", statut="VERT"); r += 1
            _kpi(ws6, r, "Méthode", geo.get('methode', 'N/A')); r += 1
            _kpi(ws6, r, "Colonne géographique", geo.get('col_geo', 'N/A')); r += 1
            _kpi(ws6, r, "Nombre de zones", geo.get('n_zones', 0), fmt=FMT_NB); r += 1
            _kpi(ws6, r, "Source de la prime", geo.get('source_prime', 'N/A')); r += 1
            if 'mu_global' in geo:
                _kpi(ws6, r, "μ global", round(geo.get('mu_global', 0), 6), fmt=FMT_DEC4); r += 1
            if 'bandwidth_h' in geo:
                _kpi(ws6, r, "Bandwidth h (krigeage)", geo.get('bandwidth_h', 0), fmt=FMT_DEC4); r += 1
            r += 1
            if geo.get('zones'):
                _section(ws6, r, "▶ PRIMES LISSÉES PAR ZONE"); r += 1
                headers = ['Zone', 'N obs', 'Exposition', 'Prime moy.', 'Z géo', 'Prime lissée']
                for ci, h in enumerate(headers, 1):
                    _header(ws6, r, ci, h, width=20)
                r += 1
                for ligne in geo.get('zones', [])[:20]:
                    vals = list(ligne.values())
                    for ci, v in enumerate(vals, 1):
                        _cell(ws6, r, ci, v, ah="center")
                    r += 1
                r += 1
            _kpi(ws6, r, "Référence", geo.get('reference', ''), wrap=True); r += 1
        else:
            _kpi(ws6, r, "Statut", "○ Non applicable", statut="AMBRE"); r += 1
            _kpi(ws6, r, "Raison", geo.get('raison', 'N/A'), wrap=True); r += 1

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    except Exception as e:
        logger.error(f"export_excel_a3 échoué : {e}", exc_info=True)
        return b''


# =============================================================================
#  EXPORT A4 — ML
#  5 onglets : Synthèse · Classement · SHAP · Hypothèses H1-H4 · Audit
# =============================================================================

def export_excel_a4(result_a4: Dict, audit_id: str = "", arrete: Optional[str] = None) -> bytes:
    """Génère le rapport Excel A4 ML (5 onglets). Retourne bytes ou b''."""
    if not OPENPYXL_OK or not result_a4 or not result_a4.get('success'):
        return b''
    try:
        wb  = Workbook()
        now = datetime.now().strftime("%d/%m/%Y %H:%M")
        aid = audit_id or result_a4.get('audit_id', 'N/A')
        val = result_a4.get('validation_ml', {})

        # ── Onglet 1 : Synthèse ───────────────────────────────────────────────
        ws1 = wb.active
        ws1.title = "1-Synthèse"
        _bandeau(ws1, "Rapport Tarification ML",
                 f"{len(result_a4.get('classement', []))} modèles comparés — Sélection multicritères",
                 "A4 — Machine Learning", aid, arrete)
        r = 7
        _section(ws1, r, "▶ MODÈLE SÉLECTIONNÉ"); r += 1
        classement = result_a4.get('classement', [])
        if classement:
            best = classement[0]
            _kpi(ws1, r, "Modèle retenu",    best.get('modele', 'N/A')); r += 1
            _kpi(ws1, r, "Gini test",         round(best.get('gini_test', 0), 4), fmt=FMT_DEC4); r += 1
            _kpi(ws1, r, "RMSE test",         round(best.get('rmse_test', 0), 4), fmt=FMT_DEC4); r += 1
            _kpi(ws1, r, "Overfit ratio",     round(best.get('overfit_ratio', 0), 3)); r += 2

        _section(ws1, r, "▶ STATUT VALIDATION"); r += 1
        statut = val.get('statut_global', 'N/A')
        _kpi(ws1, r, "Statut global",  statut, statut=statut); r += 1
        _kpi(ws1, r, "Conclusion",     val.get('conclusion', ''), wrap=True); r += 1

        # ── Onglet 2 : Classement ─────────────────────────────────────────────
        ws2 = wb.create_sheet("2-Classement modèles")
        _bandeau(ws2, "Classement des modèles ML", "Grille multicritères actuarielle",
                 "A4 — Comparaison", aid, arrete)
        r = 7
        _section(ws2, r, "▶ CLASSEMENT MULTICRITÈRES (Gini 40% · Stabilité 30% · Interprét. 20% · RMSE 10%)"); r += 1
        for col, txt, w in [(1,"#",6),(2,"Modèle",28),(3,"Famille",14),
                             (4,"Gini test",14),(5,"RMSE test",14),
                             (6,"Overfit ratio",16),(7,"Score global",16),(8,"Alerte",12)]:
            _header(ws2, r, col, txt, w)
        r += 1
        for rank, m in enumerate(classement, 1):
            bg = GRIS_L if rank % 2 == 0 else None
            alerte = "⚠ Overfit" if m.get('overfit_alerte') else "✓ OK"
            vals = [rank, m.get('modele',''), m.get('famille',''),
                    round(m.get('gini_test',0),4), round(m.get('rmse_test',0),4),
                    round(m.get('overfit_ratio',0),3), round(m.get('score_global',0),4), alerte]
            fmts = [None,None,None,FMT_DEC4,FMT_DEC4,FMT_DEC4,FMT_DEC4,None]
            for j, (v, f) in enumerate(zip(vals, fmts), 1):
                _cell(ws2, r, j, v, cf=NOIR, fill=bg, fmt=f,
                      ah="right" if isinstance(v,(int,float)) else "left",
                      bold=(rank==1))
            r += 1

        # ── Onglet 3 : SHAP Values ────────────────────────────────────────────
        ws3 = wb.create_sheet("3-SHAP Values")
        _bandeau(ws3, "Feature Importance SHAP", "Interprétabilité AI Act 2025",
                 "A4 — SHAP (Shapley 1953)", aid, arrete)
        shap_vals = result_a4.get('shap_values', {})
        r = 7
        _section(ws3, r, "▶ IMPORTANCE FEATURES (SHAP moyen |valeur|)"); r += 1
        _header(ws3, r, 1, "Feature", 32)
        _header(ws3, r, 2, "Importance SHAP", 22)
        _header(ws3, r, 3, "Rang", 10)
        r += 1
        if isinstance(shap_vals, dict) and shap_vals:
            importance = shap_vals.get('importance_globale', shap_vals)
            if isinstance(importance, dict):
                sorted_feats = sorted(importance.items(), key=lambda x: -abs(x[1]))
                for rang, (feat, imp) in enumerate(sorted_feats[:30], 1):
                    bg = GRIS_L if rang % 2 == 0 else None
                    _cell(ws3, r, 1, feat, cf=NOIR, fill=bg)
                    _cell(ws3, r, 2, round(float(imp), 6), cf=NOIR, fill=bg,
                          fmt=FMT_DEC4, ah="right")
                    _cell(ws3, r, 3, rang, cf=NOIR, fill=bg, ah="center")
                    r += 1
        else:
            _cell(ws3, r, 1, "SHAP non disponible — installer le package 'shap'",
                  cf=GRIS, fill=None, border=False)

        # ── Onglet 4 : Hypothèses H1-H4 ──────────────────────────────────────
        ws4 = wb.create_sheet("4-Hypothèses H1-H4")
        _bandeau(ws4, "Validation Hypothèses ML",
                 "H1 Overfitting · H2 PSI réel · H3 Gini · H4 Calibration",
                 "A4 — Validation", aid, arrete)
        r = 7
        hyp_map = [
            ("h1_overfitting",  "H1 — Absence d'overfitting",         "ratio"),
            ("h2_psi",          "H2 — Stabilité PSI réel",            "psi"),
            ("h3_gini",         "H3 — Performance Gini",              "gini"),
            ("h4_calibration",  "H4 — Calibration reliability",       "ecart_moy_pct"),
        ]
        for hkey, hlabel, hval_key in hyp_map:
            h = val.get(hkey, {})
            if not h:
                continue
            st = h.get('statut', 'N/A')
            _section(ws4, r, f"▶ {hlabel}"); r += 1
            _kpi(ws4, r, "Statut",  st,                   statut=st); r += 1
            # ⚠️ `round(h.get(cle, 0), 4)` FAISAIT DEUX FAUTES : il rendait
            # 0 pour une valeur absente, et il planterait sur un None. Une
            # grandeur non mesuree se declare, elle ne s'invente pas.
            _v = h.get(hval_key)
            _kpi(ws4, r, "Valeur",
                 '—' if _v is None else round(_v, 4)); r += 1
            _kpi(ws4, r, "Message", h.get('message', ''), wrap=True); r += 1
            _kpi(ws4, r, "Conseil", h.get('conseil', ''), wrap=True); r += 2

        # ── Onglet 5 : Audit Trail ────────────────────────────────────────────
        ws5 = wb.create_sheet("5-Audit Trail")
        _bandeau(ws5, "Audit Trail", "Traçabilité ACPR — Agent A4 ML",
                 "A4 — Audit", aid, arrete)
        r = 7
        _section(ws5, r, "▶ INFORMATIONS AUDIT"); r += 1
        _kpi(ws5, r, "Audit ID",       aid); r += 1
        _kpi(ws5, r, "Date",           now); r += 1
        _kpi(ws5, r, "Agent",          "A4 — Machine Learning"); r += 1
        _kpi(ws5, r, "Branche",        result_a4.get('branche', 'N/A')); r += 1
        _kpi(ws5, r, "Statut RAG",     result_a4.get('statut_rag', 'N/A'),
             statut=result_a4.get('statut_rag')); r += 1
        _kpi(ws5, r, "Nb modèles",     len(classement), fmt=FMT_NB); r += 1
        if classement:
            _kpi(ws5, r, "Modèle retenu", classement[0].get('modele','')); r += 1
            _kpi(ws5, r, "Gini retenu",   round(classement[0].get('gini_test',0),4),
                 fmt=FMT_DEC4); r += 1
        opt = result_a4.get('rapport', {}).get('optuna_xgboost')
        if opt:
            _kpi(ws5, r, "Optuna XGBoost — trials", opt.get('trials', 0)); r += 1
            _kpi(ws5, r, "Optuna XGBoost — best Gini",
                 round(opt.get('best_gini', 0), 4), fmt=FMT_DEC4); r += 1

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    except Exception as e:
        logger.error(f"export_excel_a4 échoué : {e}", exc_info=True)
        return b''


# =============================================================================
#  EXPORT A5 — DEEP LEARNING (CANN Wüthrich + TabNet)
#  4 onglets : Synthèse · Classement · Fidélité CANN Wüthrich · Audit
#  Audit V7 MINEUR #2 — A5 était le seul agent sans export Excel.
# =============================================================================

def export_excel_a5(result_a5: Dict, audit_id: str = "", arrete: Optional[str] = None) -> bytes:
    """
    Génère le rapport Excel A5 Deep Learning (4 onglets). Retourne bytes ou b''.

    Audit V7 MINEUR #2 : A5 était le SEUL agent de la direction Tarification
    sans export Excel (aucune fonction export_excel_a5 n'existait, et A5 ne
    référençait aucun export dans son propre run()). Incohérence architecturale
    signalée par l'audit V7 catégorie F — corrigée ici en suivant le même
    gabarit que export_excel_a4 (4 onglets au lieu de 5 : pas de SHAP, non
    calculé pour les modèles Deep Learning).
    """
    if not OPENPYXL_OK or not result_a5 or not result_a5.get('success'):
        return b''
    try:
        wb  = Workbook()
        now = datetime.now().strftime("%d/%m/%Y %H:%M")
        aid = audit_id or result_a5.get('audit_id', 'N/A')
        val = result_a5.get('validation_dl', {})
        classement = result_a5.get('classement', [])
        metriques  = result_a5.get('metriques', {})

        # ── Onglet 1 : Synthèse ───────────────────────────────────────────────
        ws1 = wb.active
        ws1.title = "1-Synthèse"
        _bandeau(ws1, "Rapport Tarification Deep Learning", "CANN Wüthrich + TabNet",
                 "A5 — Deep Learning", aid, arrete)
        r = 7
        _section(ws1, r, "▶ MODÈLE SÉLECTIONNÉ"); r += 1
        if classement:
            best = classement[0]
            _kpi(ws1, r, "Modèle retenu",    best.get('modele', 'N/A')); r += 1
            _kpi(ws1, r, "Gini test",         round(best.get('gini_test', 0), 4), fmt=FMT_DEC4); r += 1
            _kpi(ws1, r, "RMSE test",         round(best.get('rmse_test', 0), 4), fmt=FMT_DEC4); r += 1
            _kpi(ws1, r, "Overfit ratio",     round(best.get('overfit_ratio', 0), 3)); r += 2

        _section(ws1, r, "▶ STATUT VALIDATION"); r += 1
        statut = val.get('statut_global', 'N/A')
        _kpi(ws1, r, "Statut global",  statut, statut=statut); r += 1
        _kpi(ws1, r, "Conclusion",     val.get('conclusion', ''), wrap=True); r += 1

        # ── Onglet 2 : Classement CANN / TabNet ────────────────────────────────
        ws2 = wb.create_sheet("2-Classement CANN-TabNet")
        _bandeau(ws2, "Classement des modèles Deep Learning", "CANN (Wüthrich) vs TabNet",
                 "A5 — Comparaison", aid, arrete)
        r = 7
        _section(ws2, r, "▶ CLASSEMENT"); r += 1
        for col, txt, w in [(1,"#",6),(2,"Modèle",20),(3,"Type",18),
                             (4,"Gini test",14),(5,"RMSE test",14),
                             (6,"Overfit ratio",16),(7,"GLM gelé",12)]:
            _header(ws2, r, col, txt, w)
        r += 1
        for rank, m in enumerate(classement, 1):
            bg = GRIS_L if rank % 2 == 0 else None
            glm_gele_txt = (
                '✓ Oui' if m.get('glm_gele') else
                ('✗ Non' if 'glm_gele' in m else '—')
            )
            vals = [rank, m.get('modele',''), m.get('type',''),
                    round(m.get('gini_test',0),4), round(m.get('rmse_test',0),4),
                    round(m.get('overfit_ratio',0),3), glm_gele_txt]
            fmts = [None,None,None,FMT_DEC4,FMT_DEC4,FMT_DEC4,None]
            for j, (v, f) in enumerate(zip(vals, fmts), 1):
                _cell(ws2, r, j, v, cf=NOIR, fill=bg, fmt=f,
                      ah="right" if isinstance(v,(int,float)) else "left",
                      bold=(rank==1))
            r += 1

        # ── Onglet 3 : Fidélité CANN Wüthrich ──────────────────────────────────
        # Spécifique à A5 (n'existe dans aucun autre agent) : traçabilité du
        # gel de la couche GLM et de la vérification numérique d'équivalence
        # CANN(époque 0) ≡ GLM Tweedie — le cœur de la promesse Wüthrich
        # (interprétabilité S2 préservée malgré le résiduel non-linéaire).
        ws3 = wb.create_sheet("3-Fidélité CANN Wüthrich")
        _bandeau(ws3, "Fidélité CANN Wüthrich (2019)", "Gel GLM + résiduel init zéro",
                 "A5 — CANN", aid, arrete)
        r = 7
        met_cann = metriques.get('cann', {})
        _section(ws3, r, "▶ GEL DE LA COUCHE GLM (offset non-entraînable)"); r += 1
        glm_gele = met_cann.get('glm_gele')
        _kpi(ws3, r, "GLM gelé", '✓ Oui' if glm_gele else '✗ Non — CANN non conforme Wüthrich',
             statut='VERT' if glm_gele else 'ROUGE'); r += 1
        if 'n_vars_glm_matchees' in met_cann:
            n_match = met_cann.get('n_vars_glm_matchees', 0)
            n_total = met_cann.get('n_vars_glm_total', 0)
            _kpi(ws3, r, "Variables GLM appariées",
                 f"{n_match} / {n_total}",
                 statut='VERT' if n_match == n_total and n_total > 0 else 'AMBRE'); r += 1
        if 'glm_verification_error' in met_cann:
            err = met_cann.get('glm_verification_error')
            _kpi(ws3, r, "Écart vérification CANN(ép.0) vs GLM",
                 f"{err:.2e}" if err is not None else 'N/A',
                 statut='VERT' if (err is not None and err < 1e-3) else 'AMBRE'); r += 1
        r += 1
        _section(ws3, r, "▶ RÉFÉRENCE MÉTHODOLOGIQUE"); r += 1
        # ⚠️ Référence VÉRIFIÉE EN EXTERNE le 27/08/2026 (Cambridge Core, le site
        # officiel d'ASTIN Bulletin — DOI 10.1017/asb.2018.42), PAS par simple
        # appariement à ce que l'agent A5 cite. C'est bien l'article CANN (gel GLM
        # + résiduel neuronal), pas un article de provisionnement chain-ladder.
        _kpi(ws3, r, "Référence",
             "Wüthrich, M.V. & Merz, M. (2019), 'Editorial: Yes, we CANN!', "
             "ASTIN Bulletin 49(1), 1-3, doi:10.1017/asb.2018.42 — principe de "
             "gel GLM adapté à la tarification (offset + résiduel).",
             wrap=True); r += 1

        # ── Onglet 4 : Audit Trail ──────────────────────────────────────────────
        ws4 = wb.create_sheet("4-Audit Trail")
        _bandeau(ws4, "Audit Trail", "Traçabilité ACPR — Agent A5 Deep Learning",
                 "A5 — Audit", aid, arrete)
        r = 7
        _section(ws4, r, "▶ INFORMATIONS AUDIT"); r += 1
        _kpi(ws4, r, "Audit ID",       aid); r += 1
        _kpi(ws4, r, "Date",           now); r += 1
        _kpi(ws4, r, "Agent",          "A5 — Deep Learning (CANN + TabNet)"); r += 1
        _kpi(ws4, r, "Branche",        result_a5.get('branche', 'N/A')); r += 1
        _kpi(ws4, r, "Statut RAG",     result_a5.get('statut_rag', 'N/A'),
             statut=result_a5.get('statut_rag')); r += 1
        _kpi(ws4, r, "Nb modèles",     len(classement), fmt=FMT_NB); r += 1
        if classement:
            _kpi(ws4, r, "Modèle retenu", classement[0].get('modele','')); r += 1
            _kpi(ws4, r, "Gini retenu",   round(classement[0].get('gini_test',0),4),
                 fmt=FMT_DEC4); r += 1

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    except Exception as e:
        logger.error(f"export_excel_a5 échoué : {e}", exc_info=True)
        return b''


# =============================================================================
#  EXPORT A6 — COMPARAISON FINALE
#  5 onglets : Synthèse · Classement final · Backtesting A/E · Fiche décision · Audit
# =============================================================================

def export_excel_a6(result_a6: Dict, audit_id: str = "", arrete: Optional[str] = None) -> bytes:
    """Génère le rapport Excel A6 Comparaison (5 onglets). Retourne bytes ou b''."""
    if not OPENPYXL_OK or not result_a6 or not result_a6.get('success'):
        return b''
    try:
        wb  = Workbook()
        now = datetime.now().strftime("%d/%m/%Y %H:%M")
        aid = audit_id or result_a6.get('audit_id', 'N/A')
        classement   = result_a6.get('classement', [])
        modele_prod  = result_a6.get('modele_production', {})
        backtest     = result_a6.get('backtest', {})
        val_sel      = result_a6.get('validation_selection', {})
        fiche        = result_a6.get('fiche_decision', {})
        ctrl_effet   = result_a6.get('controle_effet', {})
        reserve_arb  = result_a6.get('reserve_arbitrage')
        reserve_vra  = result_a6.get('reserve_vraisemblance')
        reserve_bas  = result_a6.get('reserve_bases_gini')
        arb_contest  = result_a6.get('arbitrage_contestable')
        audit_trail  = result_a6.get('audit_trail', {})

        # ── Onglet 1 : Synthèse ───────────────────────────────────────────────
        ws1 = wb.active
        ws1.title = "1-Synthèse"
        _bandeau(ws1, "Rapport Comparaison Finale", "Sélection modèle de production",
                 "A6 — Comparaison & Validation", aid, arrete)
        r = 7
        _section(ws1, r, "▶ MODÈLE DE PRODUCTION RETENU"); r += 1
        if modele_prod:
            _kpi(ws1, r, "Modèle",         modele_prod.get('modele','')); r += 1
            _kpi(ws1, r, "Famille",         modele_prod.get('famille','')); r += 1
            _kpi(ws1, r, "Score global",    round(modele_prod.get('score_global',0),4),
                 fmt=FMT_DEC4); r += 1
            _kpi(ws1, r, "Gini test",       round(modele_prod.get('gini_test',0),4),
                 fmt=FMT_DEC4); r += 1
            _kpi(ws1, r, "Overfit ratio",   round(modele_prod.get('overfit_ratio',0),3)); r += 1
            _kpi(ws1, r, "Interprétabilité",round(modele_prod.get('interpretabilite',0),2)); r += 2

        _section(ws1, r, "▶ BACKTESTING"); r += 1
        if backtest.get('disponible'):
            # ⚠️ `round(None, 4)` leve. Un A/E non calcule se dit, il ne
            # s'arrondit pas — et il ne peut rendre ni VERT ni ROUGE.
            ae = backtest.get('ae_ratio')
            _kpi(ws1, r, "A/E ratio (N-1→N)",
                 '— non calcule' if ae is None else round(ae, 4),
                 statut='AMBRE' if ae is None else
                        'VERT' if 0.95 <= ae <= 1.05 else
                        'AMBRE' if 0.90 <= ae <= 1.10 else 'ROUGE'); r += 1
            _kpi(ws1, r, "Sinistres de la fenetre de test",
                 backtest.get('n_sinistres_test', '—'), fmt=FMT_NB); r += 1
            _kpi(ws1, r, "Interprétation",   backtest.get('interpretation','')); r += 1
            _kpi(ws1, r, "Stabilité WF",     backtest.get('stabilite_wf','N/A')); r += 1
            _kpi(ws1, r, "Fenêtres testées", backtest.get('n_fenetres', 0), fmt=FMT_NB); r += 2

        _section(ws1, r, "▶ VALIDATION SÉLECTION"); r += 1
        statut = val_sel.get('statut_global', 'N/A')
        _kpi(ws1, r, "Statut global", statut, statut=statut); r += 1

        # ── Onglet 2 : Classement final ───────────────────────────────────────
        ws2 = wb.create_sheet("2-Classement final")
        _bandeau(ws2, "Classement Final", "Tous modèles — Score multicritères",
                 "A6 — Sélection", aid, arrete)
        r = 7
        _section(ws2, r, "▶ CLASSEMENT GLOBAL (GLM · ML · DL)"); r += 1
        for col, txt, w in [(1,"#",6),(2,"Modèle",28),(3,"Famille",12),
                             (4,"Gini test",14),(5,"RMSE test",14),
                             (6,"Overfit",12),(7,"Interprét.",12),(8,"Score final",14)]:
            _header(ws2, r, col, txt, w)
        r += 1
        for rank, m in enumerate(classement, 1):
            bg = GRIS_L if rank % 2 == 0 else None
            prod = rank == 1
            vals = [rank, m.get('modele',''), m.get('famille',''),
                    round(m.get('gini_test',0),4), round(m.get('rmse_test',0),4),
                    round(m.get('overfit_ratio',0),3), round(m.get('interpretabilite',0),2),
                    round(m.get('score_global',0),4)]
            fmts = [None,None,None,FMT_DEC4,FMT_DEC4,FMT_DEC4,FMT_DEC4,FMT_DEC4]
            for j,(v,f) in enumerate(zip(vals,fmts),1):
                _cell(ws2, r, j, v, cf=NOIR, fill=bg, fmt=f, bold=prod,
                      ah="right" if isinstance(v,(int,float)) else "left")
            if prod:
                ws2.cell(row=r, column=2).fill = _fill(OR)
                ws2.cell(row=r, column=2).font = _font(bold=True, color=BLANC, size=10)
            r += 1

        # ── Onglet 3 : Backtesting A/E ────────────────────────────────────────
        ws3 = wb.create_sheet("3-Backtesting AE")
        _bandeau(ws3, "Backtesting Walk-Forward", "Test A/E global et par segment",
                 "A6 — Backtesting", aid, arrete)
        r = 7
        # ── Résumé recalibration ──────────────────────────────────────────────
        _section(ws3, r, "▶ RECALIBRATION DU MODÈLE (backtesting réel)"); r += 1
        # ⚠ AUDIT V10 (BLOQUANT B3) : le statut était calculé par
        #   statut="VERT" if backtest.get('modele_recalibre') else "AMBRE"
        # Or 'modele_recalibre' est une CHAÎNE — y compris quand elle vaut
        # "GLM_POISSON → proxy GBM (...)", qui est truthy. Le livrable Excel
        # estampillait donc « ✓ Conforme » en VERT sur la ligne même qui
        # annonçait une recalibration sur proxy, pendant que le gate RAG, lui,
        # plafonnait à AMBRE. Le rapport client contredisait la certification.
        # On lit désormais le FLAG booléen de fidélité, seule source de vérité.
        # Source UNIQUE de vérité sur la portée de la validation temporelle,
        # partagée par Excel, Word, HTML et le rapport d'équipe (audit V11).
        _avert_wf = avertissement_walk_forward(backtest)
        _wf_fidele = bool(backtest.get('modele_recalibre_fidele', False))
        _kpi(ws3, r, "Modèle recalibré par fenêtre",
             backtest.get('modele_recalibre') or "Non disponible",
             statut=("VERT" if _wf_fidele
                     else ("AMBRE" if backtest.get('modele_recalibre') else "ROUGE"))); r += 1
        if _avert_wf:
            _kpi(ws3, r, "⚠ Portée de la validation temporelle", _avert_wf,
                 statut="AMBRE", wrap=True); r += 1
        # ⚠️⚠️ LE GARDE-FOU N°4 SE DÉCLARE — constat `conformite/C7`.
        # La propriété `controle_effet_execute` portait « À REMONTER DANS
        # LES RAPPORTS » depuis l'audit V14 : mesuré, aucun livrable ne la
        # lisait. Elle n'a pu être publiée qu'une fois rendue VÉRIDIQUE
        # (constat `conformite/C1`) — la publier avant aurait attesté un
        # contrôle qui n'avait examiné aucune colonne.
        # ⚠️⚠️ LA RÉSERVE D'ARBITRAGE SE PUBLIE — un arbitrage entre un seul
        # candidat n'est pas un arbitrage, et l'actuaire doit le lire avant
        # de signer « modèle de production retenu ».
        if reserve_arb:
            _kpi(ws3, r, "⚠ Réserve sur l'arbitrage", reserve_arb,
                 statut="AMBRE", wrap=True); r += 1
        # ⚠️ Une absence de contrôle se PUBLIE : sans cette ligne, l'actuaire
        # lirait un statut sans savoir qu'un garde-fou n'a pas pu s'exercer.
        if reserve_vra:
            _kpi(ws3, r, "⚠ Réserve sur la vraisemblance", reserve_vra,
                 statut="AMBRE", wrap=True); r += 1
        if reserve_bas:
            _kpi(ws3, r, "⚠ Bases de Gini mélangées", reserve_bas,
                 statut="AMBRE", wrap=True); r += 1
        # ⚠️ Le choix automatique se dit contestable — il ne bloque rien.
        if arb_contest:
            _kpi(ws3, r, "⚠ Arbitrage contestable", arb_contest,
                 statut="AMBRE", wrap=True); r += 1
        _avert_ce = avertissement_controle_effet(ctrl_effet)
        if _avert_ce:
            _kpi(ws3, r, "⚠ Contrôle anti-fuite par l'effet", _avert_ce,
                 statut="ROUGE", wrap=True); r += 1
        elif ctrl_effet:
            _kpi(ws3, r, "Contrôle anti-fuite par l'effet",
                 "exécuté sur toutes les cibles", statut="VERT"); r += 1
        if backtest.get('gini_wf_moyen') is not None:
            _kpi(ws3, r, "Gini walk-forward moyen (recalibré)",
                 round(backtest.get('gini_wf_moyen', 0), 4), fmt=FMT_DEC4); r += 1
        # ⚠️ CETTE LIGNE PUBLIAIT UNE CITATION NON SOURCÉE DANS LE LIVRABLE.
        # Elle nommait « Commission Tarification IA France (2019) §3.2.5 » et
        # en donnait une phrase ENTRE GUILLEMETS, puis demandait au lecteur de
        # faire valider le document par ce corps « avant toute diffusion
        # externe (ACPR) ». Rien ne confirme qu'un tel document ni un tel corps
        # existent. La méthode se décrit désormais par ce qu'elle FAIT.
        _kpi(ws3, r, "Méthode de backtesting",
             "Walk-forward avec recalibration : le modèle est réajusté sur "
             "chaque fenêtre d'entraînement puis évalué sur la fenêtre de test "
             "suivante (Gini et A/E hors échantillon, temporellement "
             "cohérents). Choix méthodologique du module — aucune norme "
             "externe n'est invoquée ici.",
             wrap=True); r += 1
        r += 1
        _section(ws3, r, "▶ WALK-FORWARD TEMPOREL"); r += 1
        if backtest.get('walk_forward'):
            for col, txt, w in [(1,"Année test",14),(2,"N train",12),(3,"N test",10),
                                 (4,"Moy train",14),(5,"Moy test",14),
                                 (6,"A/E ratio",14),(7,"Statut",12)]:
                _header(ws3, r, col, txt, w)
            r += 1
            for wf in backtest.get('walk_forward', []):
                st = wf.get('statut','')
                bg = {"VERT":"EAF3DE","AMBRE":"FAEEDA","ROUGE":"FCEBEB"}.get(st, GRIS_L)
                vals = [wf.get('annee_test',''), wf.get('n_train',0), wf.get('n_test',0),
                        wf.get('moy_train',0), wf.get('moy_test',0), wf.get('ae_ratio',0), st]
                fmts = [None,FMT_NB,FMT_NB,FMT_DEC4,FMT_DEC4,FMT_DEC4,None]
                for j,(v,f) in enumerate(zip(vals,fmts),1):
                    _cell(ws3, r, j, v, cf=NOIR, fill=bg, fmt=f,
                          ah="right" if isinstance(v,(int,float)) else "center")
                r += 1
        r += 1
        # A/E par segment
        segs = backtest.get('ae_par_segment', {})
        for seg_name, seg_data in segs.items():
            if isinstance(seg_data, list) and seg_data:
                _section(ws3, r, f"▶ A/E PAR SEGMENT : {seg_name.upper()}"); r += 1
                first = seg_data[0]
                keys = [k for k in first.keys() if k != 'statut']
                for j, k in enumerate(keys, 1):
                    _header(ws3, r, j, k.replace('_',' ').title(), 16)
                _header(ws3, r, len(keys)+1, "Statut", 12)
                r += 1
                for item in seg_data:
                    st = item.get('statut','')
                    bg = {"VERT":"EAF3DE","AMBRE":"FAEEDA","ROUGE":"FCEBEB"}.get(st, GRIS_L)
                    for j, k in enumerate(keys, 1):
                        v = item.get(k, '')
                        _cell(ws3, r, j, v, cf=NOIR, fill=bg,
                              fmt=FMT_DEC4 if isinstance(v, float) else FMT_NB if isinstance(v, int) else None,
                              ah="right" if isinstance(v,(int,float)) else "left")
                    _cell(ws3, r, len(keys)+1, st, cf=BLANC,
                          fill=_statut_fill(st), ah="center")
                    r += 1
                r += 1

        # ── Onglet 4 : Fiche décision ─────────────────────────────────────────
        ws4 = wb.create_sheet("4-Fiche décision")
        _bandeau(ws4, "Fiche de Décision", "À valider par l'actuaire responsable",
                 "A6 — Décision production", aid, arrete)
        r = 7
        _section(ws4, r, "▶ CONTRÔLES SÉLECTION"); r += 1
        for ckey, clabel in [("c1_nb_modeles","C1 — Nb modèles"),
                               ("c2_ecart_gini","C2 — Écart Gini"),
                               ("c3_coherence", "C3 — Cohérence")]:
            c = val_sel.get(ckey, {})
            if c:
                st = c.get('statut','')
                _kpi(ws4, r, clabel, st, statut=st); r += 1
                _kpi(ws4, r, "  → Message", c.get('message',''), wrap=True); r += 1
                _kpi(ws4, r, "  → Conseil", c.get('conseil',''), wrap=True); r += 2

        if fiche:
            _section(ws4, r, "▶ FICHE ACTUARIELLE"); r += 1
            # ⚠️⚠️ CE BLOC N'ÉCRIVAIT QUE LES VALEURS `str` — constat `a6/C11`.
            # Mesuré par exécution : 3 champs publiés sur 12, soit 25 %.
            # Les 9 muets étaient `forces`, `faiblesses`, `risques`,
            # `alternatives`, `questions_actuaire`, `justification_regl` (des
            # LISTES) et les trois numériques. Or c'est exactement ce que la
            # docstring d'A6 annonce comme contenu de l'aide à la décision :
            # « forces et faiblesses · risques identifiés · alternatives ·
            # questions à poser AVANT SIGNATURE ». L'actuaire n'en voyait
            # aucun. Ce n'était pas une décision de mise en page : c'était un
            # `isinstance` qui triait par TYPE, et personne ne le savait.
            #
            # ⚠️⚠️ ET CE CORRECTIF NE PEUT PAS ÊTRE APPLIQUÉ SEUL. En publiant
            # les listes, il publie `justification_regl` — qui portait une
            # affirmation de conformité S2 INCONDITIONNELLE (constat `a6/C5`).
            # Ce garde-fou-là a été posé D'ABORD, dans le même changement :
            # `_generer_fiche_decision` exige désormais `backtest` et
            # `statut_rag`. **Ne jamais revenir sur l'un sans l'autre.**
            for k, v in fiche.items():
                label = k.replace('_', ' ').title()
                if isinstance(v, str):
                    _kpi(ws4, r, label, v, wrap=True); r += 1
                elif isinstance(v, bool):
                    _kpi(ws4, r, label, 'oui' if v else 'non'); r += 1
                elif isinstance(v, (int, float)):
                    _kpi(ws4, r, label, v); r += 1
                elif isinstance(v, (list, tuple)):
                    # Le libellé une fois, puis chaque élément sur sa ligne.
                    # Aucun écrêtement : une liste tronquée en silence
                    # redonnerait une fiche incomplète sans le dire.
                    for i, item in enumerate(v):
                        _kpi(ws4, r, label if i == 0 else "",
                             str(item), wrap=True); r += 1
        # ── Gouvernance du profil de pondération (ACPR-2022-P-01 §4.3) ────────
        _section(ws4, r, "▶ GOUVERNANCE — PROFIL DE PONDÉRATION"); r += 1
        _gov = audit_trail.get('gouvernance_ok')
        _kpi(ws4, r, "Profil retenu",
             audit_trail.get('profil_ponderation', 'N/A')); r += 1
        _kpi(ws4, r, "Environnement", audit_trail.get('environnement', 'N/A')); r += 1
        _kpi(ws4, r, "Validé par",
             audit_trail.get('profil_valide_par') or "⚠ NON VALIDÉ",
             statut="VERT" if audit_trail.get('profil_valide_par') else "AMBRE"); r += 1
        _kpi(ws4, r, "Gouvernance conforme",
             "✓ Oui" if _gov else "✗ Non — validation actuarielle requise",
             statut="VERT" if _gov else "AMBRE"); r += 1
        r += 1

        _section(ws4, r, "▶ DÉCISION FINALE"); r += 1
        _kpi(ws4, r, "Décision", "À VALIDER PAR L'ACTUAIRE RESPONSABLE",
             statut="AMBRE"); r += 1
        _kpi(ws4, r, "Modèle proposé",
             modele_prod.get('modele','') if modele_prod else 'N/A'); r += 1
        _kpi(ws4, r, "Date rapport", now); r += 1
        _kpi(ws4, r, "Signature actuaire", ""); r += 1

        # ── Onglet 5 : Audit Trail ────────────────────────────────────────────
        ws5 = wb.create_sheet("5-Audit Trail")
        _bandeau(ws5, "Audit Trail", "Traçabilité ACPR — Agent A6 Comparaison",
                 "A6 — Audit", aid, arrete)
        r = 7
        _section(ws5, r, "▶ INFORMATIONS AUDIT"); r += 1
        _kpi(ws5, r, "Audit ID",       aid); r += 1
        _kpi(ws5, r, "Date",           now); r += 1
        _kpi(ws5, r, "Agent",          "A6 — Comparaison & Validation Finale"); r += 1
        _kpi(ws5, r, "Branche",        result_a6.get('branche','N/A')); r += 1
        _kpi(ws5, r, "Statut RAG",     result_a6.get('statut_rag','N/A'),
             statut=result_a6.get('statut_rag')); r += 1
        _kpi(ws5, r, "Nb modèles comparés", len(classement), fmt=FMT_NB); r += 1
        _kpi(ws5, r, "Modèle production",
             modele_prod.get('modele','') if modele_prod else 'N/A'); r += 1
        # Audit V4 point #11 — défaut 0 trompeur : un A/E ratio non calculé
        # (backtest indisponible) affichait "0.0000", indiscernable d'un
        # vrai A/E=0 (résultat en soi alarmant en usage réel). "N/A" lève
        # toute ambiguïté d'interprétation pour l'actuaire lecteur.
        # ── EXCLUSIONS DE CONFORMITÉ (audit V11 / I5) ────────────────────────
        # Une exclusion silencieuse est un défaut en soi : c'est ce silence qui a
        # rendu le BLOQUANT B5 si coûteux. Source UNIQUE partagée avec Word/HTML.
        _synth_exc = synthese_exclusions(result_a6.get('exclusions_conformite'))
        if _synth_exc:
            _kpi(ws5, r, "Colonnes écartées de la matrice X", _synth_exc,
                 statut=("AMBRE" if "ACTION REQUISE" in _synth_exc else "VERT"),
                 wrap=True); r += 1
        # ── MODÈLE DEEP LEARNING EN PRODUCTION — validation humaine (2026-07) ─
        # Même exigence de visibilité qu'une exclusion : un DL en production ne
        # doit pas être un fait invisible à l'actuaire qui relit le dossier.
        _synth_dl = synthese_modele_dl(result_a6.get('modele_production'),
                                       result_a6.get('valide_par_actuaire_dl'),
                                       audit_trail.get('timestamp'))
        if _synth_dl:
            _kpi(ws5, r, "Modèle Deep Learning — validation actuarielle", _synth_dl,
                 statut=("AMBRE" if "ACTION REQUISE" in _synth_dl else "VERT"),
                 wrap=True); r += 1
        # ── QUALITÉ DES DONNÉES (couche générique, chemin déclaratif) ─────────
        # Lignes exclues (impossible) / corrigées (implausible) / signalées
        # (ambigu) + escalade validée : jamais un traitement silencieux. None
        # (donc rien affiché) si la couche n'a pas tourné (ex. chemin agent A6).
        _synth_q = synthese_qualite_donnees(result_a6.get('rapport_qualite'))
        if _synth_q:
            _kpi(ws5, r, "Qualité des données — traitements appliqués", _synth_q,
                 statut=("AMBRE" if ("EXCLUE" in _synth_q or "SIGNALEE" in _synth_q
                                     or "BLOQUE" in _synth_q) else "VERT"),
                 wrap=True); r += 1
        # ── COLONNES DU PLAN NON PRODUITES (modèle amputé) ────────────────────
        # Un fichier client incomplet fait tourner les modèles SANS certains
        # facteurs déclarés. Même discipline que ci-dessus : jamais silencieux.
        # None (rien affiché) si le plan a été honoré intégralement.
        _synth_cp = synthese_colonnes_plan_manquantes(
            result_a6.get('colonnes_plan_manquantes'))
        if _synth_cp:
            _kpi(ws5, r, "Colonnes du plan non produites — modèle amputé",
                 _synth_cp, statut="AMBRE", wrap=True); r += 1
        # ⚠️ Constat `conformite/C15` — la colonne EXISTE, un filtre l'a
        # retirée avant la conformité : ni retenue, ni écartée explicitement.
        _synth_ce = synthese_colonnes_plan_ecartees(
            result_a6.get('colonnes_plan_ecartees'))
        if _synth_ce:
            _kpi(ws5, r, "Colonnes du plan écartées avant le filtre",
                 _synth_ce, statut="AMBRE", wrap=True); r += 1

        # Mapping client (couche 2) — renommage du fichier avant A1. Même
        # mécanisme source-unique. Rien affiché si aucun mapping n'a été appliqué.
        _synth_map = synthese_mapping(result_a6.get('rapport_mapping'))
        if _synth_map:
            _kpi(ws5, r, "Mapping client appliqué", _synth_map,
                 statut=("AMBRE" if "⚠" in _synth_map else "VERT"), wrap=True); r += 1
        # Alertes d'expérience passée conservée (audit V13 / B7). Ce n'est pas une
        # exclusion : c'est une VÉRIFICATION demandée à l'actuaire. Ce qui n'est
        # que dans les logs n'existe pas (constat I5).
        _synth_alertes = synthese_alertes_experience(
            result_a6.get('alertes_conformite'))
        if _synth_alertes:
            _kpi(ws5, r, "Sinistralité passée conservée — à vérifier",
                 _synth_alertes, statut="AMBRE", wrap=True); r += 1
        _kpi(ws5, r, "A/E ratio final",
             (round(backtest['ae_ratio'], 4)
              if backtest.get('disponible') and backtest.get('ae_ratio') is not None
              else "N/A"),
             fmt=FMT_DEC4 if backtest.get('disponible') else None); r += 1

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    except Exception as e:
        logger.error(f"export_excel_a6 échoué : {e}", exc_info=True)
        return b''


# =============================================================================
#  EXPORT A1 — INGESTION
#  4 onglets : Synthèse · Qualité & Aberrants · Coercition Types · Audit Trail
# =============================================================================

def export_excel_a1(result_a1: Dict, audit_id: str = "", arrete: Optional[str] = None) -> bytes:
    """Génère le rapport Excel A1 Ingestion (4 onglets). Retourne bytes ou b''."""
    if not OPENPYXL_OK or not result_a1 or not result_a1.get('success'):
        return b''
    try:
        wb  = Workbook()
        now = datetime.now().strftime("%d/%m/%Y %H:%M")
        aid = audit_id or result_a1.get('audit_id', 'N/A')
        qualite = result_a1.get('qualite', {})
        rapport = result_a1.get('rapport', {})
        coercition = rapport.get('coercition_types', {})

        # ── Onglet 1 : Synthèse ───────────────────────────────────────────────
        ws1 = wb.active
        ws1.title = "1-Synthèse"
        _bandeau(ws1, "Rapport Ingestion & Validation", "Synthèse qualité des données",
                 "A1 — Ingestion", aid, arrete)
        r = 7
        _section(ws1, r, "▶ QUALITÉ DES DONNÉES"); r += 1
        _kpi(ws1, r, "Score global qualité", round(qualite.get('score_global', 0), 1),
             statut=result_a1.get('statut_rag'), fmt=FMT_DEC4); r += 1
        _kpi(ws1, r, "Nb lignes", qualite.get('nb_lignes', 0), fmt=FMT_NB); r += 1
        _kpi(ws1, r, "Nb colonnes", qualite.get('nb_colonnes', 0), fmt=FMT_NB); r += 1
        _kpi(ws1, r, "Taux complétude", qualite.get('taux_completude', 0) / 100, fmt=FMT_PCT); r += 1
        _kpi(ws1, r, "Nb doublons", qualite.get('nb_doublons', 0), fmt=FMT_NB); r += 1
        _kpi(ws1, r, "Taux doublons", qualite.get('taux_doublons', 0) / 100, fmt=FMT_PCT); r += 1
        _kpi(ws1, r, "Exposition conforme [0,1]", qualite.get('expo_ok_pct', 0) / 100, fmt=FMT_PCT); r += 1
        _kpi(ws1, r, "Nb types d'anomalies détectées", qualite.get('nb_types_aberrants', 0),
             statut="VERT" if qualite.get('nb_types_aberrants', 0) == 0 else "AMBRE",
             fmt=FMT_NB); r += 1

        # ── Onglet 2 : Qualité & Aberrants ────────────────────────────────────
        # ⚠️ « §4.2 » NE RENVOYAIT À RIEN. Le numéro apparaissait ici, dans le
        # sous-titre du bandeau et dans le libellé du KPI, sous un titre de
        # document que rien ne permet de retrouver — et le même titre portait
        # §3.2.4 dans A3 et §3.2.5 dans A6. Les contrôles restent ce qu'ils
        # sont : des contrôles du module, énoncés comme tels.
        ws2 = wb.create_sheet("2-Aberrants")
        _bandeau(ws2, "Valeurs Aberrantes", "Contrôles actuariels de cohérence",
                 "A1 — Validation Qualité", aid, arrete)
        r = 7
        _kpi(ws2, r, "Portée de ces contrôles",
             "Contrôles de cohérence propres au module (bornes de plausibilité "
             "par variable, types, valeurs extrêmes). Aucune norme externe "
             "n'est invoquée : ils ne remplacent pas la revue de l'actuaire.",
             wrap=True); r += 1
        _section(ws2, r, "▶ ANOMALIES DÉTECTÉES"); r += 1
        aberrants = qualite.get('aberrants', {})
        if aberrants:
            for col, txt, w in [(1, "Type d'anomalie", 30), (2, "Nombre de valeurs", 18)]:
                _header(ws2, r, col, txt, width=w)
            r += 1
            for cle, val in aberrants.items():
                _cell(ws2, r, 1, cle.replace('_', ' ').title(), cf=NOIR, fill=GRIS_L)
                _cell(ws2, r, 2, val, cf=NOIR, ah="right", fmt=FMT_NB)
                r += 1
            r += 1
        else:
            _kpi(ws2, r, "Statut", "✓ Aucune anomalie détectée", statut="VERT"); r += 1
            r += 1

        _section(ws2, r, "▶ ALERTES DÉTAILLÉES"); r += 1
        for alerte in qualite.get('alertes_aberrants', []):
            _cell(ws2, r, 1, f"• {alerte}", cf=NOIR, wrap=True)
            ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
            ws2.row_dimensions[r].height = 30
            r += 1
        if not qualite.get('alertes_aberrants'):
            _kpi(ws2, r, "Alertes", "Aucune", statut="VERT"); r += 1

        # ── Onglet 3 : Coercition de Types ────────────────────────────────────
        ws3 = wb.create_sheet("3-Coercition Types")
        _bandeau(ws3, "Coercition de Types", "Forçage explicite des types de données",
                 "A1 — Coercition", aid, arrete)
        r = 7
        _section(ws3, r, "▶ COLONNES FORCÉES"); r += 1
        cols_forcees = coercition.get('colonnes_forcees', [])
        if cols_forcees:
            _kpi(ws3, r, "Nb colonnes forcées", len(cols_forcees), fmt=FMT_NB); r += 1
            r += 1
            for col, txt, w in [(1, "Colonne", 30), (2, "Valeurs perdues", 18)]:
                _header(ws3, r, col, txt, width=w)
            r += 1
            for col_nom in cols_forcees:
                n_perdu = coercition.get('nb_valeurs_perdues', {}).get(col_nom, 0)
                _cell(ws3, r, 1, col_nom, cf=NOIR, fill=GRIS_L)
                _cell(ws3, r, 2, n_perdu, cf=NOIR, ah="right", fmt=FMT_NB)
                r += 1
        else:
            _kpi(ws3, r, "Statut", "✓ Aucune coercition nécessaire (types déjà corrects)",
                 statut="VERT"); r += 1
        r += 1
        _section(ws3, r, "▶ ALERTES"); r += 1
        for alerte in coercition.get('alertes', []):
            _cell(ws3, r, 1, f"• {alerte}", cf=NOIR, wrap=True)
            ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
            ws3.row_dimensions[r].height = 30
            r += 1
        if not coercition.get('alertes'):
            _kpi(ws3, r, "Alertes", "Aucune", statut="VERT"); r += 1

        # ── Onglet 4 : Audit Trail ─────────────────────────────────────────────
        ws4 = wb.create_sheet("4-Audit Trail")
        _bandeau(ws4, "Audit Trail", "Traçabilité ACPR — Agent A1 Ingestion",
                 "A1 — Audit", aid, arrete)
        r = 7
        _section(ws4, r, "▶ INFORMATIONS AUDIT"); r += 1
        _kpi(ws4, r, "Audit ID", aid); r += 1
        _kpi(ws4, r, "Date", now); r += 1
        _kpi(ws4, r, "Agent", "A1 — Ingestion & Validation"); r += 1
        _kpi(ws4, r, "Branche", result_a1.get('branche', 'N/A')); r += 1
        _kpi(ws4, r, "Statut RAG", result_a1.get('statut_rag', 'N/A'),
             statut=result_a1.get('statut_rag')); r += 1
        _kpi(ws4, r, "Hash MD5", result_a1.get('hash_md5', 'N/A')); r += 1

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    except Exception as e:
        logger.error(f"export_excel_a1 échoué : {e}", exc_info=True)
        return b''


# =============================================================================
#  EXPORT A2 — PREPROCESSING
#  3 onglets : Synthèse · Data Dictionnaire · Audit Trail
# =============================================================================

def export_excel_a2(result_a2: Dict, audit_id: str = "", arrete: Optional[str] = None) -> bytes:
    """Génère le rapport Excel A2 Preprocessing (3 onglets). Retourne bytes ou b''."""
    if not OPENPYXL_OK or not result_a2 or not result_a2.get('success'):
        return b''
    try:
        wb  = Workbook()
        now = datetime.now().strftime("%d/%m/%Y %H:%M")
        aid = audit_id or result_a2.get('audit_id', 'N/A')
        rapport = result_a2.get('rapport', {})
        data_dict = result_a2.get('data_dictionnaire', {})

        # ── Onglet 1 : Synthèse ───────────────────────────────────────────────
        ws1 = wb.active
        ws1.title = "1-Synthèse"
        _bandeau(ws1, "Rapport Preprocessing", "Synthèse des transformations",
                 "A2 — Preprocessing", aid, arrete)
        r = 7
        _section(ws1, r, "▶ TRANSFORMATIONS APPLIQUÉES"); r += 1
        _kpi(ws1, r, "Statut RAG", result_a2.get('statut_rag', 'N/A'),
             statut=result_a2.get('statut_rag')); r += 1
        _kpi(ws1, r, "Branche", result_a2.get('branche', 'N/A')); r += 1
        _kpi(ws1, r, "Étapes exécutées",
             ', '.join(rapport.get('etapes', [])), wrap=True); r += 1
        if rapport.get('alertes'):
            r += 1
            _section(ws1, r, "▶ ALERTES"); r += 1
            for alerte in rapport.get('alertes', []):
                _cell(ws1, r, 1, f"• {alerte}", cf=NOIR, wrap=True)
                ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
                ws1.row_dimensions[r].height = 30
                r += 1

        # ── Onglet 2 : Data Dictionnaire ───────────────────────────────────────
        # Réf. : ACPR-2022-P-01 §3.2 — traçabilité des variables dérivées
        # ⚠ Référence non vérifiée formellement (audit V4) — voir caveat KPI
        # inséré en corps de feuille ci-dessous.
        ws2 = wb.create_sheet("2-Data Dictionnaire")
        _bandeau(ws2, "Data Dictionnaire", "Traçabilité des variables dérivées — ACPR-2022-P-01 §3.2",
                 "A2 — Traçabilité", aid, arrete)
        r = 7
        _kpi(ws2, r, "Portée de cette traçabilité",
             "Chaque variable dérivée est documentée par sa source, son "
             "opération et son usage — traçabilité exigible en revue ACPR. "
             "Le détail des opérations relève du module, pas d'une norme "
             "externe.",
             wrap=True); r += 1
        _section(ws2, r, "▶ VARIABLES DÉRIVÉES DOCUMENTÉES"); r += 1
        headers = ['Variable', 'Source', 'Opération', 'Usage', 'Justification']
        widths  = [26, 26, 34, 22, 50]
        for ci, (h, w) in enumerate(zip(headers, widths), 1):
            _header(ws2, r, ci, h, width=w)
        r += 1
        for var_nom, meta in data_dict.items():
            source = meta.get('source', '')
            source_txt = ', '.join(source) if isinstance(source, list) else str(source)
            _cell(ws2, r, 1, var_nom, cf=NOIR, fill=GRIS_L, bold=True)
            _cell(ws2, r, 2, source_txt, cf=NOIR, wrap=True)
            _cell(ws2, r, 3, meta.get('operation', ''), cf=NOIR, wrap=True)
            _cell(ws2, r, 4, meta.get('usage', ''), cf=NOIR)
            _cell(ws2, r, 5, meta.get('justification', ''), cf=NOIR, wrap=True)
            ws2.row_dimensions[r].height = 45
            r += 1
        if not data_dict:
            _kpi(ws2, r, "Statut", "Aucune variable dérivée documentée", statut="AMBRE"); r += 1

        # ── Onglet 3 : Audit Trail ─────────────────────────────────────────────
        ws3 = wb.create_sheet("3-Audit Trail")
        _bandeau(ws3, "Audit Trail", "Traçabilité ACPR — Agent A2 Preprocessing",
                 "A2 — Audit", aid, arrete)
        r = 7
        _section(ws3, r, "▶ INFORMATIONS AUDIT"); r += 1
        _kpi(ws3, r, "Audit ID", aid); r += 1
        _kpi(ws3, r, "Date", now); r += 1
        _kpi(ws3, r, "Agent", "A2 — Preprocessing & Feature Engineering"); r += 1
        _kpi(ws3, r, "Branche", result_a2.get('branche', 'N/A')); r += 1
        _kpi(ws3, r, "Statut RAG", result_a2.get('statut_rag', 'N/A'),
             statut=result_a2.get('statut_rag')); r += 1
        _kpi(ws3, r, "Nb variables dérivées tracées", len(data_dict), fmt=FMT_NB); r += 1

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    except Exception as e:
        logger.error(f"export_excel_a2 échoué : {e}", exc_info=True)
        return b''
