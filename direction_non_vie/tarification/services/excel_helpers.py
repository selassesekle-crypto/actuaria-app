"""
ActuarIA — Helpers Excel partagés (openpyxl)
Direction Non-Vie · Équipe Tarification

Factorisation (audit V4 point #12) : ces constantes et fonctions étaient
dupliquées à l'identique entre `tarif_excel.py` (fonctions module-level)
et `rapport_equipe_tarif.py` (mêmes fonctions redéfinies en closures
locales dans `export_excel_equipe`). Risque de dérive silencieuse entre
les deux copies (palette, format, épaisseur de bordure...) si l'une est
modifiée sans répercuter sur l'autre — désormais impossible par
construction : une seule source de vérité.

Utilisé par : tarif_excel.py, rapport_equipe_tarif.py
"""

import logging
from typing import Optional

logger = logging.getLogger('actuaria.tarif.excel_helpers')

try:
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False
    logger.warning("openpyxl non disponible — helpers Excel désactivés")

# ── La date d'arrêté et la date de génération : SOURCE UNIQUE ────────────────
# Le bandeau ne calcule aucune date lui-même — il dérive tout d'`entete_livrable`
# (lui-même sur `core/arrete.py`). C'est ce qui empêche « Arrêté : » de porter la
# date du jour (constat services/C1).
try:
    from .entete_livrable import libelle_arrete, genere_le
except ImportError:
    from direction_non_vie.tarification.services.entete_livrable import (
        libelle_arrete, genere_le)

# ── Palette ActuarIA (hexadécimal sans '#', format attendu par openpyxl) ─────
NAVY   = "0F2E52"
OR     = "C9A84C"
BLANC  = "F0F4F8"
GRIS   = "8A9AB0"
VERT_H = "2ECC71"
AMBRE  = "F39C12"
ROUGE  = "E74C3C"
NAVY_L = "1B3A5C"
GRIS_L = "EAF0F6"
NOIR   = "1A1A1A"

# Alias '_HEX' — conservés pour compatibilité avec le nommage historique
# de rapport_equipe_tarif.py (même valeurs, noms différents).
NAVY_HEX   = NAVY
GOLD_HEX   = OR
VERT_HEX   = VERT_H
AMBRE_HEX  = AMBRE
ROUGE_HEX  = ROUGE
GRIS_HEX   = GRIS
GRIS_L_HEX = GRIS_L
NOIR_HEX   = NOIR
BLANC_HEX  = BLANC

FMT_EUR  = '# ##0 €;-# ##0 €'
FMT_PCT  = '0.00%'
FMT_DEC4 = '0.0000'
FMT_NB   = '# ##0'


# ── Helpers ───────────────────────────────────────────────────────────────────

def _font(bold=False, color=BLANC, size=10, italic=False):
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _statut_fill(statut: Optional[str]) -> str:
    return {"VERT": VERT_H, "AMBRE": AMBRE, "ROUGE": ROUGE}.get(
        (statut or '').upper(), GRIS)

def _col_w(ws, col: int, width: float):
    ws.column_dimensions[get_column_letter(col)].width = width

def _cell(ws, row, col, val, bold=False, cf=BLANC, fill=None,
          ah="left", fmt=None, border=True, wrap=False):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = _font(bold=bold, color=cf)
    c.alignment = _align(h=ah, wrap=wrap)
    if fill:
        c.fill = _fill(fill)
    if border:
        c.border = _border()
    if fmt:
        c.number_format = fmt
    return c

def _header(ws, row, col, text, width=18):
    _cell(ws, row, col, text, bold=True, cf=BLANC, fill=NAVY, ah="center")
    _col_w(ws, col, width)

def _section(ws, row, titre, n_cols=8):
    c = ws.cell(row=row, column=1, value=f"  {titre}")
    c.font      = _font(bold=True, color=OR, size=11)
    c.fill      = _fill(NAVY_L)
    c.alignment = _align(h="left")
    c.border    = _border()
    if n_cols > 1:
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=n_cols)
    ws.row_dimensions[row].height = 20

def _kpi(ws, row, label, value, statut=None, fmt=None, wrap=False):
    _cell(ws, row, 1, label, bold=True, cf=NOIR, fill=GRIS_L, wrap=wrap)
    _col_w(ws, 1, 38)
    _cell(ws, row, 2, value, bold=True, cf=NOIR, fill=None, ah="left" if wrap else "right",
          fmt=fmt, wrap=wrap)
    _col_w(ws, 2, 50 if wrap else 22)
    if statut:
        txt = {"VERT": "✓ Conforme", "AMBRE": "△ À surveiller", "ROUGE": "✗ Attention"}.get(statut, statut)
        _cell(ws, row, 3, txt, bold=True, cf=BLANC,
              fill=_statut_fill(statut), ah="center")
        _col_w(ws, 3, 18)

def _bandeau(ws, titre, sous_titre, agent, audit_id, arrete=None, n_cols=8):
    # ⚠️ DEUX DATES, JAMAIS CONFONDUES (constat services/C1) :
    #   · « Arrêté » = date de RÉFÉRENCE déclarée ; absente → « non déclaré »,
    #     VISIBLE, jamais la date du jour glissée sous cette étiquette ;
    #   · « Généré le » = date d'IMPRESSION, honnête, sur TOUT document.
    # Les deux dérivent de la source unique `entete_livrable` (sur `core/arrete`).
    for _r, _h in ((1, 32), (2, 22), (3, 18), (4, 16), (5, 14), (6, 14)):
        ws.row_dimensions[_r].height = _h
    for r, (txt, sz, bold) in enumerate([
        (f"ActuarIA — {titre}",                14, True),
        (sous_titre,                            11, False),
        (f"Agent : {agent}",                    10, False),
        (f"Arrêté : {libelle_arrete(arrete)}",   9, False),
        (f"Généré le : {genere_le()}",           9, False),
        (f"Audit ID : {audit_id}",               9, False),
    ], 1):
        c = ws.cell(row=r, column=1, value=txt)
        c.font      = _font(bold=bold, color=OR if bold else BLANC, size=sz)
        c.fill      = _fill(NAVY)
        c.alignment = _align(h="left")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n_cols)
