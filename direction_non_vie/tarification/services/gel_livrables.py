# -*- coding: utf-8 -*-
"""
ActuarIA — Tarification · LE GEL DES LIVRABLES : comparer deux productions
=========================================================================

À quoi il sert : rendre MESURABLE la phrase « aucun euro n'a bougé ». Un
correctif se prouve en produisant les livrables signés avant et après, puis en
comparant leur CONTENU. Sans cet outil la preuve se refait à la main à chaque
lot, et c'est exactement ce qui produit des constats fermés « sur leur
assiette » — vrais là où on a regardé, muets ailleurs.

⚠️⚠️ IL NE PEUT PAS SE FAIRE PAR EMPREINTE D'OCTETS. **Mesuré** le 05/09/2026 :
deux exécutions identiques rendent un `.docx` de contenu identique mais
d'octets différents — le ZIP horodate chacune de ses entrées. La taille elle
aussi bouge (41 588 puis 41 589 octets) parce que `deflate` comprime
différemment un horodatage qui a changé de minute. *Comparer des octets, ou
des tailles, c'est se condamner à un rouge par minute.*

⚠️⚠️ ET IL NE PEUT PAS NEUTRALISER TOUTE DATE. `entete_livrable.genere_le()`
rend « 05/09/2026 01:46 » — une métadonnée d'impression. Mais
`entete_livrable.libelle_arrete()` rend « 30/06/2026 » — la date d'ARRÊTÉ,
c'est-à-dire du contenu signé, de forme identique. Effacer l'une efface
l'autre, et le gel devient aveugle au champ le plus lourd du document. La
distinction retenue n'est ni l'étiquette ni la position, elle est
STRUCTURELLE :

    une date SUIVIE D'UNE HEURE est une impression  -> neutralisée
    une date SEULE est du métier                    -> conservée, comparée

Les tests de `test_gel_livrables.py` ne recopient pas ces formats : ils
appellent les VRAIS producteurs (`genere_le`, `libelle_arrete`,
`rapport_modeles_tarif.valeur_audit`) et vérifient le résultat. Une liste
recopiée se met à jour à moitié ; un contrôle qui appelle la source suit.

⚠️ CE QUI EST COMPARÉ SE DÉCLARE. `empreinte()` rend, à côté du contenu, la
liste des surfaces qu'il n'a PAS su lire, avec la raison. Un « 0 écart » qui
tairait une surface illisible serait le pire des trois défauts possibles ici.
"""
from __future__ import annotations

import io
import re
import zipfile
from collections.abc import Mapping
from dataclasses import dataclass, field
from typing import Any

#: Ce qu'on inscrit quand une surface existe d'un côté et pas de l'autre.
#: ⚠️ Une ABSENCE est une valeur, jamais un silence : deux `None` fidèles l'un
#: à l'autre restent une absence, et un contrôle qui les compare reste vert
#: en n'ayant rien surveillé.
ABSENT = '<livrable absent>'

#: Les motifs d'IMPRESSION, dans l'ordre — le plus spécifique d'abord.
#: ⚠️ L'ordre porte le sens : « 05/09/2026 à 01 h 46 » doit être pris ENTIER
#: par le motif 2, sinon le motif 5 n'en mangerait que l'heure et laisserait
#: une date d'impression se faire passer pour une date d'arrêté.
MOTIFS_IMPRESSION: tuple[tuple[re.Pattern, str], ...] = (
    # identifiants d'audit et noms de modèles : A3_20260905_014612
    # ⚠️⚠️ PAS `\b\d{8}_\d{6}\b`. **Mesuré** : `\b` ne s'ouvre pas entre `A3_`
    # et `20260905` — le tiret bas et le chiffre sont tous deux des caractères
    # de mot. Le motif mordait `20260905_015505` seul et laissait passer
    # `A3_20260905_015505`, la forme réelle des identifiants d'audit. C'est la
    # sentinelle GEL-11 qui l'a trouvé, pas la relecture.
    (re.compile(r'(?<!\d)\d{8}_\d{6}(?!\d)'), '<horodatage>'),
    # rendu français de la piste d'audit : 05/09/2026 à 01 h 46
    (re.compile(r'\b\d{2}/\d{2}/\d{4}\s+à\s+\d{1,2}\s*h\s*\d{2}(\s*m?i?n?)?'),
     '<horodatage>'),
    # « Généré le : 05/09/2026 01:46 »
    (re.compile(r'\b\d{2}/\d{2}/\d{4}[\sT]+\d{1,2}:\d{2}(:\d{2})?'), '<horodatage>'),
    # ISO : 2026-09-05T01:46:12 / 2026-09-05 01:46
    (re.compile(r'\b\d{4}-\d{2}-\d{2}[\sT]+\d{1,2}:\d{2}(:\d{2})?(\.\d+)?'),
     '<horodatage>'),
    # une heure seule, dans une cellule qui ne porte que ça
    (re.compile(r'(?<![\d.,])\d{1,2}:\d{2}(:\d{2})?(?![\d.,])'), '<heure>'),
)


def neutraliser(texte: str) -> str:
    """Efface les horodatages d'IMPRESSION, laisse les dates MÉTIER.

    ⚠️ « 30/06/2026 » ressort tel quel : c'est une date d'arrêté. Si elle
    change, le gel doit rougir.
    """
    for motif, remplacement in MOTIFS_IMPRESSION:
        texte = motif.sub(remplacement, texte)
    return texte


# =============================================================================
#  LIRE UN LIVRABLE — par son CONTENU, jamais par le nom de sa clé
# =============================================================================
#
# ⚠️ Le format se DÉRIVE des octets. Une clé nommée `word_bytes` qui porterait
# un HTML serait lue comme un HTML — et l'anomalie deviendrait visible, au lieu
# d'être maquillée par une étiquette qui rassure.

def format_livrable(octets: bytes | None) -> str:
    """'xlsx' | 'docx' | 'pdf' | 'html' | 'texte' | 'vide'."""
    if not octets:
        return 'vide'
    if octets[:4] == b'PK\x03\x04':
        try:
            with zipfile.ZipFile(io.BytesIO(octets)) as z:
                noms = set(z.namelist())
        except zipfile.BadZipFile:
            return 'texte'
        if 'xl/workbook.xml' in noms:
            return 'xlsx'
        if 'word/document.xml' in noms:
            return 'docx'
        return 'texte'
    if octets[:5] == b'%PDF-':
        return 'pdf'
    tete = octets[:400].lstrip().lower()
    if tete.startswith((b'<!doctype html', b'<html')):
        return 'html'
    return 'texte'


def contenu_xlsx(octets: bytes) -> dict:
    """Toutes les cellules non vides, feuille par feuille, coordonnée en clé.

    ⚠️ La coordonnée VOYAGE AVEC LA VALEUR. Comparer deux listes par leur
    position dirait « la ligne 12 a changé » le jour où une ligne est insérée
    en 3 — un écart de contenu et un décalage se ressemblent alors trait pour
    trait.
    """
    from openpyxl import load_workbook
    classeur = load_workbook(io.BytesIO(octets), data_only=True)
    feuilles: dict[str, dict] = {}
    for feuille in classeur.worksheets:
        cellules: dict[str, Any] = {}
        for ligne in feuille.iter_rows():
            for cellule in ligne:
                valeur = cellule.value
                if valeur is None:
                    continue
                cellules[cellule.coordinate] = _valeur_cellule(valeur)
        feuilles[feuille.title] = cellules
    classeur.close()
    return feuilles


def _valeur_cellule(valeur: Any) -> Any:
    """Une valeur de cellule, comparable d'un run à l'autre.

    ⚠️ Une vraie date Excel SANS heure reste une date (métier). Avec une
    heure, c'est une impression.
    """
    if hasattr(valeur, 'isoformat'):
        heure = getattr(valeur, 'hour', 0) or 0
        minute = getattr(valeur, 'minute', 0) or 0
        seconde = getattr(valeur, 'second', 0) or 0
        if heure or minute or seconde:
            return '<horodatage>'
        return valeur.isoformat()[:10]
    if isinstance(valeur, str):
        return neutraliser(valeur)
    if isinstance(valeur, float):
        # ⚠️ Le dernier bit d'un flottant n'est pas du contenu signé : deux
        # sommations dans un ordre différent le font bouger sans qu'aucun
        # euro ne change. Neuf décimales gardent le centime très large.
        return round(valeur, 9)
    return valeur


#: Les parties d'un .docx qui portent du texte lu par un humain.
#: ⚠️ Pas seulement `document.xml` : un en-tête ou un pied de page porte la
#: date d'arrêté et le nom du signataire.
_PARTIES_DOCX = re.compile(r'^word/(document|header\d*|footer\d*|footnotes|'
                           r'endnotes)\.xml$')

#: ⚠️⚠️ `<w:t[^>]*>` MORD AUSSI `<w:tcPr>`, `<w:tcW>`, `<w:tbl>`. Mesuré le
#: 05/09/2026 : l'extraction remontait des largeurs de colonne et des couleurs
#: de fond comme si c'était du contenu signé — un changement de style aurait
#: été annoncé comme un changement de tarif. La balise se ferme donc sur un
#: espace ou sur `>`, jamais sur n'importe quelle lettre.
_TEXTE_W = re.compile(r'<w:t(?:\s[^>]*)?>(.*?)</w:t>', re.DOTALL)


def contenu_docx(octets: bytes) -> dict:
    """Le texte du document, partie par partie, dans l'ordre de lecture."""
    parties: dict[str, list[str]] = {}
    with zipfile.ZipFile(io.BytesIO(octets)) as z:
        for nom in sorted(z.namelist()):
            if not _PARTIES_DOCX.match(nom):
                continue
            xml = z.read(nom).decode('utf-8', 'replace')
            parties[nom] = [neutraliser(_desechapper(m))
                            for m in _TEXTE_W.findall(xml)]
    return parties


def _desechapper(texte: str) -> str:
    for avant, apres in (('&lt;', '<'), ('&gt;', '>'), ('&quot;', '"'),
                         ('&apos;', "'"), ('&amp;', '&')):
        texte = texte.replace(avant, apres)
    return texte


_BALISE = re.compile(r'<(script|style)\b.*?</\1>|<[^>]+>',
                     re.DOTALL | re.IGNORECASE)


def contenu_html(octets: bytes) -> list[str]:
    """Le texte visible d'une page, ligne à ligne, balises retirées."""
    texte = octets.decode('utf-8', 'replace')
    texte = _BALISE.sub('\n', texte)
    texte = _desechapper(texte).replace('&nbsp;', ' ')
    return [neutraliser(ligne.strip()) for ligne in texte.split('\n')
            if ligne.strip()]


def contenu_pdf(octets: bytes) -> list[str]:
    """Le texte de chaque page."""
    from pypdf import PdfReader
    lecteur = PdfReader(io.BytesIO(octets))
    return [neutraliser((page.extract_text() or '').strip())
            for page in lecteur.pages]


_LECTEURS = {'xlsx': contenu_xlsx, 'docx': contenu_docx,
             'html': contenu_html, 'pdf': contenu_pdf}


def lire_livrable(octets: bytes | None) -> tuple[Any, str | None]:
    """(contenu comparable, raison de non-comparaison).

    Exactement l'un des deux est non nul. ⚠️ Une surface illisible ne
    disparaît pas de la mesure : elle en ressort NOMMÉE.
    """
    forme = format_livrable(octets)
    if forme == 'vide':
        return ABSENT, None
    lecteur = _LECTEURS.get(forme)
    if lecteur is None:
        return None, f'format non lu ({forme})'
    try:
        return lecteur(octets), None
    except Exception as erreur:                                    # noqa: BLE001
        return None, f'lecture {forme} impossible : {type(erreur).__name__}'


# =============================================================================
#  L'EMPREINTE D'UN RÉSULTAT D'AGENT
# =============================================================================

#: Les clés d'octets d'un résultat d'agent. ⚠️ Cette table n'est pas la
#: référence : `livrables_d_un_resultat` prend TOUTE clé finissant par
#: `_bytes`, connue ou non. La table ne sert qu'à donner un nom lisible.
NOMS_LISIBLES = {
    'excel_bytes': 'Excel',
    'word_bytes': 'Word',
    'html_bytes': 'HTML',
    'pdf_bytes': 'PDF',
}


def livrables_d_un_resultat(resultat: Mapping[str, Any],
                            prefixe: str = '') -> dict[str, bytes | None]:
    """Les livrables portés par un résultat d'agent, nom lisible en clé.

    ⚠️⚠️ ELLE ÉNUMÈRE, ELLE NE DÉCLARE PAS. Toute clé finissant par `_bytes`
    entre dans la mesure, y compris une clé ajoutée demain par un export
    nouveau. Une table tenue à la main aurait divergé le jour de cet ajout —
    et l'assiette se serait rétrécie sans que personne le voie.

    `rapport_equipe` porte un dictionnaire de quatre formats : il est ouvert
    d'un cran, sous le même principe.
    """
    livrables: dict[str, bytes | None] = {}
    for cle, valeur in resultat.items():
        if not isinstance(cle, str):
            continue
        if cle.endswith('_bytes'):
            nom = NOMS_LISIBLES.get(cle, cle[:-len('_bytes')])
            livrables[f'{prefixe}{nom}'] = valeur if isinstance(valeur, bytes) else None
        elif cle == 'rapport_equipe' and isinstance(valeur, Mapping):
            livrables.update(livrables_d_un_resultat(
                valeur, prefixe=f'{prefixe}Rapport equipe '))
    return livrables


@dataclass
class Empreinte:
    """Le contenu comparable d'un jeu de livrables, et ce qui ne l'est pas."""

    contenus: dict[str, Any] = field(default_factory=dict)
    #: surface -> raison. ⚠️ À LIRE AVEC LE VERDICT : un « 0 écart » sur une
    #: empreinte qui n'a pas su lire trois surfaces ne dit rien de ces trois.
    non_lues: dict[str, str] = field(default_factory=dict)

    def surfaces(self) -> list[str]:
        return sorted(set(self.contenus) | set(self.non_lues))


def empreinte(livrables: Mapping[str, bytes | None]) -> Empreinte:
    """L'empreinte normalisée d'un ensemble de livrables."""
    emp = Empreinte()
    for nom in sorted(livrables):
        contenu, raison = lire_livrable(livrables[nom])
        if raison is None:
            emp.contenus[nom] = contenu
        else:
            emp.non_lues[nom] = raison
    return emp


# =============================================================================
#  COMPARER
# =============================================================================

@dataclass(frozen=True)
class Ecart:
    """Un écart de contenu, désigné par son chemin dans le livrable."""

    surface: str
    emplacement: str
    avant: Any
    apres: Any

    def __str__(self) -> str:
        lieu = f'{self.surface}{self.emplacement}'
        return f'{lieu}\n    avant : {self.avant!r}\n    apres : {self.apres!r}'


def _aplatir(objet: Any, chemin: str = '') -> dict[str, Any]:
    if isinstance(objet, Mapping):
        plat: dict[str, Any] = {}
        for cle, valeur in objet.items():
            plat.update(_aplatir(valeur, f'{chemin}[{cle}]'))
        return plat or {chemin: '<vide>'}
    if isinstance(objet, (list, tuple)):
        plat = {}
        for indice, valeur in enumerate(objet):
            plat.update(_aplatir(valeur, f'{chemin}[{indice}]'))
        return plat or {chemin: '<vide>'}
    return {chemin: objet}


def comparer(avant: Empreinte, apres: Empreinte) -> list[Ecart]:
    """Les écarts de CONTENU entre deux empreintes.

    ⚠️ Une surface présente d'un seul côté est un écart, pas une omission.
    ⚠️ Une surface devenue illisible est un écart, pas un silence : sinon un
    export cassé passerait pour « rien n'a changé ».
    """
    ecarts: list[Ecart] = []
    for surface in sorted(set(avant.surfaces()) | set(apres.surfaces())):
        raison_av = avant.non_lues.get(surface)
        raison_ap = apres.non_lues.get(surface)
        if raison_av or raison_ap:
            if raison_av != raison_ap:
                ecarts.append(Ecart(surface, '', raison_av or 'lisible',
                                    raison_ap or 'lisible'))
            continue
        if surface not in avant.contenus or surface not in apres.contenus:
            ecarts.append(Ecart(
                surface, '',
                avant.contenus.get(surface, ABSENT) if surface in avant.contenus
                else ABSENT,
                apres.contenus.get(surface, ABSENT) if surface in apres.contenus
                else ABSENT))
            continue
        plat_av = _aplatir(avant.contenus[surface])
        plat_ap = _aplatir(apres.contenus[surface])
        for chemin in sorted(set(plat_av) | set(plat_ap)):
            val_av = plat_av.get(chemin, ABSENT)
            val_ap = plat_ap.get(chemin, ABSENT)
            if val_av != val_ap:
                ecarts.append(Ecart(surface, chemin, val_av, val_ap))
    return ecarts


def rapport_ecarts(ecarts: list[Ecart], avant: Empreinte, apres: Empreinte,
                   maximum: int = 40) -> str:
    """Le verdict, avec son assiette — jamais un chiffre seul."""
    lignes = [f'{len(ecarts)} ecart(s) de contenu']
    surfaces = sorted(set(avant.surfaces()) | set(apres.surfaces()))
    lignes.append(f'assiette : {len(surfaces)} surface(s) '
                  f'-- {", ".join(surfaces) or "aucune"}')
    non_lues = {**avant.non_lues, **apres.non_lues}
    if non_lues:
        lignes.append(f'NON LUES ({len(non_lues)}) : ' + '; '.join(
            f'{nom} -> {raison}' for nom, raison in sorted(non_lues.items())))
    for ecart in ecarts[:maximum]:
        lignes.append(f'  {ecart}')
    if len(ecarts) > maximum:
        lignes.append(f'  ... et {len(ecarts) - maximum} autre(s)')
    return '\n'.join(lignes)
