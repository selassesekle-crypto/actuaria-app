# -*- coding: utf-8 -*-
"""A3 CODAIT LA FAMILLE GAMMA EN DUR, LE PLAN LA DECLARAIT.

Le plan porte `famille_severite` depuis longtemps, et le chemin DECLARATIF
sait ajuster les trois -- gamma, inverse-gaussienne, lognormale (moindres
carres sur log(cout) avec correction de Duan, 1983). A3, lui, ecrivait
`families.Gamma(link=Log())` a DEUX sites. C'est la seule capacite que le
chemin a supprimer possedait et que le moteur retenu n'avait pas : fusionner
sans elle aurait ete une REGRESSION.

MESURE DU 05/09/2026, A3 sur `_portefeuille_auto(1500)` :

  | famille declaree    | statut | Gini    | AIC      | deviance |
  |---------------------|--------|---------|----------|----------|
  | gamma               | ROUGE  | -0,0399 | 7 567,66 |  244,23  |
  | inverse_gaussienne  | ROUGE  | -0,0381 | 7 869,54 |    0,21  |
  | lognormal           | VERT   |  0,0298 | 1 098,34 |   None   |

⚠️⚠️ ET LA LOGNORMALE FAISAIT ECHOUER LE RUN ENTIER. Un OLS n'a ni deviance
ni deviance nulle : `round(float(modele.deviance), 4)` levait. *Une grandeur
qui n'a pas de sens pour ce modele n'a pas de valeur -- et surtout pas zero,
qui se lirait comme un ajustement parfait.*

⚠️ LES PRIMITIVES ONT DEMENAGE DANS `core/severite.py`. Elles vivaient dans
`pipeline_tarifaire` -- une direction -- et A3 ne pouvait pas les atteindre.
*Deux chemins qui ajustent la meme grandeur avec deux codes finissent par
diverger* : c'est exactement ce que ce module documente en tete.

Ce que cette sentinelle exige :
  FS-1  A3 ajuste REELLEMENT la famille declaree, les trois ;
  FS-2  trois familles -> trois modeles distincts (elle n'est pas ignoree) ;
  FS-3  aucune famille de severite codee en dur dans A3 ;
  FS-4  la deviance absente vaut `None`, jamais 0, et le classeur le DIT ;
  FS-5  la famille du run est REINITIALISEE a chaque appel ;
  FS-6  les primitives sont PARTAGEES entre les deux moteurs, pas dupliquees ;
  FS-7  les 20 plans declarent `gamma` -- la production ne bouge pas, et
        c'est MESURE.

Tout en `unittest.TestCase` : la gate lance `unittest discover`.
"""
import ast
import dataclasses
import glob
import io
import logging
import os
import pathlib
import re
import sys
import unittest
import warnings

_ICI = os.path.dirname(os.path.abspath(__file__))
_RACINE = os.path.dirname(os.path.dirname(_ICI))
for _c in (_RACINE, _ICI):
    if _c not in sys.path:
        sys.path.insert(0, _c)

import numpy as np

from core.plan_tarifaire import PlanTarifaire
from core.severite import ModeleCout, ajuster_glm_cout

_FAMILLES = ('gamma', 'inverse_gaussienne', 'lognormal')


class TestPrimitivesPartagees(unittest.TestCase):

    def test_FS6_les_deux_moteurs_lisent_LA_MEME_primitive(self):
        """⚠️⚠️ Elles vivaient dans `pipeline_tarifaire`, hors d'atteinte
        d'A3 -- qui codait donc la Gamma en dur. Deux codes pour la meme
        grandeur, c'est une divergence en attente."""
        from direction_non_vie.tarification import pipeline_tarifaire as PT
        self.assertIs(PT.ajuster_glm_cout, ajuster_glm_cout,
                      'le pipeline a sa PROPRE fonction d ajustement')
        self.assertIs(PT.ModeleCout, ModeleCout)

    def test_FS3_aucune_famille_codee_en_dur_dans_A3(self):
        """⚠️ Au TEXTE, hors commentaires et docstrings : c'est une LECTURE
        qu'on traque, et un releve par AST ne distingue pas une mention
        d'un usage."""
        chemin = os.path.join(_ICI, 'a3_glm', 'agent.py')
        source = pathlib.Path(chemin).read_text(encoding='utf-8')
        arbre = ast.parse(source)
        docs = set()
        for n in ast.walk(arbre):
            if (isinstance(n, ast.Expr) and isinstance(n.value, ast.Constant)
                    and isinstance(n.value.value, str)):
                docs.update(range(n.lineno, (n.end_lineno or n.lineno) + 1))
        motif = re.compile(r'families\.(Gamma|InverseGaussian)\b')
        fautes = [f'a3:{i} {l.strip()[:70]}'
                  for i, l in enumerate(source.splitlines(), 1)
                  if i not in docs and not l.strip().startswith('#')
                  and motif.search(l)]
        self.assertEqual(fautes, [],
                         'la famille de severite est codee en dur :\n  '
                         + '\n  '.join(fautes))

    def test_FS3b_aucun_appel_a_ajuster_glm_cout_ne_fige_la_famille(self):
        """⚠️⚠️ TROUVE PAR LE SCEAU. A3 appelle `ajuster_glm_cout` a DEUX
        sites : la selection descendante, et un REPLI quand la selection ne
        converge pas. Ce repli n'est pas atteint par la fixture -- le plant
        qui y recodait `'gamma'` ne faisait donc rougir personne.

        Ce controle ne depend d'aucune fixture : il exige que la famille soit
        DERIVEE a chaque appel, jamais un litteral.
        """
        chemin = os.path.join(_ICI, 'a3_glm', 'agent.py')
        arbre = ast.parse(pathlib.Path(chemin).read_text(encoding='utf-8'))
        appels, figes = 0, []
        for n in ast.walk(arbre):
            if not (isinstance(n, ast.Call)
                    and getattr(n.func, 'id', None) == 'ajuster_glm_cout'):
                continue
            appels += 1
            for argument in list(n.args) + [k.value for k in n.keywords]:
                if isinstance(argument, ast.Constant) and isinstance(
                        argument.value, str):
                    figes.append(f'a3:{n.lineno} famille figee '
                                 f'a {argument.value!r}')
        self.assertGreaterEqual(appels, 2,
                                f'{appels} appel(s) releve(s) : A3 en a deux, '
                                'le controle ne mesure pas tout')
        self.assertEqual(figes, [], '\n  '.join(figes))

    def test_FS2b_la_correction_de_Duan_est_REELLEMENT_appliquee(self):
        """⚠️⚠️ AUTRE TROU TROUVE PAR LE SCEAU. Le smearing de Duan (1983)
        corrige le retour de log(cout) a l'echelle des euros : sans lui, la
        prediction est BIAISEE VERS LE BAS -- le tarif sous-estime. Or il
        n'agit que sur `predict`, pas sur l'AIC : FS-2, qui compare des AIC,
        ne pouvait pas le voir.
        """
        import pandas as pd
        rng = np.random.default_rng(11)
        X = pd.DataFrame({'const': np.ones(120), 'x': rng.normal(size=120)})
        y = pd.Series(np.exp(3 + 0.4 * X['x'] + rng.normal(scale=0.6, size=120)))
        modele = ajuster_glm_cout(X, y, 'lognormal')
        self.assertGreater(modele._duan, 1.0,
                           'le smearing vaut 1 : la correction est inerte sur '
                           'ce jeu, le test ne surveille rien')
        sans_duan = ModeleCout('lognormal', modele._res, duan=1.0)
        avec = modele.predict(X)
        sans = sans_duan.predict(X)
        self.assertFalse(np.allclose(avec, sans),
                         'la prediction ne depend pas du smearing')
        self.assertGreater(float(np.mean(avec)), float(np.mean(sans)),
                           'sans Duan la prediction devrait etre plus BASSE : '
                           'le sens de la correction est inverse')

    def test_FS4_la_deviance_absente_vaut_None_jamais_zero(self):
        """Un OLS n'a ni deviance ni deviance nulle. *Zero se lirait comme un
        ajustement parfait.*"""
        import pandas as pd
        rng = np.random.default_rng(7)
        X = pd.DataFrame({'const': np.ones(80),
                          'x': rng.normal(size=80)})
        y = pd.Series(np.exp(3 + 0.4 * X['x'] + rng.normal(scale=0.2, size=80)))
        for famille in _FAMILLES:
            with self.subTest(famille=famille):
                modele = ajuster_glm_cout(X, y, famille)
                self.assertEqual(modele.famille_severite, famille)
                self.assertIsNotNone(modele.aic)
                if famille == 'lognormal':
                    self.assertIsNone(modele.deviance,
                                      'un OLS publie une deviance')
                    self.assertIsNone(modele.null_deviance)
                else:
                    self.assertIsNotNone(modele.deviance)
                # ... et la surface de lecture d'A3 est complete.
                for attribut in ('params', 'pvalues', 'aic', 'bic'):
                    self.assertTrue(hasattr(modele, attribut), attribut)
                self.assertTrue(callable(modele.conf_int))

    def test_FS7_les_plans_du_depot_declarent_tous_gamma(self):
        """⚠️ LA CONDITION 4, MESUREE. Si un plan declarait une autre famille,
        ce lot deplacerait un prix -- et ce test le dirait."""
        familles = {}
        for chemin in sorted(glob.glob(os.path.join(_RACINE, 'plans', '*.yaml'))):
            try:
                plan = PlanTarifaire.depuis_yaml(chemin)
            except Exception as _e:                               # noqa: BLE001
                print(f'    (plan ignore : {os.path.basename(chemin)} -- {_e})')
                continue
            familles.setdefault(plan.famille_severite, []).append(
                os.path.basename(chemin))
        self.assertEqual(
            sorted(familles), ['gamma'],
            'des plans declarent une autre famille que gamma : la production '
            f'CHANGE et il faut le mesurer -- { {k: v for k, v in familles.items() if k != "gamma"} }')
        self.assertGreaterEqual(len(familles['gamma']), 15,
                                'trop peu de plans lus : le test ne mesure '
                                'presque rien')


class TestA3AjusteLaFamilleDeclaree(unittest.TestCase):
    """⚠️⚠️ PAR EXECUTION, SUR LE MOTEUR D'A3. `INV-5c` et `INV-5d` prouvent
    deja la propriete sur le chemin DECLARATIF ; c'est le moteur d'A3 qui la
    manquait, et c'est lui qui sera retenu a la fusion."""

    @classmethod
    def setUpClass(cls):
        warnings.filterwarnings('ignore')
        logging.disable(logging.CRITICAL)

    @classmethod
    def tearDownClass(cls):
        logging.disable(logging.NOTSET)

    def _a3(self, famille, agent=None):
        from core.qualite_donnees import preambule_qualite
        from direction_non_vie.tarification import test_pipeline_agents as T
        from direction_non_vie.tarification.a1_ingestion.agent import (
            AgentA1Ingestion,
        )
        from direction_non_vie.tarification.a2_preprocessing.agent import (
            AgentA2Preprocessing,
        )
        from direction_non_vie.tarification.a3_glm.agent import AgentA3GLM

        plan = dataclasses.replace(T._PLAN_AUTO, famille_severite=famille)
        np.random.seed(7)
        donnees = T._portefeuille_auto(1200)
        base = {'audit_path': '/tmp', 'verbose': False}
        r1 = AgentA1Ingestion(**base).run(branche='non_vie',
                                          sous_branche='auto',
                                          dataframe=donnees)
        qualite = preambule_qualite(r1.get('dataframe'), plan,
                                    qualite_validee_par='Test',
                                    horodatage=None)
        r2 = AgentA2Preprocessing(**base).run(
            result_a1={**r1, 'dataframe': qualite.dataframe_propre}, plan=plan)
        agent = agent or AgentA3GLM(models_path='/tmp', audit_path='/tmp')
        return agent, agent.run(
            result_a2=r2, plan=plan, col_frequence=plan.cible_frequence,
            col_cout=plan.cible_cout, generer_graphiques=False)

    def test_FS1_les_trois_familles_ABOUTISSENT_dans_A3(self):
        """⚠️⚠️ La lognormale faisait ECHOUER LE RUN ENTIER."""
        for famille in _FAMILLES:
            with self.subTest(famille=famille):
                _, r3 = self._a3(famille)
                self.assertTrue(r3.get('success'),
                                f"A3 echoue sur {famille} : {r3.get('erreur')}")
                self.assertTrue(r3.get('excel_bytes'),
                                'aucun classeur produit')

    def test_FS2_trois_familles_donnent_trois_ajustements_DISTINCTS(self):
        """Preuve que la famille n'est pas ignoree — le pendant d'INV-5d, sur
        le moteur d'A3."""
        aic = {}
        for famille in _FAMILLES:
            _, r3 = self._a3(famille)
            aic[famille] = (r3.get('metriques') or {}).get('gamma', {}).get('aic')
        self.assertEqual(len(set(aic.values())), 3,
                         f'les trois familles donnent le meme ajustement : {aic}')
        self.assertNotIn(None, aic.values())

    def test_FS4b_le_classeur_A3_DIT_la_deviance_non_mesuree(self):
        """⚠️ Elle n'a pas de sens pour un OLS : le classeur signe doit
        l'ecrire, pas publier un zero."""
        from openpyxl import load_workbook
        _, r3 = self._a3('lognormal')
        met = (r3.get('metriques') or {}).get('gamma', {})
        self.assertIsNone(met.get('deviance'))
        self.assertIsNone(met.get('pseudo_r2'))
        classeur = load_workbook(io.BytesIO(r3['excel_bytes']), data_only=True)
        texte = '\n'.join(str(c.value) for f in classeur.worksheets
                          for ligne in f.iter_rows() for c in ligne
                          if c.value is not None)
        self.assertIn('non mesuré', texte,
                      'le classeur publie un chiffre la ou la grandeur '
                      "n'existe pas")

    def test_FS5_la_famille_du_run_est_REINITIALISEE(self):
        """⚠️⚠️ MEME RISQUE QUE LA CIBLE DU LOT 5 : la famille est rangee sur
        l'instance pour `_calibrer_gamma`, qui ne recoit pas le plan. Un meme
        agent relance sur un autre plan doit changer de loi."""
        from direction_non_vie.tarification.a3_glm.agent import AgentA3GLM
        agent = AgentA3GLM(models_path='/tmp', audit_path='/tmp')
        self._a3('lognormal', agent)
        self.assertEqual(agent._famille_severite_run, 'lognormal')
        _, second = self._a3('gamma', agent)
        self.assertEqual(agent._famille_severite_run, 'gamma',
                         'le MEME agent relance garde la famille du run '
                         'precedent')
        self.assertIsNotNone(
            (second.get('metriques') or {}).get('gamma', {}).get('deviance'),
            "le second run n'a pas de deviance : il ajuste encore un OLS")


if __name__ == '__main__':
    unittest.main(verbosity=2)
