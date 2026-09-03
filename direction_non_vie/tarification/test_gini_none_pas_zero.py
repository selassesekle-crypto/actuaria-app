# -*- coding: utf-8 -*-
"""UN GINI QUI N'EXISTE PAS SE DIT `None`, JAMAIS `0.0` — sentinelle du lot.

Constat (03/09/2026) : `_calculer_gini` d'A3, A4 et A5 rendait `0.0` sur un
jeu de test sans sinistre, sur des tableaux vides et sur une exception. Ce
zero traversait toute la chaine comme une MESURE : il colorait un statut
(« Gini < 0.05 → ROUGE »), il s'ecrivait « 0.0000 » dans le commentaire,
l'Excel, le Word, et il alimentait la selection d'A6. Personne n'avait
mesure un pouvoir discriminant nul — la mesure n'existait pas.

Contrat verifie ici, de la methode jusqu'au refus d'A6 :
  NZ-1  les trois `_calculer_gini` ne rendent plus aucune constante (AST) ;
  NZ-2  vide, somme nulle, exception → `None` (methodes reelles) ;
  NZ-3  `gini_texte` / `gini_arrondi` : « non mesuré », jamais 0 ;
        et un Gini mesure s'ecrit EXACTEMENT comme avant ;
  NZ-4  A5 : sans Gini mesure, le statut est AMBRE (pas de couleur fabriquee) ;
  NZ-5  la chaine A1→A6 sur un test sans sinistre : A3/A4/A5 TERMINENT, tous
        leurs Ginis sont None, aucun 0.0 nulle part, leurs Excel EXISTENT et
        portent « non mesuré », la prose d'A3 ne vante plus un pouvoir
        discriminant, et A6 REFUSE l'arbitrage.

⚠️ Tout en `unittest.TestCase` : la gate lance `unittest discover`.
"""
import ast
import importlib.util
import inspect
import io
import logging
import math
import os
import sys
import unittest
import warnings
from unittest.mock import patch

import numpy as np

_RACINE = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..'))
if _RACINE not in sys.path:
    sys.path.insert(0, _RACINE)

from core.conformite_reglementaire import gini_arrondi, gini_texte
from direction_non_vie.tarification.a3_glm.agent import AgentA3GLM
from direction_non_vie.tarification.a4_ml.agent import AgentA4ML
from direction_non_vie.tarification.a5_deep_learning.agent import (
    AgentA5DeepLearning,
)

_AGENTS = (('A3', AgentA3GLM), ('A4', AgentA4ML), ('A5', AgentA5DeepLearning))
# Les cles cote TEST : sans sinistre en test, elles ne peuvent etre que None.
_CLES_TEST = ('gini', 'gini_test', 'gini_max', 'gini_val', 'gini_poisson',
              'gini_gamma', 'gini_tweedie', 'gini_reference', 'gini_actuel',
              'gini_dl', 'gini_glm', 'best_gini')
# `gini_train` est MESURE sur l'entrainement, qui porte des sinistres : il
# a le droit d'etre un nombre — jamais un 0.0 de repli.
_CLES_GINI = _CLES_TEST + ('gini_train',)


def _retours_constants(fonction):
    """Les `return <constante numerique>` d'une fonction, par AST."""
    arbre = ast.parse(_dedent(inspect.getsource(fonction)))
    return [n.value.value for n in ast.walk(arbre)
            if isinstance(n, ast.Return) and isinstance(n.value, ast.Constant)
            and isinstance(n.value.value, (int, float))
            and not isinstance(n.value.value, bool)]


def _dedent(src):
    import textwrap
    return textwrap.dedent(src)


def _valeurs_gini(obj, chemin='', out=None):
    """Toutes les valeurs publiees sous une cle de Gini, avec leur chemin."""
    if out is None:
        out = []
    if isinstance(obj, dict):
        for k, v in obj.items():
            if k in _CLES_GINI and not isinstance(v, (dict, list)):
                out.append((f'{chemin}/{k}', v))
            elif isinstance(v, (dict, list)):
                _valeurs_gini(v, f'{chemin}/{k}', out)
    elif isinstance(obj, list):
        for i, v in enumerate(obj):
            _valeurs_gini(v, f'{chemin}[{i}]', out)
    return out


class T1_LaMethode(unittest.TestCase):

    def test_nz_1_aucun_calculer_gini_ne_rend_une_constante(self):
        """NZ-1 : par AST, aucun `return 0.0` (ni autre constante) dans les trois."""
        for nom, cls in _AGENTS:
            constantes = _retours_constants(cls._calculer_gini)
            self.assertEqual(
                constantes, [],
                f"{nom}._calculer_gini rend une constante {constantes} : un "
                f"Gini qui n'existe pas se dit None, jamais un nombre")

    def test_nz_2_vide_somme_nulle_et_exception_rendent_None(self):
        """NZ-2 : les trois methodes REELLES, sur les trois cas degeneres."""
        pred = np.random.RandomState(0).rand(50)
        for nom, cls in _AGENTS:
            a = cls.__new__(cls)
            with self.subTest(agent=nom, cas='vide'):
                self.assertIsNone(a._calculer_gini(np.array([]), np.array([])))
            with self.subTest(agent=nom, cas='aucun sinistre'):
                self.assertIsNone(a._calculer_gini(np.zeros(50), pred))
            # une exception AU MILIEU du calcul (le tri des predictions
            # echoue) : l'except ne doit pas fabriquer un nombre
            with self.subTest(agent=nom, cas='exception'), \
                    patch.object(np, 'argsort', side_effect=RuntimeError('sonde')):
                self.assertIsNone(a._calculer_gini(np.ones(50), pred))
            with self.subTest(agent=nom, cas='mesure'):
                y = np.random.RandomState(1).poisson(0.3, 400).astype(float)
                g = a._calculer_gini(y, y + np.random.RandomState(2).rand(400) * 0.1)
                self.assertIsInstance(g, float)
                self.assertTrue(math.isfinite(g))


class T2_LeMot(unittest.TestCase):

    def test_nz_3_non_mesure_jamais_zero_et_un_nombre_s_ecrit_comme_avant(self):
        """NZ-3 : None et nan → « non mesuré » ; 0.12345 → '0.1235' (identique)."""
        self.assertEqual(gini_texte(None), 'non mesuré')
        self.assertEqual(gini_texte(float('nan')), 'non mesuré')
        self.assertEqual(gini_texte(float('inf')), 'non mesuré')
        self.assertEqual(gini_arrondi(None), 'non mesuré')
        self.assertEqual(gini_arrondi(float('nan')), 'non mesuré')
        for v in (0.12345, -0.0298, 0.0, 0.19125, 1.0):
            self.assertEqual(gini_texte(v), f'{v:.4f}')
            self.assertEqual(gini_texte(v, 3), f'{v:.3f}')
            self.assertEqual(gini_arrondi(v), round(v, 4))
            self.assertEqual(gini_arrondi(v, 3), round(v, 3))
        self.assertNotIn('0.0000', gini_texte(None))
        self.assertNotEqual(gini_arrondi(None), 0)


class T3_LeStatut(unittest.TestCase):

    def test_nz_4_A5_sans_gini_mesure_est_AMBRE_pas_ROUGE(self):
        """NZ-4 : un Gini absent reserve, il ne colore pas (arbitrage 03/09)."""
        a5 = AgentA5DeepLearning.__new__(AgentA5DeepLearning)
        self.assertEqual(
            a5._calculer_statut_rag([
                {'type': 'Deep Learning', 'modele': 'cann', 'gini_test': None},
                {'type': 'Deep Learning', 'modele': 'tabnet', 'gini_test': None},
            ]), 'AMBRE')
        # et un Gini MESURE garde exactement l'echelle d'avant
        self.assertEqual(a5._calculer_statut_rag(
            [{'type': 'Deep Learning', 'gini_test': 0.16}]), 'VERT')
        self.assertEqual(a5._calculer_statut_rag(
            [{'type': 'Deep Learning', 'gini_test': 0.06}]), 'AMBRE')
        self.assertEqual(a5._calculer_statut_rag(
            [{'type': 'Deep Learning', 'gini_test': 0.01}]), 'ROUGE')
        # un DL non mesure a cote d'un DL mesure : c'est le mesure qui decide
        self.assertEqual(a5._calculer_statut_rag([
            {'type': 'Deep Learning', 'gini_test': None},
            {'type': 'Deep Learning', 'gini_test': 0.16},
        ]), 'VERT')


class T4_LaChaine(unittest.TestCase):
    """NZ-5 : la chaine REELLE sur un test sans sinistre (decoupage temporel)."""

    @classmethod
    def setUpClass(cls):
        from core.qualite_donnees import preambule_qualite
        from direction_non_vie.tarification import test_pipeline_agents as T
        from direction_non_vie.tarification.a1_ingestion.agent import AgentA1Ingestion
        from direction_non_vie.tarification.a2_preprocessing.agent import (
            AgentA2Preprocessing,
        )
        from direction_non_vie.tarification.a6_comparaison.agent import (
            AgentA6Comparaison,
        )

        logging.disable(logging.CRITICAL)
        warnings.filterwarnings('ignore')
        np.random.seed(3)
        df = T._portefeuille_auto(1600)
        masque = df['annee_souscription'].isin([2022, 2023])
        df.loc[masque, 'nb_sinistres'] = 0.0
        df.loc[masque, 'cout_total_sinistres'] = 0.0
        assert df.loc[~masque, 'nb_sinistres'].sum() > 0, 'fixture : train sans sinistre'
        plan = T._PLAN_AUTO
        _a = {'models_path': '/tmp', 'audit_path': '/tmp'}
        _ab = {'audit_path': '/tmp', 'verbose': False}
        r1 = AgentA1Ingestion(**_ab).run(branche='non_vie', sous_branche='auto',
                                         dataframe=df)
        rq = preambule_qualite(r1['dataframe'], plan,
                               qualite_validee_par='Actuaire Test', horodatage=None)
        r1 = {**r1, 'dataframe': rq.dataframe_propre}
        r2 = AgentA2Preprocessing(**_ab).run(result_a1=r1, plan=plan)
        cls.r3 = AgentA3GLM(**_a).run(
            result_a2=r2, plan=plan, col_frequence=plan.cible_frequence,
            col_cout=plan.cible_cout, generer_graphiques=False)
        cls.r4 = AgentA4ML(**_a).run(
            result_a2=r2, result_a3=cls.r3, plan=plan, col_cible='nb_sinistres',
            ponderer_par_exposition=True, calcul_shap=False,
            generer_graphiques=False)
        if importlib.util.find_spec('torch') is None:
            cls.r5 = None      # sans PyTorch, A5 ne se calibre pas : on le dit
        else:
            cls.r5 = AgentA5DeepLearning(**_a).run(
                result_a2=r2, result_a3=cls.r3, result_a4=cls.r4, plan=plan,
                col_cible='nb_sinistres', modeles=('cann', 'tabnet'), n_epochs=1,
                batch_size=512, generer_graphiques=False)
        cls.r6 = AgentA6Comparaison(**_a).run(
            result_a2=r2, result_a3=cls.r3, result_a4=cls.r4,
            result_a5=cls.r5 if cls.r5 and cls.r5.get('success') else None,
            col_cible='nb_sinistres', plan=plan, environnement='production',
            profil_valide_par='Actuaire Test', generer_graphiques=False,
            generer_rapport_equipe=False)

    @classmethod
    def tearDownClass(cls):
        logging.disable(logging.NOTSET)

    def _resultats(self):
        out = [('A3', self.r3), ('A4', self.r4)]
        if self.r5 is not None:
            out.append(('A5', self.r5))
        return out

    def test_nz_5a_les_trois_agents_TERMINENT_avec_un_diagnostic(self):
        for nom, r in self._resultats():
            with self.subTest(agent=nom):
                self.assertTrue(r.get('success'), f"{nom} : {r.get('erreur')}")
                self.assertTrue(r.get('diagnostic_evaluation'),
                                f"{nom} ne publie pas le diagnostic d'evaluation")
                self.assertIn(r.get('statut_rag'), ('VERT', 'AMBRE', 'ROUGE'))

    def test_nz_5b_tous_les_ginis_publies_sont_None_et_AUCUN_0_0(self):
        for nom, r in self._resultats():
            valeurs = _valeurs_gini(r)
            with self.subTest(agent=nom):
                self.assertTrue(valeurs, f"{nom} ne publie aucune cle de Gini ?")
                zeros = [(c, v) for c, v in valeurs if v == 0]
                self.assertEqual(
                    zeros, [],
                    f"{nom} publie un Gini 0.0 fabrique sur un test sans sinistre")
                mesures = [(c, v) for c, v in valeurs if v is not None
                           and c.rsplit('/', 1)[-1] in _CLES_TEST]
                self.assertEqual(
                    mesures, [],
                    f"{nom} publie un Gini de TEST 'mesure' sur un test sans "
                    f"sinistre : {mesures}")

    def test_nz_5c_les_Excel_existent_et_disent_non_mesure(self):
        from openpyxl import load_workbook
        for nom, r in self._resultats():
            with self.subTest(agent=nom):
                octets = r.get('excel_bytes') or b''
                self.assertTrue(
                    octets, f"l'Excel de {nom} a DISPARU : un except a avale un "
                            f"round(None) et rendu b''")
                wb = load_workbook(io.BytesIO(octets), data_only=True)
                textes = [str(c.value) for ws in wb.worksheets
                          for row in ws.iter_rows() for c in row
                          if c.value is not None]
                self.assertTrue(any('non mesuré' in t for t in textes),
                                f"l'Excel de {nom} n'ecrit nulle part « non mesuré »")

    def test_nz_5d_la_prose_d_A3_ne_vante_plus_un_pouvoir_discriminant(self):
        com = self.r3.get('commentaire', '')
        self.assertIn('ÉVALUATION est impossible', com)
        self.assertNotIn('bon pouvoir discriminant', com)
        self.assertNotIn('0.0000', com)
        self.assertIn('non mesuré', com)
        self.assertNotIn('0€', com.replace('0 €', '0€').split('Coût obs/pred')[-1]
                         .split('\n')[0],
                         "le cout moyen predit s'affiche « 0€ » : mesure fabriquee")

    def test_nz_5e_A4_et_A5_sont_AMBRE_et_A6_REFUSE(self):
        self.assertEqual(self.r4.get('statut_rag'), 'AMBRE',
                         "A4 colore sans Gini mesure")
        # ⚠️ Les modeles doivent etre LA : un lecteur qui leve dans la boucle
        # de calibration est avale par l'except de chaque modele, et un
        # classement reduit a la ligne GLM passerait tout le reste en vert.
        # Le sceau P3 l'a montre : sans cette assertion, le plant restait vert.
        _ml = [c.get('modele') for c in self.r4.get('classement', [])
               if c.get('famille') != 'GLM']
        self.assertGreaterEqual(
            len(_ml), 2, f"A4 n'a garde que {_ml} : des modeles ont ete perdus "
                         f"en silence (alertes : {self.r4.get('rapport', {}).get('alertes')})")
        for c in self.r4.get('classement', []):
            self.assertIsNone(c.get('gini_test'))
            self.assertIsNone(c.get('overfit_alerte'),
                              f"{c.get('modele')} : alerte de sur-apprentissage "
                              f"fabriquee sans ratio")
            self.assertIsNone(c.get('overfit_ratio'),
                              f"{c.get('modele')} : ratio de sur-apprentissage "
                              f"fabrique sans Gini")
            if c.get('famille') != 'GLM':   # la ligne GLM porte un libelle, pas un verdict
                self.assertIn('non évaluable', c.get('recommandation', ''))
        if self.r5 is not None:
            self.assertEqual(self.r5.get('statut_rag'), 'AMBRE',
                             "A5 colore sans Gini mesure")
            self.assertEqual(sorted(self.r5.get('metriques', {})), ['cann', 'tabnet'],
                             "A5 a perdu un modele en silence")
        self.assertFalse(self.r6.get('success'))
        self.assertIn('ARBITRAGE IMPOSSIBLE', str(self.r6.get('erreur')))
        self.assertEqual(self.r6.get('modele_production'), {})


if __name__ == '__main__':
    unittest.main()
