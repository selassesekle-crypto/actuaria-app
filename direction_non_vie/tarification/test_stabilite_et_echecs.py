# -*- coding: utf-8 -*-
"""UNE STABILITE NON MESUREE NE VAUT PAS 1.0, ET UNE PANNE N'EST PAS UN RESULTAT.

Trois constats fermes le 03/09/2026, tous sur le chemin qui CHOISIT le modele
de production, donc le tarif :

  1. A6 posait `overfit_ratio: 1.0` pour tout GLM, sans l'avoir mesure. Or la
     normalisation de la stabilite fait `1 - (r - min)/(max - min)` : les ML
     sur-apprennent (ratios > 1), donc 1.0 etait le MINIMUM du catalogue et le
     GLM recevait `s_stab = 1.0`, la MEILLEURE note possible, sur un critere
     qui pese **30 %** de la selection. Mesure : le vrai ratio vaut 1,0842
     (Poisson), 1,2751 (Tweedie), 1,7468 (Gamma).
  2. La ligne << GLM Poisson (reference A3) >> du classement d'A4 portait le
     meme 1.0 et `overfit_alerte: False`, non mesures.
  3. L'`except` par modele d'A4 avalait une PANNE TECHNIQUE en
     `WARNING`+`rapport['alertes']` -- un canal qu'aucune surface signee ne
     lit -- puis le commentaire concluait sur les DONNEES (<< aucun modele ML
     n'ameliore le GLM >>). Mesure : deux pannes injectees, statut A6 **VERT**,
     temoin absent des huit livrables.

Ce que cette sentinelle exige :
  ST-1  A3 MESURE `gini_train` et `overfit_ratio` pour ses trois GLM ;
  ST-2  la formule du ratio est PARTAGEE par A3, A4 et A5 (source unique) ;
  ST-3  A6 ne fabrique plus aucun ratio : ni `1.0` litteral, ni repli ;
  ST-4  un ratio non mesure RETIRE le critere du score au lieu d'inventer une
        valeur, et le score reste identique quand tout est mesure ;
  ST-5  une stabilite non mesuree sur le modele retenu PLAFONNE le statut ;
  ST-6  la ligne GLM d'A4 porte la mesure d'A3, jamais 1.0 ;
  ST-7  une panne technique est distinguee d'une librairie absente ;
  ST-8  une panne technique PLAFONNE A4 et A6 et interdit la conclusion sur
        les donnees ;
  ST-9  une panne technique ATTEINT les surfaces signees.

⚠️ Tout en `unittest.TestCase` : la gate lance `unittest discover`.
"""
import ast
import inspect
import io
import logging
import os
import sys
import textwrap
import unittest
import warnings
import zipfile

import numpy as np

_RACINE = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..'))
if _RACINE not in sys.path:
    sys.path.insert(0, _RACINE)

from core.conformite_reglementaire import ratio_sur_apprentissage
from direction_non_vie.tarification.a3_glm.agent import AgentA3GLM
from direction_non_vie.tarification.a4_ml import agent as A4MOD
from direction_non_vie.tarification.a5_deep_learning import agent as A5MOD
from direction_non_vie.tarification.a6_comparaison.agent import (
    PROFILS_PONDERATION,
    AgentA6Comparaison,
)

_TEMOIN = 'PANNE_TECHNIQUE_SENTINELLE'


def _src(objet):
    return textwrap.dedent(inspect.getsource(objet))


def _catalogue(ratios):
    """Un catalogue minimal : seul le ratio varie d'un modele a l'autre."""
    return [{'modele': f'M{i}', 'famille': 'ML', 'gini_test': 0.20,
             'rmse_test': 1.0, 'interpretabilite': 0.6, 'overfit_ratio': r}
            for i, r in enumerate(ratios)]


class T1_LaMesure(unittest.TestCase):
    """ST-1, ST-2 : la stabilite du GLM se MESURE, avec la formule partagee."""

    @classmethod
    def setUpClass(cls):
        from core.qualite_donnees import preambule_qualite
        from direction_non_vie.tarification import test_pipeline_agents as T
        from direction_non_vie.tarification.a1_ingestion.agent import AgentA1Ingestion
        from direction_non_vie.tarification.a2_preprocessing.agent import (
            AgentA2Preprocessing,
        )

        logging.disable(logging.CRITICAL)
        warnings.filterwarnings('ignore')
        np.random.seed(11)
        df = T._portefeuille_auto(2500)
        plan = T._PLAN_AUTO
        _ab = {'audit_path': '/tmp', 'verbose': False}
        r1 = AgentA1Ingestion(**_ab).run(branche='non_vie', sous_branche='auto',
                                         dataframe=df)
        rq = preambule_qualite(r1['dataframe'], plan,
                               qualite_validee_par='Actuaire Test', horodatage=None)
        r1 = {**r1, 'dataframe': rq.dataframe_propre}
        r2 = AgentA2Preprocessing(**_ab).run(result_a1=r1, plan=plan)
        cls.r3 = AgentA3GLM(models_path='/tmp', audit_path='/tmp').run(
            result_a2=r2, plan=plan, col_frequence=plan.cible_frequence,
            col_cout=plan.cible_cout, generer_graphiques=False)

    @classmethod
    def tearDownClass(cls):
        logging.disable(logging.NOTSET)

    def test_st_1_A3_publie_une_stabilite_MESUREE_pour_ses_trois_GLM(self):
        for nom in ('poisson', 'gamma', 'tweedie'):
            met = self.r3['metriques'].get(nom, {})
            with self.subTest(glm=nom):
                self.assertIn('gini_train', met,
                              f"{nom} ne publie pas son Gini d'entrainement")
                self.assertIn('overfit_ratio', met)
                r = met['overfit_ratio']
                self.assertIsNotNone(
                    r, f"{nom} : ratio non mesure sur une fixture normale")
                # La valeur est MESUREE, donc differente du 1.0 fabrique.
                self.assertNotEqual(
                    round(float(r), 6), 1.0,
                    f"{nom} publie exactement 1.0 : c'est la valeur fabriquee "
                    f"que ce lot supprime, pas une mesure")
                # ⚠️ Les deux Ginis sont ARRONDIS a 4 decimales dans le dict :
                # le ratio recalcule ne peut coincider qu'a l'arrondi pres, et
                # l'ecart relatif grandit quand le Gini de test est petit
                # (0,0318 pour le Gamma). On verifie la DERIVATION, pas une
                # egalite au bit.
                _attendu = ratio_sur_apprentissage(met['gini_train'], met['gini'])
                self.assertAlmostEqual(
                    float(r), _attendu, delta=0.01 * abs(_attendu),
                    msg=f"{nom} : le ratio publie ne derive pas de ses deux Ginis")

    def test_st_2_les_trois_agents_partagent_LA_formule(self):
        """ST-2 : A6 normalise ENTRE les modeles — un ratio calcule autrement
        ne se compare pas. Le controle porte sur l'APPEL, par AST, pas sur le
        texte : `if False and ratio_sur_apprentissage(...)` ne passerait pas."""
        for nom, methode in (('A3', AgentA3GLM._stabilite_train),
                             ('A4', A4MOD.AgentA4ML._calculer_metriques),
                             ('A5', A5MOD.AgentA5DeepLearning._calculer_metriques)):
            arbre = ast.parse(_src(methode))
            appels = [n for n in ast.walk(arbre)
                      if isinstance(n, ast.Call)
                      and getattr(n.func, 'id', None) == 'ratio_sur_apprentissage']
            with self.subTest(agent=nom):
                self.assertTrue(
                    appels,
                    f"{nom} calcule son ratio sans passer par la source unique")

    def test_st_2b_la_formule_rend_None_et_jamais_1_0(self):
        self.assertIsNone(ratio_sur_apprentissage(None, 0.2))
        self.assertIsNone(ratio_sur_apprentissage(0.2, None))
        self.assertIsNone(ratio_sur_apprentissage(float('nan'), 0.2))
        self.assertIsNone(ratio_sur_apprentissage(0.2, float('inf')))
        self.assertAlmostEqual(ratio_sur_apprentissage(0.25, 0.20), 1.25)


class T2_A6NeFabriquePlus(unittest.TestCase):
    """ST-3, ST-4, ST-5 : le catalogue et le score d'A6."""

    def test_st_3_aucun_ratio_fabrique_dans_l_agregation(self):
        """ST-3 : par AST, aucune constante numerique n'alimente
        `overfit_ratio` ni `gini_train` dans `_agreger_resultats`."""
        arbre = ast.parse(_src(AgentA6Comparaison._agreger_resultats))
        fautes = []
        for n in ast.walk(arbre):
            if not isinstance(n, ast.Dict):
                continue
            for cle, val in zip(n.keys, n.values):
                if getattr(cle, 'value', None) in ('overfit_ratio', 'gini_train'):
                    if isinstance(val, ast.Constant):
                        fautes.append((cle.value, val.value))
                    # `met.get('x', <defaut>)` : un defaut est une fabrication
                    if (isinstance(val, ast.Call)
                            and getattr(val.func, 'attr', None) == 'get'
                            and len(val.args) > 1):
                        fautes.append((cle.value, 'valeur par defaut'))
        self.assertEqual(
            fautes, [],
            f"A6 fabrique encore une stabilite : {fautes}. Chez A6 un ratio de "
            f"1.0 est la MEILLEURE note possible — 30 % de la selection.")

    def test_st_4_un_ratio_absent_RETIRE_le_critere_sans_rien_inventer(self):
        a6 = AgentAtelier()
        poids = PROFILS_PONDERATION['equilibre']
        # (a) tout mesure : le score est celui d'avant ce lot, au bit pres
        cat = _catalogue([1.0, 1.5, 2.0])
        note = AgentA6Comparaison._calculer_scores_multicriteres(a6, cat, poids)
        attendus = []
        for m in note:
            s_stab = m['score_stabilite']
            attendus.append(round(
                poids['gini'] * m['score_gini']
                + poids['stabilite'] * s_stab
                + poids['interpretabilite'] * m['score_interpretabilite']
                + poids['rmse'] * m['score_rmse'], 4))
        self.assertEqual([m['score_global'] for m in note], attendus,
                         "le score a bouge alors que tout est mesure")
        # (b) un ratio absent : le critere sort, il n'est pas remplace
        cat = _catalogue([1.0, 1.5, 2.0])
        cat[1]['overfit_ratio'] = None
        note = AgentA6Comparaison._calculer_scores_multicriteres(a6, cat, poids)
        sans = note[1]
        self.assertIsNone(sans['score_stabilite'],
                          "une stabilite absente a recu une note")
        self.assertEqual(sans['criteres_non_mesures'], ('stabilite',))
        _reste = poids['gini'] + poids['interpretabilite'] + poids['rmse']
        self.assertAlmostEqual(
            sans['score_global'],
            round((poids['gini'] * sans['score_gini']
                   + poids['interpretabilite'] * sans['score_interpretabilite']
                   + poids['rmse'] * sans['score_rmse']) / _reste, 4),
            places=4,
            msg="le score du modele sans stabilite n'est pas renormalise")

    def test_st_4b_un_ratio_absent_ne_deplace_PAS_le_score_des_autres(self):
        """Le modele non mesure sort de la normalisation : les autres gardent
        exactement le score qu'ils avaient. Sinon une absence deplacerait
        `min_of`/`max_of`, donc la note de TOUT le catalogue."""
        a6 = AgentAtelier()
        poids = PROFILS_PONDERATION['equilibre']
        avant = AgentA6Comparaison._calculer_scores_multicriteres(
            a6, _catalogue([1.0, 1.5, 2.0]), poids)
        cat = _catalogue([1.0, 1.5, 2.0, 3.0])
        cat[3]['overfit_ratio'] = None       # un 4e modele, sans mesure
        apres = AgentA6Comparaison._calculer_scores_multicriteres(a6, cat, poids)
        for a, b in zip(avant, apres[:3]):
            self.assertEqual(a['score_global'], b['score_global'],
                             f"{a['modele']} : son score a bouge a cause d'un "
                             f"AUTRE modele dont la stabilite n'est pas mesuree")

    def test_st_5_une_stabilite_non_mesuree_PLAFONNE_le_statut(self):
        a6 = AgentAtelier()
        base = {'modele': 'M', 'famille': 'ML', 'gini_test': 0.30,
                'score_global': 0.90, 'overfit_ratio': 1.05}
        commun = {'classement': [base], 'profil_valide_par': 'Actuaire',
                  'environnement': 'production',
                  'backtest': {'disponible': True, 'walk_forward_fidele': True,
                               'ae_ratio': 1.0, 'n_fenetres_rouge': 0,
                               'gini_wf_moyen': 0.28, 'stabilite_wf': 'Stable'}}
        avec = AgentA6Comparaison._calculer_statut_rag(a6, dict(base), **commun)
        sans = AgentA6Comparaison._calculer_statut_rag(
            a6, {**base, 'overfit_ratio': None}, **commun)
        self.assertNotEqual(
            sans, 'VERT',
            "un modele dont la stabilite n'est PAS mesuree ressort VERT : "
            "retirer le critere du score ne doit pas devenir un avantage")
        self.assertTrue(
            any('NON MESUR' in r.upper() for r in a6._raisons_plafond),
            f"la cause n'est pas nommee : {a6._raisons_plafond}")
        self.assertIsInstance(avec, str)


class AgentAtelier(AgentA6Comparaison):
    """Une instance d'A6 sans effet de bord (pas de dossier, pas de journal)."""

    def __init__(self):
        self.models_path = self.audit_path = None
        self.verbose = False


class T3_LaPanneNEstPasUnResultat(unittest.TestCase):
    """ST-6..ST-9 : la chaine REELLE, avec une panne technique injectee."""

    @classmethod
    def setUpClass(cls):
        from core.qualite_donnees import preambule_qualite
        from direction_non_vie.tarification import test_pipeline_agents as T
        from direction_non_vie.tarification.a1_ingestion.agent import AgentA1Ingestion
        from direction_non_vie.tarification.a2_preprocessing.agent import (
            AgentA2Preprocessing,
        )

        logging.disable(logging.CRITICAL)
        warnings.filterwarnings('ignore')
        np.random.seed(11)
        df = T._portefeuille_auto(2000)
        plan = T._PLAN_AUTO
        _a = {'models_path': '/tmp', 'audit_path': '/tmp'}
        _ab = {'audit_path': '/tmp', 'verbose': False}
        r1 = AgentA1Ingestion(**_ab).run(branche='non_vie', sous_branche='auto',
                                         dataframe=df)
        rq = preambule_qualite(r1['dataframe'], plan,
                               qualite_validee_par='Actuaire Test', horodatage=None)
        r1 = {**r1, 'dataframe': rq.dataframe_propre}
        r2 = AgentA2Preprocessing(**_ab).run(result_a1=r1, plan=plan)
        cls.r3 = AgentA3GLM(**_a).run(
            result_a2=r2, plan=plan, col_frequence=plan.cible_frequence,
            col_cout=plan.cible_cout, generer_graphiques=False)

        _vraie = A4MOD.creer_modele_ml_pour_nom

        def _en_panne(nom, *a, **k):
            if nom == 'gbm':
                raise TypeError(_TEMOIN)
            return _vraie(nom, *a, **k)

        A4MOD.creer_modele_ml_pour_nom = _en_panne
        try:
            cls.r4 = A4MOD.AgentA4ML(**_a).run(
                result_a2=r2, result_a3=cls.r3, plan=plan,
                col_cible='nb_sinistres', ponderer_par_exposition=True,
                calcul_shap=False, generer_graphiques=False)
        finally:
            A4MOD.creer_modele_ml_pour_nom = _vraie
        cls.r6 = AgentA6Comparaison(**_a).run(
            result_a2=r2, result_a3=cls.r3, result_a4=cls.r4, result_a5=None,
            col_cible='nb_sinistres', plan=plan, environnement='production',
            profil_valide_par='Actuaire Test', generer_graphiques=False,
            generer_rapport_equipe=True)

    @classmethod
    def tearDownClass(cls):
        logging.disable(logging.NOTSET)

    def test_st_6_la_ligne_GLM_d_A4_porte_la_MESURE_d_A3(self):
        glm = [c for c in self.r4.get('classement', [])
               if c.get('famille') == 'GLM']
        self.assertTrue(glm, "la ligne GLM de reference a disparu du classement")
        ligne = glm[0]
        attendu = self.r3['metriques']['poisson']['overfit_ratio']
        # ⚠️ EXIGER LA MESURE, PAS SEULEMENT LE RELAIS. Le sceau l'a montre :
        # un plant qui empeche A3 de mesurer laissait ce test VERT, les deux
        # cotes valant None. Un relais fidele a une absence est encore une
        # absence -- meme famille que << verifier des statuts, pas la
        # presence >>. La fixture est normale : la mesure DOIT exister.
        self.assertIsNotNone(
            attendu, "premisse : A3 doit mesurer sa stabilite sur cette fixture")
        self.assertIsNotNone(
            ligne['overfit_ratio'],
            "la ligne GLM d'A4 ne porte AUCUNE stabilite alors qu'A3 l'a mesuree")
        self.assertEqual(
            ligne['overfit_ratio'], attendu,
            f"la ligne GLM d'A4 publie {ligne['overfit_ratio']} au lieu de la "
            f"mesure d'A3 ({attendu})")
        self.assertNotEqual(ligne['overfit_ratio'], 1.0,
                            "le 1.0 fabrique est revenu dans le classement A4")
        self.assertIsNotNone(ligne['gini_train'],
                             "le Gini d'entrainement de la ligne GLM a disparu")

    def test_st_7_une_panne_TECHNIQUE_est_distinguee_d_une_absence(self):
        echecs = self.r4.get('echecs_modeles') or []
        tech = [e for e in echecs if e.get('nature') == 'technique']
        self.assertTrue(tech, f"la panne n'est pas collectee : {echecs}")
        self.assertEqual(tech[0]['modele'], 'gbm')
        self.assertEqual(tech[0]['type'], 'TypeError')
        self.assertIn(_TEMOIN, tech[0]['message'])
        for e in echecs:
            self.assertIn(e['nature'], ('technique', 'absence'))

    def test_st_8_la_panne_PLAFONNE_et_interdit_la_conclusion_sur_les_donnees(self):
        self.assertNotEqual(self.r4.get('statut_rag'), 'VERT',
                            "A4 certifie VERT malgre un modele plante")
        self.assertNotEqual(self.r6.get('statut_rag'), 'VERT',
                            "A6 certifie VERT malgre un modele plante")
        # ⚠️ Les causes vivent dans `audit_trail`, pas au premier niveau : on
        # passe par le LECTEUR PUBLIC, celui que les livrables utilisent.
        from direction_non_vie.tarification.services.rapport_modeles_tarif import (
            raisons_plafond,
        )
        raisons = ' | '.join(raisons_plafond(self.r6))
        self.assertIn('ECHEC TECHNIQUE', raisons,
                      f"la cause du plafond ne nomme pas la panne : {raisons}")
        com = self.r4.get('commentaire', '')
        self.assertIn('ECHEC TECHNIQUE', com)
        self.assertIn('gbm', com)
        # la fausse conclusion ne doit pas coexister avec une panne
        if "Aucun modèle ML n'améliore" in com:
            self.fail("A4 conclut encore sur les DONNEES alors qu'un modele "
                      "a plante : un echec technique n'est pas un resultat")

    def test_st_9_la_panne_ATTEINT_les_surfaces_signees(self):
        from openpyxl import load_workbook

        def cellules(b):
            if not b:
                return ''
            wb = load_workbook(io.BytesIO(b), data_only=True)
            return '\n'.join(str(c.value) for ws in wb.worksheets
                             for row in ws.iter_rows() for c in row
                             if c.value is not None)

        def word(b):
            if not b:
                return ''
            with zipfile.ZipFile(io.BytesIO(b)) as z:
                return z.read('word/document.xml').decode('utf-8', 'replace')

        eq = self.r6.get('rapport_equipe') or {}
        surfaces = {
            'a4.excel':     cellules(self.r4.get('excel_bytes')),
            'a6.excel':     cellules(self.r6.get('excel_bytes')),
            'equipe.html':  (eq.get('html_bytes') or b'').decode('utf-8', 'replace'),
            'equipe.word':  word(eq.get('word_bytes')),
            'equipe.excel': cellules(eq.get('excel_bytes')),
        }
        for nom, contenu in surfaces.items():
            with self.subTest(surface=nom):
                self.assertTrue(contenu, f"{nom} : livrable VIDE")
                self.assertIn(
                    'ECHEC TECHNIQUE', contenu,
                    f"{nom} ne dit rien de la panne : l'erreur y est avalee")


if __name__ == '__main__':
    unittest.main()
