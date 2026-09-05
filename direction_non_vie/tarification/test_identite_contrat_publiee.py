# -*- coding: utf-8 -*-
"""A1 DEVINAIT L'IDENTITE DU CONTRAT, ET NE LE DISAIT A PERSONNE (A1.5).

A1 calcule `source_identifiant` (`plan` / `devinee` / `aucune`),
`note_identite` et `granularite` depuis le 24/08. Mesure du 05/09/2026 :
**seuls des tests les lisaient**. Le classeur A1 signe publiait
<< Nb doublons >> et << Taux doublons >> sans jamais dire SUR QUELLE CLE le
compte portait -- alors que le titre du bloc annonce, lui,
<< lisibilite, completude, IDENTITE >>.

  *Un calcul qui n'atteint aucun livrable n'existe pas ; et un titre qui
  annonce plus que son bloc ne porte est une dette.*

⚠️⚠️ ET LE MOTIF DU REPLI A CHANGE, CE QUI CHANGE LE CONSTAT. Le code disait
<< aucun des vingt plans ne declare d'identifiant >> : re-mesure le
05/09/2026, **les vingt declarent `identifiant_contrat` ET `echeance`**. Le
repli ne vient donc plus d'un plan incomplet -- il vient de ce qu'**aucun
appelant de production ne passe le plan a `A1.run`** (releve par AST :
`pipeline_agents.py:346`, les demos, les scripts). *L'information existe, elle
n'atteint pas l'agent qui en a besoin.*

⚠️⚠️ LE CABLAGE N'EST PAS FAIT, ET C'EST DELIBERE. Passer le plan ferait
dedoublonner sur (identifiant, echeance) au lieu du seul identifiant : mesure
portee par l'arbre de decision -- sur un historique de renouvellement de 200
contrats sur 3 exercices, **400 faux doublons (66,7 %) et statut ROUGE sans le
plan, contre 0 et VERT avec**. Cela change les lignes retenues, donc le
modele, donc le prix. **C'est une decision, pas un correctif.** Ce lot rend le
fait VISIBLE ; il ne le tranche pas.

⚠️ ET LE STATUT N'EST PAS PLAFONNE, POUR UNE RAISON MESUREE. Aucun appelant de
production ne passant le plan, plafonner mettrait TOUS les fichiers en AMBRE :
*un avertissement permanent est un avertissement qu'on cesse de lire* -- le
meme principe qui gouverne `phrase_unite_non_declaree`. Le badge de la LIGNE
d'identite, lui, suit la source.

Ce que cette sentinelle exige :
  ID-1  la note d'identite atteint le classeur A1 SIGNE ;
  ID-2  son badge suit la SOURCE : declaree -> VERT, devinee -> AMBRE ;
  ID-3  la granularite du dedoublonnage est publiee ;
  ID-4  second sens : plan transmis -> plus aucune mention de << devinee >> ;
  ID-5  la note DIT quoi faire, et ne se contredit pas avec le referentiel.

Tout en `unittest.TestCase` : la gate lance `unittest discover`.
"""
import io
import logging
import os
import pathlib
import sys
import tempfile
import unittest
import warnings

_ICI = os.path.dirname(os.path.abspath(__file__))
_RACINE = os.path.dirname(os.path.dirname(_ICI))
for _c in (_RACINE, _ICI):
    if _c not in sys.path:
        sys.path.insert(0, _c)

import numpy as np


def _run_a1(**kwargs):
    from direction_non_vie.tarification import test_pipeline_agents as T
    from direction_non_vie.tarification.a1_ingestion.agent import (
        AgentA1Ingestion,
    )
    np.random.seed(7)
    return AgentA1Ingestion(audit_path='/tmp', verbose=False).run(
        branche='non_vie', sous_branche='auto',
        dataframe=T._portefeuille_auto(900), **kwargs)


def _texte_classeur(result_a1):
    from openpyxl import load_workbook
    octets = result_a1.get('excel_bytes') or b''
    if not octets:
        return ''
    classeur = load_workbook(io.BytesIO(octets), data_only=True)
    return '\n'.join(str(c.value) for f in classeur.worksheets
                     for ligne in f.iter_rows() for c in ligne
                     if c.value is not None)


class TestLIdentiteAtteintLeClasseur(unittest.TestCase):

    @classmethod
    def setUpClass(cls):
        warnings.filterwarnings('ignore')
        logging.disable(logging.CRITICAL)
        from direction_non_vie.tarification import test_pipeline_agents as T
        cls.sans = _run_a1()                      # le cas de PRODUCTION
        cls.avec = _run_a1(plan=T._PLAN_AUTO)

    @classmethod
    def tearDownClass(cls):
        logging.disable(logging.NOTSET)

    def test_ID1_la_note_d_identite_atteint_le_classeur_SIGNE(self):
        """⚠️⚠️ *Un calcul qui n'atteint aucun livrable n'existe pas.*

        ⚠️ `cls.sans` N'EST PLUS << LE CAS DE PRODUCTION >> DEPUIS LE
        06/09/2026 : `pipeline_agents` transmet desormais le plan. Il reste le
        cas d'un appelant qui ne le transmet pas -- dont `actuaria_app.py`,
        ferme par decision. Ce test verifie donc que le REPLI reste correct,
        pas qu'il est ce que la production fait. Voir `ID-7`, qui mesure les
        appelants reels.
        """
        qualite = self.sans.get('qualite') or {}
        self.assertEqual(qualite.get('source_identifiant'), 'devinee',
                         "sans plan, A1 devrait DEVINER l'identite : ce test "
                         'ne mesure plus le bon chemin')
        texte = _texte_classeur(self.sans)
        self.assertTrue(texte, 'aucun classeur A1')
        self.assertIn('Identité du contrat', texte,
                      "le classeur publie un compte de doublons sans dire sur "
                      'quelle CLE il porte')
        self.assertIn('DEVINÉE', texte)

    def test_ID2_le_badge_suit_la_SOURCE_pas_le_compte(self):
        """⚠️ Une identite DECLAREE n'est pas un fait a verifier ; une
        identite DEVINEE en est un. Le badge doit distinguer les deux."""
        from openpyxl import load_workbook
        for resultat, attendu in ((self.sans, 'devinee'), (self.avec, 'plan')):
            with self.subTest(source=attendu):
                qualite = resultat.get('qualite') or {}
                self.assertEqual(qualite.get('source_identifiant'), attendu)
                classeur = load_workbook(
                    io.BytesIO(resultat['excel_bytes']), data_only=True)
                trouve = None
                for feuille in classeur.worksheets:
                    for ligne in feuille.iter_rows():
                        for cellule in ligne:
                            if str(cellule.value or '').startswith(
                                    'Identité du contrat'):
                                trouve = feuille.cell(row=cellule.row,
                                                      column=3).value
                self.assertIsNotNone(
                    trouve, "la ligne d'identite n'est pas dans le classeur")
                # ⚠️⚠️ L'ATTENDU EST LE MOT PUBLIE, PAS LE CODE INTERNE.
                # `_kpi` rend « ✓ Conforme » / « △ À surveiller » -- c'est ce
                # que l'actuaire LIT. Et il se DERIVE de sa source unique
                # (`_MOT_RAG`, `GLYPHE_RAG_EXCEL`) : le recopier ici en
                # creerait une seconde, qui divergerait le jour ou la charte
                # change. *Mon premier attendu etait `VERT` -- le classeur ne
                # l'ecrit nulle part.*
                from direction_non_vie.tarification.services.excel_helpers import (
                    MOT_RAG,
                )
                code = 'VERT' if attendu == 'plan' else 'AMBRE'
                self.assertIn(
                    MOT_RAG[code], str(trouve),
                    f"identite {attendu} : le badge publie est {trouve!r}, "
                    f"attendu un badge {code}")

    def test_ID3_la_granularite_du_dedoublonnage_est_publiee(self):
        for resultat in (self.sans, self.avec):
            with self.subTest():
                self.assertIn('Granularité déclarée',
                              _texte_classeur(resultat))
                self.assertTrue((resultat.get('qualite') or {})
                                .get('granularite'))

    def test_ID4_SECOND_SENS_plan_transmis_plus_aucune_mention_devinee(self):
        """⚠️ Sans ce sens, on aurait pu ecrire << DEVINEE >> en dur."""
        texte = _texte_classeur(self.avec)
        self.assertNotIn('DEVINÉE', texte,
                         "le plan est transmis et le classeur dit encore que "
                         "l'identite est devinee")
        self.assertIn('DÉCLARÉE au plan', texte)

    def test_ID5_la_note_dit_QUOI_FAIRE_et_ne_ment_pas_sur_le_referentiel(self):
        """⚠️⚠️ LE TEXTE SE RELIT QUAND LE FAIT CHANGE. La note disait
        << Non declaree au plan >> ; re-mesure le 05/09/2026, **les vingt
        plans declarent `identifiant_contrat` ET `echeance`**. Le repli ne
        vient donc pas d'un plan incomplet mais d'un plan NON TRANSMIS --
        affirmer le contraire enverrait l'actuaire corriger un fichier qui
        n'a rien."""
        note = (self.sans.get('qualite') or {}).get('note_identite') or ''
        self.assertIn('DEVINÉE', note)
        self.assertIn("n'a pas été transmis", note,
                      'la note attribue le repli au plan alors que le plan '
                      'porte deja l information')
        self.assertIn('échéance', note.lower() + note)


class TestLeReferentielPorteBienLInformation(unittest.TestCase):
    """⚠️ Ce test FIGE la mesure sur laquelle repose la note ci-dessus. Le
    jour ou un plan cesse de declarer son identifiant, la note redevient
    fausse -- et c'est ici qu'on l'apprend."""

    def test_ID6_les_vingt_plans_declarent_identifiant_ET_echeance(self):
        import pathlib

        import yaml
        racine = pathlib.Path(_RACINE)
        lus = sans_id = sans_ech = 0
        manquants = []
        illisibles = []
        for fichier in sorted(racine.rglob('*.y*ml')):
            chemin = fichier.as_posix()
            if '.venv' in chemin or 'plan' not in chemin.lower():
                continue
            try:
                contenu = yaml.safe_load(fichier.read_text(encoding='utf-8'))
            except Exception as erreur:                      # noqa: BLE001
                # ⚠️⚠️ UN PLAN ILLISIBLE NE S'IGNORE PAS. Ma premiere version
                # faisait `except: continue` : un YAML casse aurait fait
                # BAISSER le compte de plans, donc rendu ce controle plus
                # facile a satisfaire. *Un test qui avale l'erreur qu'il
                # devrait voir se rend lui-meme complaisant.*
                illisibles.append(f'{chemin} ({type(erreur).__name__})')
                continue
            if not isinstance(contenu, dict) or 'cible_frequence' not in contenu:
                continue
            lus += 1
            if not contenu.get('identifiant_contrat'):
                sans_id += 1
                manquants.append(f"{contenu.get('lob', fichier.stem)}:id")
            if not contenu.get('echeance'):
                sans_ech += 1
                manquants.append(f"{contenu.get('lob', fichier.stem)}:echeance")
        self.assertEqual(illisibles, [],
                         'des fichiers de plan sont illisibles : le compte '
                         'de plans ci-dessous est donc incomplet')
        self.assertGreaterEqual(lus, 20, 'le referentiel de plans a retreci')
        self.assertEqual(
            manquants, [],
            f"{sans_id} plan(s) sans `identifiant_contrat` et {sans_ech} sans "
            f"`echeance` : la note d'identite d'A1 affirme que les vingt les "
            f"declarent. Corrigez le plan, ou corrigez la note.")


class TestQuiTransmetLePlanAA1(unittest.TestCase):
    """⚠️⚠️ LE CONSTAT `A1.5`, DEVENU CONTROLE — lot 1, 06/09/2026.

    A1 etait le SEUL des six agents dont `run()` ne recevait pas le plan,
    c'est-a-dire exactement l'agent qui DEVINAIT l'identite du contrat. Le
    mecanisme existait depuis toujours et les vingt plans declarent
    `identifiant_contrat` ET `echeance` (`ID-6`) : *l'information existait,
    elle n'atteignait pas l'agent qui en avait besoin.*

    ⚠️ Releve PAR AST, jamais au grep : un alias ou un `self.` ne se voit pas
    au texte.
    """

    #: Les appelants de production de `A1.run`, et ce qu'on exige de chacun.
    #: ⛔ `actuaria_app.py` est FERME par decision : il ne transmet pas le
    #: plan, et ce fait est ATTENDU ici plutot que decouvert plus tard.
    #: *Nommer ce qu'un lot ne couvre pas vaut mieux que de laisser croire
    #: qu'il a tout pris.*
    _ATTENDU = {
        'direction_non_vie/tarification/pipeline_agents.py': True,
        'actuaria_app.py': False,
    }

    @staticmethod
    def _analyser(source):
        """Les appels a `A1.run` d'UN source, et s'ils transmettent le plan.

        ⚠️ Isole du parcours de fichiers pour etre testable sur un extrait en
        memoire (`ID-7b`) : *un detecteur dont on n'a jamais vu le second sens
        peut rendre `False` par accident.*
        """
        import ast
        try:
            arbre = ast.parse(source)
        except SyntaxError:                         # pragma: no cover
            return []
        trouves = []
        for n in ast.walk(arbre):
            if not (isinstance(n, ast.Call)
                    and isinstance(n.func, ast.Attribute)
                    and n.func.attr == 'run'):
                continue
            cible = ast.unparse(n.func.value)
            # ⚠️ On accepte l'appel direct `AgentA1Ingestion(...).run(...)`
            # ET la variable `a1.run(...)` : deux formes, un seul constat.
            if 'AgentA1Ingestion' not in cible and cible not in ('a1',):
                continue
            trouves.append({
                'ligne': n.lineno,
                'plan': any(k.arg == 'plan' for k in (n.keywords or [])),
            })
        return trouves

    @classmethod
    def _appels_a1(cls):
        """Les appels a `A1.run` de tout le depot, par AST."""
        racine = pathlib.Path(_RACINE)
        sites = []
        for chemin in sorted(racine.rglob('*.py')):
            s = chemin.as_posix()
            if ('/.venv/' in s or '/audit_2026_08/' in s
                    or chemin.name.startswith('test_')):
                continue
            source = chemin.read_text(encoding='utf-8', errors='replace')
            if 'AgentA1Ingestion' not in source:
                continue
            for t in cls._analyser(source):
                sites.append({'fichier': s[len(racine.as_posix()) + 1:], **t})
        return sites

    def test_ID7b_le_detecteur_DISTINGUE_les_deux_formes(self):
        """⚠️⚠️ LE SECOND SENS DU DETECTEUR, SUR DES EXTRAITS EN MEMOIRE.

        `ID-7` affirme que l'app NE transmet PAS le plan. Cette affirmation
        serait vraie sans rien mesurer si le detecteur rendait toujours
        `False`. On le lui fait dire dans les deux sens -- sans ecrire une
        ligne dans `actuaria_app.py`, **fermee par decision**.
        """
        cas = (
            ("r = AgentA1Ingestion(audit_path='x').run(branche='b')", False),
            ("r = AgentA1Ingestion(audit_path='x').run(branche='b', "
             "plan=plan)", True),
            ("r = a1.run(branche='b', dataframe=df)", False),
            ("r = a1.run(branche='b', dataframe=df, plan=p)", True),
        )
        for source, attendu in cas:
            with self.subTest(extrait=source[:44]):
                trouves = self._analyser(source)
                self.assertEqual(len(trouves), 1,
                                 f'le detecteur ne voit pas l appel : {source}')
                self.assertEqual(
                    trouves[0]['plan'], attendu,
                    f'le detecteur rend {trouves[0]["plan"]} au lieu de '
                    f'{attendu} sur : {source}')

    def test_ID7_le_plan_atteint_A1_par_le_pipeline_et_PAS_par_l_app(self):
        """⚠️⚠️ LES DEUX SENS DANS UN SEUL CONTROLE. Il exige que
        `pipeline_agents` transmette le plan, ET que l'app ne le fasse pas --
        parce qu'elle est fermee. Si l'un des deux change, ce test le dit
        plutot que de laisser la situation deriver en silence."""
        sites = self._appels_a1()
        self.assertGreater(len(sites), 0,
                           'aucun appel a `A1.run` trouve : le releve AST ne '
                           'traverse plus rien')
        mesure = {}
        for site in sites:
            mesure.setdefault(site['fichier'], []).append(site['plan'])
        for fichier, attendu in self._ATTENDU.items():
            with self.subTest(appelant=fichier):
                self.assertIn(
                    fichier, mesure,
                    f'{fichier} n appelle plus `A1.run` : cette liste doit '
                    f'etre relue')
                for transmet in mesure[fichier]:
                    if attendu:
                        self.assertTrue(
                            transmet,
                            f'{fichier} ne transmet PAS le plan a A1 : '
                            f'l identite du contrat y redevient DEVINEE, et '
                            f'un historique de renouvellement repasse en '
                            f'ROUGE sur 66,67 % de faux doublons')
                    else:
                        self.assertFalse(
                            transmet,
                            f'{fichier} transmet desormais le plan a A1. '
                            f'C est peut-etre une bonne chose -- mais l app '
                            f'est FERMEE par decision, donc ce changement n a '
                            f'pas ete arbitre. Mettre `_ATTENDU` a jour APRES '
                            f'l avoir fait arbitrer.')

    def test_ID8_AUCUNE_ligne_ne_disparait_quand_le_plan_est_transmis(self):
        """⚠️⚠️ LA CRAINTE ECRITE DANS `a1` ETAIT FAUSSE, ET ON LA MESURE.

        Le code disait : << passer le plan ferait dedoublonner sur
        (identifiant, echeance) au lieu du seul identifiant, ce qui change les
        lignes retenues, donc le modele, donc le prix >>. **A1 SCORE la
        qualite, il n'exclut rien** -- c'est ecrit dans le meme fichier. Le
        statut bascule ; aucune ligne ne bouge.
        """
        import pandas as pd

        from core.plan_tarifaire import PlanTarifaire
        from direction_non_vie.tarification import test_plan_invariants as T
        plan = PlanTarifaire.depuis_yaml(
            os.path.join(_RACINE, 'plans', 'auto.yaml'))
        col_id, col_ech = plan.identifiant_contrat, plan.echeance
        np.random.seed(7)
        base = T.portefeuille_auto(400, 1)
        morceaux = []
        for an in (2023, 2024, 2025):
            m = base.copy()
            m[col_id] = [f'P{n:06d}' for n in range(len(m))]
            m[col_ech] = pd.Timestamp(f'{an}-01-01')
            morceaux.append(m)
        hist = pd.concat(morceaux, ignore_index=True)

        from direction_non_vie.tarification.a1_ingestion.agent import (
            AgentA1Ingestion,
        )
        agent = AgentA1Ingestion(audit_path=tempfile.mkdtemp(), verbose=False)
        sans = agent.run(branche='non_vie', sous_branche='auto',
                         dataframe=hist)
        avec = agent.run(branche='non_vie', sous_branche='auto',
                         dataframe=hist, plan=plan)
        # ⚠️ La fixture doit TRAVERSER le cas : sans historique, les deux
        # cotes seraient VERTS et ce test ne prouverait rien.
        self.assertEqual((sans.get('qualite') or {}).get('source_identifiant'),
                         'devinee')
        self.assertEqual((avec.get('qualite') or {}).get('source_identifiant'),
                         'plan')
        self.assertEqual(sans.get('statut_rag'), 'ROUGE',
                         'la fixture ne declenche plus le ROUGE : elle ne '
                         'porte plus d historique de renouvellement')
        self.assertEqual(avec.get('statut_rag'), 'VERT')
        # ⚠️⚠️ ET LE POINT QUI COMPTE : AUCUNE LIGNE NE DISPARAIT.
        self.assertEqual(
            len(sans['dataframe']), len(hist),
            'A1 a EXCLU des lignes sans le plan : il devait seulement SCORER')
        self.assertEqual(
            len(avec['dataframe']), len(hist),
            'A1 a EXCLU des lignes AVEC le plan : le statut devait basculer '
            'sans qu aucune ligne ne bouge -- un prix pourrait avoir change')


if __name__ == '__main__':
    unittest.main(verbosity=2)
