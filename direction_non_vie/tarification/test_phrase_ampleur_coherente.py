# -*- coding: utf-8 -*-
"""DEUX PHRASES CONTRADICTOIRES DANS LE DOCUMENT QU'ON SIGNE.

`phrase_ampleur_exclusion` affirmait, SANS CONDITION, que << ces retraits-ci
ne declenchent PAS cette exigence >>. Or on n'atteint ce point que dans deux
etats -- la synthese s'arrete plus haut sur << CONTROLE QUALITE BLOQUE >>
quand l'escalade n'a PAS ete validee :

  · l'escalade ne s'est pas declenchee            -> la phrase est vraie ;
  · elle S'EST declenchee et un actuaire l'a validee -> la phrase est FAUSSE.

MESURE DU 05/09/2026, sur un portefeuille reduit de 67,3 % :

    ⚠ AMPLEUR — cette part (67,3 %) vaut 13,5 fois le seuil de 5 % ...
      Ces retraits-ci ne declenchent PAS cette exigence ...
    ✔ Poursuite malgre anomalie(s) >= 5% VALIDEE par « Selasse Sekle ».

*L'une dit que l'exigence ne s'est pas declenchee, l'autre qu'elle s'est
declenchee et qu'un actuaire nomme l'a assumee -- dans le meme document, sous
les yeux de celui qui va signer une assiette reduite des deux tiers.*

⚠️⚠️ ET UN TEST EPINGLAIT LA PHRASE FAUSSE. `test_liste_disqualifiante`
exigeait `assertIn('ne declenchent PAS', p)` SANS CONDITION : tant qu'il
tenait, la correction etait interdite par la gate. Il a donc ete retourne
EN PREMIER -- il exige desormais la phrase qui CORRESPOND A L'ETAT.

Ce que cette sentinelle exige :
  AC-1  jamais les deux affirmations ensemble, dans AUCUN etat ;
  AC-2  sans escalade, la phrase dit que l'exigence ne s'est pas declenchee ;
  AC-3  sous escalade, elle dit qu'elle S'EST declenchee et NOMME qui l'a
        assumee ;
  AC-4  l'etat est OBLIGATOIRE : un defaut reproduirait la phrase fausse des
        qu'un appelant l'oublierait ;
  AC-5  ce qui NE dependait pas de l'etat n'a pas bouge (ratio, seuil,
        assiette restante, la question rendue) ;
  AC-6  et cela atteint les TROIS surfaces signees qui publient la synthese.

Tout en `unittest.TestCase` : la gate lance `unittest discover`.
"""
import dataclasses
import io
import os
import sys
import unittest
import zipfile

_ICI = os.path.dirname(os.path.abspath(__file__))
_RACINE = os.path.dirname(os.path.dirname(_ICI))
for _c in (_RACINE, _ICI):
    if _c not in sys.path:
        sys.path.insert(0, _c)

from core.qualite_donnees import (
    SEUIL_ESCALADE,
    Anomalie,
    RapportQualite,
    phrase_ampleur_exclusion,
    synthese_qualite_donnees,
)

_NIE = 'ne declenchent PAS'
_AFFIRME = "S'EST declenchee"


def _anomalie(n, total):
    champs = {f.name for f in dataclasses.fields(Anomalie)}
    valeurs = {'code': 'doublon_identifiant', 'regle': 'doublon',
               'role': 'identifiant', 'colonne': 'id_contrat',
               'nb_lignes': n, 'proportion': n / total,
               'index': tuple(range(n)),
               'description': 'identifiant en double', 'correction': '',
               'effet_agrege': None, 'resume_charges_negatives': None}
    return Anomalie(**{k: v for k, v in valeurs.items() if k in champs})


def _rapport(n=6730, total=10000, *, escalade, valide_par):
    champs = {f.name for f in dataclasses.fields(RapportQualite)}
    valeurs = {
        'lignes_initiales': total, 'lignes_retenues': total - n,
        'exclusions': [_anomalie(n, total)], 'corrections': [],
        'signalements': [], 'escalade_declenchee': escalade,
        'anomalies_au_dela_seuil': ['doublon_identifiant'] if escalade else [],
        'seuil': SEUIL_ESCALADE, 'validee_par': valide_par,
        'horodatage': None, 'bloque': False, 'dataframe_propre': None}
    return RapportQualite(**{k: v for k, v in valeurs.items() if k in champs})


class TestJamaisLesDeux(unittest.TestCase):

    def test_AC1_aucun_etat_ne_produit_les_DEUX_affirmations(self):
        """⚠️⚠️ LE COEUR. La synthese ne peut jamais nier et affirmer la meme
        escalade."""
        for escalade, qui in ((False, None), (True, 'Selasse Sekle'),
                              (True, None)):
            with self.subTest(escalade=escalade, valide_par=qui):
                texte = synthese_qualite_donnees(
                    _rapport(escalade=escalade, valide_par=qui)) or ''
                nie = _NIE in texte
                valide = 'VALIDEE par' in texte
                self.assertFalse(
                    nie and valide,
                    "la synthese NIE l'escalade et la declare VALIDEE dans le "
                    f'meme document :\n{texte[:400]}')

    def test_AC2_sans_escalade_la_phrase_NIE_correctement(self):
        texte = synthese_qualite_donnees(
            _rapport(escalade=False, valide_par=None)) or ''
        self.assertIn('AMPLEUR', texte, "la phrase n'atteint pas la synthese")
        self.assertIn(_NIE, texte)
        self.assertNotIn(_AFFIRME, texte)

    def test_AC3_sous_escalade_elle_AFFIRME_et_NOMME(self):
        texte = synthese_qualite_donnees(
            _rapport(escalade=True, valide_par='Selasse Sekle')) or ''
        self.assertIn(_AFFIRME, texte)
        self.assertNotIn(_NIE, texte)
        self.assertIn('Selasse Sekle', texte,
                      "la phrase d'ampleur ne nomme pas qui a assume "
                      "l'escalade")

    def test_AC3b_une_escalade_validee_par_PERSONNE_le_dit(self):
        """⚠️ Un cas qui ne devrait pas exister -- la synthese bloque plus haut
        sans validation -- mais que la fonction doit savoir ecrire sans
        inventer un nom."""
        phrase = phrase_ampleur_exclusion(0.673, escalade_declenchee=True,
                                          validee_par=None) or ''
        self.assertIn(_AFFIRME, phrase)
        self.assertIn('non nomme', phrase,
                      "l'absence de signataire est passee sous silence")

    def test_AC4_l_etat_est_OBLIGATOIRE(self):
        """⚠️⚠️ Un defaut a `False` reproduirait la phrase FAUSSE des qu'un
        appelant oublierait de passer l'etat -- exactement la famille de
        defauts que ce chantier ferme."""
        with self.assertRaises(TypeError):
            phrase_ampleur_exclusion(0.673)
        with self.assertRaises(TypeError):
            phrase_ampleur_exclusion(0.673, escalade_declenchee=True)


class TestCeQuiNeDevaitPasBouger(unittest.TestCase):
    """⚠️ *Chaque test doit prouver EXACTEMENT ce qu'il prouvait.* Le reste de
    la phrase ne depend pas de l'etat, et ne doit pas avoir change."""

    def test_AC5_le_ratio_le_seuil_et_l_assiette_restante_tiennent(self):
        for escalade, qui in ((False, None), (True, 'X')):
            with self.subTest(escalade=escalade):
                phrase = phrase_ampleur_exclusion(
                    0.226, escalade_declenchee=escalade,
                    validee_par=qui) or ''
                self.assertIn('4,5 fois', phrase)
                self.assertIn(f"{SEUIL_ESCALADE:.0%}".replace('%', ' %'),
                              phrase)
                self.assertIn('77,4 %', phrase,
                              "l'assiette restante n'est plus publiee")
                self.assertIn('VERIFIEZ', phrase,
                              "la question n'est plus rendue a l'actuaire")

    def test_AC5b_sous_le_repere_elle_se_TAIT_toujours(self):
        for escalade in (False, True):
            self.assertIsNone(phrase_ampleur_exclusion(
                0.001, escalade_declenchee=escalade, validee_par='X'))
            self.assertIsNone(phrase_ampleur_exclusion(
                None, escalade_declenchee=escalade, validee_par='X'))

    def test_AC5c_elle_ne_JUGE_toujours_pas(self):
        for escalade, qui in ((False, None), (True, 'X')):
            phrase = (phrase_ampleur_exclusion(
                0.673, escalade_declenchee=escalade, validee_par=qui)
                or '').lower()
            for interdit in ('non conforme', 'invalide', 'inacceptable',
                             'refuse'):
                self.assertNotIn(interdit, phrase,
                                 f'la phrase JUGE (<< {interdit} >>)')


class TestLesSurfacesSignees(unittest.TestCase):
    """⚠️⚠️ *Le correctif doit atteindre la surface, pas la froler.* Trois
    surfaces publient la synthese qualite : le classeur A6, le Word/HTML A6 et
    le rapport d'equipe."""

    def _resultat_a6(self, *, escalade, qui):
        return {
            'success': True, 'statut_rag': 'AMBRE', 'branche': 'auto',
            'rapport_qualite': _rapport(escalade=escalade, valide_par=qui),
            'classement': [{'modele': 'GLM_TWEEDIE', 'famille': 'GLM',
                            'gini_test': 0.19, 'rmse_test': 0.71,
                            'overfit_ratio': 0.97, 'score_global': 1.0,
                            'interpretabilite': 1.0}],
            'modele_production': {'modele': 'GLM_TWEEDIE', 'famille': 'GLM',
                                  'gini_test': 0.19, 'score_global': 1.0,
                                  'interpretabilite': 1.0},
            'metriques': {}, 'hypotheses': {}, 'audit_trail': {},
            'backtest': {}, 'commentaire': '',
        }

    def _texte_xlsx(self, octets):
        from openpyxl import load_workbook
        classeur = load_workbook(io.BytesIO(octets), data_only=True)
        return '\n'.join(str(c.value) for f in classeur.worksheets
                         for ligne in f.iter_rows() for c in ligne
                         if c.value is not None)

    def test_AC6_le_classeur_A6_porte_la_phrase_CORRESPONDANT_a_l_etat(self):
        from direction_non_vie.tarification.services.tarif_excel import (
            export_excel_a6,
        )
        for escalade, attendu, interdit in ((False, _NIE, _AFFIRME),
                                            (True, _AFFIRME, _NIE)):
            with self.subTest(escalade=escalade):
                octets = export_excel_a6(self._resultat_a6(
                    escalade=escalade, qui='Selasse Sekle'))
                self.assertTrue(octets, "l'export A6 n'a rien produit")
                texte = self._texte_xlsx(octets)
                self.assertIn('AMPLEUR', texte,
                              "la phrase n'atteint pas le classeur signe")
                self.assertIn(attendu, texte)
                self.assertNotIn(interdit, texte)

    def test_AC6b_le_rapport_equipe_porte_la_meme_phrase(self):
        from direction_non_vie.tarification.services.rapport_equipe_tarif import (
            export_html_equipe,
        )
        for escalade, attendu, interdit in ((False, _NIE, _AFFIRME),
                                            (True, _AFFIRME, _NIE)):
            with self.subTest(escalade=escalade):
                html = export_html_equipe(
                    {'a6': self._resultat_a6(escalade=escalade, qui='S. S.')},
                    branche='auto', audit_id='TEST')
                texte = html if isinstance(html, str) else html.decode(
                    'utf-8', 'replace')
                self.assertIn('AMPLEUR', texte,
                              "la phrase n'atteint pas le rapport d'equipe")
                self.assertIn(attendu, texte)
                self.assertNotIn(interdit, texte)

    def test_AC6c_aucune_surface_ne_porte_les_DEUX(self):
        """La contre-epreuve d'assiette : on cherche la contradiction dans les
        OCTETS PRODUITS, pas dans le code."""
        from direction_non_vie.tarification.services.rapport_equipe_tarif import (
            export_word_equipe,
        )
        octets = export_word_equipe(
            {'a6': self._resultat_a6(escalade=True, qui='Selasse Sekle')},
            branche='auto', audit_id='TEST')
        with zipfile.ZipFile(io.BytesIO(octets)) as z:
            xml = z.read('word/document.xml').decode('utf-8', 'replace')
        self.assertFalse(_NIE in xml and 'VALIDEE par' in xml,
                         'le Word signe NIE et VALIDE la meme escalade')


if __name__ == '__main__':
    unittest.main(verbosity=2)
