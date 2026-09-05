# -*- coding: utf-8 -*-
"""UN AGENT QUI PLANTE NE DOIT PAS SORTIR DU STATUT CONSOLIDE.

Le rapport d'equipe part au commissaire aux comptes. Son bandeau
<< Statut consolide >> est ce qu'on lit en premier.

MESURE DU 05/09/2026, et les deux defauts allaient dans le sens RASSURANT --
celui qui ne se remarque pas :

  | scenario                                | consolide publie |
  |-----------------------------------------|------------------|
  | cinq agents VERT, le sixieme EXECUTE     | **VERT**         |
  | puis ECHOUE en ROUGE                     |                  |
  | les SIX agents en echec                  | **AMBRE**        |
  | un agent echoue sans prononcer de statut | **VERT**         |

La cause : `_collecter_statuts` filtrait sur `success`. Un agent en echec
sortait de la liste -- et du tableau par agent, sans meme un << N/A >>. Son
ROUGE n'atteignait jamais `_statut_global`. Et la liste vide retombait sur
un `return 'AMBRE'` : un rapport ou rien n'a tourne n'est pas << a
surveiller >>, il n'a pas de statut.

*Un echec qui ne laisse aucune trace dans le livrable est indiscernable
d'un succes.*

Trois etats, et ils ne se confondent plus :
    absent / vide  -> NON FOURNI     (l'agent n'a pas tourne)
    statut publie  -> ce statut
    aucun statut   -> NON DETERMINE  (il a tourne, il n'a rien dit)

Ce que cette sentinelle exige :
  SC-1  un agent EXECUTE ET ECHOUE en ROUGE rend le consolide ROUGE ;
  SC-2  il FIGURE dans le tableau par agent, avec son statut ;
  SC-3  << non fourni >> n'abaisse rien -- le Deep Learning est facultatif ;
  SC-4  aucun agent -> NON DETERMINE, jamais un RAG par defaut ;
  SC-5  un agent muet PLAFONNE a AMBRE : on ne certifie pas sur un silence ;
  SC-6  et tout cela ATTEINT LES TROIS FORMATS SIGNES, pas seulement la
        fonction -- Excel, HTML et Word passent tous les trois par la.

Tout en `unittest.TestCase` : la gate lance `unittest discover`.
"""
import io
import os
import sys
import unittest
import zipfile

_ICI = os.path.dirname(os.path.abspath(__file__))
_RACINE = os.path.dirname(os.path.dirname(_ICI))
for _c in (_RACINE, _ICI):
    if _c not in sys.path:
        sys.path.insert(0, _c)

from direction_non_vie.tarification.services.rapport_equipe_tarif import (
    STATUT_NON_DETERMINE,
    STATUT_NON_FOURNI,
    _collecter_statuts,
    _statut_global,
    export_excel_equipe,
    export_html_equipe,
    export_word_equipe,
)


def _verts():
    return {f'a{i}': {'success': True, 'statut_rag': 'VERT',
                      'commentaire': ''} for i in range(1, 7)}


def _consolide(results):
    return _statut_global(list(_collecter_statuts(results).values()))


def _texte_docx(octets):
    with zipfile.ZipFile(io.BytesIO(octets)) as z:
        return z.read('word/document.xml').decode('utf-8', 'replace')


def _texte_xlsx(octets):
    from openpyxl import load_workbook
    classeur = load_workbook(io.BytesIO(octets), data_only=True)
    morceaux = []
    for feuille in classeur.worksheets:
        for ligne in feuille.iter_rows():
            for cellule in ligne:
                if cellule.value is not None:
                    morceaux.append(str(cellule.value))
    return '\n'.join(morceaux)


class TestStatutConsolide(unittest.TestCase):

    def test_SC1_un_agent_execute_et_echoue_rend_le_consolide_ROUGE(self):
        """⚠️⚠️ MESURE : ce cas donnait **VERT**."""
        results = _verts()
        results['a4'] = {'success': False, 'statut_rag': 'ROUGE',
                         'erreur': 'plantage du lecteur de classement'}
        self.assertEqual(_consolide(results), 'ROUGE')

    def test_SC2_l_agent_en_echec_FIGURE_dans_le_tableau_par_agent(self):
        """Il disparaissait entierement -- pas meme un << N/A >>."""
        results = _verts()
        results['a4'] = {'success': False, 'statut_rag': 'ROUGE'}
        table = _collecter_statuts(results)
        self.assertIn('a4', table, 'l agent en echec a disparu de la table')
        self.assertEqual(table['a4'], 'ROUGE')
        self.assertEqual(len(table), 6, table)

    def test_SC3_non_fourni_n_est_pas_un_echec(self):
        """Le Deep Learning est facultatif : ne pas le lancer n'est pas une
        anomalie, et ne doit rien abaisser."""
        results = _verts()
        results['a5'] = None
        table = _collecter_statuts(results)
        self.assertEqual(table['a5'], STATUT_NON_FOURNI)
        self.assertEqual(_statut_global(list(table.values())), 'VERT')

    def test_SC3c_TOUS_non_fournis_ne_certifient_rien(self):
        """⚠️⚠️ TROU TROUVE PAR LE SCEAU, PAS PAR LA RELECTURE. Le plant qui
        retirait le filtre `NON FOURNI` ne faisait rougir personne : avec UN
        SEUL agent absent, << NON FOURNI >> n'est ni ROUGE ni AMBRE, donc le
        consolide retombe sur VERT dans les deux cas, filtre ou pas.

        Le filtre ne decide que dans CE cas-ci -- et c'est le plus grave :
        sans lui, six agents non fournis publieraient **VERT**. Un rapport ou
        rien n'a ete produit certifierait que tout va bien.
        """
        aucun = {f'a{i}': None for i in range(1, 7)}
        table = _collecter_statuts(aucun)
        self.assertEqual(set(table.values()), {STATUT_NON_FOURNI}, table)
        self.assertEqual(_statut_global(list(table.values())),
                         STATUT_NON_DETERMINE,
                         'un rapport dont AUCUN agent na ete fourni se '
                         'prononce quand meme')

    def test_SC3b_non_fourni_et_echoue_ne_se_confondent_pas(self):
        """⚠️ LA DISTINCTION EST TOUT CE LOT. Avant, les deux cas donnaient
        exactement le meme resultat : l'agent absent de la table."""
        absent = _verts()
        absent['a4'] = None
        echoue = _verts()
        echoue['a4'] = {'success': False, 'statut_rag': 'ROUGE'}
        self.assertNotEqual(_consolide(absent), _consolide(echoue),
                            'un agent absent et un agent en echec donnent '
                            'encore le meme statut consolide')

    def test_SC4_aucun_agent_donne_NON_DETERMINE_pas_un_RAG(self):
        """⚠️⚠️ MESURE : les six agents en echec donnaient **AMBRE**, par le
        repli de la liste vide. Un rapport ou rien n'a tourne n'a pas de
        statut -- il ne se dit pas << a surveiller >>."""
        self.assertEqual(_statut_global([]), STATUT_NON_DETERMINE)
        self.assertNotIn(_statut_global([]), ('VERT', 'AMBRE', 'ROUGE'))
        # ... et six agents en echec REEL donnent bien ROUGE, pas AMBRE.
        tous = {f'a{i}': {'success': False, 'statut_rag': 'ROUGE'}
                for i in range(1, 7)}
        self.assertEqual(_consolide(tous), 'ROUGE')

    def test_SC5_un_agent_muet_PLAFONNE_a_AMBRE(self):
        """Il a tourne, il n'a rien dit : on ne certifie pas VERT dessus."""
        results = _verts()
        results['a4'] = {'success': False}          # aucun `statut_rag`
        table = _collecter_statuts(results)
        self.assertEqual(table['a4'], STATUT_NON_DETERMINE)
        self.assertEqual(_statut_global(list(table.values())), 'AMBRE')

    def test_SC7_un_agent_en_echec_ne_peut_pas_certifier_VERT(self):
        """⚠️⚠️ TROUVE PAR UNE FIXTURE ECRITE POUR UNE TOUT AUTRE RAISON.
        `test_c6_syntheses` porte `{'success': False, 'statut_rag': 'VERT'}` :
        un agent qui a plante TOUT EN declarant VERT. Une premiere version de
        ce lot le croyait sur parole et publiait VERT -- le correctif etait
        donc inutile dans le seul cas ou il compte. L'echec l'emporte."""
        results = _verts()
        results['a4'] = {'success': False, 'statut_rag': 'VERT'}
        table = _collecter_statuts(results)
        self.assertEqual(table['a4'], STATUT_NON_DETERMINE,
                         'un agent en echec certifie encore VERT')
        self.assertEqual(_statut_global(list(table.values())), 'AMBRE')

    def test_SC7b_un_agent_en_echec_qui_dit_ROUGE_garde_son_ROUGE(self):
        """La contre-epreuve : on ne remplace pas TOUT statut d'un agent en
        echec -- seulement celui qui certifie."""
        for declare, attendu in (('ROUGE', 'ROUGE'), ('AMBRE', 'AMBRE')):
            with self.subTest(declare=declare):
                results = _verts()
                results['a4'] = {'success': False, 'statut_rag': declare}
                self.assertEqual(_collecter_statuts(results)['a4'], attendu)

    def test_SC5b_un_statut_vide_est_traite_comme_un_silence(self):
        results = _verts()
        results['a4'] = {'success': True, 'statut_rag': '   '}
        self.assertEqual(_collecter_statuts(results)['a4'],
                         STATUT_NON_DETERMINE)


class TestAtteintLesTroisFormats(unittest.TestCase):
    """⚠️⚠️ LE CORRECTIF DOIT ATTEINDRE LA SURFACE SIGNEE, PAS LA FROLER.
    Les trois exportateurs appellent `_collecter_statuts` puis
    `_statut_global` -- on le VERIFIE en produisant les trois documents et en
    y cherchant le ROUGE, pas en le supposant."""

    @classmethod
    def setUpClass(cls):
        cls.results = _verts()
        cls.results['a4'] = {'success': False, 'statut_rag': 'ROUGE',
                             'erreur': 'plantage du lecteur de classement'}

    def test_SC6_le_classeur_signe_porte_le_ROUGE(self):
        octets = export_excel_equipe(self.results, branche='auto',
                                     audit_id='TEST')
        self.assertTrue(octets, "l'export Excel n'a rien produit")
        texte = _texte_xlsx(octets)
        self.assertIn('ROUGE', texte,
                      'le classeur equipe ne porte aucun ROUGE alors qu un '
                      'agent a echoue')

    def test_SC6b_le_HTML_signe_porte_le_ROUGE(self):
        html = export_html_equipe(self.results, branche='auto',
                                  audit_id='TEST')
        self.assertTrue(html, "l'export HTML n'a rien produit")
        self.assertIn('ROUGE', html if isinstance(html, str)
                      else html.decode('utf-8', 'replace'))

    def test_SC6c_le_Word_signe_porte_le_ROUGE(self):
        octets = export_word_equipe(self.results, branche='auto',
                                    audit_id='TEST')
        self.assertTrue(octets, "l'export Word n'a rien produit")
        self.assertIn('ROUGE', _texte_docx(octets))

    def test_SC6d_contre_epreuve_six_verts_ne_font_pas_un_ROUGE(self):
        """Sans cette contre-epreuve, les trois tests ci-dessus passeraient
        aussi sur un document qui ecrit << ROUGE >> dans une legende."""
        verts = _verts()
        html = export_html_equipe(verts, branche='auto', audit_id='TEST')
        texte = html if isinstance(html, str) else html.decode('utf-8', 'replace')
        self.assertNotIn('ROUGE', texte,
                         'le mot ROUGE figure dans un rapport ou tous les '
                         'agents sont VERT : les tests SC-6 ne prouvent rien')


if __name__ == '__main__':
    unittest.main(verbosity=2)
