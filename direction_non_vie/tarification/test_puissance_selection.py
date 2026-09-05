# -*- coding: utf-8 -*-
"""LA SELECTION PUBLIAIT SON RESULTAT, JAMAIS SA PUISSANCE (G.4 et G.11).

Mesure du 05/09/2026, rendue a l'actuaire signataire et arbitree. Sur le plan
`mrh`, la selection descendante retient UNE OU DEUX variables, et **elles
changent a chaque taille d'echantillon** :

    3 000 contrats (235 sinistres)  ->  `alarme` SEULE
    8 000 contrats (561 sinistres)  ->  les deux vraies variables
    20 000 contrats (1 406 sin.)    ->  les trois vraies

Or `alarme` est **du bruit pur** : le generateur la tire a pile ou face, sans
aucun effet sur la frequence. *Un tarif a zero facteur se voit ; un tarif
segmente sur `alarme` a l'air d'un vrai tarif.*

⚠️⚠️ ET LE SEUIL DE 5 % N'EST PAS EN CAUSE. Mesure du contre-test sur `auto` :
le relacher a 20 % achete UNE vraie variable (`puissance_fiscale`) pour TROIS
fausses (`csp_employe`, `csp_retraite`, `usage_enc`). **Recommandation rendue
et validee : ne pas toucher au 0,05 ; publier la PUISSANCE.** Ce module ne
change donc AUCUN seuil et AUCUN prix -- il publie trois nombres.

Ce que cette sentinelle exige :
  PS-1  la phrase se TAIT quand la selection retient assez de variables ;
  PS-2  elle PARLE a zero et a une seule retenue -- le cas << une seule >>
        est le cas trompeur, et c'est pour lui qu'elle existe ;
  PS-3  elle nomme les TROIS nombres : retenues, candidates, SINISTRES ;
  PS-4  le denominateur est le nombre de SINISTRES, jamais de contrats ;
  PS-5  la p-value de la meilleure REJETEE se DERIVE de `vars_exclues`, et
        un retrait NON TESTE n'y entre pas ;
  PS-6  les TROIS moteurs la publient -- pas d'asymetrie entre voisins ;
  PS-7  cela atteint le classeur A3 SIGNE ;
  PS-8  ⚠️ le seuil vient de l'APPELANT : une seconde source du 0,05 dans le
        socle serait le defaut meme que ce chantier ferme.

⚠️⚠️ G.11 -- CE QUI ETAIT DECRIT COMME UN SILENCE EST DEJA FERME. L'arbre de
decision dit du coefficient d'equilibre k : << aujourd'hui k reste a 1,0 en
silence >>. C'est vrai du CODE et **faux du COMPORTEMENT**, mesure le
05/09/2026 : les deux voies vers une somme nulle sont REFUSEES en amont, et la
somme predite sort d'un `exp()`, donc strictement positive. Il n'y avait donc
rien a declarer -- une declaration inatteignable serait un controle qui atteste
sans surveiller. `TestEquilibrageNonSilencieux` FIGE cette raison : retirer un
refus fait tomber le test.

Tout en `unittest.TestCase` : la gate lance `unittest discover`.
"""
import io
import logging
import os
import sys
import unittest
import warnings

_ICI = os.path.dirname(os.path.abspath(__file__))
_RACINE = os.path.dirname(os.path.dirname(_ICI))
for _c in (_RACINE, _ICI):
    if _c not in sys.path:
        sys.path.insert(0, _c)

import numpy as np

from core.conformite_reglementaire import (
    SELECTION_PEU_SEGMENTANTE,
    meilleure_rejetee,
    phrase_puissance_selection,
)

#: Les trois moteurs GLM d'A3. ⚠️ Les trois, pas un : ne poser la declaration
#: que sur l'un d'eux creerait l'asymetrie entre voisins que ce chantier
#: traque partout ailleurs.
_MOTEURS = ('poisson', 'gamma', 'tweedie')


class TestLaPhrase(unittest.TestCase):

    def test_PS1_elle_se_TAIT_quand_la_selection_retient_assez(self):
        """*Une phrase qui s'affiche toujours ne se lit plus.*"""
        self.assertIsNone(phrase_puissance_selection(
            SELECTION_PEU_SEGMENTANTE, 22, 563, seuil=0.05))
        self.assertIsNone(phrase_puissance_selection(6, 22, 441, seuil=0.05))
        # ⚠️ Et sur une absence de mesure, elle se tait aussi : elle ne
        # fabrique pas un verdict a partir d'un `None`.
        self.assertIsNone(phrase_puissance_selection(None, 22, 563,
                                                     seuil=0.05))

    def test_PS2_elle_PARLE_a_zero_ET_a_une_seule_retenue(self):
        """⚠️⚠️ LE CAS << UNE SEULE >> EST LE CAS TROMPEUR. Un tarif a zero
        facteur se voit ; un tarif segmente sur une variable de bruit a l'air
        d'un vrai tarif. C'est pour lui que la phrase existe."""
        zero = phrase_puissance_selection(0, 22, 563, seuil=0.05)
        une = phrase_puissance_selection(1, 22, 563, seuil=0.05,
                                         vars_retenues=['alarme'])
        for phrase in (zero, une):
            self.assertTrue(phrase, 'la selection ne se declare pas')
        self.assertIn('AUCUNE', zero)
        self.assertIn('UNE SEULE', une)
        self.assertIn('alarme', une,
                      "la variable retenue n'est pas nommee : l'actuaire ne "
                      "peut pas juger si c'est du bruit")

    def test_PS3_elle_nomme_les_TROIS_nombres(self):
        p = phrase_puissance_selection(1, 22, 563, seuil=0.05,
                                       vars_retenues=['jeune_conducteur'])
        self.assertIn('22', p, 'le nombre de candidates manque')
        self.assertIn('563', p, 'le nombre de sinistres manque')
        self.assertIn('5%', p.replace(' ', ''), 'le seuil manque')

    def test_PS4_le_denominateur_est_le_nombre_de_SINISTRES(self):
        """⚠️ 3 000 contrats a une frequence de 0,09 ne donnent que ~235
        sinistres. C'est le nombre de SINISTRES qui gouverne la puissance
        d'un GLM de frequence, jamais le nombre de contrats."""
        p = phrase_puissance_selection(1, 11, 235, seuil=0.05)
        self.assertIn('sinistres', p.lower())
        self.assertIn('235', p)
        self.assertNotIn('contrats', p.lower().split('sinistres')[0],
                         'la phrase annonce des contrats la ou elle compte '
                         'des sinistres')

    def test_PS4b_une_grandeur_absente_est_DITE_absente(self):
        """⚠️ Jamais remplacee par un nombre -- la regle de tout le module."""
        p = phrase_puissance_selection(1, None, None, seuil=0.05)
        self.assertTrue(p)
        self.assertNotIn('0 candidates', p)
        self.assertNotIn('0 sinistres', p)
        self.assertIn('non mesur', p.lower())

    def test_PS5_la_meilleure_rejetee_se_DERIVE_de_vars_exclues(self):
        exclues = [
            {'variable': 'a', 'pvalue': 0.42},
            {'variable': 'b', 'pvalue': 0.0969},
            {'variable': 'c', 'pvalue': 0.31},
        ]
        self.assertAlmostEqual(meilleure_rejetee(exclues), 0.0969)
        self.assertIn('0.0969', phrase_puissance_selection(
            1, 22, 563, seuil=0.05, pvalue_meilleure_rejetee=0.0969))

    def test_PS5b_un_retrait_NON_TESTE_n_entre_PAS_dans_la_derivation(self):
        """⚠️⚠️ Une variable ecartee parce que l'ajustement a ECHOUE porte
        `pvalue: None` et `pvalue_non_testee: True` (constat `a3/C14`). La
        confondre avec une p-value mesuree ferait dire a cette phrase quelque
        chose que PERSONNE n'a calcule."""
        self.assertIsNone(meilleure_rejetee([
            {'variable': 'x', 'pvalue': None, 'pvalue_non_testee': True}]))
        self.assertIsNone(meilleure_rejetee([]))
        self.assertIsNone(meilleure_rejetee(None))
        # ...et une p-value MESUREE reste vue meme melangee a un non-teste.
        self.assertAlmostEqual(meilleure_rejetee([
            {'variable': 'x', 'pvalue': None, 'pvalue_non_testee': True},
            {'variable': 'y', 'pvalue': 0.07}]), 0.07)

    def test_PS5c_le_DRAPEAU_prime_sur_la_valeur(self):
        """⚠️⚠️ CE TEST EXISTE PARCE QUE LE SCEAU L'A EXIGE. Le plant qui
        retirait `not e.get('pvalue_non_testee')` restait MUET : sur les
        entrees d'aujourd'hui, `pvalue` vaut `None` et le controle de type
        suffisait a les ecarter. Les deux garde-fous se recouvraient, donc un
        seul etait mesure.

        Ici la p-value est un NOMBRE et le drapeau dit qu'elle n'a pas ete
        testee. C'est le seul cas qui separe les deux, et il decide : *un
        nombre qu'on n'a pas mesure ne devient pas une mesure parce qu'il a
        la forme d'un nombre.*
        """
        self.assertIsNone(meilleure_rejetee([
            {'variable': 'x', 'pvalue': 0.01, 'pvalue_non_testee': True}]),
            "une p-value marquee NON TESTEE a ete prise pour une mesure")
        self.assertAlmostEqual(meilleure_rejetee([
            {'variable': 'x', 'pvalue': 0.01, 'pvalue_non_testee': True},
            {'variable': 'y', 'pvalue': 0.33}]), 0.33)

    def test_PS8_le_seuil_vient_de_L_APPELANT_pas_du_socle(self):
        """⚠️⚠️ `SEUIL_PVALUE` vit dans A3, et un test l'y fige. Lui donner un
        defaut dans le socle creerait une SECONDE SOURCE du meme nombre."""
        with self.assertRaises(TypeError):
            phrase_puissance_selection(1, 22, 563)      # sans `seuil=`
        # ...et le seuil PUBLIE est bien celui qu'on passe.
        self.assertIn('20%', phrase_puissance_selection(
            1, 22, 563, seuil=0.20).replace(' ', ''))


class TestA3LaPublie(unittest.TestCase):

    @classmethod
    def setUpClass(cls):
        warnings.filterwarnings('ignore')
        logging.disable(logging.CRITICAL)
        from core.qualite_donnees import preambule_qualite
        from direction_non_vie.tarification import test_pipeline_agents as T
        from direction_non_vie.tarification.a1_ingestion.agent import (
            AgentA1Ingestion,
        )
        from direction_non_vie.tarification.a2_preprocessing.agent import (
            AgentA2Preprocessing,
        )
        from direction_non_vie.tarification.a3_glm.agent import AgentA3GLM

        np.random.seed(7)
        donnees = T._portefeuille_auto(1500)
        cls.plan = T._PLAN_AUTO
        base = {'audit_path': '/tmp', 'verbose': False}
        r1 = AgentA1Ingestion(**base).run(branche='non_vie',
                                          sous_branche='auto',
                                          dataframe=donnees)
        qualite = preambule_qualite(r1.get('dataframe'), cls.plan,
                                    qualite_validee_par='Test',
                                    horodatage=None)
        r2 = AgentA2Preprocessing(**base).run(
            result_a1={**r1, 'dataframe': qualite.dataframe_propre},
            plan=cls.plan)
        cls.r3 = AgentA3GLM(models_path='/tmp', audit_path='/tmp').run(
            result_a2=r2, plan=cls.plan,
            col_frequence=cls.plan.cible_frequence,
            col_cout=cls.plan.cible_cout, generer_graphiques=False)

    @classmethod
    def tearDownClass(cls):
        logging.disable(logging.NOTSET)

    def test_PS6_les_TROIS_moteurs_publient_la_meme_chose(self):
        """⚠️⚠️ L'ASYMETRIE ENTRE VOISINS EST LE REVELATEUR LE MOINS CHER.
        Avant ce lot, le Tweedie ne publiait NI ses variables ecartees NI la
        taille de son jeu d'entrainement, la ou Poisson et Gamma publiaient
        les deux."""
        metriques = self.r3.get('metriques') or {}
        for moteur in _MOTEURS:
            with self.subTest(moteur=moteur):
                m = metriques.get(moteur)
                self.assertTrue(m, f'{moteur} absent des metriques')
                for cle in ('nb_vars_retenues', 'nb_vars_exclues',
                            'vars_exclues', 'nb_sinistres_train',
                            'puissance_selection'):
                    self.assertIn(cle, m, f'{moteur} ne publie pas {cle!r}')

    def test_PS4c_nb_sinistres_train_compte_des_SINISTRES(self):
        """⚠️ `nb_obs_train` compte des CONTRATS pour Poisson et Tweedie ; le
        confondre avec la puissance serait surestimer celle-ci d'un facteur
        1/frequence. Pour Gamma les deux coincident : il ne s'ajuste QUE sur
        les sinistres."""
        m = self.r3['metriques']
        for moteur in ('poisson', 'tweedie'):
            with self.subTest(moteur=moteur):
                self.assertLess(m[moteur]['nb_sinistres_train'],
                                m[moteur]['nb_obs_train'],
                                'autant de sinistres que de contrats : la '
                                'cle compte-t-elle vraiment des sinistres ?')
        self.assertEqual(m['gamma']['nb_sinistres_train'],
                         m['gamma']['nb_obs_train'])

    def test_PS2b_le_cas_TROMPEUR_se_produit_sur_la_fixture(self):
        """⚠️ Ce n'est pas un cas theorique : sur la fixture `auto` standard,
        le Tweedie ne retient QU'UNE variable sur 22 candidates. Si un jour ce
        n'est plus vrai, ce test le dira -- et il faudra une autre fixture
        pour continuer a exercer le chemin."""
        twe = self.r3['metriques']['tweedie']
        self.assertLess(twe['nb_vars_retenues'], SELECTION_PEU_SEGMENTANTE,
                        "le Tweedie retient assez de variables : le chemin de "
                        "la declaration n'est plus exerce par cette fixture")
        self.assertTrue(twe['puissance_selection'],
                        'la selection est peu segmentante et ne le dit pas')

    def test_PS7_le_classeur_A3_signe_porte_la_declaration(self):
        from openpyxl import load_workbook
        octets = self.r3.get('excel_bytes') or b''
        self.assertTrue(octets, 'aucun classeur A3')
        classeur = load_workbook(io.BytesIO(octets), data_only=True)
        texte = '\n'.join(str(c.value) for f in classeur.worksheets
                          for ligne in f.iter_rows() for c in ligne
                          if c.value is not None)
        self.assertIn('SELECTION PEU SEGMENTANTE', texte,
                      'la puissance de la selection n atteint pas le '
                      'classeur signe')
        self.assertIn('Puissance de la', texte)


class TestEquilibrageNonSilencieux(unittest.TestCase):
    """⚠️⚠️ G.11 : POURQUOI IL N'Y A RIEN A DECLARER, ET CE QUI LE GARANTIT.

    L'arbre de decision dit << aujourd'hui k reste a 1,0 en silence >>. La
    mesure du 05/09/2026 dit l'inverse au niveau du COMPORTEMENT : les deux
    voies vers une somme nulle sont refusees en amont, dans `pipeline_complet`
    lui-meme. Ce test fige les DEUX refus. Les retirer rouvrirait le silence.
    """

    def test_PS9_les_deux_refus_amont_existent_et_MORDENT(self):
        from direction_non_vie.tarification import (
            test_pipeline_agents as T,
        )
        from direction_non_vie.tarification.pipeline_tarifaire import (
            CalculImpossibleBloquant,
            pipeline_complet,
        )
        warnings.filterwarnings('ignore')
        logging.disable(logging.CRITICAL)
        try:
            np.random.seed(7)
            donnees = T._portefeuille_auto(800)
            plan = T._PLAN_AUTO
            # ① aucune frequence -> aucune charge observee
            sans_sinistre = donnees.copy()
            sans_sinistre[plan.cible_frequence] = 0
            sans_sinistre[plan.cible_cout] = 0.0
            with self.assertRaises(CalculImpossibleBloquant):
                pipeline_complet(sans_sinistre, plan)
            # ② des sinistres COMPTES, mais aucun cout -> charge nulle aussi
            sans_cout = donnees.copy()
            sans_cout[plan.cible_cout] = 0.0
            with self.assertRaises(CalculImpossibleBloquant):
                pipeline_complet(sans_cout, plan)
        finally:
            logging.disable(logging.NOTSET)

    def test_PS10_la_somme_predite_ne_peut_pas_etre_nulle(self):
        """⚠️ La prime pure sort d'un lien log, donc d'un `exp()` : elle est
        strictement positive. Une somme nulle exigerait un portefeuille VIDE,
        lui-meme refuse plus haut. C'est ce qui rend le second cas
        inatteignable -- et donc toute declaration inutile."""
        from direction_non_vie.tarification import (
            test_pipeline_agents as T,
        )
        from direction_non_vie.tarification.pipeline_tarifaire import (
            pipeline_complet,
        )
        warnings.filterwarnings('ignore')
        logging.disable(logging.CRITICAL)
        try:
            np.random.seed(7)
            donnees = T._portefeuille_auto(800)
            tarif = pipeline_complet(donnees, T._PLAN_AUTO)
            primes = tarif.predire_portefeuille(donnees)['prime_pure']
            self.assertGreater(float(primes.min()), 0.0,
                               'une prime pure nulle ou negative : le lien '
                               "log n'est plus ce qui produit la prediction, "
                               'et le raisonnement de G.11 tombe')
            self.assertGreater(float(primes.sum()), 0.0)
            # ...et k a bien ete applique, donc different de 1,0.
            self.assertNotEqual(tarif.coefficient_equilibre, 1.0,
                                "k vaut exactement 1,0 : l'equilibrage n'a "
                                'peut-etre pas eu lieu')
        finally:
            logging.disable(logging.NOTSET)


if __name__ == '__main__':
    unittest.main(verbosity=2)
