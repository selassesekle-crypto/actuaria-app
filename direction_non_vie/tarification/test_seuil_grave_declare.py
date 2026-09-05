# -*- coding: utf-8 -*-
"""LE PORTEFEUILLE DEFINISSAIT CE QUI, EN LUI, EST ANORMAL.

Le seuil au-dela duquel un sinistre est dit GRAVE retombait TOUJOURS sur le
quantile 0,995 des couts observes -- et rien ne le disait. Or ce seuil est
une donnee de REASSURANCE : il vient d'un traite, d'une politique de
souscription, d'une note du client. Le calculer sur les donnees rend le tarif
circulaire -- un portefeuille tres sinistre se donne un seuil eleve, donc
ecrete peu, donc charge la prime pure de sinistres que le traite aurait pris.

MESURE DU 05/09/2026, sur `_portefeuille_auto(1500)` :

  | plan                    | seuil    | origine  | graves | prime grave |
  |-------------------------|----------|----------|--------|-------------|
  | sans declaration        |  7 243 E | quantile |    3   | 2,4924/expo |
  | `seuil_grave: 8 000 E`  |  8 000 E | declare  |    1   | 1,0174/expo |

⚠️⚠️ ET LE SEUIL ETAIT ACCEPTE SANS AUCUNE VERIFICATION. `construire_cible_
severite` prenait deja un `seuil=` : un montant negatif ou nul traversait
`float(seuil)` puis desactivait silencieusement l'ecretement au `if seuil >
0`. L'actuaire aurait cru tarifer sous son traite de reassurance pendant que
le code n'ecretait rien.
  *Un seuil declare puis ignore en silence est pire que pas de seuil.*

⚠️⚠️ ET LE BLOC N'ATTEIGNAIT AUCUN LIVRABLE. `ecretement_severite` n'allait
que dans `_construire_contexte_tarif` -- LE PROMPT -- exactement le defaut
ferme la veille sur l'elasticite. Pire : `_tmp`, le dictionnaire passe a
`export_excel_a3`, ne le portait pas du tout. *Le correctif doit atteindre la
surface, pas la froler.*

Ce que cette sentinelle exige :
  SG-1  un seuil declare donne EXACTEMENT ce seuil, sur les trois chemins ;
  SG-2  un montant illisible ou <= 0 est un REFUS, jamais un repli ;
  SG-3  sans declaration, le quantile s'applique et la supposition SE DIT ;
  SG-4  aucun grave -> prime grave = 0, ET la publication le dit (zero est
        une MESURE, pas un calcul absent) ;
  SG-5  tout cela atteint le classeur A3 SIGNE ;
  SG-6  le seuil est dans l'EMPREINTE du plan : il decide du prix ;
  SG-7  aucun plan du depot n'en declare aujourd'hui -- la production ne
        bouge donc pas, et ce fait est MESURE, pas suppose.

Tout en `unittest.TestCase` : la gate lance `unittest discover`.
"""
import dataclasses
import glob
import io
import logging
import os
import sys
import unittest
import warnings

_ICI = os.path.dirname(os.path.abspath(__file__))
_RACINE = os.path.dirname(os.path.dirname(_ICI))
for _c in (_RACINE, _ICI):
    if _c not in sys.path:
        sys.path.insert(0, _c)

import numpy as np
import pandas as pd

from core.plan_tarifaire import EMPREINTE_SCHEMA, PlanTarifaire, SeuilGrave
from core.severite import (
    construire_cible_severite,
    phrase_aucun_grave,
    phrase_seuil_suppose,
    seuil_declare,
)

_COUTS = pd.Series([0.0, 1200.0, 800.0, 25000.0, 3000.0, 0.0, 60000.0, 500.0])
_NB = pd.Series([0.0, 1.0, 1.0, 2.0, 1.0, 0.0, 3.0, 1.0])
_EXPO = pd.Series([1.0] * 8)


def _cible(**kw):
    return construire_cible_severite(_COUTS, _NB, _EXPO, **kw)


class TestSeuilDeclare(unittest.TestCase):

    def test_SG1_un_seuil_declare_est_le_seuil_APPLIQUE(self):
        cible = _cible(seuil=10000.0)
        self.assertEqual(cible.seuil_ecretement, 10000.0)
        self.assertEqual(cible.source_seuil, 'declare')
        # ... et il ECRETE reellement : deux couts depassent 10 000.
        self.assertEqual(cible.n_graves, 2)
        self.assertGreater(cible.prime_grave_unitaire, 0)

    def test_SG1b_le_seuil_declare_l_emporte_sur_le_quantile(self):
        """La contre-epreuve : le quantile donnerait autre chose."""
        quantile = _cible().seuil_ecretement
        self.assertNotEqual(round(quantile), 10000,
                            'le temoin ne distingue pas les deux chemins')
        self.assertEqual(_cible(seuil=10000.0).seuil_ecretement, 10000.0)

    def test_SG2_un_montant_illisible_ou_negatif_est_un_REFUS(self):
        """⚠️⚠️ IL TRAVERSAIT EN SILENCE et desactivait l'ecretement."""
        for mauvais in (0, -1, -25000.0, float('nan'), float('inf')):
            with self.subTest(seuil=mauvais), self.assertRaises(ValueError):
                _cible(seuil=mauvais)
        with self.assertRaises(ValueError):
            _cible(seuil='beaucoup')

    def test_SG2b_le_plan_refuse_AUSSI_a_la_declaration(self):
        """Deux gardes, et c'est voulu : le plan refuse une declaration
        fautive, la construction refuse un seuil fautif. Un plan valide peut
        etre lu par un autre chemin ; un seuil peut venir d'ailleurs."""
        for cas in ({'montant': 0, 'source': 't'},
                    {'montant': -5, 'source': 't'},
                    {'montant': 5000, 'assiette': 'inventee', 'source': 't'},
                    {'montant': 5000, 'source': '   '}):
            with self.subTest(**cas), self.assertRaises((TypeError, ValueError)):
                SeuilGrave(**cas)
        with self.assertRaises(TypeError):
            SeuilGrave(montant='beaucoup', source='t')

    def test_SG3_sans_declaration_la_supposition_SE_DIT(self):
        cible = _cible()
        self.assertEqual(cible.source_seuil, 'quantile')
        phrase = phrase_seuil_suppose(cible, PlanTarifaire.depuis_yaml(
            os.path.join(_RACINE, 'plans', 'auto.yaml')))
        self.assertIsNotNone(phrase, "la supposition n'est pas dite")
        self.assertIn('NON DECLARE', phrase)
        self.assertIn('auto', phrase, 'la phrase ne nomme pas le plan')
        self.assertIn('seuil_grave', phrase,
                      "la phrase ne dit pas COMMENT declarer")

    def test_SG3b_avec_declaration_la_phrase_se_TAIT(self):
        """*Un avertissement permanent est un avertissement qu'on cesse de
        lire.* Il n'y a plus d'hypothese a signaler."""
        self.assertIsNone(phrase_seuil_suppose(_cible(seuil=10000.0)))

    def test_SG4_aucun_grave_donne_zero_ET_LE_DIT(self):
        """⚠️⚠️ ZERO EST ICI UNE MESURE. Une prime grave nulle publiee sans
        phrase est indiscernable d'un ecretement qui n'a pas tourne."""
        cible = _cible(seuil=1_000_000.0)
        self.assertEqual(cible.n_graves, 0)
        self.assertEqual(cible.prime_grave_unitaire, 0.0)
        phrase = phrase_aucun_grave(cible)
        self.assertIsNotNone(phrase, 'zero est publie sans un mot')
        self.assertIn('AUCUN SINISTRE GRAVE', phrase)
        self.assertIn('MESURE', phrase,
                      "la phrase ne distingue pas la mesure du calcul absent")

    def test_SG4b_avec_des_graves_la_phrase_se_TAIT(self):
        self.assertIsNone(phrase_aucun_grave(_cible(seuil=10000.0)))
        self.assertIsNone(phrase_aucun_grave(_cible(seuil=None)))

    def test_SG6_le_seuil_est_dans_l_EMPREINTE_du_plan(self):
        """⚠️⚠️ IL DECIDE DU PRIX : deux plans qui n'en different que par lui
        ecretent differemment, donc n'ont pas les memes relativites de cout,
        donc ne tarifent pas pareil. Meme argument que `chargements`."""
        plan = PlanTarifaire.depuis_yaml(
            os.path.join(_RACINE, 'plans', 'auto.yaml'))
        avec = dataclasses.replace(plan, seuil_grave=SeuilGrave(
            montant=50000.0, assiette='total_contrat', source='traite'))
        autre = dataclasses.replace(plan, seuil_grave=SeuilGrave(
            montant=80000.0, assiette='total_contrat', source='traite'))
        self.assertNotEqual(plan.empreinte(), avec.empreinte(),
                            "declarer un seuil ne change pas l'empreinte")
        self.assertNotEqual(avec.empreinte(), autre.empreinte(),
                            "deux seuils differents portent la MEME empreinte")
        # ⚠️ LE NUMERO SE DERIVE, IL NE S'EPINGLE PAS. Seul le golden de
        # `test_plan_invariants` a le droit de le figer : ailleurs, un
        # numero en dur ferait rougir au prochain bump un test dont le sujet
        # n'a pas bouge. Un controle du depot (PL-8) tient cette regle -- et
        # il a attrape ce test-ci.
        self.assertTrue(
            plan.empreinte().startswith(f's{EMPREINTE_SCHEMA}:'),
            f'empreinte non prefixee par le schema : {plan.empreinte()[:6]}')

    def test_SG7_aucun_plan_du_depot_n_en_declare_aujourd_hui(self):
        """⚠️ LA CONDITION 4, MESUREE ET NON SUPPOSEE. Si un plan en declarait
        un, ce lot deplacerait un prix -- et ce test le dirait."""
        declarants = []
        for chemin in sorted(glob.glob(os.path.join(_RACINE, 'plans', '*.yaml'))):
            try:
                plan = PlanTarifaire.depuis_yaml(chemin)
            except Exception as _e:                           # noqa: BLE001
                # ⚠️ Un plan illisible n'est pas le sujet de ce test, mais il
                # ne disparait pas en silence : d'autres controles le tiennent.
                print(f'    (plan ignore : {os.path.basename(chemin)} -- {_e})')
                continue
            if seuil_declare(plan) is not None:
                declarants.append(os.path.basename(chemin))
        self.assertEqual(
            declarants, [],
            'ces plans declarent un seuil grave : la production CHANGE, et '
            f'il faut le mesurer avant de conclure -- {declarants}')


class TestSurLaChaineEtLeClasseur(unittest.TestCase):

    @classmethod
    def setUpClass(cls):
        warnings.filterwarnings('ignore')
        logging.disable(logging.CRITICAL)

    @classmethod
    def tearDownClass(cls):
        logging.disable(logging.NOTSET)

    def _a3(self, plan):
        from core.qualite_donnees import preambule_qualite
        from direction_non_vie.tarification import test_pipeline_agents as T
        from direction_non_vie.tarification.a1_ingestion.agent import (
            AgentA1Ingestion,
        )
        from direction_non_vie.tarification.a2_preprocessing.agent import (
            AgentA2Preprocessing,
        )
        from direction_non_vie.tarification.a3_glm.agent import AgentA3GLM

        np.random.seed(7)
        donnees = T._portefeuille_auto(1200)
        base = {'audit_path': '/tmp', 'verbose': False}
        r1 = AgentA1Ingestion(**base).run(branche='non_vie',
                                          sous_branche='auto',
                                          dataframe=donnees)
        qualite = preambule_qualite(r1.get('dataframe'), plan,
                                    qualite_validee_par='Test',
                                    horodatage=None)
        r2 = AgentA2Preprocessing(**base).run(
            result_a1={**r1, 'dataframe': qualite.dataframe_propre}, plan=plan)
        return AgentA3GLM(models_path='/tmp', audit_path='/tmp').run(
            result_a2=r2, plan=plan, col_frequence=plan.cible_frequence,
            col_cout=plan.cible_cout, generer_graphiques=False)

    def _texte(self, octets):
        from openpyxl import load_workbook
        classeur = load_workbook(io.BytesIO(octets), data_only=True)
        return '\n'.join(str(c.value) for f in classeur.worksheets
                         for ligne in f.iter_rows() for c in ligne
                         if c.value is not None)

    def test_SG5_le_classeur_A3_porte_le_bloc_et_l_origine(self):
        """⚠️⚠️ `ecretement_severite` n'atteignait AUCUN livrable : il n'allait
        que dans le prompt, et `_tmp` -- le dictionnaire passe a
        `export_excel_a3` -- ne le portait meme pas."""
        from direction_non_vie.tarification import test_pipeline_agents as T
        r3 = self._a3(T._PLAN_AUTO)
        octets = r3.get('excel_bytes') or b''
        self.assertTrue(octets, "aucun classeur A3")
        texte = self._texte(octets)
        self.assertIn('SEUIL DE SINISTRE GRAVE', texte,
                      'le classeur signe ne publie pas le seuil')
        self.assertIn('supposé', texte,
                      "le classeur ne dit pas que le seuil est SUPPOSE")
        self.assertIn('NON DECLARE', texte,
                      "la phrase d'hypothese n'atteint pas le classeur")

    def test_SG5b_un_seuil_DECLARE_change_le_classeur_ET_la_prime(self):
        from direction_non_vie.tarification import test_pipeline_agents as T
        plan = dataclasses.replace(T._PLAN_AUTO, seuil_grave=SeuilGrave(
            montant=8000.0, assiette='total_contrat', source='traite XL 2026'))
        r3 = self._a3(plan)
        ecr = r3.get('ecretement_severite') or {}
        self.assertEqual(ecr.get('seuil'), 8000.0,
                         "le seuil declare n'est pas celui applique")
        self.assertEqual(ecr.get('source_seuil'), 'declare')
        texte = self._texte(r3.get('excel_bytes') or b'')
        self.assertIn('déclaré au plan', texte)
        self.assertNotIn('NON DECLARE', texte,
                         "le classeur signale une supposition alors que le "
                         "seuil est declare")


if __name__ == '__main__':
    unittest.main(verbosity=2)
