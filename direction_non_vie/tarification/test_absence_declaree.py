# -*- coding: utf-8 -*-
"""UNE VALEUR ABSENTE NE VAUT PAS UN LITTERAL — LA CAUSE COMMUNE DU MODULE.

Le defaut central de la tarification n'est pas une erreur de calcul : c'est
une valeur qu'on n'a pas mesuree, remplacee par un nombre. Elle traverse
ensuite tout, parce qu'un nombre ne se distingue pas d'une mesure.

CE QUI A ETE MESURE LE 05/09/2026, sur la fixture `_portefeuille_auto` :

  1. `999` COMME RMSE. Le Tweedie ne publie PAS de `rmse_test` (la cle est
     ABSENTE de sa famille). A6 posait `met.get('rmse_test', 999)`. Sur la
     cible `prime_pure`, le Tweedie est le SEUL modele du catalogue : le
     commentaire signe publiait donc, sur le modele de PRODUCTION,
     << RMSE test : 999.00 >>. Et sur un catalogue peuple, ce 999 devenait le
     maximum qui normalise le critere pour tous les autres.

  2. UN RATIO QUI EXPLOSE AU LIEU DE REFUSER. `ratio_sur_apprentissage`
     faisait `gini_train / max(gini_test, 1e-6)`. Avec le Gamma mesure a
     **-0,0298**, le ratio valait **39 534,31** ; avec TabNet a -0,0201, il
     valait **-9 387,61**. Ce -9 387 devenait le `min_of` de la
     normalisation `1 - (r - min) / (max - min)` : TabNet recevait
     `s_stab = 1.0`, **la MEILLEURE note de stabilite du catalogue**, et tous
     les autres tombaient a ~0. TabNet se classait 4e ; mesure apres
     correction, il est 9e.

  3. UN GINI D'ENTRAINEMENT FABRIQUE, QUI NEUTRALISAIT SA PROPRE GARDE. A4
     posait `gini_train = gini_test * 1.10` quand la mesure manquait. Le
     ratio valait alors MECANIQUEMENT 1/1,10 = 0,909, donc >= 0,90, donc H1
     publiait << Pas d'overfitting >>. La branche `if ratio_of is None`,
     ecrite juste en dessous pour dire << NON MESURABLE >>, etait
     inatteignable des que `gini_test` existait.

  4. UN GINI DE REFERENCE D'ALLURE CREDIBLE. A5 posait `gini_glm_ref = 0.10`
     quand A3 manquait. Un chiffre plausible est plus dangereux qu'un zero :
     personne ne le remarque en relecture, et il decide le statut de H3.

  5. TROIS ZEROS DANS UN CLASSEUR SIGNE. Le Gamma et le Tweedie ne publient
     pas `deviance_nulle`, le Tweedie pas `pseudo_r2` : la feuille
     << 1-Synthese >> du classeur A3 affichait << GLM Tweedie -- Pseudo-R2 :
     0 >>, c'est-a-dire *le modele n'explique rien*, a chaque run normal.

  6. UNE GARDE QUI SE TESTAIT A TRAVERS SON PROPRE DEFAUT.
     `meilleur.get('gini_test', 0) is not None` est VRAI meme quand la cle
     manque.

Ce que cette sentinelle exige :
  AD-1   la primitive generale rend le MOT, jamais un nombre invente, et
         `gini_texte` / `gini_arrondi` n'en sont que des noms ;
  AD-2   un ratio ne se calcule pas sur un Gini de test <= 0 ;
  AD-3   **LE CONTROLE QUI DERIVE** : aucun site de la tarification ne lit
         une cle de mesure avec un defaut numerique -- et l'ensemble des cles
         se DERIVE des producteurs, il ne s'enumere pas ;
  AD-4   une valeur non mesuree SORT du score et le modele le DECLARE ;
  AD-5   une valeur non mesuree ne deplace ni le min ni le max de la
         normalisation, donc pas le score des autres ;
  AD-6   H1 d'A4 dit << non mesurable >>, jamais << pas d'overfitting >> ;
  AD-7   H3 d'A5 dit << non evaluable >> sans reference mesuree ;
  AD-8   le classeur SIGNE ecrit le mot, pas un zero ;
  AD-9   A6 refuse de prononcer un statut sur un Gini de production absent.

Tout en `unittest.TestCase` : la gate lance `unittest discover`.
"""
import ast
import glob
import io
import os
import pathlib
import sys
import unittest

_ICI = os.path.dirname(os.path.abspath(__file__))
_RACINE = os.path.dirname(os.path.dirname(_ICI))
for _c in (_RACINE, _ICI):
    if _c not in sys.path:
        sys.path.insert(0, _c)

from core.conformite_reglementaire import (
    NON_MESURE,
    gini_arrondi,
    gini_texte,
    mesure_arrondie,
    mesure_texte,
    ratio_sur_apprentissage,
)
from direction_non_vie.tarification.a6_comparaison.agent import AgentA6Comparaison

# =============================================================================
#  AD-1, AD-2 — LE SOCLE
# =============================================================================

class TestPrimitiveGenerale(unittest.TestCase):

    def test_AD1_le_mot_remplace_toute_valeur_inexistante(self):
        for absente in (None, float('nan'), float('inf'), float('-inf')):
            self.assertEqual(mesure_texte(absente), NON_MESURE, repr(absente))
            self.assertEqual(mesure_arrondie(absente), NON_MESURE, repr(absente))

    def test_AD1b_une_vraie_mesure_reste_un_nombre(self):
        self.assertEqual(mesure_texte(0.26514, 4), '0.2651')
        self.assertEqual(mesure_arrondie(0.26514, 4), 0.2651)
        # ⚠️ ZERO EST UNE MESURE. Le confondre avec l'absence serait le meme
        # defaut dans l'autre sens : un modele qui ne discrimine pas doit
        # pouvoir le dire.
        self.assertEqual(mesure_texte(0.0, 4), '0.0000')
        self.assertEqual(mesure_arrondie(0.0, 4), 0.0)
        self.assertEqual(mesure_texte(-0.0298, 4), '-0.0298')

    def test_AD1c_gini_texte_et_gini_arrondi_ne_sont_QUE_des_noms(self):
        """Source unique : deux contrats identiques ne peuvent pas diverger
        s'il n'y en a qu'un."""
        for valeur in (None, float('nan'), 0.0, -0.0298, 0.2651, 1.75):
            self.assertEqual(gini_texte(valeur), mesure_texte(valeur), repr(valeur))
            self.assertEqual(gini_arrondi(valeur), mesure_arrondie(valeur),
                             repr(valeur))

    def test_AD2_un_ratio_ne_se_calcule_pas_sur_un_gini_de_test_negatif(self):
        """⚠️⚠️ MESURE : avec `max(gini_test, 1e-6)`, le Gamma a -0,0298
        rendait 39 534,31 et TabNet a -0,0201 rendait -9 387,61. Ces nombres
        devenaient les bornes de la normalisation de stabilite."""
        self.assertIsNone(ratio_sur_apprentissage(0.0395, -0.0298),
                          'un Gini de test negatif produit encore un ratio')
        self.assertIsNone(ratio_sur_apprentissage(0.0395, 0.0),
                          'un Gini de test nul produit encore un ratio')
        self.assertIsNone(ratio_sur_apprentissage(None, 0.20))
        self.assertIsNone(ratio_sur_apprentissage(0.20, None))
        self.assertIsNone(ratio_sur_apprentissage(float('nan'), 0.20))

    def test_AD2b_un_ratio_mesurable_reste_mesure_a_l_identique(self):
        """La correction ne doit PAS changer le ratio quand il existe."""
        self.assertAlmostEqual(ratio_sur_apprentissage(0.2100, 0.1912),
                               0.2100 / 0.1912, places=12)


# =============================================================================
#  AD-3 — LE CONTROLE QUI DERIVE
# =============================================================================

#: Les repertoires ou la regle s'applique. ⚠️ `audit_2026_08` en est exclu :
#: ce sont des scripts de PREUVE d'audit, pas du code de production.
_ZONES = ('direction_non_vie/tarification/**/*.py',)


def _fichiers_de_production():
    vus = set()
    for motif in _ZONES:
        for f in glob.glob(os.path.join(_RACINE, motif), recursive=True):
            chemin = f.replace('\\', '/')
            nom = pathlib.Path(chemin).name
            if ('__pycache__' in chemin or 'audit_2026_08' in chemin
                    or nom.startswith('test_')):
                continue
            vus.add(chemin)
    return sorted(vus)


def _arbres():
    for chemin in _fichiers_de_production():
        try:
            yield chemin, ast.parse(pathlib.Path(chemin).read_text(encoding='utf-8'))
        except (SyntaxError, UnicodeDecodeError):                   # pragma: no cover
            continue


def _fonctions_pouvant_rendre_none(arbres):
    """Les fonctions dont un `return` vaut `None`, explicitement."""
    noms = set()
    for _, arbre in arbres:
        for noeud in ast.walk(arbre):
            if not isinstance(noeud, (ast.FunctionDef, ast.AsyncFunctionDef)):
                continue
            for sous in ast.walk(noeud):
                if isinstance(sous, ast.Return) and (
                        sous.value is None
                        or (isinstance(sous.value, ast.Constant)
                            and sous.value.value is None)):
                    noms.add(noeud.name)
                    break
    return noms


def _cles_que_les_producteurs_laissent_vides(arbres):
    """⚠️⚠️ IL DERIVE, IL N'ENUMERE PAS. Est << cle de mesure >> toute cle a
    laquelle un producteur de la tarification affecte, quelque part, soit le
    litteral `None`, soit le RESULTAT D'UNE FONCTION QUI PEUT RENDRE `None`.

    ⚠️ Le second cas est le principal, et l'oublier vide le controle : les
    agents n'ecrivent presque jamais `'gini': None`, ils ecrivent
    `'gini': self._calculer_gini(...)`. Une premiere version de ce controle ne
    regardait que le litteral et derivait ZERO cle -- elle serait restee verte
    en ne surveillant rien.

    Une liste tenue a la main aurait diverge le jour ou un agent publie une
    metrique nouvelle -- et l'assiette du controle se serait retrecie sans que
    personne le voie.
    """
    nullables = _fonctions_pouvant_rendre_none(arbres)

    def _source_nullable(noeud):
        if isinstance(noeud, ast.Constant) and noeud.value is None:
            return True
        if isinstance(noeud, ast.Call):
            nom = (noeud.func.id if isinstance(noeud.func, ast.Name)
                   else getattr(noeud.func, 'attr', None))
            return nom in nullables
        if isinstance(noeud, ast.IfExp):
            return _source_nullable(noeud.body) or _source_nullable(noeud.orelse)
        return False

    cles = set()
    for _, arbre in arbres:
        for noeud in ast.walk(arbre):
            if isinstance(noeud, ast.Dict):
                for cle, valeur in zip(noeud.keys, noeud.values):
                    if (isinstance(cle, ast.Constant)
                            and isinstance(cle.value, str)
                            and _source_nullable(valeur)):
                        cles.add(cle.value)
            elif isinstance(noeud, ast.Assign) and len(noeud.targets) == 1:
                cible = noeud.targets[0]
                if (isinstance(cible, ast.Subscript)
                        and isinstance(cible.slice, ast.Constant)
                        and isinstance(cible.slice.value, str)
                        and _source_nullable(noeud.value)):
                    cles.add(cible.slice.value)
    return cles


def _est_numerique(noeud):
    if isinstance(noeud, ast.Constant):
        return (isinstance(noeud.value, (int, float))
                and not isinstance(noeud.value, bool))
    return (isinstance(noeud, ast.UnaryOp)
            and isinstance(noeud.operand, ast.Constant)
            and isinstance(noeud.operand.value, (int, float)))


#: Les grandeurs dont l'absence FABRIQUE une mesure. Ce sont celles que les
#: agents publient comme resultat d'un calcul statistique -- pas les libelles
#: (`branche`, `note`, `erreur`), pour lesquels un defaut textuel est honnete.
#: ⚠️ Ce filtre restreint l'assiette DERIVEE : il est donc verifie par
#: `test_AD3b`, qui exige que chacune de ces cles soit effectivement derivee.
_MESURES = ('gini', 'gini_test', 'gini_train', 'rmse_test', 'overfit_ratio',
            'cout_moyen_pred')


class TestControleQuiDerive(unittest.TestCase):

    def test_AD3_aucun_defaut_numerique_sur_une_cle_de_mesure(self):
        """⚠️⚠️ ZERO EXEMPTION, ET C'EST LE POINT. Un controle a liste
        d'exceptions grandit ; celui-ci ne le peut pas. Mesure du
        05/09/2026 : 101 sites portaient un tel defaut avant ce lot."""
        arbres = list(_arbres())
        self.assertGreater(len(arbres), 10,
                           "aucun fichier de production trouve : le controle "
                           "ne mesure rien")
        derivees = _cles_que_les_producteurs_laissent_vides(arbres)
        surveillees = {c for c in _MESURES if c in derivees}
        self.assertTrue(surveillees, 'aucune cle de mesure derivee')

        fautes = []
        for chemin, arbre in arbres:
            lignes = pathlib.Path(chemin).read_text(encoding='utf-8').splitlines()
            for noeud in ast.walk(arbre):
                if not (isinstance(noeud, ast.Call)
                        and isinstance(noeud.func, ast.Attribute)
                        and noeud.func.attr == 'get' and len(noeud.args) == 2):
                    continue
                cle, defaut = noeud.args
                if (isinstance(cle, ast.Constant) and cle.value in surveillees
                        and _est_numerique(defaut)):
                    fautes.append(
                        f'{pathlib.Path(chemin).name}:{noeud.lineno} '
                        f'{cle.value} = {ast.unparse(defaut)} | '
                        f'{lignes[noeud.lineno - 1].strip()[:80]}')
        self.assertEqual(fautes, [], 'une valeur non mesuree recevrait un '
                                     'nombre :\n  ' + '\n  '.join(fautes))

    def test_AD3b_les_cles_surveillees_sont_bien_DERIVEES_des_producteurs(self):
        """⚠️ Sans ceci, `_MESURES` deviendrait une liste tenue a la main : le
        jour ou un agent cesse de publier `None` sur une cle, elle sortirait
        de l'assiette EN SILENCE et le controle resterait vert."""
        derivees = _cles_que_les_producteurs_laissent_vides(list(_arbres()))
        manquantes = [c for c in _MESURES if c not in derivees]
        self.assertEqual(manquantes, [],
                         'ces grandeurs ne sont plus publiees comme pouvant '
                         'valoir None par aucun producteur : '
                         f'{manquantes} -- le controle AD-3 ne les surveille '
                         'donc plus. Relire le lot avec ce changement.')

    def test_AD3c_le_999_n_existe_plus_nulle_part(self):
        """Le littéral nomme, cherche au texte : il ne doit plus etre le
        repli d'une RMSE dans aucun agent."""
        for chemin in _fichiers_de_production():
            texte = pathlib.Path(chemin).read_text(encoding='utf-8')
            self.assertNotIn("rmse_test', 999", texte, chemin)
            self.assertNotIn('rmse_test", 999', texte, chemin)


# =============================================================================
#  AD-4, AD-5 — LE SCORE : UNE ASSIETTE REDUITE, ET DECLAREE
# =============================================================================

def _catalogue(**remplacements):
    """Trois modeles comparables, dont un modifiable par le test."""
    base = [
        {'modele': 'GLM_POISSON', 'famille': 'GLM', 'gini_test': 0.19,
         'rmse_test': 0.71, 'overfit_ratio': 0.97, 'interpretabilite': 1.0},
        {'modele': 'ML_GBM', 'famille': 'ML', 'gini_test': 0.13,
         'rmse_test': 0.74, 'overfit_ratio': 3.58, 'interpretabilite': 0.5},
        {'modele': 'ML_XGBOOST', 'famille': 'ML', 'gini_test': 0.10,
         'rmse_test': 0.74, 'overfit_ratio': 5.48, 'interpretabilite': 0.5},
    ]
    for i, modif in remplacements.items():
        base[int(i[1:])].update(modif)
    return base


def _scorer(catalogue):
    agent = AgentA6Comparaison.__new__(AgentA6Comparaison)
    poids = {'gini': 0.40, 'stabilite': 0.30, 'interpretabilite': 0.20,
             'rmse': 0.10}
    return agent._calculer_scores_multicriteres(catalogue, poids)


class TestScoreSurAssietteReduite(unittest.TestCase):

    def test_AD4_une_rmse_absente_sort_du_score_et_se_DECLARE(self):
        avec = _scorer(_catalogue())
        sans = _scorer(_catalogue(m0={'rmse_test': None}))
        self.assertEqual(avec[0]['criteres_non_mesures'], ())
        self.assertEqual(sans[0]['criteres_non_mesures'], ('rmse',),
                         "le modele ne declare pas ce qu'il n'a pas mesure")
        self.assertIsNone(sans[0]['score_rmse'])
        self.assertIsNotNone(sans[0]['score_global'])

    def test_AD5_une_absence_se_comporte_comme_une_ABSENCE(self):
        """⚠️⚠️ LE COEUR DU DEFAUT, ET J'AI D'ABORD ECRIT L'INVARIANT FAUX.

        Ma premiere version exigeait que le score des AUTRES ne bouge pas
        quand un ratio devient non mesure. C'est impossible et ce serait
        meme faux : la normalisation est un min-max, retirer une valeur de
        l'ensemble change legitimement l'etendue.

        Le bon invariant est celui-ci : un ratio NON MESURE doit produire
        exactement le meme effet qu'un modele ABSENT de ce critere -- ni plus
        (il ne borne rien), ni moins (il ne disparait pas du classement). Un
        litteral, lui, BORNE.
        """
        non_mesure = {m['modele']: m['score_stabilite']
                      for m in _scorer(_catalogue(m0={'overfit_ratio': None}))}
        absent = {m['modele']: m['score_stabilite']
                  for m in _scorer(_catalogue()[1:])}
        for nom in ('ML_GBM', 'ML_XGBOOST'):
            self.assertEqual(non_mesure[nom], absent[nom],
                             f"la stabilite de {nom} differe selon qu'un autre "
                             "modele a un ratio NON MESURE ou est ABSENT : "
                             "l'absence borne encore quelque chose")
        self.assertIsNone(non_mesure['GLM_POISSON'])
        self.assertIn('GLM_POISSON', non_mesure,
                      "le modele a disparu du classement au lieu d'y figurer "
                      'sans note de stabilite')

    def test_AD5c_un_litteral_BORNE_la_normalisation(self):
        """La contre-epreuve : c'est bien le LITTERAL qui deplacait les
        autres, pas l'absence."""
        avec_litteral = {m['modele']: m['score_stabilite']
                         for m in _scorer(_catalogue(m0={'overfit_ratio': 1.0}))}
        non_mesure = {m['modele']: m['score_stabilite']
                      for m in _scorer(_catalogue(m0={'overfit_ratio': None}))}
        self.assertNotEqual(avec_litteral['ML_GBM'], non_mesure['ML_GBM'],
                            'le plant ne demontre rien : le littéral ne change '
                            'pas le score des autres')

    def test_AD5b_un_ratio_absurde_aurait_ecrase_la_normalisation(self):
        """La contre-epreuve, avec le nombre REELLEMENT mesure : -9 387,61.

        Ce test ne fait pas passer le correctif, il MONTRE ce qu'il evite --
        si un tel ratio revenait dans le catalogue, la stabilite de tous les
        autres modeles s'effondrerait, et celle du fautif serait parfaite."""
        pollue = _scorer(_catalogue(m0={'overfit_ratio': -9387.61}))
        par_nom = {m['modele']: m for m in pollue}
        self.assertAlmostEqual(par_nom['GLM_POISSON']['score_stabilite'], 1.0,
                               places=3)
        self.assertLess(par_nom['ML_GBM']['score_stabilite'], 0.01)
        # ... et c'est bien pourquoi `ratio_sur_apprentissage` REFUSE de le
        # produire : aucun agent ne peut plus faire entrer ce nombre.
        self.assertIsNone(ratio_sur_apprentissage(0.0094, -0.0201))

    def test_AD4b_tout_mesure_donne_le_MEME_score_qu_avant_le_lot(self):
        """⚠️ La renormalisation ne doit rien changer quand rien ne manque :
        le facteur vaut exactement 1,0. Sans ce test, le correctif pourrait
        deplacer un prix en croyant ne rien faire."""
        scores = _scorer(_catalogue())
        attendu = (0.40 * (0.19 / 0.19) + 0.30 * 1.0 + 0.20 * 1.0
                   + 0.10 * (0.71 / 0.71))
        self.assertAlmostEqual(scores[0]['score_global'], round(attendu, 4),
                               places=4)


# =============================================================================
#  AD-6, AD-7 — LES HYPOTHESES NE SE CONCLUENT PAS SANS MESURE
# =============================================================================

def _h1(modele):
    """L'hypothese H1 d'A4 sur un classement d'un seul modele.

    ⚠️ On appelle la VRAIE methode (`_valider_modele_ml`), pas une copie de sa
    logique : un controle qui recopie le calcul ne surveille que sa copie.
    """
    import numpy as np

    from direction_non_vie.tarification.a4_ml.agent import AgentA4ML
    agent = AgentA4ML.__new__(AgentA4ML)
    zeros = np.zeros((4, 2))
    val = agent._valider_modele_ml([modele], {'psi': 0.05}, 100, 25, zeros,
                                   zeros, np.zeros(4))
    return val.get('h1_overfitting', {})


class TestHypothesesSansMesure(unittest.TestCase):

    def test_AD6_H1_dit_non_mesurable_jamais_pas_d_overfitting(self):
        """⚠️⚠️ `gini_train = gini_test * 1.10` donnait un ratio de 0,909,
        donc >= 0,90, donc << Pas d'overfitting >> -- sur un modele dont
        l'entrainement n'avait jamais ete mesure."""
        h1 = _h1({'modele': 'ML_GBM', 'gini_test': 0.13, 'gini_train': None,
                  'rmse_test': 0.74, 'overfit_ratio': None})
        self.assertEqual(h1.get('statut'), 'AMBRE', h1)
        self.assertIn('NON MESURABLE', h1.get('message', '').upper(), h1)
        self.assertNotIn("pas d'overfitting", h1.get('message', '').lower(), h1)

    def test_AD6b_H1_conclut_normalement_quand_les_DEUX_ginis_existent(self):
        h1 = _h1({'modele': 'ML_GBM', 'gini_test': 0.19, 'gini_train': 0.20,
                  'rmse_test': 0.74, 'overfit_ratio': 1.05})
        self.assertEqual(h1.get('statut'), 'VERT', h1)
        self.assertIn('0.95', h1.get('message', ''), h1)

    def test_AD6c_un_gini_d_entrainement_NEGATIF_ne_produit_plus_de_ratio(self):
        """`max(gini_train, 0.001)` bornait le denominateur : le ratio
        explosait au lieu de se declarer non mesurable."""
        h1 = _h1({'modele': 'ML_GBM', 'gini_test': 0.13, 'gini_train': -0.02,
                  'rmse_test': 0.74, 'overfit_ratio': None})
        self.assertEqual(h1.get('statut'), 'AMBRE', h1)
        self.assertIn('NON MESURABLE', h1.get('message', '').upper(), h1)

    def test_AD7_le_gini_de_reference_0_10_n_existe_plus(self):
        """⚠️ Un chiffre PLAUSIBLE est plus dangereux qu'un zero : personne ne
        le remarque en relecture, et il decidait le statut de H3."""
        chemin = os.path.join(_ICI, 'a5_deep_learning', 'agent.py')
        texte = pathlib.Path(chemin).read_text(encoding='utf-8')
        self.assertNotIn(".get('gini', 0.10)", texte,
                         'A5 fabrique encore un Gini GLM de reference')
        self.assertNotIn('gini_glm_ref = 0.10', texte)


# =============================================================================
#  AD-8, AD-9 — LES SURFACES SIGNEES
# =============================================================================

class TestSurfacesSignees(unittest.TestCase):

    def test_AD8_le_classeur_A3_ecrit_le_MOT_pas_un_zero(self):
        """⚠️⚠️ MESURE DU 05/09/2026 : sur un run NORMAL, la feuille
        << 1-Synthese >> publiait << GLM Tweedie -- Pseudo-R2 : 0 >> et deux
        << Deviance nulle : 0 >>. Le Gamma et le Tweedie ne publient pas ces
        metriques : le zero etait une affirmation que personne n'avait faite,
        dans le document que l'actuaire signe."""
        from openpyxl import load_workbook

        from direction_non_vie.tarification.services.tarif_excel import (
            export_excel_a3,
        )
        # Un resultat A3 fidele a la forme reelle : le Tweedie n'a NI
        # `deviance_nulle` NI `pseudo_r2`, le Gamma n'a pas `deviance_nulle`.
        resultat = {
            'success': True, 'statut_rag': 'AMBRE', 'sous_branche': 'auto',
            'metriques': {
                'poisson': {'gini': 0.1912, 'aic': 3421.61,
                            'deviance_nulle': 1990.87, 'pseudo_r2': 0.0396,
                            'nb_vars_retenues': 7},
                'gamma': {'gini': -0.0298, 'aic': 11816.09,
                          'pseudo_r2': 0.0079, 'nb_vars_retenues': 3},
                'tweedie': {'gini': 0.1901, 'aic': 14954.35,
                            'nb_vars_retenues': 5},
            },
        }
        octets = export_excel_a3(resultat)
        self.assertTrue(octets, "l'export n'a produit aucun classeur")
        feuille = load_workbook(io.BytesIO(octets), data_only=True).worksheets[0]
        publie = {}
        etiquette = None
        for ligne in feuille.iter_rows():
            for cellule in ligne:
                if isinstance(cellule.value, str) and '—' in str(cellule.value):
                    etiquette = cellule.value
                elif etiquette is not None and cellule.value is not None:
                    publie[etiquette] = cellule.value
                    etiquette = None
        for nom in ('GLM Gamma — Déviance nulle', 'GLM Tweedie — Déviance nulle',
                    'GLM Tweedie — Pseudo-R²'):
            self.assertIn(nom, publie, f'{nom} absent du classeur : le test ne '
                                       f'mesure rien. Vu : {sorted(publie)[:8]}')
            self.assertEqual(publie[nom], NON_MESURE,
                             f'{nom} publie {publie[nom]!r} alors que la '
                             'metrique n existe pas')
        # ... et ce qui EST mesure reste un nombre.
        self.assertEqual(publie['GLM Poisson — Pseudo-R²'], 0.0396)
        self.assertEqual(publie['GLM Gamma — Pseudo-R²'], 0.0079)

    def test_AD9_A6_refuse_un_statut_sur_un_gini_de_production_absent(self):
        """Le catalogue ecarte deja ces modeles ; si l'un passe quand meme, on
        le DIT au lieu de prononcer un statut sur une mesure inexistante."""
        agent = AgentA6Comparaison.__new__(AgentA6Comparaison)
        with self.assertRaises(ValueError) as capture:
            agent._calculer_statut_rag(
                {'modele': 'X', 'score_global': 0.9, 'gini_test': None},
                {}, [], 'production', 'Actuaire', None)
        self.assertIn('Gini', str(capture.exception))

    def test_AD9b_le_commentaire_ne_publie_plus_999(self):
        chemin = os.path.join(_ICI, 'a6_comparaison', 'agent.py')
        texte = pathlib.Path(chemin).read_text(encoding='utf-8')
        self.assertNotIn("mp['rmse_test']:.2f", texte,
                         'le commentaire signe formate encore la RMSE sans '
                         'savoir dire quand elle manque')
        self.assertIn("mesure_texte(mp['rmse_test']", texte)


if __name__ == '__main__':
    unittest.main(verbosity=2)
