# -*- coding: utf-8 -*-
"""DES FAITS FAUX DANS LES LIVRABLES SIGNES — CHERCHES DANS LES OCTETS.

⚠️⚠️ LA METHODE EST LE CONTROLE. Chaque litteral fautif est cherche DANS LES
DOCUMENTS PRODUITS, pas dans le code : un correctif peut atterrir a cote de
la surface signee sans que rien ne le dise. Ce fichier prend les livrables
comme ASSIETTE.

CINQ FAITS MESURES LE 05/09/2026 SUR UN RUN AMBRE :

  1. << le meilleur modele vaut toujours . 1,0000 >> -- imprime a cote d'un
     score mesure a **0,9062**. La normalisation est par CRITERE : chaque
     dimension est ramenee au meilleur SUR ELLE, puis ponderee. Un modele
     n'atteint 1,0000 que s'il est le meilleur sur TOUS les criteres a la
     fois. *Un lecteur qui croit la phrase interprete le 0,91 comme un ecart
     de 9 % a un maximum atteignable ; ce n'en est pas un.*

  2. << -> Deployer GLM_POISSON comme modele de tarification >> sous un
     statut **AMBRE**, dans QUATRE surfaces -- deux paragraphes apres un
     diagnostic qui dit << presente des points d'attention >>.

  3. LA PISTE D'AUDIT PUBLIAIT UNE REFERENCE REGLEMENTAIRE COUPEE.
     `note_score_global` fait 205 caracteres, `valeur_audit` tronquait a
     120 : tout ce qui suit << RMSE) >> disparaissait -- dont le profil
     retenu et **ACPR-2022-P-01 §4.3**.

  4. L'ELASTICITE N'ATTEIGNAIT NI LE WORD NI LE HTML D'A6 : 7 occurrences
     dans le classeur A6, 7 dans le rapport d'equipe, **0** dans les deux
     formats du rapport de modeles. Elle y etait pourtant calculee -- dans
     `_construire_contexte_tarif`, c'est-a-dire dans LE PROMPT. Et la
     narration rendue n'en portait aucune trace.
     *Un calcul qui n'atteint que le prompt n'atteint pas un livrable.*

  5. SHAP ETAIT DEMANDE SUR LA LIGNE DE REFERENCE GLM. `classement[0]` peut
     etre << GLM Poisson (reference A3) >>, une ligne RELAYEE d'A3 qui n'est
     pas dans `self.modeles` : le classeur signe publiait << SHAP non
     calcule : Modele ... non trouve >>. Avant le lot precedent, ce meme
     echec faisait disparaitre le classeur ENTIER.

CE QUI N'EST **PAS** UN DEFAUT, ET QUI A ETE VERIFIE. << Cette narration
engage l'actuaire signataire >> reste, et c'est une decision DOCUMENTEE du
depot (`core/traitement_ia.py` l. 62-70) : la phrase dit l'ENGAGEMENT, pas
le FAIT. Certifier << relue et validee >> serait invérifiable ; c'est
`trace_relecture` qui dit le fait, a cote. On ne rouvre pas un arbitrage
motive.

Ce que cette sentinelle exige :
  FP-1  aucune surface ne publie << toujours . 1,0000 >> ;
  FP-2  << Deployer >> ne se dit QUE sous VERT ;
  FP-3  la reference reglementaire SURVIT dans la piste d'audit publiee ;
  FP-4  l'elasticite atteint les QUATRE surfaces, pas seulement deux ;
  FP-5  SHAP porte sur un modele CALIBRE par A4 ;
  FP-6  et le nom du plan atteint reellement les octets (contre-preuve du
        lot 3, dont le decompte de surfaces etait faux).

Tout en `unittest.TestCase` : la gate lance `unittest discover`.
"""
import io
import logging
import os
import re
import sys
import unittest
import warnings
import zipfile

_ICI = os.path.dirname(os.path.abspath(__file__))
_RACINE = os.path.dirname(os.path.dirname(_ICI))
for _c in (_RACINE, _ICI):
    if _c not in sys.path:
        sys.path.insert(0, _c)

import numpy as np


def _texte_docx(octets):
    if not octets:
        return ''
    with zipfile.ZipFile(io.BytesIO(octets)) as z:
        xml = z.read('word/document.xml').decode('utf-8', 'replace')
    return ' '.join(re.findall(r'<w:t(?:\s[^>]*)?>(.*?)</w:t>', xml,
                               re.DOTALL))


def _texte_xlsx(octets):
    if not octets:
        return ''
    from openpyxl import load_workbook
    classeur = load_workbook(io.BytesIO(octets), data_only=True)
    return '\n'.join(str(c.value) for f in classeur.worksheets
                     for ligne in f.iter_rows() for c in ligne
                     if c.value is not None)


def _texte(octets):
    if not octets:
        return ''
    return (octets.decode('utf-8', 'replace') if isinstance(octets, bytes)
            else str(octets))


class TestFaitsPublies(unittest.TestCase):
    """Une seule chaine, tous les livrables, puis on cherche dedans."""

    @classmethod
    def setUpClass(cls):
        warnings.filterwarnings('ignore')
        logging.disable(logging.CRITICAL)
        from core.qualite_donnees import preambule_qualite
        from direction_non_vie.tarification import test_pipeline_agents as T
        from direction_non_vie.tarification.a1_ingestion.agent import (
            AgentA1Ingestion,
        )
        from direction_non_vie.tarification.a2_preprocessing.agent import (
            AgentA2Preprocessing,
        )
        from direction_non_vie.tarification.a3_glm.agent import AgentA3GLM
        from direction_non_vie.tarification.a4_ml.agent import AgentA4ML
        from direction_non_vie.tarification.a6_comparaison.agent import (
            AgentA6Comparaison,
        )

        np.random.seed(7)
        donnees = T._portefeuille_auto(1500)
        plan = T._PLAN_AUTO
        base = {'audit_path': '/tmp', 'verbose': False}
        r1 = AgentA1Ingestion(**base).run(branche='non_vie',
                                          sous_branche='auto',
                                          dataframe=donnees)
        qualite = preambule_qualite(r1.get('dataframe'), plan,
                                    qualite_validee_par='Test',
                                    horodatage=None)
        r1 = {**r1, 'dataframe': qualite.dataframe_propre}
        r2 = AgentA2Preprocessing(**base).run(result_a1=r1, plan=plan)
        r3 = AgentA3GLM(models_path='/tmp', audit_path='/tmp').run(
            result_a2=r2, plan=plan, col_frequence=plan.cible_frequence,
            col_cout=plan.cible_cout, generer_graphiques=False)
        cls.r4 = AgentA4ML(models_path='/tmp', audit_path='/tmp').run(
            result_a2=r2, result_a3=r3, plan=plan, col_cible='nb_sinistres',
            ponderer_par_exposition=True, calcul_shap=True,
            generer_graphiques=False)
        cls.r6 = AgentA6Comparaison(models_path='/tmp', audit_path='/tmp').run(
            result_a2=r2, result_a3=r3, result_a4=cls.r4, result_a5=None,
            col_cible='nb_sinistres', plan=plan, environnement='production',
            profil_valide_par='Test', generer_graphiques=False,
            generer_rapport_equipe=True)
        equipe = cls.r6.get('rapport_equipe') or {}
        cls.surfaces = {
            'Excel A4': _texte_xlsx(cls.r4.get('excel_bytes')),
            'Excel A6': _texte_xlsx(cls.r6.get('excel_bytes')),
            'Word A6': _texte_docx(cls.r6.get('word_bytes')),
            'HTML A6': _texte(cls.r6.get('html_bytes')),
            'Equipe HTML': _texte(equipe.get('html_bytes')),
            'Equipe Word': _texte_docx(equipe.get('word_bytes')),
        }

    @classmethod
    def tearDownClass(cls):
        logging.disable(logging.NOTSET)

    def test_FP0_les_surfaces_existent_vraiment(self):
        """⚠️ Sans ceci, TOUS les `assertNotIn` passeraient sur du vide."""
        vides = [nom for nom, t in self.surfaces.items() if len(t) < 500]
        self.assertEqual(vides, [], f'surfaces vides ou minuscules : {vides} '
                                    '-- les controles suivants ne mesureraient '
                                    'rien')

    def test_FP1_aucune_surface_ne_dit_toujours_1_0000(self):
        """⚠️⚠️ Le score reel du modele de production vaut 0,9062."""
        motif = re.compile(r'toujours\s*[=~≈]?\s*1[.,]0000')
        fautives = [nom for nom, t in self.surfaces.items() if motif.search(t)]
        self.assertEqual(fautives, [],
                         'ces surfaces affirment encore que le meilleur '
                         f'modele vaut toujours 1,0000 : {fautives}')
        # ... et la qualification JUSTE, elle, est bien publiee.
        for nom in ('Word A6', 'HTML A6'):
            self.assertIn('meilleur sur', self.surfaces[nom].lower(),
                          f'{nom} ne qualifie plus le score composite')

    def test_FP2_Deployer_ne_se_dit_pas_sous_AMBRE(self):
        """⚠️⚠️ Le statut de ce run est AMBRE : la recommandation ne peut pas
        etre une autorisation."""
        self.assertEqual(self.r6.get('statut_rag'), 'AMBRE',
                         'ce run n est plus AMBRE : le test ne surveille plus '
                         'le cas qu il vise')
        for nom, texte in self.surfaces.items():
            with self.subTest(surface=nom):
                for m in re.finditer(r'[→>-]\s*D[ée]ployer', texte):
                    debut = max(0, m.start() - 30)
                    self.assertIn('Ne PAS', texte[debut:m.end()],
                                  f'{nom} autorise le deploiement sous AMBRE : '
                                  f'...{texte[debut:m.end() + 60]}...')

    def test_FP3_la_reference_reglementaire_survit_a_la_piste_d_audit(self):
        """⚠️⚠️ `note_score_global` fait 205 caracteres ; la troncature a 120
        coupait juste avant << Ref. : ACPR-2022-P-01 §4.3 >>."""
        for nom in ('Word A6', 'HTML A6'):
            with self.subTest(surface=nom):
                self.assertIn('ACPR-2022-P-01', self.surfaces[nom],
                              'la reference reglementaire est coupee de la '
                              'piste d audit publiee')
                self.assertNotIn('RMSE)…', self.surfaces[nom],
                                 'la note est encore tronquee')

    def test_FP3b_rien_de_ce_que_nous_tronquons_n_est_reellement_tronque(self):
        """⚠️⚠️ IL DERIVE, ET SON ASSIETTE EST JUSTE. Ma premiere version
        interdisait le caractere << … >> dans TOUT le document : elle aurait
        rougi sur une narration LLM qui en contient un legitimement --
        l'assiette trop LARGE, la forme miroir du defaut habituel.

        Elle confronte maintenant les VALEURS REELLES du run aux bornes que
        nous leur appliquons. Trois textes etaient coupes dans le Word signe,
        en plein milieu d'une conclusion actuarielle :
            << Le GLM explique bien la sinistralite -- defendable devant… >>
        """
        from direction_non_vie.tarification.services.rapport_modeles_tarif import (
            LIMITE_TEXTE_HYPOTHESE,
            LIMITE_VALEUR_AUDIT,
        )
        coupes = []
        for cle, valeur in (self.r6.get('audit_trail') or {}).items():
            if len(str(valeur)) > LIMITE_VALEUR_AUDIT:
                coupes.append(f'audit_trail[{cle}] : {len(str(valeur))} car.')
        for source, nom in ((self.r4, 'A4'), (self.r6, 'A6')):
            for cle, hyp in (source.get('hypotheses') or {}).items():
                if not isinstance(hyp, dict):
                    continue
                for champ in ('message', 'conseil'):
                    n = len(str(hyp.get(champ) or ''))
                    if n > LIMITE_TEXTE_HYPOTHESE:
                        coupes.append(f'{nom}.{cle}.{champ} : {n} car.')
        self.assertEqual(coupes, [],
                         'ces textes sont coupes dans les livrables signes :\n'
                         '  ' + '\n  '.join(coupes))

    def test_FP3c_les_conclusions_d_hypotheses_sont_ENTIERES(self):
        """La contre-epreuve dans les OCTETS : les phrases mesurees comme
        coupees doivent maintenant se terminer."""
        for extrait in ('defendable devant', 'fragile(s)'):
            for nom in ('Word A6',):
                texte = self.surfaces[nom]
                if extrait not in texte:
                    continue
                i = texte.index(extrait) + len(extrait)
                self.assertNotEqual(
                    texte[i:i + 1], '…',
                    f'{nom} coupe encore la phrase juste apres {extrait!r}')

    def test_FP4_l_elasticite_atteint_les_QUATRE_surfaces(self):
        """⚠️⚠️ Elle n'atteignait que le classeur A6 et le rapport d'equipe.
        Dans le rapport de modeles, elle n'allait que dans LE PROMPT."""
        for nom in ('Excel A6', 'Word A6', 'HTML A6', 'Equipe HTML',
                    'Equipe Word'):
            with self.subTest(surface=nom):
                self.assertRegex(
                    self.surfaces[nom], r'(?i)[EÉ]lasticit',
                    f"{nom} ne publie AUCUNE trace de l'elasticite")

    def test_FP5_SHAP_porte_sur_un_modele_CALIBRE(self):
        """⚠️ Sur la chaine reelle : aucun modele introuvable ne doit sortir."""
        shap_vals = self.r4.get('shap_values') or {}
        erreur = str(shap_vals.get('erreur') or '')
        self.assertNotIn('non trouv', erreur,
                         f'SHAP demande sur un modele absent : {erreur}')
        self.assertNotIn('non trouv', self.surfaces['Excel A4'],
                         'le classeur A4 publie encore un modele introuvable')

    def test_FP5b_la_SELECTION_ecarte_la_ligne_de_reference_GLM(self):
        """⚠️⚠️ MON PREMIER FP-5 ATTESTAIT SANS SURVEILLER, ET LE SCEAU L'A
        MONTRE. Sur la fixture de ce fichier (1 500 lignes), `classement[0]`
        est deja un modele calibre : le plant qui remettait
        `classement[0]['modele']` ne changeait donc rien. C'est sur un
        portefeuille de 2 000 lignes que la ligne << GLM Poisson (reference
        A3) >> arrive PREMIERE. La regle se teste donc sur elle-meme, pas sur
        une fixture qui ne la sollicite pas."""
        from direction_non_vie.tarification.a4_ml.agent import (
            modele_a_expliquer,
        )
        classement = [
            {'modele': 'GLM Poisson (référence A3)', 'famille': 'GLM'},
            {'modele': 'gbm', 'famille': 'Arbres / Boosting'},
        ]
        self.assertEqual(
            modele_a_expliquer(classement, {'gbm': object()}), 'gbm',
            'SHAP porterait sur la ligne de reference relayee d A3')
        # ... et un nom absent de `modeles` est ecarte lui aussi.
        self.assertIsNone(
            modele_a_expliquer(classement, {}),
            'un modele qui n existe pas serait quand meme demande a SHAP')
        # ... la contre-epreuve : un ML calibre en tete reste choisi.
        self.assertEqual(
            modele_a_expliquer(list(reversed(classement)),
                               {'gbm': object()}), 'gbm')

    def test_FP6_le_nom_du_plan_atteint_les_OCTETS(self):
        """⚠️⚠️ CONTRE-PREUVE DU LOT 3, dont le decompte etait faux. Le
        libelle atteint le classeur A6 et le rapport d'equipe -- PAS le
        Word/HTML d'A6, qui ne le publient sur aucun chemin. On mesure donc
        la ou il arrive."""
        porteuses = [nom for nom, t in self.surfaces.items()
                     if 'ACTION REQUISE' in t]
        if not porteuses:
            self.skipTest("aucune colonne du plan ecartee sur cette fixture")
        for nom in porteuses:
            with self.subTest(surface=nom):
                self.assertNotIn("plan '?'", self.surfaces[nom],
                                 'le libelle ne nomme toujours pas le plan')
                self.assertIn("plan 'auto'", self.surfaces[nom])


if __name__ == '__main__':
    unittest.main(verbosity=2)
