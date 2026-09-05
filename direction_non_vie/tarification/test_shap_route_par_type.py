# -*- coding: utf-8 -*-
"""LE CLASSEUR A4 DISPARAISSAIT SUR UN SHAP QUI ECHOUE.

Trois defauts en chaine, mesures le 05/09/2026.

  1. LE ROUTAGE SHAP BRANCHAIT SUR LE NOM DU MODELE. Cinq des dix modeles du
     catalogue -- `lineaire_regularise`, `quantile_50`, `quantile_90`,
     `xgboost_optuna`, `xgboost_tweedie` -- tombaient dans le `else` et
     recevaient un `LinearExplainer`, y compris des modeles d'arbres.

  2. LE MEILLEUR MODELE DE LA FIXTURE EST UN `Pipeline`
     (`StandardScaler -> PoissonRegressor`). SHAP rendait
     << An unknown model type was passed: Pipeline >>.

  3. LE LECTEUR EXCEL FAISAIT `abs()` SUR LE MESSAGE D'ERREUR. Il lisait
     `shap_vals.get('importance_globale', shap_vals)` : le repli renvoyait LE
     DICTIONNAIRE LUI-MEME, donc `{'erreur': '<message>'}` passait le
     `isinstance(..., dict)` et `abs('<message>')` levait un TypeError, que
     l'exportateur transformait en `return b''`.

  | classeur A4 signe            | avant           | apres         |
  |------------------------------|-----------------|---------------|
  | SHAP reussi                  | 9 895 octets    | 9 895 octets  |
  | SHAP en echec                | **0** (perdu)   | 9 851 octets  |
  | SHAP non installe            | **0** (perdu)   | 9 826 octets  |
  | chaine reelle, cible primaire| **0** (perdu)   | 11 698 octets |

*Sur la cible PRIMAIRE de la fixture par defaut, le classeur A4 valait zero
octet -- et le statut restait AMBRE, sans que rien ne dise le livrable
disparu.*

ET UN QUATRIEME, PLUS DISCRET : `shap_absent` se derivait de la PROSE d'une
alerte (`any('SHAP ABSENT' in a for a in rapport['alertes'])`). Le drapeau ne
surveillait pas SHAP, il surveillait une chaine de caracteres dans un canal
annexe. *Un controle qui lit un texte ne surveille pas un comportement.*

Ce que cette sentinelle exige :
  SH-1  la route se DERIVE du modele (capacites), jamais de son nom ;
  SH-2  un `Pipeline` est deballe ET sa matrice transformee -- expliquer le
        dernier maillon avec la matrice d'ENTREE donnerait des valeurs
        calculees dans le mauvais espace ;
  SH-3  un modele sans explicateur possible est REFUSE nommement ;
  SH-4  le classeur SURVIT a un SHAP en echec, et l'onglet DIT la cause ;
  SH-5  le message ne diagnostique plus << installer le package >> quand shap
        est installe ;
  SH-6  `shap_absent` se derive du RESULTAT ;
  SH-7  sur la chaine reelle, le classeur A4 EXISTE.

Tout en `unittest.TestCase` : la gate lance `unittest discover`.
"""
import io
import logging
import os
import sys
import unittest
import warnings
from types import SimpleNamespace

_ICI = os.path.dirname(os.path.abspath(__file__))
_RACINE = os.path.dirname(os.path.dirname(_ICI))
for _c in (_RACINE, _ICI):
    if _c not in sys.path:
        sys.path.insert(0, _c)

import numpy as np

from direction_non_vie.tarification.a4_ml.agent import (
    _CAPACITES_ARBRE,
    _estimateur_et_matrice,
    _shap_est_absent,
)
from direction_non_vie.tarification.services.tarif_excel import export_excel_a4

_SOCLE_A4 = {
    'success': True, 'statut_rag': 'AMBRE', 'branche': 'auto',
    'col_cible': 'nb_sinistres',
    'classement': [{'modele': 'gbm', 'famille': 'Arbres / Boosting',
                    'gini_test': 0.13, 'rmse_test': 0.74,
                    'overfit_ratio': 3.58, 'score_global': 0.5,
                    'recommandation': 'ok'}],
    'metriques': {'gbm': {'gini_test': 0.13}},
    'validation_ml': {}, 'hypotheses': {}, 'audit_trail': {},
    'monitoring': {}, 'rapport': {}, 'commentaire': '',
}


def _cellules(octets):
    from openpyxl import load_workbook
    classeur = load_workbook(io.BytesIO(octets), data_only=True)
    textes = []
    for feuille in classeur.worksheets:
        for ligne in feuille.iter_rows():
            for cellule in ligne:
                if cellule.value is not None:
                    textes.append(str(cellule.value))
    return textes


class TestRoutageParType(unittest.TestCase):

    def test_SH1_un_pipeline_est_deballe_et_sa_matrice_TRANSFORMEE(self):
        """⚠️⚠️ DEBALLER NE SUFFIT PAS. Le pipeline vaut
        `StandardScaler -> PoissonRegressor` : expliquer l'estimateur final
        avec la matrice d'ENTREE donnerait des valeurs SHAP calculees dans le
        mauvais espace -- un chiffre faux publie comme importance de facteur
        dans un livrable signe."""
        from sklearn.linear_model import PoissonRegressor
        from sklearn.pipeline import Pipeline
        from sklearn.preprocessing import StandardScaler

        X = np.array([[1.0, 100.0], [2.0, 200.0], [3.0, 300.0], [4.0, 400.0]])
        y = np.array([0.0, 1.0, 0.0, 2.0])
        tuyau = Pipeline([('scaler', StandardScaler()),
                          ('poisson', PoissonRegressor(alpha=1.0))]).fit(X, y)

        estimateur, matrice, prefixe = _estimateur_et_matrice(tuyau, X)
        self.assertIsInstance(estimateur, PoissonRegressor,
                              'le pipeline n a pas ete deballe')
        self.assertIsNotNone(prefixe, 'aucun prefixe applique')
        self.assertFalse(np.allclose(matrice, X),
                         'la matrice n a PAS ete transformee : les valeurs '
                         'SHAP seraient calculees dans le mauvais espace')
        # ... et elle l'est bien PAR le prefixe du pipeline.
        np.testing.assert_allclose(matrice, tuyau[:-1].transform(X))

    def test_SH1b_l_enveloppe_d_exposition_est_deballee_aussi(self):
        """`_ModeleFrequenceExposition` expose l'estimateur nu sous `.base`."""
        nu = SimpleNamespace(coef_=np.array([1.0]))
        estimateur, _, prefixe = _estimateur_et_matrice(
            SimpleNamespace(base=nu), np.array([[1.0]]))
        self.assertIs(estimateur, nu, "l'enveloppe n a pas ete deballee")
        self.assertIsNone(prefixe)

    def test_SH1c_les_capacites_ne_sont_pas_des_NOMS_de_modeles(self):
        """⚠️ Une capacite ne se renomme pas ; un nom, si. Le routage
        branchait sur `nom_modele in ['gbm', 'random_forest']` -- et cinq
        modeles du catalogue n'y figuraient pas."""
        for capacite in _CAPACITES_ARBRE:
            self.assertTrue(capacite.endswith('_') or capacite.startswith('get_'),
                            f'{capacite!r} ressemble a un nom de modele, pas a '
                            "une capacite d'objet")
        for nom in ('gbm', 'xgboost', 'lightgbm', 'catboost', 'random_forest'):
            self.assertNotIn(nom, _CAPACITES_ARBRE)


class TestLaRouteEstEMPRUNTEE(unittest.TestCase):
    """⚠️⚠️ TROUVE PAR LE SCEAU. Verifier qu'un modele EXPOSE une capacite
    d'arbre ne prouve pas que `_calculer_shap` l'EMPRUNTE : le plant qui
    rebranchait sur le nom ne faisait rougir personne, parce que la fonction
    n'est appelee que sur le MEILLEUR modele -- un `Pipeline` sur cette
    fixture. Les modeles d'arbres au nom hors liste n'etaient jamais
    explique. On appelle donc la fonction directement."""

    def _agent_avec(self, nom, modele):
        from direction_non_vie.tarification.a4_ml.agent import AgentA4ML
        agent = AgentA4ML.__new__(AgentA4ML)
        agent.modeles = {nom: modele}
        return agent

    def test_SH8_un_arbre_au_nom_HORS_LISTE_recoit_un_TreeExplainer(self):
        """`xgboost_tweedie` ne figurait dans AUCUNE branche nommee : il
        recevait un `LinearExplainer`, qui leve sur un modele d'arbres."""
        try:
            from xgboost import XGBRegressor
        except ImportError:                                    # pragma: no cover
            self.skipTest('xgboost absent')
        X = np.random.default_rng(7).normal(size=(60, 3))
        y = (X[:, 0] > 0).astype(float)
        modele = XGBRegressor(n_estimators=5, max_depth=2,
                              verbosity=0).fit(X, y)
        agent = self._agent_avec('xgboost_tweedie', modele)
        resultat = agent._calculer_shap('xgboost_tweedie', X, ['a', 'b', 'c'])
        self.assertNotIn('erreur', resultat,
                         f"SHAP a echoue : {resultat.get('erreur')}")
        self.assertEqual(resultat.get('explicateur'), 'TreeExplainer',
                         'un modele d arbres a recu '
                         f"{resultat.get('explicateur')!r}")
        self.assertTrue(resultat.get('importance_globale'))

    def test_SH8b_un_lineaire_recoit_bien_un_LinearExplainer(self):
        """La contre-epreuve : la route ne bascule pas tout vers les arbres."""
        from sklearn.linear_model import PoissonRegressor
        X = np.abs(np.random.default_rng(7).normal(size=(60, 3)))
        y = np.random.default_rng(8).poisson(0.5, 60).astype(float)
        agent = self._agent_avec('lineaire_regularise',
                                 PoissonRegressor(alpha=1.0).fit(X, y))
        resultat = agent._calculer_shap('lineaire_regularise', X,
                                        ['a', 'b', 'c'])
        self.assertEqual(resultat.get('explicateur'), 'LinearExplainer',
                         resultat.get('erreur'))

    def test_SH8c_un_modele_sans_explicateur_est_REFUSE_NOMMEMENT(self):
        """Il ne doit pas recevoir un explicateur au hasard : la levee qui
        s'ensuivait coutait le classeur entier."""
        class _Opaque:
            def predict(self, X):
                return np.zeros(len(X))

        agent = self._agent_avec('opaque', _Opaque())
        resultat = agent._calculer_shap('opaque', np.zeros((4, 2)), ['a', 'b'])
        self.assertIn('erreur', resultat)
        self.assertIn('_Opaque', resultat['erreur'],
                      "le refus ne NOMME pas le type : l'actuaire ne peut pas "
                      'savoir quoi corriger')


class TestDrapeauShapAbsent(unittest.TestCase):
    """⚠️⚠️ SH-6 ATTESTAIT SANS SURVEILLER, ET LE SCEAU L'A MONTRE. Sa
    premiere version verifiait qu'aucune alerte ne dit << SHAP ABSENT >> sur
    un run ou SHAP a REUSSI -- vrai avec l'ancienne derivation comme avec la
    nouvelle. Le plant qui remettait la lecture de prose ne faisait donc
    rougir personne. La regle est desormais une FONCTION, et c'est elle qu'on
    interroge, sur les cas ou les deux derivations DIVERGENT."""

    def test_SH6a_un_SHAP_en_echec_declare_l_absence(self):
        """⚠️ Le cas decisif : l'ancienne derivation rendait `False` ici (il
        n'y a AUCUNE alerte), donc le statut pouvait passer VERT sans la
        moindre interpretabilite."""
        for echec in ({'erreur': 'An unknown model type was passed'},
                      {'erreur': 'SHAP non installé'},
                      {}, None, 'pas un dictionnaire'):
            with self.subTest(cas=repr(echec)[:40]):
                self.assertTrue(_shap_est_absent(echec))

    def test_SH6b_une_importance_reelle_ne_declare_pas_l_absence(self):
        self.assertFalse(_shap_est_absent(
            {'importance_globale': {'age': 0.31}, 'explicateur': 'TreeExplainer'}))

    def test_SH6c_une_importance_VIDE_est_une_absence(self):
        """Un dictionnaire vide n'est pas une importance : le publier
        laisserait le statut monter au VERT sur zero facteur explique."""
        self.assertTrue(_shap_est_absent({'importance_globale': {}}))
        self.assertTrue(_shap_est_absent({'importance_globale': None}))

    def test_SH6d_le_drapeau_INTERDIT_reellement_le_VERT(self):
        """La contre-epreuve du cablage : sans elle, la derivation pourrait
        etre juste et le parametre ignore."""
        from direction_non_vie.tarification.a4_ml.agent import AgentA4ML
        agent = AgentA4ML.__new__(AgentA4ML)
        agent._cible_run = 'nb_sinistres'
        classement = [{'modele': 'gbm', 'famille': 'Arbres / Boosting',
                       'gini_test': 0.40, 'overfit_alerte': False}]
        r_a3 = {'success': True,
                'metriques': {'poisson': {'cible': 'nb_sinistres',
                                          'gini': 0.10}}}
        self.assertEqual(
            agent._calculer_statut_rag(classement, r_a3, shap_absent=False),
            'VERT')
        self.assertEqual(
            agent._calculer_statut_rag(classement, r_a3, shap_absent=True),
            'AMBRE',
            'le drapeau est calcule mais ne change RIEN au statut')


class TestLeClasseurSurvit(unittest.TestCase):

    def test_SH4_un_SHAP_en_echec_ne_coute_plus_le_classeur(self):
        """⚠️⚠️ MESURE : il valait `b''`, c'est-a-dire RIEN."""
        octets = export_excel_a4({
            **_SOCLE_A4,
            'shap_values': {'erreur': 'An unknown model type was passed: '
                                      "<class 'sklearn.pipeline.Pipeline'>"}})
        self.assertTrue(octets, 'le classeur A4 a disparu sur un SHAP en echec')
        textes = _cellules(octets)
        self.assertTrue(any('SHAP non calcule' in t for t in textes),
                        "l'onglet ne DIT pas l'absence")
        self.assertTrue(any('unknown model type' in t for t in textes),
                        "l'onglet ne dit pas la CAUSE : l'actuaire ne peut "
                        'pas savoir quoi corriger')

    def test_SH5_le_message_ne_diagnostique_plus_a_tort(self):
        """<< installer le package 'shap' >> etait FAUX des que shap etait
        installe mais que l'explicateur avait echoue."""
        octets = export_excel_a4({**_SOCLE_A4,
                                  'shap_values': {'erreur': 'Pipeline inconnu'}})
        textes = _cellules(octets)
        self.assertFalse(any("installer le package" in t for t in textes),
                         'le classeur conseille encore d installer shap alors '
                         "que la cause est ailleurs")

    def test_SH4b_les_quatre_etats_de_SHAP_laissent_un_classeur(self):
        etats = {
            'reussi': {'modele': 'gbm', 'n_contrats': 1000,
                       'explicateur': 'TreeExplainer',
                       'importance_globale': {'age': 0.31, 'bm': 0.22},
                       'top_feature': 'age'},
            'en echec': {'erreur': 'explicateur inconnu'},
            'vide': {},
            'non installe': {'erreur': 'SHAP non installé'},
        }
        for nom, valeurs in etats.items():
            with self.subTest(etat=nom):
                self.assertTrue(
                    export_excel_a4({**_SOCLE_A4, 'shap_values': valeurs}),
                    f'classeur perdu sur un SHAP {nom}')

    def test_SH4c_une_importance_NON_NUMERIQUE_ne_passe_pas_pour_une_mesure(self):
        """⚠️ Le repli `.get('importance_globale', shap_vals)` rendait le
        dictionnaire d'erreur lui-meme. Un dict dont les valeurs ne sont pas
        des nombres n'est pas une importance."""
        octets = export_excel_a4({
            **_SOCLE_A4,
            'shap_values': {'importance_globale': {'age': 'beaucoup'}}})
        self.assertTrue(octets)
        self.assertTrue(any('SHAP non calcule' in t for t in _cellules(octets)))

    def test_SH4d_l_explicateur_utilise_est_PUBLIE(self):
        """Une importance calculee par un explicateur lineaire sur un modele
        d'arbres n'a pas le meme sens : le livrable doit dire lequel a servi."""
        octets = export_excel_a4({
            **_SOCLE_A4,
            'shap_values': {'importance_globale': {'age': 0.31},
                            'explicateur': 'TreeExplainer'}})
        self.assertTrue(any('TreeExplainer' in t for t in _cellules(octets)))


class TestSurLaChaineReelle(unittest.TestCase):

    @classmethod
    def setUpClass(cls):
        warnings.filterwarnings('ignore')
        logging.disable(logging.CRITICAL)
        from core.qualite_donnees import preambule_qualite
        from direction_non_vie.tarification import test_pipeline_agents as T
        from direction_non_vie.tarification.a1_ingestion.agent import (
            AgentA1Ingestion,
        )
        from direction_non_vie.tarification.a2_preprocessing.agent import (
            AgentA2Preprocessing,
        )
        from direction_non_vie.tarification.a3_glm.agent import AgentA3GLM
        from direction_non_vie.tarification.a4_ml.agent import AgentA4ML

        np.random.seed(7)
        donnees = T._portefeuille_auto(1500)
        plan = T._PLAN_AUTO
        base = {'audit_path': '/tmp', 'verbose': False}
        r1 = AgentA1Ingestion(**base).run(branche='non_vie',
                                          sous_branche='auto',
                                          dataframe=donnees)
        qualite = preambule_qualite(r1.get('dataframe'), plan,
                                    qualite_validee_par='Test',
                                    horodatage=None)
        r1 = {**r1, 'dataframe': qualite.dataframe_propre}
        r2 = AgentA2Preprocessing(**base).run(result_a1=r1, plan=plan)
        r3 = AgentA3GLM(models_path='/tmp', audit_path='/tmp').run(
            result_a2=r2, plan=plan, col_frequence=plan.cible_frequence,
            col_cout=plan.cible_cout, generer_graphiques=False)
        cls.agent = AgentA4ML(models_path='/tmp', audit_path='/tmp')
        cls.r4 = cls.agent.run(
            result_a2=r2, result_a3=r3, plan=plan, col_cible='nb_sinistres',
            ponderer_par_exposition=True, calcul_shap=True,
            generer_graphiques=False)

    @classmethod
    def tearDownClass(cls):
        logging.disable(logging.NOTSET)

    def test_SH2_SHAP_aboutit_sur_le_modele_retenu(self):
        """⚠️⚠️ Sur cette fixture le meilleur modele est un `Pipeline` --
        exactement le cas qui echouait."""
        shap_vals = self.r4.get('shap_values')
        self.assertIsInstance(shap_vals, dict)
        self.assertNotIn('erreur', shap_vals,
                         f"SHAP a echoue : {shap_vals.get('erreur')}")
        self.assertTrue(shap_vals.get('importance_globale'))
        self.assertIn(shap_vals.get('explicateur'),
                      ('TreeExplainer', 'LinearExplainer'))

    def test_SH3_chaque_modele_calibre_recoit_un_explicateur(self):
        """Aucun ne doit tomber dans un `else` qui devine."""
        self.assertTrue(self.agent.modeles, 'aucun modele calibre')
        orphelins = []
        for nom, modele in self.agent.modeles.items():
            # ⚠️ La matrice se dimensionne SUR LE MODELE : un scaler ajuste
            # sur 23 variables refuse une matrice 2x2, et le test echouerait
            # sur son propre outillage plutot que sur le code mesure.
            nu = getattr(modele, 'base', modele)
            largeur = int(getattr(nu, 'n_features_in_', 0)
                          or getattr(getattr(nu, 'steps', [(None, nu)])[0][1],
                                     'n_features_in_', 0) or 2)
            estimateur, _, _ = _estimateur_et_matrice(
                modele, np.zeros((2, largeur)))
            arbre = any(hasattr(estimateur, c) for c in _CAPACITES_ARBRE)
            lineaire = hasattr(estimateur, 'coef_')
            if not (arbre or lineaire):
                orphelins.append(f'{nom} ({type(estimateur).__name__})')
        self.assertEqual(orphelins, [],
                         'ces modeles n ont aucun explicateur derivable : '
                         + ', '.join(orphelins))

    def test_SH6_le_run_reel_ne_declare_pas_SHAP_absent_quand_il_aboutit(self):
        self.assertFalse(_shap_est_absent(self.r4['shap_values']),
                         'SHAP a abouti mais le drapeau le dit absent')

    def test_SH3b_un_modele_d_arbres_HORS_ANCIENNE_LISTE_est_bien_route(self):
        """⚠️⚠️ TROUVE PAR LE SCEAU. Le plant qui rebranche sur le NOM ne
        faisait rougir personne : `_estimateur_et_matrice` deballe le
        `Pipeline` AVANT le test, donc `lineaire_regularise` retombait sur le
        `elif coef_` et fonctionnait quand meme. C'est le deballage qui
        sauvait la chaine -- pas le passage du nom a la capacite.

        Le cas que seul le nom decide est celui-ci : un modele D'ARBRES dont
        le nom ne figurait PAS dans l'ancienne liste
        `['gbm','random_forest','xgboost','lightgbm','catboost']`.
        `xgboost_tweedie` en est un, et il est reellement calibre.
        """
        self.assertIn('xgboost_tweedie', self.agent.modeles,
                      'ce modele n est pas calibre : le test ne surveille pas '
                      'le cas qu il vise')
        modele = self.agent.modeles['xgboost_tweedie']
        nu = getattr(modele, 'base', modele)
        largeur = int(getattr(nu, 'n_features_in_', 0) or 2)
        estimateur, _, _ = _estimateur_et_matrice(modele,
                                                  np.zeros((2, largeur)))
        self.assertTrue(
            any(hasattr(estimateur, c) for c in _CAPACITES_ARBRE),
            f'{type(estimateur).__name__} n est pas reconnu comme un modele '
            "d'arbres : il recevrait un explicateur lineaire")
        self.assertFalse(hasattr(estimateur, 'coef_'),
                         'ce modele expose `coef_` : le test ne distingue pas '
                         'les deux routes')

    def test_SH7_le_classeur_A4_EXISTE_sur_la_chaine_reelle(self):
        """⚠️⚠️ IL VALAIT ZERO OCTET. Sur la cible PRIMAIRE."""
        octets = self.r4.get('excel_bytes') or b''
        self.assertTrue(octets,
                        'le classeur A4 de la cible primaire est vide : le '
                        'livrable signe a disparu')
        self.assertGreater(len(octets), 5000, f'{len(octets)} octets seulement')
        textes = _cellules(octets)
        self.assertTrue(any('SHAP' in t for t in textes))
        self.assertFalse(any('SHAP non calcule' in t for t in textes),
                         "l'onglet dit que SHAP n'a pas ete calcule alors "
                         "qu'il l'a ete")


if __name__ == '__main__':
    unittest.main(verbosity=2)
