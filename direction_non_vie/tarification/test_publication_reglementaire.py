# -*- coding: utf-8 -*-
"""PUBLIER CE QU'ON POSSEDE, DECLARER CE QU'ON NE PRODUIT PAS.

Quatre agents de la reglementation lisent `result_a6` -- A8, A9, A10, A11.
Trois des cles qu'A8 y cherche N'EXISTENT PAS, et il les remplace par des
litteraux (releve par AST le 05/09/2026) :

    a8:328  result_a6.get('gini', 0.25)
    a8:329  result_a6.get('loss_ratio_attendu', 0.72)
    a8:330  result_a6.get('modele_retenu', 'N/A')

Le 0,72 se pose donc a CHAQUE run, il est journalise sous << A6 branche >>,
et `a8:1084` fait `if lr_attendu > 0.85` pour declencher l'action
<< Resserrer les criteres de souscription >> : cette recommandation est
STRUCTURELLEMENT inatteignable.

  *Declarer ce qu'on ne produit pas vaut mieux que laisser un consommateur
  l'inventer.*

⚠️⚠️ ET LA PUBLICATION NE CABLE RIEN. Elle vit sous SA PROPRE CLE
(`publication_reglementaire`), jamais au premier niveau : publier `gini` et
`modele_retenu` EN HAUT ferait trouver ces cles a A8 et cablerait la
frontiere par la bande. **Mesure : le cablage deplace +36,2 % de SCR, soit
1 435 571 EUR.** Cette decision n'est pas prise -- la porte s'ouvre, on ne
la franchit pas. PR-4 rend cette interdiction MECANIQUE.

Ce que cette sentinelle exige :
  PR-1  la publication porte ce qu'A6 POSSEDE, chaque valeur avec sa
        PROVENANCE ;
  PR-2  elle DECLARE ce que la tarification ne produit pas, avec le motif ;
  PR-3  une grandeur absente n'est pas remplie d'un repli -- elle manque ;
  PR-4  **AUCUNE des cles lues par la reglementation n'apparait au premier
        niveau de `result_a6`** : la frontiere reste non cablee ;
  PR-5  cela atteint le classeur A6 SIGNE.

Tout en `unittest.TestCase` : la gate lance `unittest discover`.
"""
import io
import logging
import os
import sys
import unittest
import warnings

_ICI = os.path.dirname(os.path.abspath(__file__))
_RACINE = os.path.dirname(os.path.dirname(_ICI))
for _c in (_RACINE, _ICI):
    if _c not in sys.path:
        sys.path.insert(0, _c)

import numpy as np

from direction_non_vie.tarification.contrat_sortie import (
    NON_PRODUIT_PAR_LA_TARIFICATION,
    publication_reglementaire,
)

#: Les cles que la reglementation lit sur `result_a6` et qu'A6 NE PUBLIE PAS.
#: ⚠️ Relevees par AST, pas recopiees de memoire — voir `test_contrat_sortie`
#: (CS-5), qui fige la meme liste sous l'angle du contrat de sortie.
_LUES_ET_ABSENTES = ('gini', 'loss_ratio_attendu', 'modele_retenu')


class TestPublication(unittest.TestCase):

    def _a6(self):
        return {
            'statut_rag': 'AMBRE', 'branche': 'auto', 'audit_id': 'A6_TEST',
            'modele_production': {'modele': 'GLM_POISSON',
                                  'gini_test': 0.1406,
                                  'score_global': 0.805,
                                  'cible': 'nb_sinistres'},
        }

    def test_PR1_chaque_valeur_publiee_porte_sa_PROVENANCE(self):
        """*Un chiffre reglementaire sans l'endroit d'ou il vient n'est pas
        contestable* — meme exigence que la `source` d'un seuil grave."""
        pub = publication_reglementaire(self._a6())
        possede = pub['possede']
        self.assertTrue(possede, 'rien de publie')
        for nom, entree in possede.items():
            with self.subTest(champ=nom):
                self.assertIn('valeur', entree)
                self.assertTrue(str(entree.get('provenance') or '').strip(),
                                f'{nom} est publie SANS provenance')
        self.assertEqual(possede['modele_retenu']['valeur'], 'GLM_POISSON')
        self.assertEqual(possede['gini_test']['provenance'],
                         'A6.modele_production.gini_test')

    def test_PR2_elle_declare_ce_qu_elle_ne_produit_PAS(self):
        pub = publication_reglementaire(self._a6())
        for nom in ('primes_acquises', 'loss_ratio_attendu'):
            with self.subTest(champ=nom):
                self.assertIn(nom, pub['non_produit'])
                self.assertGreater(len(pub['non_produit'][nom]), 60,
                                   'le motif ne dit pas POURQUOI')
        self.assertIn('non', pub['avertissement'].lower())

    def test_PR2b_une_grandeur_declaree_non_produite_n_est_PAS_publiee(self):
        """⚠️ Sinon la declaration se contredirait elle-meme."""
        pub = publication_reglementaire(self._a6())
        for nom in NON_PRODUIT_PAR_LA_TARIFICATION:
            self.assertNotIn(nom, pub['possede'],
                             f'{nom} est a la fois publie et declare non '
                             'produit')

    def test_PR3_une_grandeur_absente_MANQUE_au_lieu_d_etre_remplie(self):
        """⚠️⚠️ Sans A3, il n'y a pas de prime pure moyenne : elle ne doit
        pas apparaitre avec un repli. Et sans plan, pas d'empreinte."""
        pub = publication_reglementaire(self._a6(), result_a3=None, plan=None)
        self.assertNotIn('prime_pure_moyenne', pub['possede'])
        self.assertNotIn('empreinte_plan', pub['possede'])
        # ... et elle APPARAIT quand la source existe.
        avec = publication_reglementaire(
            self._a6(),
            result_a3={'metriques': {'tweedie': {'prime_pure_moy_pred': 910.92}}})
        self.assertEqual(avec['possede']['prime_pure_moyenne']['valeur'],
                         910.92)
        self.assertIn('A3.metriques.tweedie',
                      avec['possede']['prime_pure_moyenne']['provenance'])

    def test_PR3b_un_resultat_vide_ne_fabrique_rien(self):
        pub = publication_reglementaire(None)
        for entree in pub['possede'].values():
            self.assertIsNone(entree['valeur'],
                              'une valeur est apparue sans source')


class TestLaFrontiereResteNonCablee(unittest.TestCase):
    """⚠️⚠️ L'INTERDICTION, RENDUE MECANIQUE. Publier ouvre la porte ; le
    cablage deplacerait **+36,2 % de SCR, soit 1 435 571 EUR**, sans qu'aucune
    decision ne soit prise. Ce test tombera le jour ou quelqu'un remontera une
    de ces cles au premier niveau -- volontairement ou non."""

    @classmethod
    def setUpClass(cls):
        warnings.filterwarnings('ignore')
        logging.disable(logging.CRITICAL)
        from core.qualite_donnees import preambule_qualite
        from direction_non_vie.tarification import test_pipeline_agents as T
        from direction_non_vie.tarification.a1_ingestion.agent import (
            AgentA1Ingestion,
        )
        from direction_non_vie.tarification.a2_preprocessing.agent import (
            AgentA2Preprocessing,
        )
        from direction_non_vie.tarification.a3_glm.agent import AgentA3GLM
        from direction_non_vie.tarification.a4_ml.agent import AgentA4ML
        from direction_non_vie.tarification.a6_comparaison.agent import (
            AgentA6Comparaison,
        )

        np.random.seed(7)
        donnees = T._portefeuille_auto(1200)
        cls.plan = T._PLAN_AUTO
        base = {'audit_path': '/tmp', 'verbose': False}
        r1 = AgentA1Ingestion(**base).run(branche='non_vie',
                                          sous_branche='auto',
                                          dataframe=donnees)
        qualite = preambule_qualite(r1.get('dataframe'), cls.plan,
                                    qualite_validee_par='Test',
                                    horodatage=None)
        r2 = AgentA2Preprocessing(**base).run(
            result_a1={**r1, 'dataframe': qualite.dataframe_propre},
            plan=cls.plan)
        r3 = AgentA3GLM(models_path='/tmp', audit_path='/tmp').run(
            result_a2=r2, plan=cls.plan,
            col_frequence=cls.plan.cible_frequence,
            col_cout=cls.plan.cible_cout, generer_graphiques=False)
        r4 = AgentA4ML(models_path='/tmp', audit_path='/tmp').run(
            result_a2=r2, result_a3=r3, plan=cls.plan,
            col_cible='nb_sinistres', ponderer_par_exposition=True,
            calcul_shap=False, generer_graphiques=False)
        cls.r6 = AgentA6Comparaison(models_path='/tmp',
                                    audit_path='/tmp').run(
            result_a2=r2, result_a3=r3, result_a4=r4, result_a5=None,
            col_cible='nb_sinistres', plan=cls.plan,
            environnement='production', profil_valide_par='Test',
            generer_graphiques=False, generer_rapport_equipe=False)

    @classmethod
    def tearDownClass(cls):
        logging.disable(logging.NOTSET)

    def test_PR4_aucune_cle_lue_par_la_reglementation_au_PREMIER_NIVEAU(self):
        """⚠️⚠️ LE COEUR DE L'INTERDICTION."""
        remontees = [c for c in _LUES_ET_ABSENTES if c in self.r6]
        self.assertEqual(
            remontees, [],
            f'ces cles sont remontees au premier niveau de `result_a6` : '
            f'{remontees}. `a8_stress_testing` les lit avec un repli -- les '
            f'publier ici CABLE la frontiere, ce qui deplace +36,2 % de SCR '
            f'(1 435 571 EUR) sans decision. Voir le lot 15.')

    def test_PR4b_la_publication_EXISTE_pourtant(self):
        """La porte est bien ouverte : c'est la franchir qui est interdit."""
        pub = self.r6.get('publication_reglementaire') or {}
        self.assertTrue(pub.get('possede'), 'rien n est publie')
        self.assertEqual(pub['possede']['modele_retenu']['valeur'],
                         (self.r6.get('modele_production') or {}).get('modele'))
        self.assertIn('empreinte_plan', pub['possede'],
                      "l'empreinte du plan ne scelle pas la publication")
        self.assertTrue(
            pub['possede']['empreinte_plan']['valeur'].startswith('s'),
            'empreinte non versionnee')

    def test_PR5_le_classeur_A6_signe_porte_la_publication(self):
        from openpyxl import load_workbook
        octets = self.r6.get('excel_bytes') or b''
        self.assertTrue(octets, 'aucun classeur A6')
        classeur = load_workbook(io.BytesIO(octets), data_only=True)
        texte = '\n'.join(str(c.value) for f in classeur.worksheets
                          for ligne in f.iter_rows() for c in ligne
                          if c.value is not None)
        self.assertIn('PUBLICATION À L', texte,
                      'la publication n atteint pas le classeur signe')
        self.assertIn('A6.modele_production.gini_test', texte,
                      'la PROVENANCE n est pas publiee')
        self.assertIn('NON PRODUIT', texte,
                      'le classeur ne declare pas ce qui n est pas produit')


if __name__ == '__main__':
    unittest.main(verbosity=2)
