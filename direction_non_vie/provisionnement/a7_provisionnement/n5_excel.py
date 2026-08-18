# =============================================================================
#  ActuarIA — Agent A7 Ibrahim
#  n5_excel.py  —  Export Excel 10 onglets
# =============================================================================
#
#  10 onglets :
#  1. Synthèse              — KPIs + statuts couleur
#  2. Triangle brut         — triangle cumulé avec mise en forme heatmap
#  3. Facteurs CL           — facteurs + cumulés + % développé
#  4. Résultats méthodes    — CL / Mack / BF / CC / Bootstrap / BE S2
#  5. IBNR par année        — détail CL / Mack / BF / CC
#  6. Validation hypothèses — H1/H2/H3/H4 avec statuts colorés
#  7. Bootstrap ODP         — percentiles + distribution (stats)
#  8. Comparatif N-1/N      — variation BE / IBNR / LR entre arrêtés
#  9. SCR formule standard  — Art. 115 S2 avec facteurs EIOPA
# 10. Sensibilités          — tornado des impacts sur BE
#
#  Style : palette Navy/Gold ActuarIA, polices Arial, formats €/%
#
# =============================================================================

import io
import logging
from datetime import datetime
from typing import Dict, Optional

import numpy as np

# Formatage du loss ratio : source unique, pour qu'une méthode non calculée
# n'apparaisse jamais comme un loss ratio de 0 %.
from .methodes_be import ORDRE_AFFICHAGE, libelle, reserve
from .n2_hypotheses_clm import lignes_correlations_h1
from .n2_hypotheses_bfcc import lignes_hypotheses_bfcc
from .n2_hypotheses_clm import lignes_hypotheses_clm
from .n2_hypotheses_bootstrap import lignes_hypotheses_bootstrap
from .n2_hypotheses_munich import lignes_hypotheses_munich
from .n3.bf_cape_cod import libelle_loss_ratio
from .n3.bootstrap_odp import libelle_incertitude
# Source UNIQUE d'affichage de Munich CL — la même que HTML et Word.
from .n3.munich_cl import lignes_munich_rapport

logger = logging.getLogger('actuaria.a7')

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False
    logger.warning("openpyxl non disponible — export Excel désactivé")


# ── Couleurs ActuarIA ────────────────────────────────────────────────────────
NAVY       = "0F2E52"
GOLD       = "C9A84C"
BLANC      = "FFFFFF"
GRIS_CLAIR = "F0F4F8"
GRIS_MED   = "D0D8E4"
VERT_S2    = "1D7A3A"
AMBRE_S2   = "B87A00"
ROUGE_S2   = "C0392B"
BLEU_SOFT  = "1B3A5C"
NOIR       = "1A1A1A"

# ── Formats numériques ───────────────────────────────────────────────────────
FMT_EUR   = '#\u202f##0\u202f€;-#\u202f##0\u202f€'
FMT_EUR2  = '# ##0 €;-# ##0 €'
FMT_PCT   = '0.0%'
FMT_PCT2  = '0.00%'
FMT_DEC4  = '0.0000'
FMT_DEC6  = '0.000000'
FMT_NB    = '# ##0'


# =============================================================================
#  HELPERS DE STYLE
# =============================================================================

def _font(bold=False, color=NOIR, size=10, italic=False):
    return Font(name='Arial', bold=bold, color=color, size=size, italic=italic)

def _fill(hex_color):
    return PatternFill('solid', start_color=hex_color, fgColor=hex_color)

def _align(h='left', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _border_thin():
    s = Side(style='thin', color=GRIS_MED)
    return Border(left=s, right=s, top=s, bottom=s)

def _border_bottom():
    s = Side(style='medium', color=GOLD)
    return Border(bottom=s)

def _col_w(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width

def _row_h(ws, row, height):
    ws.row_dimensions[row].height = height

def _header(ws, row, col, text, width=None):
    """Cellule d'en-tête Navy/Or."""
    c = ws.cell(row=row, column=col, value=text)
    c.font      = _font(bold=True, color=GOLD, size=10)
    c.fill      = _fill(NAVY)
    c.alignment = _align(h='center')
    c.border    = _border_thin()
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return c

def _kpi(ws, row, col, label, value, fmt=None, statut=None):
    """Ligne KPI : label + valeur + pastille statut optionnelle."""
    cl = ws.cell(row=row, column=col, value=label)
    cl.font      = _font(bold=True, color=BLEU_SOFT, size=10)
    cl.fill      = _fill(GRIS_CLAIR)
    cl.alignment = _align()
    cl.border    = _border_thin()

    cv = ws.cell(row=row, column=col+1, value=value)
    cv.font      = _font(bold=True, color=NOIR, size=10)
    cv.alignment = _align(h='right')
    cv.border    = _border_thin()
    if fmt:
        cv.number_format = fmt

    if statut:
        cs = ws.cell(row=row, column=col+2)
        cs.value     = {'VERT':'✓ Conforme','AMBRE':'△ A surveiller','ROUGE':'✗ Attention'}.get(statut, statut)
        cs.font      = _font(bold=True, color={
            'VERT': VERT_S2, 'AMBRE': AMBRE_S2, 'ROUGE': ROUGE_S2
        }.get(statut, NOIR), size=10)
        cs.alignment = _align(h='center')
        cs.fill      = _fill({'VERT':'EAF3DE','AMBRE':'FAEEDA','ROUGE':'FCEBEB'}.get(statut, BLANC))
        cs.border    = _border_thin()

def _titre_section(ws, row, col, text, n_cols=5):
    """Titre de section fusionné Navy/Or."""
    c = ws.cell(row=row, column=col, value=text)
    c.font      = _font(bold=True, color=GOLD, size=11)
    c.fill      = _fill(NAVY)
    c.alignment = _align(h='left')
    c.border    = _border_thin()
    if n_cols > 1:
        ws.merge_cells(
            start_row=row, start_column=col,
            end_row=row,   end_column=col + n_cols - 1
        )
    _row_h(ws, row, 20)
    return c

def _data_row(ws, row, values, formats=None, bold=False, bg=None):
    """Ligne de données standard."""
    for j, val in enumerate(values):
        c = ws.cell(row=row, column=j+1, value=val)
        c.font      = _font(bold=bold, color=NOIR, size=10)
        c.alignment = _align(h='right' if isinstance(val, (int,float)) else 'left')
        c.border    = _border_thin()
        if formats and j < len(formats) and formats[j]:
            c.number_format = formats[j]
        if bg:
            c.fill = _fill(bg)
    return ws


# =============================================================================
#  ONGLET 1 — SYNTHÈSE
# =============================================================================

def _ong1_synthese(wb, n1, n2, n3, n4, ref_client, date_str):
    ws = wb.create_sheet("1. Synthèse")
    ws.sheet_view.showGridLines = False

    # En-tête
    _titre_section(ws, 1, 1, f"ACTUARIA — Rapport de Provisionnement Non-Vie | {ref_client} | {date_str}", 6)
    _row_h(ws, 1, 24)

    # Bloc BE S2
    _titre_section(ws, 3, 1, "BEST ESTIMATE S2 (Art. 77)", 6)
    be     = n4['best_estimate']
    statut = n4['statut']
    kpis = [
        ("Best Estimate S2",       be,                      FMT_NB, statut),
        ("Provision prudentielle P75", n4['reserve_p75'],   FMT_NB, None),
        ("Provision stress test P90",  n4['reserve_p90'],   FMT_NB, None),
        ("Provision extrême P99.5",    n4['reserve_p99_5'], FMT_NB, None),
        ("Incertitude Mack (σ)",       n4['sigma_mack'],    FMT_NB, None),
        ("CV inter-méthodes",          n4['cv_inter_methodes'] / 100, FMT_PCT, None),
    ]
    for i, (lbl, val, fmt, st) in enumerate(kpis):
        _kpi(ws, 4+i, 1, lbl, val, fmt, st)

    # Bloc SCR
    _titre_section(ws, 11, 1, "SCR PROVISIONS (Art. 115 S2)", 6)
    scr = n4.get('scr', {})
    _kpi(ws, 12, 1, "SCR Provisions",            scr.get('scr_provisions', 0), FMT_NB)
    _kpi(ws, 13, 1, "Facteur σ(LoB) EIOPA",     scr.get('sigma_eiopa', 0),   FMT_PCT2)
    _kpi(ws, 14, 1, "Ratio SCR/BE",              scr.get('ratio_scr_be', 0),  FMT_PCT)
    _kpi(ws, 15, 1, "Branche (LoB)",             scr.get('lob_label', '—'),   None)

    # Bloc méthodes
    _titre_section(ws, 17, 1, "RÉSULTATS PAR MÉTHODE", 6)
    # ⚠️ LE FAUX ZÉRO (lot C3a). La liste était écrite en dur et lisait
    # `n3['bf']['reserve_totale']` sans garde : sans exposition, la ligne
    # affichait « Bornhuetter-Ferguson | 0 | ✗ Attention », ce qui se lit
    # « BF donne une réserve nulle » là où la vérité est « BF n'a pas pu
    # être calculée ». Le motif vient de N4, qui le publie depuis ce lot.
    motifs = n4.get('methodes_exclues_motifs', {})
    for i, m in enumerate(ORDRE_AFFICHAGE):
        r    = reserve(n3, m)
        poid = n4.get('poids', {}).get(m, 0)
        if r is None:
            _kpi(ws, 18+i, 1, libelle(m),
                 motifs.get(m, 'non calculée'), None, '—')
        else:
            st = 'VERT' if m in n4.get('methodes_incluses', []) else 'ROUGE'
            _kpi(ws, 18+i, 1, libelle(m), r, FMT_NB, st)
        w = ws.cell(row=18+i, column=4, value=poid)
        w.number_format = FMT_PCT
        w.font      = _font(size=10)
        w.alignment = _align(h='center')
        w.border    = _border_thin()

    # Bloc hypothèses
    _titre_section(ws, 23, 1, "VALIDATION HYPOTHÈSES", 6)
    for i, (key, lbl) in enumerate([
        ('h1_independance','H1 Indépendance'),
        ('h2_stabilite','H2 Stabilité'),
    ]):
        h = n2.get(key, {})
        # ⚠️ SECOND SITE EXCEL, ET LE LOT F3 N'AVAIT FERME QUE L'AUTRE (onglet
        # « Validation hypothèses »). Ici, en synthese, une hypothese non
        # testable sortait « ✓ Conforme » EN VERT avec son score par defaut.
        _st = str(h.get('statut', 'VALIDÉE' if h.get('ok') else 'REJETÉE'))
        _nt = _st == 'NON TESTABLE'
        # ⚠️ TABLE PLUTOT QUE BRANCHE SUR `ok` : un statut inconnu tombe sur
        # AMBRE (« à surveiller »), jamais sur VERT. Le filet a refuse la
        # premiere version, qui rebranchait sur `ok` pour choisir la couleur.
        _kpi(ws, 24+i, 1, lbl, '—' if _nt else h.get('score', 0), None,
             {'VALIDÉE': 'VERT', 'REJETÉE': 'ROUGE'}.get(_st, 'AMBRE'))
    # Synthèse BFCC en une ligne : le pire des six statuts, et son code.
    _bfcc = lignes_hypotheses_bfcc(n2)
    _pire = next((l for l in _bfcc if l['statut'] == 'NON VALIDÉE'),
                 next((l for l in _bfcc if l['statut'] == 'À JUSTIFIER'), None))
    _kpi(ws, 26, 1, "BF / Cape Cod",
         f"{_pire['code']} {_pire['statut']}" if _pire else "H1-H5 OK",
         None, 'ROUGE' if (_pire and _pire['statut'] == 'NON VALIDÉE')
                else 'AMBRE' if _pire else 'VERT')
    # Synthèse BOOT sur le même modèle. La ligne qui occupait cette place
    # affichait le score de l'ancienne H4 — un `/100` calculé sur le CV des
    # variances des facteurs, rouge sur les trois triangles de référence.
    _boot = lignes_hypotheses_bootstrap(n2)
    _pb   = next((l for l in _boot if l['statut'] == 'NON VALIDÉE'),
                 next((l for l in _boot if l['statut'] == 'À JUSTIFIER'), None))
    _kpi(ws, 27, 1, "Bootstrap ODP",
         f"{_pb['code']} {_pb['statut']}" if _pb else "H1-H4 OK",
         None, 'ROUGE' if (_pb and _pb['statut'] == 'NON VALIDÉE')
                else 'AMBRE' if _pb else 'VERT')

    # Bloc triangle
    _titre_section(ws, 29, 1, "DONNÉES", 6)
    _kpi(ws, 30, 1, "Taille triangle", n1.get('taille','—'))
    _kpi(ws, 31, 1, "Mode détecté",    n1.get('mode_detecte','—'))
    _kpi(ws, 32, 1, "Méthode CL",      n3.get('methode_cl','—'))
    _kpi(ws, 33, 1, "Tail factor",     n3['chain_ladder'].get('tail_factor',{}).get('tail_factor',1) if isinstance(n3['chain_ladder'].get('tail_factor'),dict) else 1, FMT_DEC4)
    _kpi(ws, 34, 1, "Date rapport",    date_str)
    _kpi(ws, 35, 1, "Référence",       ref_client)

    # Largeurs colonnes
    for col, w in [(1,32),(2,18),(3,16),(4,12),(5,12),(6,12)]:
        ws.column_dimensions[get_column_letter(col)].width = w


# =============================================================================
#  ONGLET 2 — TRIANGLE BRUT
# =============================================================================

def _ong2_triangle(wb, C):
    ws = wb.create_sheet("2. Triangle brut")
    ws.sheet_view.showGridLines = False
    n, m = C.shape

    _titre_section(ws, 1, 1, "Triangle de développement — Sinistres cumulés payés (€)", m+1)

    # En-têtes colonnes
    ws.cell(row=2, column=1, value="Année \\ Dév.").font = _font(bold=True, color=BLANC)
    ws.cell(row=2, column=1).fill = _fill(NAVY)
    ws.cell(row=2, column=1).alignment = _align(h='center')
    ws.cell(row=2, column=1).border    = _border_thin()

    for j in range(m):
        _header(ws, 2, j+2, f"{(j+1)*12}M", width=14)

    # Données avec dégradé couleur
    max_val = float(C[C > 0].max()) if C[C > 0].size > 0 else 1
    for i in range(n):
        c_lbl = ws.cell(row=3+i, column=1, value=f"An. {i}")
        c_lbl.font      = _font(bold=True, color=BLANC, size=10)
        c_lbl.fill      = _fill(NAVY)
        c_lbl.alignment = _align(h='center')
        c_lbl.border    = _border_thin()

        for j in range(m):
            val = float(C[i, j])
            c   = ws.cell(row=3+i, column=j+2)
            in_zone = (j <= n-i-1) and val > 0
            if in_zone:
                c.value          = val
                c.number_format  = FMT_NB
                # Dégradé : valeur faible = bleu clair, haute = or
                ratio = val / max_val
                r = int(0x1B + (0xC9-0x1B)*ratio)
                g = int(0x3A + (0xA8-0x3A)*ratio)
                b = int(0x5C + (0x4C-0x5C)*ratio)
                c.fill = PatternFill('solid',
                    fgColor=f"{r:02X}{g:02X}{b:02X}",
                    start_color=f"{r:02X}{g:02X}{b:02X}"
                )
                c.font      = _font(color=BLANC if ratio > 0.5 else NOIR, size=9)
                c.alignment = _align(h='right')
            else:
                c.fill      = _fill(GRIS_CLAIR)
                c.font      = _font(color=GRIS_MED, size=9)
                c.alignment = _align(h='center')
                c.value     = "—"
            c.border = _border_thin()

    _row_h(ws, 2, 20)
    ws.column_dimensions['A'].width = 12


# =============================================================================
#  ONGLET 3 — FACTEURS CL
# =============================================================================

def _ong3_facteurs(wb, n3):
    ws = wb.create_sheet("3. Facteurs CL")
    ws.sheet_view.showGridLines = False

    _titre_section(ws, 1, 1, "Facteurs de développement Chain Ladder", 6)
    hdrs = ["Transition", "Facteur f_j", "Facteur cumulé F_j", "% Développé", "Tail", "Statut"]
    widths = [14, 14, 16, 14, 10, 14]
    for j, (h, w) in enumerate(zip(hdrs, widths)):
        _header(ws, 2, j+1, h, width=w)

    facteurs   = n3['chain_ladder'].get('facteurs', [])
    f_cum      = n3['chain_ladder'].get('facteurs_cumules', [])
    tail_info  = n3['chain_ladder'].get('tail_factor', {})
    tail_val   = tail_info.get('tail_factor', 1.0) if isinstance(tail_info, dict) else 1.0

    moy = float(np.mean(facteurs)) if facteurs else 1
    std = float(np.std(facteurs))  if facteurs else 0

    for i, f in enumerate(facteurs):
        row = 3 + i
        bg  = BLANC if i % 2 == 0 else GRIS_CLAIR
        ecart = abs(f - moy) / max(std, 1e-9)
        statut = '⚠️ Outlier >2σ' if ecart > 2 else '△ Ecart >1σ' if ecart > 1 else '✓ Normal'
        st_col = ROUGE_S2 if ecart > 2 else AMBRE_S2 if ecart > 1 else VERT_S2
        st_bg  = 'FCEBEB' if ecart > 2 else 'FAEEDA' if ecart > 1 else 'EAF3DE'

        vals = [
            f"j={i}→{i+1}",
            f,
            f_cum[i] if i < len(f_cum) else None,
            (1/f_cum[i]) if i < len(f_cum) and f_cum[i] > 0 else None,
            tail_val if i == len(facteurs)-1 else None,
        ]
        fmts = [None, FMT_DEC6, FMT_DEC6, FMT_PCT, FMT_DEC6]

        for j, (val, fmt) in enumerate(zip(vals, fmts)):
            c = ws.cell(row=row, column=j+1, value=val)
            c.font      = _font(size=10)
            c.alignment = _align(h='right' if j > 0 else 'center')
            c.border    = _border_thin()
            c.fill      = _fill(bg)
            if fmt and val is not None:
                c.number_format = fmt

        # Statut
        cs = ws.cell(row=row, column=6, value=statut)
        cs.font      = _font(bold=True, color=st_col, size=9)
        cs.fill      = _fill(st_bg)
        cs.alignment = _align(h='center')
        cs.border    = _border_thin()

    # Ligne tail
    row_tail = 3 + len(facteurs)
    ws.cell(row=row_tail, column=1, value="Tail factor").font = _font(bold=True, color=GOLD)
    ws.cell(row=row_tail, column=1).fill = _fill(NAVY)
    ws.cell(row=row_tail, column=1).border = _border_thin()
    ws.cell(row=row_tail, column=1).alignment = _align(h='center')
    c_tail = ws.cell(row=row_tail, column=2, value=tail_val)
    c_tail.number_format = FMT_DEC6
    c_tail.font      = _font(bold=True, color=GOLD)
    c_tail.fill      = _fill(NAVY)
    c_tail.border    = _border_thin()
    c_tail.alignment = _align(h='right')


# =============================================================================
#  ONGLET 4 — RÉSULTATS MÉTHODES
# =============================================================================

def _ong4_methodes(wb, n3, n4):
    ws = wb.create_sheet("4. Résultats méthodes")
    ws.sheet_view.showGridLines = False

    _titre_section(ws, 1, 1, "Résultats par méthode actuarielle — Réserves IBNR", 6)
    hdrs = ["Méthode", "Réserve (€)", "Poids BE", "Statut", "LR / σ", "Notes"]
    widths = [28, 16, 10, 18, 14, 30]
    for j, (h, w) in enumerate(zip(hdrs, widths)):
        _header(ws, 2, j+1, h, width=w)

    methodes_inc = n4.get('methodes_incluses', [])
    poids        = n4.get('poids', {})
    be           = n4.get('best_estimate', 0)

    rows = [
        ('Chain Ladder',           n3['chain_ladder']['reserve_totale'],   'chain_ladder',
         n3['chain_ladder'].get('methode',''), '—'),
        ('Mack 1993 (stochastique)', n3['mack']['reserve_best_estimate'],  'mack',
         f"σ={n3['mack']['sigma_total']:,.0f}€", f"CV={n3['mack']['cv_pct']:.1f}%"),
        ('Bornhuetter-Ferguson',   n3['bf']['reserve_totale'],             'bornhuetter_ferguson',
         f"LR={libelle_loss_ratio(n3['bf'])}", n3['bf']['source_lr']),
        ('Cape Cod',               n3['cape_cod']['reserve_totale'],       'cape_cod',
         f"LR_CC={libelle_loss_ratio(n3['cape_cod'], 'lr_cape_cod')}", n3['cape_cod']['source_exposition']),
        ('Bootstrap ODP',          n3['bootstrap'].get('be_bootstrap',0), 'bootstrap_odp',
         f"φ={libelle_incertitude(n3['bootstrap'], 'phi')}",
         f"n={n3['bootstrap'].get('n_simulations',0):,}"),
    ]
    # Benktander — INFORMATIF, poids nul par construction. Il est deja un
    # melange de Chain Ladder et de BF : l'inclure au Best Estimate ferait
    # peser Chain Ladder 44,2 % pour 25 % affiches.
    _gb = n3.get('benktander') or {}
    if _gb.get('disponible'):
        rows.append(('Benktander (1976)', _gb.get('reserve_totale', 0.0),
                     'benktander',
                     f"α moyen = {_gb.get('alpha_moyen', 0):.4f}",
                     'INFORMATIF — hors Best Estimate'))

    for i, (nom, res, key, detail, note) in enumerate(rows):
        inc  = key in methodes_inc
        poid = poids.get(key, 0)
        _mack = (key == 'mack')   # Mack : volatilité (σ), non pondérée dans le BE (point = CL)
        # Benktander : ni incluse ni EXCLUE. Il est informatif PAR
        # CONSTRUCTION -- l'etiqueter << Exclue >> en rouge le ferait
        # passer pour une methode disqualifiee par une hypothese.
        _gb_l = (key == 'benktander')
        bg   = 'EAF3DE' if inc else GRIS_CLAIR
        st   = ('σ (volatilité)' if _mack else
                'ⓘ Informatif' if _gb_l else
                ('✅ Incluse' if inc else '❌ Exclue'))
        st_c = NAVY if _gb_l else (VERT_S2 if inc else ROUGE_S2)

        vals = [nom, res, poid, st, detail, note]
        fmts = [None, FMT_NB, FMT_PCT, None, None, None]
        for j, (val, fmt) in enumerate(zip(vals, fmts)):
            c = ws.cell(row=3+i, column=j+1, value=val)
            c.font      = _font(size=10,
                               bold=(j==0),
                               color=st_c if j==3 else NOIR)
            c.fill      = _fill(bg)
            c.alignment = _align(h='right' if j==1 else 'center' if j in (2,3) else 'left')
            c.border    = _border_thin()
            if fmt and isinstance(val, (int,float)):
                c.number_format = fmt

    # Ligne BE S2 — sans merge pour éviter MergedCell
    row_be = 3 + len(rows)
    for col_be in range(1, 7):
        c = ws.cell(row=row_be, column=col_be)
        c.fill   = _fill(NAVY)
        c.border = _border_thin()
    ws.cell(row=row_be, column=1, value="BEST ESTIMATE S2").font = _font(bold=True, color=GOLD, size=11)
    ws.cell(row=row_be, column=1).fill      = _fill(NAVY)
    ws.cell(row=row_be, column=1).alignment = _align()
    c_be = ws.cell(row=row_be, column=2, value=be)
    c_be.number_format = FMT_NB
    c_be.font      = _font(bold=True, color=GOLD, size=12)
    c_be.fill      = _fill(NAVY)
    c_be.border    = _border_thin()
    c_be.alignment = _align(h='right')
    c_p = ws.cell(row=row_be, column=3, value=1.0)
    c_p.number_format = FMT_PCT
    c_p.font      = _font(bold=True, color=GOLD)
    c_p.fill      = _fill(NAVY)
    c_p.border    = _border_thin()
    c_p.alignment = _align(h='center')

    # Bloc phare — distribution log-normale composée (percentiles retenus)
    _titre_section(ws, row_be+2, 1, "Distribution log-normale — incertitude composée (σ Mack ⊕ σ modèle)", 4)
    pcts = [
        ("P75 — Provision prudentielle (composé)", n4['reserve_p75']),
        ("P90 — Provision stress test (composé)",  n4['reserve_p90']),
        # ⚠️ Règle 2 du lot C3b : un percentile se nomme par son percentile.
        ("P99.5 — Provision extrême (composé)",    n4['reserve_p99_5']),
        ("σ composé",                              n4.get('sigma_total_compose', n4['sigma_mack'])),
    ]
    for i, (lbl, val) in enumerate(pcts):
        _kpi(ws, row_be+3+i, 1, lbl, val, FMT_NB)

    # Bloc diagnostic — décomposition de l'incertitude (4 approches, colonne Centre)
    _mk = n3.get('mack', {}); _bo = n3.get('bootstrap', {})
    row_diag = row_be + 3 + len(pcts) + 1
    _titre_section(ws, row_diag, 1,
                   "Diagnostic — décomposition de l'incertitude (outil analytique interne, non destiné au bilan)", 4)
    for j, h in enumerate(["Approche", "P90 (€)", "σ (€)", "Centre"]):
        _header(ws, row_diag+1, j+1, h)
    diag_rows = [
        ("Incertitude composée (retenue)", n4['reserve_p90'],             n4.get('sigma_total_compose', 0), "BE pondéré"),
        ("Mack recentré",                  n4.get('reserve_p90_mack', 0), n4.get('sigma_mack', 0),          "BE pondéré"),
        ("Mack natif",                     _mk.get('reserve_p90', 0),     _mk.get('sigma_total', 0),        "réserve Mack"),
        ("Bootstrap ODP",                  _bo.get('p90') or 0,           _bo.get('std_bootstrap') or 0,      "réserve Bootstrap"),
    ]
    for i, (appr, p90v, sigv, centre) in enumerate(diag_rows):
        r   = row_diag + 2 + i
        bgd = BLANC if i % 2 == 0 else GRIS_CLAIR
        for j, (val, fmt) in enumerate([(appr, None), (p90v, FMT_NB), (sigv, FMT_NB), (centre, None)]):
            c = ws.cell(row=r, column=j+1, value=val)
            c.font      = _font(size=10)
            c.fill      = _fill(bgd)
            c.alignment = _align(h='left' if j in (0, 3) else 'right')
            c.border    = _border_thin()
            if fmt and isinstance(val, (int, float)):
                c.number_format = fmt


# =============================================================================
#  ONGLET 5 — IBNR PAR ANNÉE
# =============================================================================

def _ong5_ibnr(wb, n3):
    ws = wb.create_sheet("5. IBNR par année")
    ws.sheet_view.showGridLines = False

    _titre_section(ws, 1, 1, "IBNR par année de survenance — Comparaison méthodes", 6)
    hdrs = ["Année", "IBNR CL (€)", "IBNR BF (€)",
            "IBNR CC (€)", "Ultimate CL (€)", "% Développé"]
    widths = [10, 16, 16, 16, 18, 14]
    for j, (h, w) in enumerate(zip(hdrs, widths)):
        _header(ws, 2, j+1, h, width=w)

    cl   = n3['chain_ladder']
    bf   = n3['bf']
    cc   = n3['cape_cod']

    ibnr_cl   = cl.get('ibnr_par_annee', [])
    ibnr_bf   = bf.get('ibnr_par_annee', [])
    ibnr_cc   = cc.get('ibnr_par_annee', [])
    ult_cl    = cl.get('ultimates', [])
    pct_dev   = cl.get('pct_developpe', [])

    tot_cl = tot_bf = tot_cc = 0
    for i in range(len(ibnr_cl)):
        bg   = BLANC if i % 2 == 0 else GRIS_CLAIR
        vals = [
            f"An. {i}",
            ibnr_cl[i] if i < len(ibnr_cl) else 0,
            ibnr_bf[i]  if i < len(ibnr_bf)  else 0,
            ibnr_cc[i]  if i < len(ibnr_cc)  else 0,
            ult_cl[i]   if i < len(ult_cl)   else 0,
            pct_dev[i]  if i < len(pct_dev)  else 0,
        ]
        fmts = [None, FMT_NB, FMT_NB, FMT_NB, FMT_NB, FMT_PCT]
        tot_cl += vals[1]; tot_bf += vals[2]; tot_cc += vals[3]

        for j, (val, fmt) in enumerate(zip(vals, fmts)):
            c = ws.cell(row=3+i, column=j+1, value=val)
            c.font      = _font(size=10)
            c.fill      = _fill(bg)
            c.alignment = _align(h='right' if j > 0 else 'center')
            c.border    = _border_thin()
            if fmt and isinstance(val, (int,float)):
                c.number_format = fmt

    # Ligne total
    row_tot = 3 + len(ibnr_cl)
    for j, (val, fmt) in enumerate([
        ("TOTAL", None), (tot_cl, FMT_NB),
        (tot_bf, FMT_NB), (tot_cc, FMT_NB), (None, None), (None, None)
    ]):
        c = ws.cell(row=row_tot, column=j+1, value=val)
        c.font      = _font(bold=True, color=GOLD)
        c.fill      = _fill(NAVY)
        c.border    = _border_thin()
        c.alignment = _align(h='right' if j > 0 else 'left')
        if fmt and val:
            c.number_format = fmt

    # Note bas de table (F2) : colonne « IBNR Mack » retirée car redondante
    note = ws.cell(row=row_tot + 2, column=1,
                   value="Note : Mack (1993) reprend les ultimates du Chain Ladder — "
                         "l'IBNR par année est identique à la colonne CL. "
                         "L'incertitude σ Mack figure en section Diagnostic (onglet Méthodes).")
    note.font      = _font(size=9, italic=True, color=BLEU_SOFT)
    note.alignment = _align(wrap=True)
    ws.merge_cells(start_row=row_tot + 2, start_column=1, end_row=row_tot + 2, end_column=6)


# =============================================================================
#  ONGLET 6 — VALIDATION HYPOTHÈSES
# =============================================================================

def _ong6_hypotheses(wb, n2, n4):
    ws = wb.create_sheet("6. Hypothèses")
    ws.sheet_view.showGridLines = False

    _titre_section(ws, 1, 1, "Validation des hypothèses actuarielles — Mack (1993) & England-Verrall (2002)", 5)
    # ⚠️ SIXIÈME COLONNE : LA PUISSANCE DISPONIBLE. Un verdict dit ce qu'on a
    # trouvé ; il ne dit pas ce qu'on aurait pu trouver. Les deux ensemble font
    # une conclusion opposable, et cette colonne porte la seconde moitié.
    hdrs = ["Hypothèse", "Validée", "Score /100", "Seuil utilisé", "Message",
            "Puissance disponible"]
    widths = [28, 12, 12, 16, 60, 60]
    for j, (h, w) in enumerate(zip(hdrs, widths)):
        _header(ws, 2, j+1, h, width=w)

    hyps = [
        ('h1_independance',     'H1 — Indépendance des années (Spearman)'),
        ('h2_stabilite',        'H2 — Stabilité des facteurs (CV + dérive)'),
    ]
    rangs = []
    for key, lbl in hyps:
        h   = n2.get(key, {})
        ok  = h.get('ok', True)
        # ⚠️ LA FORME DU LOT F3 ETAIT << DERIVER DE `ok`, PUIS RATTRAPER >>, ET
        # LE FILET DU LOT C L'A REFUSEE. Elle marchait pour NON TESTABLE parce
        # qu'un correctif la reecrivait plus bas ; tout statut FUTUR serait
        # retombe sur la valeur deduite de `ok`. Le libelle vient desormais du
        # seul `statut`, par table, et `ok` n'est plus que le repli du `.get`.
        _st = str(h.get('statut', 'VALIDÉE' if ok else 'REJETÉE'))
        vert = _st == 'VALIDÉE'
        statut_txt = {'VALIDÉE': '✅ OUI',
                      'REJETÉE': '❌ NON'}.get(_st, '⚠️ ' + _st)
        seuil_txt = '—'
        if key == 'h1_independance':
            seuil_txt = f"{h.get('seuil_utilise', 0.50):.2f}"
        elif key == 'h2_stabilite':
            # ⚠️ « OUI » EN VERT SUR UNE HYPOTHÈSE NON TESTÉE. `ok` vaut True par
            # défaut quand aucune colonne n'est testable ; la ligne sortait donc
            # verte, avec « CV<15% dérive<20% » — les seuils GÉNÉRIQUES, parce
            # que `_h2` ne publiait pas ceux de la branche sur ce chemin.
            # ⚠️ LE RATTRAPAGE QUI VIVAIT ICI EST RETIRE : il ne couvrait que H2
            # et que le statut NON TESTABLE. Le libelle vient maintenant du
            # `statut` pour LES DEUX hypotheses et pour tout statut.
            seuil_txt = (f"CV<{h.get('seuil_cv', 0.15):.0%} "
                         f"dérive<{h.get('seuil_derive', 0.20):.0%}")
        rangs.append((lbl, statut_txt, h.get('score', 0),
                      seuil_txt, str(h.get('message', '—'))[:200], '', vert))

    # ⚠️ CLM-H1..H4 EN TÊTE, ET C'EST UN MANQUE QUI EST RÉPARÉ ICI. Les
    # hypothèses de CHAIN LADDER ET MACK — les méthodes qui portent le Best
    # Estimate, et dont deux sont gatantes — n'atteignaient aucun des trois
    # formats. BFCC, Bootstrap et Munich avaient chacun leur affichage ; CLM
    # n'en avait pas. Le HTML et le Word ont été complétés au lot précédent,
    # l'Excel l'est ici : les trois formats disent désormais la même chose.
    # ⚠️ PLUS DE TRONCATURE À 200. Cinq messages sur quatorze y perdaient leur
    # justification — dont la mention MCL-H4, qui porte les seuls chiffres du
    # bloc. Une cellule Excel accepte 32 767 caractères.
    for source in (lignes_hypotheses_clm, lignes_hypotheses_bfcc,
                   lignes_hypotheses_bootstrap, lignes_hypotheses_munich):
        for ligne in source(n2):
            rangs.append((ligne['libelle'], ligne['statut'], '—',
                          ligne['source'], ligne['message'],
                          ligne.get('puissance_phrase') or '', ligne['ok']))

    for i, (lbl, verdict, sc, seuil_txt, msg, puiss, ok) in enumerate(rangs):
        bg = 'EAF3DE' if ok else 'FCEBEB'
        vals = [lbl, verdict, sc, seuil_txt, msg, puiss]
        for j, val in enumerate(vals):
            c = ws.cell(row=3+i, column=j+1, value=val)
            c.font      = _font(
                bold=(j==0),
                color=VERT_S2 if (j==1 and ok) else ROUGE_S2 if (j==1 and not ok) else NOIR,
                size=10
            )
            c.fill      = _fill(bg)
            c.alignment = _align(h='center' if j in (1,2,3) else 'left', wrap=(j==4))
            c.border    = _border_thin()
        _row_h(ws, 3+i, 30)

    # ⚠️ LES BLOCS SUIVANTS ÉTAIENT ÉCRITS À DES LIGNES EN DUR — 8 ET 13 —
    # ALORS QUE LE TABLEAU DES HYPOTHÈSES EN OCCUPE 17 (lot C3b).
    # « DÉCISION MÉTHODOLOGIQUE » écrasait les lignes 8 à 11 et « ALERTES »
    # les lignes 13 à 16 : HUIT hypothèses sur dix-sept disparaissaient de
    # l'onglet, en silence. Dont BFCC-H4, BFCC-H5 et BFCC-H6 — LES TROIS
    # HYPOTHÈSES BLOQUANTES, celles qui peuvent exclure Bornhuetter-Ferguson
    # ou Cape Cod du Best Estimate. La feuille qui s'appelle « Validation des
    # hypothèses » détruisait précisément les hypothèses qui décident.
    # Un curseur remplace les constantes : le nombre d'hypothèses peut varier
    # sans que rien ne se recouvre.
    ligne = 3 + len(rangs) + 1

    # Le détail de H1, colonne par colonne — remplace le graphique g8, retiré
    # au même lot. Ces corrélations n'atteignaient AUCUN livrable.
    detail_h1 = lignes_correlations_h1(n2)
    if detail_h1:
        _titre_section(ws, ligne, 1,
                       "H1 — CORRÉLATIONS DE SPEARMAN, COLONNE PAR COLONNE", 5)
        for j, h in enumerate(["Colonnes", "|Corrélation|", "Seuil",
                               "Significative", "Verdict"]):
            _header(ws, ligne + 1, j + 1, h, width=[22, 14, 10, 14, 16][j])
        for i, d in enumerate(detail_h1):
            vals = [d['colonnes'],
                    '—' if d['corr_abs'] is None else round(d['corr_abs'], 4),
                    round(d['seuil'], 2),
                    'oui' if d['significatif'] else 'non',
                    d['statut']]
            for j, val in enumerate(vals):
                c = ws.cell(row=ligne + 2 + i, column=j + 1, value=val)
                c.font = _font(bold=(j == 0),
                               color=VERT_S2 if (j == 4 and d['ok'])
                               else ROUGE_S2 if j == 4 else NOIR, size=10)
                c.fill = _fill('EAF3DE' if d['ok'] else 'FCEBEB')
                c.alignment = _align(h='left' if j == 0 else 'center')
                c.border = _border_thin()
        ligne += 2 + len(detail_h1) + 1

    # Décision méthodologique
    _titre_section(ws, ligne, 1, "DÉCISION MÉTHODOLOGIQUE", 5)
    for i, (lbl, val) in enumerate([
        ("Méthode CL retenue",     n2.get('methode_cl_retenue','—')),
        ("Méthode recommandée",    n2.get('methode_recommandee','—')),
        ("Statut global",          n2.get('statut_global','—')),
    ]):
        _kpi(ws, ligne + 1 + i, 1, lbl, val, None,
             n2.get('statut_global') if lbl == "Statut global" else None)
    ligne += 5

    # ⚠️ LES ALERTES DE JUGEMENT, PAS SEULEMENT CELLES DE N2 (lot C1).
    # Cette feuille lisait `n2['alertes']` et elle seule. Or les alertes du
    # JUGEMENT ACTUARIEL — péremption de la courbe des taux, niveau ancré sur
    # Chain Ladder, hypothèses à justifier, et depuis le lot A2 le message du
    # filet qui porte les 87,2 % du Best Estimate — vivent dans `n4['alertes']`.
    # Aucune n'atteignait donc l'Excel, alors que le rapport HTML et le Word
    # les affichaient : le même run disait deux choses différentes selon le
    # format ouvert. N4 les reprend déjà toutes, N2 comprises.
    alertes = n4.get('alertes') or n2.get('alertes', [])
    if alertes:
        _titre_section(ws, ligne, 1, "ALERTES", 5)
        # ⚠️ LES ACCENTS NE SONT PLUS MANGÉS (lot C1). Un
        # `encode('ascii', 'ignore')` rendait « Années [9] : FILET DE SÉCURITÉ
        # déclenché » en « Annes [9] : FILET DE SCURIT dclench ». openpyxl écrit
        # de l'UTF-8 sans difficulté ; ce filtrage ne protégeait rien et rendait
        # illisible le français. Seuls les caractères de contrôle et les emoji
        # hors plan multilingue de base sont retirés, parce qu'Excel les rend
        # par un carré vide.
        for i, a in enumerate(alertes[:6]):
            clean = ''.join(c for c in str(a)
                            if c.isprintable() and ord(c) < 0x1F000).strip()
            c = ws.cell(row=ligne + 1 + i, column=1, value=clean)
            c.font      = _font(color=AMBRE_S2, size=9)
            c.fill      = _fill('FAEEDA')
            c.alignment = _align(wrap=True)
            c.border    = _border_thin()
            ws.merge_cells(start_row=ligne + 1 + i, start_column=1,
                           end_row=ligne + 1 + i, end_column=5)
            _row_h(ws, ligne + 1 + i, 25)


# =============================================================================
#  ONGLET 7 — BOOTSTRAP ODP
# =============================================================================

def _ong7_bootstrap(wb, n3):
    ws = wb.create_sheet("7. Bootstrap ODP")
    ws.sheet_view.showGridLines = False

    boot = n3.get('bootstrap', {})
    _titre_section(ws, 1, 1, f"Bootstrap ODP — England & Verrall (2002) | {boot.get('n_simulations',0):,} simulations", 4)

    hdrs = ["Statistique", "Valeur (€)", "Vs BE (%)", "Notes"]
    widths = [28, 18, 14, 30]
    for j, (h, w) in enumerate(zip(hdrs, widths)):
        _header(ws, 2, j+1, h, width=w)

    be_boot = boot.get('be_bootstrap', 0)
    rows = [
        ("BE Bootstrap (moyenne)",    be_boot,                  0,           "Espérance distribution simulée"),
        ("Écart-type σ",              boot.get('std_bootstrap') or 0, None,  "Incertitude totale Bootstrap"),
        ("CV Bootstrap",              boot.get('cv_bootstrap') or 0, None,   "Coefficient de variation"),
        ("IC 95% — Borne inférieure", boot.get('ic_95_inf',0),  None,        "Percentile 2.5%"),
        ("IC 95% — Borne supérieure", boot.get('ic_95_sup',0),  None,        "Percentile 97.5%"),
        ("P50 — Médiane",             boot.get('p50',0),         None,       "50ème percentile"),
        ("P75",                       boot.get('p75',0),         None,       "75ème percentile"),
        ("P90 — Stress test S2",      boot.get('p90',0),         None,       "90ème percentile"),
        ("P95",                       boot.get('p95',0),         None,       "95ème percentile"),
        ("P99.5 — Réserve extrême",   boot.get('p99_5',0),       None,       "99.5ème percentile (VaR S2)"),
    ]

    for i, (lbl, val, vs_be, note) in enumerate(rows):
        bg = 'EAF3DE' if lbl.startswith('P99') else BLANC if i % 2 == 0 else GRIS_CLAIR
        vs = (val/be_boot - 1) if (be_boot > 0 and val and vs_be != 0) else vs_be

        for j, (v, fmt) in enumerate([
            (lbl,  None),
            (val,  FMT_NB if not lbl.startswith('CV') else FMT_PCT),
            (vs,   FMT_PCT if vs is not None else None),
            (note, None),
        ]):
            c = ws.cell(row=3+i, column=j+1, value=v)
            c.font      = _font(bold=(i in (8,9)), size=10,
                                color=VERT_S2 if i == 9 else NOIR)
            c.fill      = _fill(bg)
            c.alignment = _align(h='right' if j in (1,2) else 'left')
            c.border    = _border_thin()
            if fmt and v is not None and isinstance(v, (int,float)):
                c.number_format = fmt

    # Paramètres Bootstrap
    _titre_section(ws, 3+len(rows)+1, 1, "PARAMÈTRES DU MODÈLE", 4)
    params = [
        ("Facteur sur-dispersion φ", boot.get('phi',0),     FMT_DEC6),
        ("Nb observations (n_obs)",  boot.get('n_obs',0),   None),
        ("Nb paramètres (n_params)", boot.get('n_params',0), None),
        ("Degrés de liberté (df)",   boot.get('df',0),       None),
        ("Simulations",              boot.get('n_simulations',0), None),
        ("Statut",                   boot.get('statut','—'), None),
    ]
    for i, (lbl, val, fmt) in enumerate(params):
        _kpi(ws, 3+len(rows)+2+i, 1, lbl, val, fmt,
             boot.get('statut') if lbl == 'Statut' else None)


# =============================================================================
#  ONGLET 8 — COMPARATIF N-1/N
# =============================================================================

def _ong_munich(wb, n3):
    """Onglet Munich CL — la seule méthode qui exploite les DEUX triangles, et
    la seule qui n'apparaissait dans aucun format de livrable.

    Lit `lignes_munich_rapport`, exactement comme le HTML et le Word : une
    méthode ne doit pas raconter trois histoires selon le fichier ouvert.
    """
    lignes = lignes_munich_rapport(n3)
    if not lignes:
        return
    ws = wb.create_sheet("11. Munich CL")
    ws.sheet_view.showGridLines = False

    _titre_section(ws, 1, 1,
                   "Munich Chain Ladder — Quarg & Mack (2004) | "
                   "rapprochement payé / engagé", 3)
    for j, (h, w) in enumerate(zip(["Indicateur", "Valeur", "Lecture"],
                                   [40, 20, 72])):
        _header(ws, 2, j + 1, h, width=w)

    # ⚠️ Renomme au lot C3b : la variable de boucle masquait l'import
    # `libelle` du referentiel (ruff F402). Inoffensif tant que cette fonction
    # ne l'appelle pas -- un piege a la premiere edition qui l'appellerait.
    for i, (intitule, valeur, lecture) in enumerate(lignes):
        alerte = str(intitule).startswith('⚠')
        bg = 'FDE9E7' if alerte else (BLANC if i % 2 == 0 else GRIS_CLAIR)
        for j, v in enumerate((intitule, valeur, lecture)):
            c = ws.cell(row=3 + i, column=j + 1, value=v)
            c.font      = _font(bold=(j == 0 or alerte), size=10)
            c.fill      = _fill(bg)
            c.border    = _border_thin()
            c.alignment = _align(h='right' if j == 1 else 'left', wrap=(j == 2))

    ws.freeze_panes = 'A3'


def _ong8_comparatif(wb, n4, resultats_precedents=None):
    ws = wb.create_sheet("8. Comparatif N-1 N")
    ws.sheet_view.showGridLines = False

    _titre_section(ws, 1, 1, "Comparatif arrêtés — Variation Best Estimate et indicateurs clés", 5)
    hdrs = ["Indicateur", "Arrêté N", "Arrêté N-1", "Variation (€ ou pts)", "Variation (%)"]
    widths = [30, 18, 18, 18, 14]
    for j, (h, w) in enumerate(zip(hdrs, widths)):
        _header(ws, 2, j+1, h, width=w)

    be_n  = n4.get('best_estimate', 0)
    p90_n = n4.get('reserve_p90',  0)
    cv_n  = n4.get('cv_inter_methodes', 0)
    sigma_n = n4.get('sigma_mack', 0)

    if resultats_precedents:
        be_nm1  = resultats_precedents.get('best_estimate', 0)
        p90_nm1 = resultats_precedents.get('reserve_p90',  0)
        cv_nm1  = resultats_precedents.get('cv_inter_methodes', 0)
        sigma_nm1 = resultats_precedents.get('sigma_mack', 0)
    else:
        be_nm1 = p90_nm1 = cv_nm1 = sigma_nm1 = None

    def _var(n, nm1):
        if nm1 is None or nm1 == 0:
            return None, None
        return n - nm1, (n - nm1) / abs(nm1)

    indicateurs = [
        ("Best Estimate S2 (€)",    be_n,    be_nm1,    FMT_NB),
        ("Provision P90 (€)",       p90_n,   p90_nm1,   FMT_NB),
        ("Incertitude Mack σ (€)",  sigma_n, sigma_nm1, FMT_NB),
        ("CV inter-méthodes",       cv_n/100, (cv_nm1/100 if cv_nm1 is not None else None), FMT_PCT),
    ]

    for i, (lbl, vn, vnm1, fmt) in enumerate(indicateurs):
        delta_abs, delta_pct = _var(vn, vnm1)
        bg = BLANC if i % 2 == 0 else GRIS_CLAIR

        # Couleur variation
        if delta_pct is not None:
            if abs(delta_pct) < 0.05:
                var_col = VERT_S2
            elif abs(delta_pct) < 0.15:
                var_col = AMBRE_S2
            else:
                var_col = ROUGE_S2
        else:
            var_col = NOIR

        for j, (val, f) in enumerate([
            (lbl,      None),
            (vn,       fmt or FMT_NB),
            (vnm1,     fmt or FMT_NB),
            (delta_abs, fmt or FMT_NB),
            (delta_pct, FMT_PCT),
        ]):
            c = ws.cell(row=3+i, column=j+1, value=val)
            c.font      = _font(bold=(j==0), size=10,
                                color=var_col if j in (3,4) else NOIR)
            c.fill      = _fill(bg)
            c.alignment = _align(h='right' if j > 0 else 'left')
            c.border    = _border_thin()
            if f and val is not None and isinstance(val, (int,float)):
                c.number_format = f

    # Message si pas de N-1
    if not resultats_precedents:
        msg = ws.cell(row=3+len(indicateurs)+1, column=1,
                      value="ℹ️ Aucun arrêté précédent fourni — passer resultats_precedents=dict(best_estimate=...) dans run()")
        msg.font      = _font(color=AMBRE_S2, size=9, italic=True)
        msg.fill      = _fill('FAEEDA')
        msg.alignment = _align(wrap=True)
        msg.border    = _border_thin()
        ws.merge_cells(start_row=3+len(indicateurs)+1, start_column=1,
                       end_row=3+len(indicateurs)+1,   end_column=5)
        _row_h(ws, 3+len(indicateurs)+1, 30)


# =============================================================================
#  ONGLET 9 — SCR FORMULE STANDARD
# =============================================================================

def _ong9_scr(wb, n4):
    ws = wb.create_sheet("9. SCR formule standard")
    ws.sheet_view.showGridLines = False

    scr = n4.get('scr', {})
    _titre_section(ws, 1, 1, "SCR Provisions — Formule standard Art. 115 Règlement Délégué (UE) 2015/35", 4)

    # Bloc calcul
    _titre_section(ws, 3, 1, "CALCUL SCR PROVISIONS (LoB unique)", 4)
    hdrs = ["Paramètre", "Valeur", "Source / Référence", "Notes"]
    widths = [32, 18, 28, 30]
    for j, (h, w) in enumerate(zip(hdrs, widths)):
        _header(ws, 4, j+1, h, width=w)

    rows_scr = [
        ("Best Estimate (BE)",          n4.get('best_estimate',0),        FMT_NB,    "Art. 77 Directive S2",         "Espérance flux futurs"),
        # La source du σ est celle du RÉSULTAT, plus une constante : elle vaut
        # « Annexe XIV » pour une LoB de santé non-SLT, où « Annexe II » mentait.
        ("Facteur σ(LoB) — réserve",    scr.get('sigma_eiopa',0),         FMT_PCT2,  scr.get('reference_s2','Annexes II / XIV Rgt 2015/35'), "Écart type risque de réserve, art. 117"),
        ("Facteur multiplicatif",        3.0,                              "0.0",     "Art. 115 Rgt 2015/35",         "Quantile 99.5% loi normale"),
        ("SCR_prov = 3 × σ × BE",       scr.get('scr_provisions',0),      FMT_NB,    "Art. 115 Rgt 2015/35",         "Exigence de capital provisions"),
        ("Ratio SCR/BE",                scr.get('ratio_scr_be',0),         FMT_PCT,   "Indicateur de pilotage",       ""),
        ("Branche (LoB)",               scr.get('lob_label','—'),          None,      "Classification EIOPA",         ""),
        ("Méthode",                     scr.get('methode','—'),             None,      "Règlement Délégué 2015/35",    ""),
    ]

    for i, (lbl, val, fmt, source, note) in enumerate(rows_scr):
        bg = 'EAF3DE' if 'SCR_prov' in lbl else BLANC if i % 2 == 0 else GRIS_CLAIR
        for j, (v, f) in enumerate([(lbl,None),(val,fmt),(source,None),(note,None)]):
            c = ws.cell(row=5+i, column=j+1, value=v)
            c.font      = _font(bold=('SCR_prov' in lbl), size=10,
                                color=GOLD if 'SCR_prov' in lbl else NOIR)
            c.fill      = _fill('0F2E52' if 'SCR_prov' in lbl else bg)
            c.alignment = _align(h='right' if j==1 else 'left')
            c.border    = _border_thin()
            if f and isinstance(v,(int,float)):
                c.number_format = f

    # Note réglementaire
    _titre_section(ws, 5+len(rows_scr)+1, 1, "NOTE RÉGLEMENTAIRE", 4)
    notes = [
        "Le SCR calculé ici est le SCR provisions d'une seule LoB (formule mono-branche).",
        "L'agrégation multi-LoB utilise la matrice de corrélation EIOPA Non-Vie 12×12",
        "(Annexe IV, Règlement Délégué 2015/35) :",
        "  SCR_NL = √(Σ_ij ρ_ij × SCR_i × SCR_j)",
        "Cette agrégation est effectuée au niveau du tableau de bord consolidé ActuarIA.",
    ]
    for i, note in enumerate(notes):
        c = ws.cell(row=5+len(rows_scr)+2+i, column=1, value=note)
        c.font      = _font(size=9, italic=True, color=BLEU_SOFT)
        c.fill      = _fill(GRIS_CLAIR)
        c.alignment = _align(wrap=True)
        c.border    = _border_thin()
        ws.merge_cells(start_row=5+len(rows_scr)+2+i, start_column=1,
                       end_row=5+len(rows_scr)+2+i,   end_column=4)
        _row_h(ws, 5+len(rows_scr)+2+i, 18)


# =============================================================================
#  ONGLET 10 — SENSIBILITÉS
# =============================================================================

def _ong10_sensibilites(wb, n4):
    ws = wb.create_sheet("10. Sensibilités")
    ws.sheet_view.showGridLines = False

    be   = n4.get('best_estimate', 0)
    sens = n4.get('sensibilites', {})
    _titre_section(ws, 1, 1, "Analyse de sensibilité — Impact sur le Best Estimate S2", 5)

    hdrs = ["Scénario", "Réserve (€)", "Δ vs BE (€)", "Δ vs BE (%)", "Interprétation"]
    widths = [32, 16, 16, 12, 40]
    for j, (h, w) in enumerate(zip(hdrs, widths)):
        _header(ws, 2, j+1, h, width=w)

    # BE de référence
    c_ref = ws.cell(row=3, column=1, value="Best Estimate de référence")
    c_ref.font      = _font(bold=True, color=GOLD, size=10)
    c_ref.fill      = _fill(NAVY)
    c_ref.border    = _border_thin()
    c_val = ws.cell(row=3, column=2, value=be)
    c_val.number_format = FMT_NB
    c_val.font      = _font(bold=True, color=GOLD, size=10)
    c_val.fill      = _fill(NAVY)
    c_val.border    = _border_thin()
    c_val.alignment = _align(h='right')
    for col in [3,4,5]:
        c = ws.cell(row=3, column=col)
        c.fill   = _fill(NAVY)
        c.border = _border_thin()

    label_map = {
        'sans_chain_ladder':         'Exclusion Chain Ladder',
        'sans_mack':                 'Exclusion Mack 1993',
        'sans_bornhuetter_ferguson': 'Exclusion Bornhuetter-Ferguson',
        'sans_cape_cod':             'Exclusion Cape Cod',
        'boot_ic_inf':               'Bootstrap IC 95% — Borne basse',
        'boot_ic_sup':               'Bootstrap IC 95% — Borne haute',
        'boot_p90':                  'Bootstrap P90',
        'boot_p99_5':                'Bootstrap P99.5',
    }
    interp_map = {
        'sans_chain_ladder':         'Sensibilité à la méthode CL',
        'sans_mack':                 'Sensibilité à la méthode Mack',
        'sans_bornhuetter_ferguson': 'Sensibilité à l\'a priori BF',
        'sans_cape_cod':             'Sensibilité à Cape Cod',
        'boot_ic_inf':               'Scénario favorable (2.5th percentile)',
        'boot_ic_sup':               'Scénario défavorable (97.5th percentile)',
        'boot_p90':                  'Provision stress test (90th percentile)',
        'boot_p99_5':                'Réserve au P99,5 — scénario extrême',
    }

    # Trier par amplitude
    items = sorted(
        [(k, float(v)) for k,v in sens.items()],
        key=lambda x: abs(x[1] - be),
        reverse=True
    )

    for i, (key, val) in enumerate(items):
        delta_abs = val - be
        delta_pct = delta_abs / max(abs(be), 1e-9)
        bg = BLANC if i % 2 == 0 else GRIS_CLAIR

        if abs(delta_pct) < 0.05:
            impact = "Impact faible (< 5%)"
            col    = VERT_S2
        elif abs(delta_pct) < 0.15:
            impact = "Impact modéré (5–15%)"
            col    = AMBRE_S2
        else:
            impact = "Impact significatif (> 15%)"
            col    = ROUGE_S2

        for j, (v, fmt, align) in enumerate([
            (label_map.get(key, key),          None,    'left'),
            (val,                              FMT_NB,  'right'),
            (delta_abs,                        FMT_NB,  'right'),
            (delta_pct,                        FMT_PCT, 'right'),
            (interp_map.get(key, impact),      None,    'left'),
        ]):
            c = ws.cell(row=4+i, column=j+1, value=v)
            c.font      = _font(size=10,
                                color=col if j in (2,3) else NOIR)
            c.fill      = _fill(bg)
            c.alignment = _align(h=align)
            c.border    = _border_thin()
            if fmt and isinstance(v,(int,float)):
                c.number_format = fmt


# =============================================================================
#  FONCTION PRINCIPALE
# =============================================================================

def export_excel(
    C:                   np.ndarray,
    n1:                  Dict,
    n2:                  Dict,
    n3:                  Dict,
    n4:                  Dict,
    ref_client:          str  = "",
    arrete:              str  = "",
    resultats_precedents: Optional[Dict] = None,
) -> bytes:
    """
    Génère le fichier Excel 10 onglets en mémoire.

    Parameters
    ----------
    C                    : triangle cumulé
    n1..n4               : dicts des niveaux N1 à N4
    ref_client           : référence client
    arrete               : libellé arrêté (ex. 'Q2 2026')
    resultats_precedents : dict {best_estimate, reserve_p90, ...} de l'arrêté N-1

    Returns
    -------
    bytes : fichier .xlsx prêt pour st.download_button()
    """
    if not OPENPYXL_OK:
        logger.warning("openpyxl non disponible — export Excel impossible")
        return b''

    try:
        date_str = arrete or datetime.now().strftime('%d/%m/%Y')
        ref      = ref_client or "ActuarIA"

        wb = Workbook()
        # Supprimer la feuille vide par défaut
        if wb.active:
            wb.remove(wb.active)

        _ong1_synthese(wb, n1, n2, n3, n4, ref, date_str)
        _ong2_triangle(wb, C)
        _ong3_facteurs(wb, n3)
        _ong4_methodes(wb, n3, n4)
        _ong5_ibnr(wb, n3)
        _ong6_hypotheses(wb, n2, n4)
        _ong7_bootstrap(wb, n3)
        _ong8_comparatif(wb, n4, resultats_precedents)
        _ong9_scr(wb, n4)
        _ong10_sensibilites(wb, n4)
        _ong_munich(wb, n3)      # silencieux si Munich n'a pas tourné

        # Activer l'onglet Synthèse par défaut
        wb.active = wb["1. Synthèse"]

        buf = io.BytesIO()
        wb.save(buf)
        logger.info(f"Excel généré : {len(wb.worksheets)} onglets")
        return buf.getvalue()

    except Exception as e:
        logger.error(f"Export Excel échoué : {e}", exc_info=True)
        return b''
