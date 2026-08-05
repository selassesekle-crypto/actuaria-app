# -*- coding: utf-8 -*-
"""
=============================================================================
  ActuarIA — A7  ·  CLÔTURE DU CHANTIER PUISSANCE  (P3 — inventaire, formats)
=============================================================================

Le chantier tenait en une phrase : « VALIDÉE » confond « j'ai cherché et il
n'y a rien » avec « je n'avais aucun moyen de voir ». P3 vérifie que plus
AUCUNE des dix-neuf hypothèses ne reste dans cette ambiguïté.

    5 portent une puissance MESURÉE sur le triangle de l'actuaire
   14 déclarent explicitement pourquoi la notion n'a pas de sens pour elles
    0 restent silencieuses

⚠️ LE SILENCE EST LE DÉFAUT QUE CE FICHIER TRAQUE. Une hypothèse sans entrée
de puissance ne se distingue pas, à la lecture, d'une hypothèse qu'on aurait
oublié de traiter. Un zéro serait pire encore : il se lirait « aucune
capacité de détection », ce qui est faux pour un contrôle de plage.

⚠️ ET LES TROIS FORMATS DOIVENT DIRE LA MÊME CHOSE. HTML, Word et Excel
lisent tous `lignes_hypotheses_*` — c'est la source unique, et ce fichier
vérifie qu'aucun n'en dévie.
"""

import io
import unittest

import numpy as np

from direction_non_vie.provisionnement.a7_provisionnement.agent import (
    AgentA7Provisionnement)
from direction_non_vie.provisionnement.a7_provisionnement.n2_hypotheses_bfcc import (
    lignes_hypotheses_bfcc)
from direction_non_vie.provisionnement.a7_provisionnement.n2_hypotheses_bootstrap import (
    lignes_hypotheses_bootstrap, puissance_bootstrap)
from direction_non_vie.provisionnement.a7_provisionnement.n2_hypotheses_clm import (
    lignes_hypotheses_clm)
from direction_non_vie.provisionnement.a7_provisionnement.n2_hypotheses_munich import (
    lignes_hypotheses_munich)
from direction_non_vie.provisionnement.a7_provisionnement.test_a7_ibrahim import (
    GENINS,
)

_G = np.asarray(GENINS, dtype=float)
_EXPO = [float(x) for x in (17_000_000, 18_000_000, 19_500_000, 20_000_000,
                            21_000_000, 22_000_000, 23_000_000, 24_000_000,
                            25_000_000, 26_000_000)]


def _engage(C, coef=0.15):
    n = C.shape[0]
    E = np.zeros_like(C)
    for i in range(n):
        for j in range(n - i):
            E[i, j] = C[i, j] * (1.0 + coef * (1.0 - j / (n - 1.0)))
    return E


_RESULTAT = None


def _run():
    """Une seule exécution de l'agent pour tout le fichier — elle coûte."""
    global _RESULTAT
    if _RESULTAT is None:
        _RESULTAT = AgentA7Provisionnement(verbose=False).run(
            source=_G, mode_declare='cumule', primes=_EXPO,
            triangle_engage=_engage(_G), generer_graphiques=False,
            generer_word=False, generer_html=True, n_sim_bootstrap=200,
            seed=42)
    return _RESULTAT


def _toutes_les_lignes(n2):
    lignes = []
    for source in (lignes_hypotheses_clm, lignes_hypotheses_bfcc,
                   lignes_hypotheses_bootstrap, lignes_hypotheses_munich):
        lignes.extend(source(n2))
    return lignes


# =============================================================================
#  P3-1 — L'INVENTAIRE EST COMPLET
# =============================================================================

class TP3_1_Inventaire(unittest.TestCase):

    def test_les_dix_neuf_hypotheses_sont_couvertes(self):
        """⚠️ AUCUNE NE RESTE SILENCIEUSE. C'est la clôture du chantier."""
        n2 = _run()['n2']
        total = mesurees = motivees = silencieuses = 0
        for cle in ('clm', 'bfcc', 'bootstrap_hyp', 'munich_hyp'):
            hyps = (n2.get(cle) or {}).get('hypotheses') or {}
            puis = (n2.get(cle) or {}).get('puissance') or {}
            for code in hyps:
                total += 1
                entree = puis.get(code)
                if entree is None:
                    silencieuses += 1
                elif entree.get('mesurable'):
                    mesurees += 1
                else:
                    motivees += 1
        self.assertEqual(total, 19, "le catalogue n'a plus dix-neuf entrées")
        self.assertEqual(silencieuses, 0,
                         "une hypothèse ne dit rien de sa puissance : à la "
                         "lecture, elle ne se distingue pas d'un oubli")
        self.assertEqual(mesurees, 5)
        self.assertEqual(motivees, 14)
        print(f"    OK P3-1a {total} hypothèses : {mesurees} avec puissance "
              f"mesurée, {motivees} sans objet motivé, {silencieuses} "
              f"silencieuse")

    def test_aucun_zero_invente(self):
        """Une hypothèse sans puissance n'en publie PAS une nulle : un zéro se
        lirait « aucune capacité de détection », ce qui est faux d'un contrôle
        de plage."""
        n2 = _run()['n2']
        for cle in ('clm', 'bfcc', 'bootstrap_hyp', 'munich_hyp'):
            for code, e in ((n2.get(cle) or {}).get('puissance') or {}).items():
                if not e.get('mesurable'):
                    self.assertNotIn('puissance', e, code)
                    self.assertIn("n'y a pas de sens", e['phrase'], code)
        print("    OK P3-1b aucune hypothèse sans puissance ne publie de "
              "chiffre — le motif remplace le nombre")

    def test_boot_h3_dit_pourquoi_son_chiffre_manque(self):
        """⚠️ LE SEUL CAS OÙ LA PUISSANCE EST MESURABLE ET N'EST PAS PUBLIÉE.

        Deux campagnes s'opposent — celles documentées en tête du module, sur
        triangles synthétiques (témoins 5,0 % et 6,4 %), et une campagne
        régénérée depuis l'ajustement des triangles de référence (21 à 30 %).
        Aucune n'est écartée : tant que l'écart n'est pas expliqué, choisir
        l'une reviendrait à publier un chiffre sans raison de le préférer.
        """
        e = puissance_bootstrap()['BOOT-H3']
        self.assertFalse(e['mesurable'])
        self.assertIn('réconcilier', e['motif'])
        self.assertNotIn('puissance', e)
        print("    OK P3-1c BOOT-H3 énonce pourquoi son chiffre n'est pas "
              "publié, au lieu de se taire ou d'en inventer un")


# =============================================================================
#  P3-2 — LES TROIS FORMATS DISENT LA MÊME CHOSE
# =============================================================================

class TP3_2_Formats(unittest.TestCase):

    def test_les_quatre_familles_ont_une_source_d_affichage(self):
        """CLM n'en avait pas : ses quatre hypothèses — celles des méthodes qui
        portent le Best Estimate — n'atteignaient aucun livrable."""
        lignes = _toutes_les_lignes(_run()['n2'])
        codes = [x['code'] for x in lignes]
        self.assertEqual(len(codes), 19)
        for prefixe, attendu in (('CLM-', 4), ('BFCC-', 6), ('BOOT-', 4),
                                 ('MCL-', 5)):
            self.assertEqual(sum(1 for c in codes if c.startswith(prefixe)),
                             attendu, prefixe)
        print("    OK P3-2a les quatre familles publient leurs lignes : "
              "4 CLM + 6 BFCC + 4 BOOT + 5 MCL = 19")

    def test_le_html_porte_les_cartes_clm_et_les_puissances(self):
        html = _run().get('html') or ''
        for libelle in ("Chain Ladder — Indépendance",
                        "Chain Ladder — Existence",
                        "Mack — Structure de variance",
                        "Mack — Incertitude"):
            self.assertIn(libelle, html, libelle)
        phrases = [x['puissance_phrase'] for x in
                   _toutes_les_lignes(_run()['n2']) if x['puissance_phrase']]
        self.assertEqual(len(phrases), 19)
        for p in phrases:
            self.assertIn(p[:50], html)
        print(f"    OK P3-2b le HTML porte les 4 cartes CLM et les "
              f"{len(phrases)} phrases de puissance")

    def test_l_excel_porte_les_memes_lignes_et_la_colonne_puissance(self):
        octets = _run().get('excel_bytes') or b''
        if not octets:
            self.skipTest("openpyxl absent — pas d'Excel à relire")
        import openpyxl
        ws = openpyxl.load_workbook(io.BytesIO(octets))['6. Hypothèses']
        self.assertEqual(ws.cell(row=2, column=6).value,
                         "Puissance disponible")
        colonne = [str(ws.cell(row=i, column=6).value or '')
                   for i in range(3, ws.max_row + 1)]
        remplies = [c for c in colonne if c]
        self.assertEqual(len(remplies), 19,
                         "l'Excel ne porte pas les dix-neuf puissances")
        libelles = [str(ws.cell(row=i, column=1).value or '')
                    for i in range(3, ws.max_row + 1)]
        self.assertTrue(any('Chain Ladder —' in x for x in libelles),
                        "les hypothèses de Chain Ladder manquent à l'Excel")
        print(f"    OK P3-2c l'Excel porte la colonne « Puissance disponible » "
              f"remplie {len(remplies)} fois, et les lignes CLM")

    def test_les_formats_ne_devient_pas_de_la_source(self):
        """HTML et Excel doivent porter EXACTEMENT les phrases produites par
        `lignes_hypotheses_*` — une reformulation locale ferait diverger les
        livrables entre eux."""
        r = _run()
        html = r.get('html') or ''
        octets = r.get('excel_bytes') or b''
        if not octets:
            self.skipTest("openpyxl absent")
        import openpyxl
        ws = openpyxl.load_workbook(io.BytesIO(octets))['6. Hypothèses']
        dans_excel = {str(ws.cell(row=i, column=6).value or '')
                      for i in range(3, ws.max_row + 1)}
        for ligne in _toutes_les_lignes(r['n2']):
            phrase = ligne['puissance_phrase']
            if not phrase:
                continue
            self.assertIn(phrase[:60], html, ligne['code'])
            self.assertIn(phrase, dans_excel, ligne['code'])
        print("    OK P3-2d les 19 phrases sont identiques dans la source, le "
              "HTML et l'Excel")


if __name__ == '__main__':
    unittest.main(verbosity=2)
