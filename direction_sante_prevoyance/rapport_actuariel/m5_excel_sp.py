"""
m5_excel_sp.py — Export Excel 8 onglets Santé-Prévoyance
Rapport Actuariel SP · Direction Santé-Prévoyance

8 onglets :
  1. Synthèse        — KPIs SP + statuts RAG
  2. Sinistralité    — LR par poste santé (DREES 2023)
  3. Provisions      — PSAP + PM + BE + RA (santé + prévoyance)
  4. Tarification    — Primes pures par garantie et CSP
  5. Morbidité       — Tables BCAC 2019 + Markov ITT/IP
  6. Solvabilité S2  — SCR NSLT + SCR SLT + consolidé EIOPA
  7. IFRS 17         — BE / RA / FCF / CSM / LC
  8. Hypothèses      — Toutes les constantes sourcées

Style : palette Navy/Gold ActuarIA, formats €/%
"""

import io
import logging
from datetime import datetime
from typing import Dict, Optional

import numpy as np

logger = logging.getLogger("actuaria.sp.rapport.excel")

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

# ── Palette ActuarIA ──────────────────────────────────────────────────────────
NAVY     = "0F2E52"
NAVY_L   = "1B3A5C"
OR       = "C9A84C"
BLANC    = "F0F4F8"
VERT     = "2ECC71"
ROUGE    = "E74C3C"
AMBRE    = "F39C12"
GRIS     = "8A9AB0"
BLEU     = "3498DB"
VIOLET   = "9B59B6"


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color=BLANC, size=11) -> Font:
    return Font(name="Arial", bold=bold, color=color, size=size)

def _border() -> Border:
    s = Side(style="thin", color="DDE4EE")
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h="left", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _header_row(ws, row, cols, fill_hex=NAVY):
    for col, val in enumerate(cols, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill    = _fill(fill_hex)
        c.font    = _font(bold=True, color=BLANC)
        c.alignment = _align("center")
        c.border  = _border()

def _data_row(ws, row, vals, fill_hex=None):
    bg = fill_hex or ("1B3A5C" if row % 2 == 0 else "243F6A")
    for col, val in enumerate(vals, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill      = _fill(bg)
        c.font      = _font(color=BLANC, size=10)
        c.alignment = _align()
        c.border    = _border()

def _rag_color(rag: str) -> str:
    return {"VERT": VERT, "AMBRE": AMBRE, "ROUGE": ROUGE}.get(rag, GRIS)

def _set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── ONGLET 1 : Synthèse ───────────────────────────────────────────────────────
def _onglet_synthese(wb, data: Dict):
    ws = wb.create_sheet("1. Synthèse")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    # Titre
    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value = f"RAPPORT ACTUARIEL SANTÉ-PRÉVOYANCE — {data.get('entite','')}"
    t.fill  = _fill(NAVY); t.font = _font(bold=True, size=14, color=OR)
    t.alignment = _align("center")

    ws.merge_cells("A2:F2")
    d = ws["A2"]
    d.value = f"Arrêté au {data.get('date_arrete','')} | Généré le {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    d.fill  = _fill(NAVY_L); d.font = _font(size=10, color=BLANC)
    d.alignment = _align("center")

    # KPIs santé
    _header_row(ws, 4, ["Indicateur", "Santé", "Prévoyance", "Consolidé", "Unité", "Statut"])
    kpis = [
        ("Primes Acquises",     data.get("pa_sante",0),    data.get("pa_prev",0),    data.get("pa_sante",0)+data.get("pa_prev",0),    "€",  ""),
        ("Best Estimate (BE)",  data.get("be_sante",0),    data.get("be_prev",0),    data.get("be_sante",0)+data.get("be_prev",0),    "€",  ""),
        ("Risk Adjustment",     data.get("ra_sante",0),    data.get("ra_prev",0),    data.get("ra_sante",0)+data.get("ra_prev",0),    "€",  ""),
        ("Loss Ratio",          data.get("lr_sante",0),    data.get("lr_prev",0),    "",    "%",  ""),
        ("SCR",                 data.get("scr_sante",0),   data.get("scr_prev",0),   data.get("scr_consolide",0),  "€",  ""),
        ("Ratio SCR",           data.get("ratio_scr_sante",0), data.get("ratio_scr_prev",0), data.get("ratio_scr",0), "%", data.get("statut_rag","?")),
        ("MCR",                 data.get("mcr_sante",0),   data.get("mcr_prev",0),   data.get("mcr_consolide",0),  "€",  ""),
        ("Fonds Propres",       "",                         "",                        data.get("fonds_propres",0), "€",  ""),
    ]
    for r, (label, s, p, c, u, st) in enumerate(kpis, 5):
        fill = _rag_color(st) if st in ("VERT","AMBRE","ROUGE") else None
        _data_row(ws, r, [label,
            f"{s:,.0f}" if isinstance(s,float) and s else s,
            f"{p:,.0f}" if isinstance(p,float) and p else p,
            f"{c:,.0f}" if isinstance(c,float) and c else c,
            u, st], fill_hex=fill if fill else None)

    _set_col_widths(ws, [28, 16, 16, 16, 8, 10])
    return ws


# ── ONGLET 2 : Sinistralité par poste ────────────────────────────────────────
def _onglet_sinistralite(wb, data: Dict):
    ws = wb.create_sheet("2. Sinistralité")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:G1")
    t = ws["A1"]
    t.value = "SINISTRALITÉ PAR POSTE — Source : DREES 2023 + FNMF 2023"
    t.fill = _fill(NAVY); t.font = _font(bold=True, color=OR)
    t.alignment = _align("center")

    _header_row(ws, 2, ["Poste", "Fréquence/an", "Coût moyen (€)", "Charge mutuelle (€)",
                          "Remb. SS (%)", "LR poste", "Source"])

    postes = data.get("sinistralite_par_poste", {})
    for r, (poste, infos) in enumerate(postes.items(), 3):
        if isinstance(infos, dict):
            _data_row(ws, r, [
                poste.capitalize(),
                f"{infos.get('frequence_an', 0):.2f}",
                f"{infos.get('cout_acte', infos.get('cout_moyen',0)):,.2f}",
                f"{infos.get('charge_mutuelle', 0):,.2f}",
                f"{infos.get('remb_ss', infos.get('taux_ss',0))*100:.1f}%",
                f"{infos.get('lr_poste', 0):.1%}" if infos.get("lr_poste") else "—",
                "DREES 2023",
            ])
        else:
            _data_row(ws, r, [poste.capitalize(), "—", "—", f"{infos:,.2f}", "—", "—", "DREES/FNMF"])

    _set_col_widths(ws, [20, 14, 16, 18, 14, 12, 14])
    return ws


# ── ONGLET 3 : Provisions ────────────────────────────────────────────────────
def _onglet_provisions(wb, data: Dict):
    ws = wb.create_sheet("3. Provisions")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:E1")
    t = ws["A1"]
    t.value = "PROVISIONS TECHNIQUES — IFRS 17 + SOLVABILITÉ 2"
    t.fill = _fill(NAVY); t.font = _font(bold=True, color=OR)
    t.alignment = _align("center")

    _header_row(ws, 2, ["Poste", "Santé", "Prévoyance", "Consolidé", "Référence"])
    lignes = [
        ("PSAP",              data.get("psap_sante",0),   data.get("psap_prev",0),   "", "Art.27 C.ass."),
        ("PM Rentes IP",      0,                          data.get("pm_rentes_ip",0),"", "BCAC 2019 + EIOPA RFR"),
        ("IBNR",              data.get("ibnr_sante",0),   data.get("ibnr_prev",0),   "", "FNMF/CTIP 2023"),
        ("Best Estimate (BE)",data.get("be_sante",0),     data.get("be_prev",0),     data.get("be_total",0), "IFRS 17 §33"),
        ("Risk Adjustment",   data.get("ra_sante",0),     data.get("ra_prev",0),     data.get("ra_sante",0)+data.get("ra_prev",0), "IFRS 17 §B91 CoC 6%"),
        ("TP = BE + RA",      data.get("tp_sante",0),     data.get("tp_prev",0),     data.get("tp_sante",0)+data.get("tp_prev",0), "IFRS 17 §33"),
    ]
    for r, (label, s, p, c, ref) in enumerate(lignes, 3):
        _data_row(ws, r, [label,
            f"{s:,.0f}" if s else "—",
            f"{p:,.0f}" if p else "—",
            f"{c:,.0f}" if c else "—",
            ref])

    _set_col_widths(ws, [24, 16, 16, 16, 28])
    return ws


# ── ONGLET 4 : Tarification ───────────────────────────────────────────────────
def _onglet_tarification(wb, data: Dict):
    ws = wb.create_sheet("4. Tarification")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:E1")
    t = ws["A1"]
    t.value = "TARIFICATION — Primes pures par garantie et CSP"
    t.fill = _fill(NAVY); t.font = _font(bold=True, color=OR)
    t.alignment = _align("center")

    _header_row(ws, 2, ["Garantie / CSP", "Prime pure (€)", "Chargement (%)", "Prime commerciale (€)", "Source"])

    tarif = data.get("tarification", {})
    r = 3
    # Santé par niveau
    for niv in ["eco","confort","premium","luxe"]:
        p = tarif.get(f"prime_{niv}", 0)
        if p:
            _data_row(ws, r, [f"Santé — {niv.capitalize()}", f"{p:,.2f}",
                               f"{tarif.get('chargement_pct',0.18)*100:.0f}%",
                               f"{p/(1-tarif.get('chargement_pct',0.18)):,.2f}",
                               "DREES 2023"])
            r += 1
    # Prévoyance par CSP
    for csp in ["ouvrier","employe","cadre","cadre_sup"]:
        p_itt = tarif.get(f"itt_{csp}", 0)
        if p_itt:
            _data_row(ws, r, [f"ITT — {csp}", f"{p_itt:,.2f}", "20%",
                               f"{p_itt/0.80:,.2f}", "BCAC 2019"])
            r += 1

    _set_col_widths(ws, [26, 16, 14, 20, 16])
    return ws


# ── ONGLET 5 : Morbidité ────────────────────────────────────────────────────
def _onglet_morbidite(wb, data: Dict):
    ws = wb.create_sheet("5. Morbidité")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value = "TABLES DE MORBIDITÉ — BCAC 2019 + TD 88-90 + Markov"
    t.fill = _fill(NAVY); t.font = _font(bold=True, color=OR)
    t.alignment = _align("center")

    _header_row(ws, 2, ["Âge", "Taux ITT (BCAC)", "Taux IP (TD 88-90)",
                          "P(maintien 6m)", "P(maintien 12m)", "Source"])
    morb = data.get("morbidite", {})
    ages = morb.get("ages", [25,30,35,40,45,50,55,60])
    taux_itt = morb.get("taux_itt_par_age", {})
    taux_ip  = morb.get("taux_ip_par_age", {})
    maintien = morb.get("maintien", {})

    for r, age in enumerate(ages, 3):
        _data_row(ws, r, [
            age,
            f"{taux_itt.get(age,0)*100:.2f}%",
            f"{taux_ip.get(age,0)*100:.3f}%",
            f"{maintien.get('mois_6',0):.3f}",
            f"{maintien.get('mois_12',0):.3f}",
            "BCAC 2019 / TD 88-90",
        ])

    _set_col_widths(ws, [8, 16, 18, 16, 16, 18])
    return ws


# ── ONGLET 6 : Solvabilité S2 ────────────────────────────────────────────────
def _onglet_solvabilite(wb, data: Dict):
    ws = wb.create_sheet("6. Solvabilité S2")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:D1")
    t = ws["A1"]
    t.value = "SOLVABILITÉ 2 — SCR/MCR consolidé (ρ=0.25 EIOPA Annexe IV)"
    t.fill = _fill(NAVY); t.font = _font(bold=True, color=OR)
    t.alignment = _align("center")

    _header_row(ws, 2, ["Module", "Montant (€)", "Ratio (%)", "Source réglementaire"])
    scr = [
        ("SCR Santé NSLT (σ=5%)",     data.get("scr_sante",0),       "", "Art.148 RD 2015/35"),
        ("SCR Invalidité SLT (+35%)",  data.get("scr_prev",0),        "", "Art.145 RD 2015/35"),
        ("Bénéfice diversification",   -data.get("diversification",0),"", "Annexe IV RD 2015/35"),
        ("SCR Consolidé",              data.get("scr_consolide",0),   "", "Formule standard EIOPA"),
        ("MCR Santé",                  data.get("mcr_sante",0),       "", "Art.252 + Art.129 S2"),
        ("MCR Prévoyance",             data.get("mcr_prev",0),        "", "Art.252 + Art.129 S2"),
        ("MCR Consolidé",              data.get("mcr_consolide",0),   "", ""),
        ("Fonds Propres Éligibles",    data.get("fonds_propres",0),   "", "Bilan S2"),
        ("Ratio SCR (%)",              "",  f"{data.get('ratio_scr',0):.1f}%", "FP/SCR — seuil 100%"),
        ("Ratio MCR (%)",              "",  f"{data.get('ratio_mcr',0):.1f}%", "FP/MCR — seuil 100%"),
    ]
    for r, (label, m, ratio, ref) in enumerate(scr, 3):
        fill = None
        if label == "Ratio SCR (%)":
            v = data.get("ratio_scr",0)
            fill = VERT if v>=130 else (AMBRE if v>=100 else ROUGE)
        _data_row(ws, r, [label, f"{m:,.0f}" if m else "—", ratio or "—", ref],
                  fill_hex=fill)

    _set_col_widths(ws, [32, 16, 12, 28])
    return ws


# ── ONGLET 7 : IFRS 17 ───────────────────────────────────────────────────────
def _onglet_ifrs17(wb, data: Dict):
    ws = wb.create_sheet("7. IFRS 17")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:E1")
    t = ws["A1"]
    t.value = "IFRS 17 — BE / RA / FCF / CSM / LC"
    t.fill = _fill(NAVY); t.font = _font(bold=True, color=OR)
    t.alignment = _align("center")

    _header_row(ws, 2, ["Composante", "Santé (PAA)", "Prévoyance (GMM)", "Total", "Référence IFRS 17"])
    i17 = data.get("ifrs17", {})
    lignes = [
        ("Best Estimate",     i17.get("be_sante",0),  i17.get("be_prev",0),  i17.get("be_total",0),  "§33"),
        ("Risk Adjustment",   i17.get("ra_sante",0),  i17.get("ra_prev",0),  i17.get("ra_total",0),  "§37 + §B91 CoC"),
        ("FCF = BE + RA",     i17.get("fcf_sante",0), i17.get("fcf_prev",0), i17.get("fcf_total",0), "§33"),
        ("CSM (profit futur)",i17.get("csm_sante",0), i17.get("csm_prev",0), i17.get("csm_total",0), "§38"),
        ("LC (perte onéros.)",i17.get("lc_sante",0),  i17.get("lc_prev",0),  i17.get("lc_total",0),  "§49"),
        ("Amort. CSM annuel", i17.get("csm_release",0),"",                    "",                      "§B119 — linéaire"),
    ]
    for r, (label, s, p, c, ref) in enumerate(lignes, 3):
        _data_row(ws, r, [label,
            f"{s:,.0f}" if isinstance(s,(int,float)) and s else "—",
            f"{p:,.0f}" if isinstance(p,(int,float)) and p else "—",
            f"{c:,.0f}" if isinstance(c,(int,float)) and c else "—",
            ref])

    _set_col_widths(ws, [22, 16, 18, 14, 24])
    return ws


# ── ONGLET 8 : Hypothèses ────────────────────────────────────────────────────
def _onglet_hypotheses(wb, data: Dict):
    ws = wb.create_sheet("8. Hypothèses")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:D1")
    t = ws["A1"]
    t.value = "HYPOTHÈSES ACTUARIELLES — Sources et valeurs"
    t.fill = _fill(NAVY); t.font = _font(bold=True, color=OR)
    t.alignment = _align("center")

    _header_row(ws, 2, ["Hypothèse", "Valeur", "Statut", "Source"])
    hyps = data.get("hypotheses_m2", [])
    for r, h in enumerate(hyps, 3):
        fill = _rag_color(h.get("statut",""))
        _data_row(ws, r, [
            h.get("hypothese",""),
            h.get("valeur",""),
            h.get("statut",""),
            h.get("source",""),
        ], fill_hex=fill if h.get("statut") in ("VALIDÉE","NON VALIDÉE","À JUSTIFIER") else None)

    _set_col_widths(ws, [36, 24, 14, 30])
    return ws


# ── POINT D'ENTRÉE ───────────────────────────────────────────────────────────
def export_excel_sp(data: Dict) -> bytes:
    """
    Génère le classeur Excel SP 8 onglets.

    Parameters
    ----------
    data : dict — données consolidées du rapport SP

    Returns
    -------
    bytes — contenu du fichier .xlsx
    """
    if not OPENPYXL_OK:
        logger.warning("openpyxl non disponible — export Excel SP désactivé")
        return b""

    wb = Workbook()
    # Supprimer la feuille par défaut
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    _onglet_synthese(wb, data)
    _onglet_sinistralite(wb, data)
    _onglet_provisions(wb, data)
    _onglet_tarification(wb, data)
    _onglet_morbidite(wb, data)
    _onglet_solvabilite(wb, data)
    _onglet_ifrs17(wb, data)
    _onglet_hypotheses(wb, data)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
