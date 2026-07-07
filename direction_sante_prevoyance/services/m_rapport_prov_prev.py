"""
Palette et helpers communs aux modules de rapport SP.
Reproduit fidèlement le style Navy/Gold de n5_rapport.py.
"""
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# Palette Navy/Gold ActuarIA
NAVY   = "0F2E52"
NAVY_L = "1B3A5C"
GOLD   = "C9A84C"
GOLD_L = "E2C97E"
GRIS   = "8A9BB0"
VERT   = "1E8449"
ROUGE  = "C0392B"
AMBRE  = "E67E22"

def _font(bold=False, size=10, color="1C2B3A", italic=False):
    return Font(name="Inter", bold=bold, size=size, color=color, italic=italic)

def _fill(hex6):
    return PatternFill("solid", start_color=hex6, end_color=hex6)

def _border_thin():
    s = Side(style="thin", color="DDE4EE")
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _header_row(ws, row, cols, fill_hex=NAVY):
    """Ligne d'en-tête Navy avec texte blanc gras."""
    for col, txt in enumerate(cols, 1):
        c = ws.cell(row=row, column=col, value=txt)
        c.font      = _font(bold=True, size=9, color="FFFFFF")
        c.fill      = _fill(fill_hex)
        c.alignment = _align("center")
        c.border    = _border_thin()

def _data_row(ws, row, values, even=True, bold=False, gold=False):
    """Ligne de données alternée."""
    fill_hex = "F5F7FA" if even else "FFFFFF"
    if gold:
        fill_hex = "F9F4E8"
    for col, v in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=v)
        c.font      = _font(bold=bold, size=9,
                            color=GOLD if gold else "1C2B3A")
        c.fill      = _fill(fill_hex)
        c.alignment = _align()
        c.border    = _border_thin()

def _title_block(ws, titre, sous_titre, rag="VERT"):
    """Bloc titre Navy en haut de la feuille."""
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value     = "ActuarIA — " + titre
    c.font      = _font(bold=True, size=14, color="FFFFFF")
    c.fill      = _fill(NAVY)
    c.alignment = _align("center")

    ws.merge_cells("A2:H2")
    c = ws["A2"]
    c.value     = sous_titre
    c.font      = _font(size=9, color=GOLD_L)
    c.fill      = _fill(NAVY_L)
    c.alignment = _align("center")

    # Statut RAG
    rag_colors = {"VERT": VERT, "AMBRE": AMBRE, "ROUGE": ROUGE}
    ws.merge_cells("A3:H3")
    c = ws["A3"]
    rag_label = {"VERT": "✅ CONFORME", "AMBRE": "⚠️ VIGILANCE", "ROUGE": "❌ ALERTE"}
    c.value     = "Statut RAG : " + rag_label.get(rag, rag)
    c.font      = _font(bold=True, size=9, color=rag_colors.get(rag, GRIS))
    c.fill      = _fill("FAFBFC")
    c.alignment = _align("center")

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 16

def _col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def _section_title(ws, row, titre, n_cols=8):
    ws.merge_cells(f"A{row}:{get_column_letter(n_cols)}{row}")
    c = ws.cell(row=row, column=1, value=titre)
    c.font      = _font(bold=True, size=10, color=GOLD)
    c.fill      = _fill(NAVY)
    c.alignment = _align("left")
    ws.row_dimensions[row].height = 20
    return row + 1

def _f(v, dec=0):
    if v is None: return "—"
    try:
        fv = float(v)
        if dec == 0:
            return f"{fv:,.0f} €".replace(",", " ")
        return f"{fv:,.{dec}f}".replace(",", " ")
    except: return "—"

def _pct(v, dec=1):
    if v is None: return "—"
    try: return f"{float(v):.{dec}f} %"
    except: return "—"


from openpyxl import Workbook
import io

def generer(result_p3: dict) -> bytes:
    """
    Rapport provisionnement prévoyance depuis P3 Élodie.
    3 onglets : Synthèse provisions · Triangle ITT · Méthodes
    """
    wb = Workbook()
    rag = result_p3.get("statut_rag", "AMBRE")

    # ── Onglet 1 : Synthèse ───────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Provisions Prévoyance"
    _title_block(ws, "Provisionnement Prévoyance", "Agent P3 Élodie — CL · Mack · BF · Bootstrap", rag)
    _col_widths(ws, [34, 18, 18, 28])

    r = 5
    r = _section_title(ws, r, "📊 PROVISIONS TECHNIQUES")
    _header_row(ws, r, ["Composante", "Montant (€)", "% Primes", "Référence"])
    r += 1

    pa   = float(result_p3.get("primes_acquises",   0) or 0)
    be   = float(result_p3.get("be_prevoyance",     0) or 0)
    be_i = float(result_p3.get("be_itt",            0) or 0)
    pm_r = float(result_p3.get("pm_rentes_ip",      0) or 0)
    psap = float(result_p3.get("psap_total",        0) or 0)
    prec = float(result_p3.get("prec",              0) or 0)
    ra   = float(result_p3.get("risk_adjustment",   0) or 0)
    tp   = float(result_p3.get("tp_prevoyance",     0) or 0)
    lr   = float(result_p3.get("loss_ratio",        0) or 0)

    prov_rows = [
        ("BE ITT (triangle Chain Ladder/Mack/BF)", _f(be_i), _pct(be_i/pa*100 if pa else 0), "Méthode pondérée"),
        ("PM Rentes IP (actuariat vie)",            _f(pm_r), _pct(pm_r/pa*100 if pa else 0), "TH0002 / TD88-90"),
        ("PSAP ITT (dossiers en cours)",            _f(psap), _pct(psap/pa*100 if pa else 0), "Sinistres déclarés"),
        ("PREC (provision risques en cours)",       _f(prec), _pct(prec/pa*100 if pa else 0), "Art. 78 S2"),
        ("Best Estimate Prévoyance",                _f(be),   _pct(be/pa*100 if pa else 0),   "Art. 77 §1 S2"),
        ("Risk Adjustment (CoC 6% floor 3%)",       _f(ra),   _pct(ra/pa*100 if pa else 0),   "IFRS 17 §B119"),
        ("TP Prévoyance (BE + RA)",                 _f(tp),   _pct(tp/pa*100 if pa else 0),   "Art. 77 §1 S2"),
    ]
    for i, row_p in enumerate(prov_rows):
        gold = i >= 4
        _data_row(ws, r, list(row_p), even=(i % 2 == 0), gold=gold, bold=gold)
        r += 1

    r += 1
    r = _section_title(ws, r, "📈 LOSS RATIO PRÉVOYANCE")
    _header_row(ws, r, ["Indicateur", "Valeur", "Norme", "Statut"])
    r += 1
    lr_statut = "✅ OK" if lr <= 0.85 else ("⚠️ VIGILANCE" if lr <= 0.95 else "❌ ALERTE")
    _data_row(ws, r, ["Loss Ratio CTIP", _pct(lr * 100), "≤ 85% (CTIP 2023)", lr_statut])
    r += 1

    # Métriques incertitude
    p90  = float(result_p3.get("reserve_p90",    0) or 0)
    p995 = float(result_p3.get("reserve_p99_5",  0) or 0)
    cv   = float(result_p3.get("cv_inter",       0) or 0)
    scr  = float(result_p3.get("scr_invalidite", 0) or 0)
    r += 1
    r = _section_title(ws, r, "📊 INCERTITUDE ET SCR")
    _header_row(ws, r, ["Indicateur", "Valeur", "Référence", ""])
    r += 1
    for i, (k, v, ref) in enumerate([
        ("CV inter-méthodes",  _pct(cv),   "< 20% norme CTIP"),
        ("Provision P90",      _f(p90),    "Percentile 90%"),
        ("Provision P99,5",    _f(p995),   "SCR Art. 105 S2"),
        ("SCR Prévoyance",     _f(scr),    "3 × σ × BE"),
    ]):
        _data_row(ws, r, [k, v, ref, ""], even=(i % 2 == 0))
        r += 1

    # ── Onglet 2 : Triangle ITT ───────────────────────────────────────────────
    ws2 = wb.create_sheet("Triangle ITT")
    _title_block(ws2, "Triangle de Développement ITT",
                 "Liquidation prévoyance : horizon 60-84 mois", rag)
    _col_widths(ws2, [14] + [12]*6)

    r2 = 5
    r2 = _section_title(ws2, r2, "🔢 TRIANGLE CUMULÉ (CTIP — semestres)")
    triangle = result_p3.get("triangle", {})
    data_tri = triangle.get("data", [])
    if data_tri:
        n_cols = max(len(row) for row in data_tri) if data_tri else 6
        _header_row(ws2, r2, ["An. Surv."] + [f"S{j+1}" for j in range(n_cols)])
        r2 += 1
        annees = triangle.get("annees_survenance", [])
        for i, row_tri in enumerate(data_tri):
            an = annees[i] if i < len(annees) else f"AS-{i}"
            _data_row(ws2, r2, [an] + [_f(v) if v else "—" for v in row_tri],
                      even=(i % 2 == 0))
            r2 += 1
    else:
        ws2.cell(row=r2, column=1,
                 value="Triangle non disponible — données insuffisantes")

    # ── Onglet 3 : Méthodes ───────────────────────────────────────────────────
    ws3 = wb.create_sheet("Méthodes Actuarielles")
    _title_block(ws3, "Résultats par Méthode", "CL · Mack · BF (CTIP) · Bootstrap ODP", rag)
    _col_widths(ws3, [28, 18, 14, 28])

    r3 = 5
    r3 = _section_title(ws3, r3, "📋 RÉSERVES PAR MÉTHODE")
    _header_row(ws3, r3, ["Méthode", "Réserve IBNR (€)", "Poids BE (%)", "Statut"])
    r3 += 1

    methodes = result_p3.get("methodes", {})
    poids    = result_p3.get("poids_methodes", {})
    meth_list = [
        ("Chain Ladder",         methodes.get("chain_ladder",    {}).get("reserve_totale")),
        ("Mack 1993",            methodes.get("mack",            {}).get("reserve_best_estimate")),
        ("Bornhuetter-Ferguson", methodes.get("bf",              {}).get("reserve_totale")),
        ("Bootstrap ODP",        methodes.get("bootstrap",       {}).get("be_bootstrap")),
    ]
    for i, (nom, res) in enumerate(meth_list):
        pds = poids.get(nom.lower().replace("-", "_").replace(" ", "_"), 0)
        _data_row(ws3, r3, [nom, _f(res), _pct(pds*100), "✓ Inclus" if pds > 0 else "—"],
                  even=(i % 2 == 0))
        r3 += 1

    _data_row(ws3, r3, ["BEST ESTIMATE S2", _f(be), "100%", "→ Bilan S2"],
              gold=True, bold=True)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
