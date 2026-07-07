"""
Palette et helpers communs aux modules de rapport SP.
Reproduit fidèlement le style Navy/Gold de n5_rapport.py.
"""
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# Palette Navy/Gold ActuarIA
NAVY   = "0F2E52"
NAVY_L = "1B3A5C"
GOLD   = "C9A84C"
GOLD_L = "E2C97E"
BLANC  = "F0F4F8"
GRIS   = "8A9BB0"
VERT   = "1E8449"
ROUGE  = "C0392B"
AMBRE  = "E67E22"
BLEU   = "3498DB"

def _font(bold=False, size=10, color="1C2B3A", italic=False):
    return Font(name="Inter", bold=bold, size=size, color=color, italic=italic)

def _fill(hex6):
    return PatternFill("solid", start_color=hex6, end_color=hex6)

def _border_thin():
    s = Side(style="thin", color="DDE4EE")
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _header_row(ws, row, cols, fill_hex=NAVY):
    """Ligne d'en-tête Navy avec texte blanc gras."""
    for col, txt in enumerate(cols, 1):
        c = ws.cell(row=row, column=col, value=txt)
        c.font      = _font(bold=True, size=9, color="FFFFFF")
        c.fill      = _fill(fill_hex)
        c.alignment = _align("center")
        c.border    = _border_thin()

def _data_row(ws, row, values, even=True, bold=False, gold=False):
    """Ligne de données alternée."""
    fill_hex = "F5F7FA" if even else "FFFFFF"
    if gold:
        fill_hex = "F9F4E8"
    for col, v in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=v)
        c.font      = _font(bold=bold, size=9,
                            color=GOLD if gold else "1C2B3A")
        c.fill      = _fill(fill_hex)
        c.alignment = _align()
        c.border    = _border_thin()

def _title_block(ws, titre, sous_titre, rag="VERT"):
    """Bloc titre Navy en haut de la feuille."""
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value     = "ActuarIA — " + titre
    c.font      = _font(bold=True, size=14, color="FFFFFF")
    c.fill      = _fill(NAVY)
    c.alignment = _align("center")

    ws.merge_cells("A2:H2")
    c = ws["A2"]
    c.value     = sous_titre
    c.font      = _font(size=9, color=GOLD_L)
    c.fill      = _fill(NAVY_L)
    c.alignment = _align("center")

    # Statut RAG
    rag_colors = {"VERT": VERT, "AMBRE": AMBRE, "ROUGE": ROUGE}
    ws.merge_cells("A3:H3")
    c = ws["A3"]
    rag_label = {"VERT": "✅ CONFORME", "AMBRE": "⚠️ VIGILANCE", "ROUGE": "❌ ALERTE"}
    c.value     = "Statut RAG : " + rag_label.get(rag, rag)
    c.font      = _font(bold=True, size=9, color=rag_colors.get(rag, GRIS))
    c.fill      = _fill("FAFBFC")
    c.alignment = _align("center")

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 16

def _col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def _section_title(ws, row, titre, n_cols=8):
    ws.merge_cells(f"A{row}:{get_column_letter(n_cols)}{row}")
    c = ws.cell(row=row, column=1, value=titre)
    c.font      = _font(bold=True, size=10, color=GOLD)
    c.fill      = _fill(NAVY)
    c.alignment = _align("left")
    ws.row_dimensions[row].height = 20
    return row + 1

def _f(v, dec=0):
    if v is None: return "—"
    try:
        fv = float(v)
        if dec == 0:
            return f"{fv:,.0f} €".replace(",", " ")
        return f"{fv:,.{dec}f}".replace(",", " ")
    except: return "—"

def _pct(v, dec=1):
    if v is None: return "—"
    try: return f"{float(v):.{dec}f} %"
    except: return "—"


from openpyxl import Workbook
import io

def generer(result_p4: dict, result_alm: dict = None, result_reg1: dict = None) -> bytes:
    """
    Rapport réglementation prévoyance : P4 (SCR/MCR/QRT) + ALM + REG1 (S2 narratif).
    3 onglets : SCR/MCR/QRT · ALM · Rapport S2
    """
    wb = Workbook()
    rag  = result_p4.get("statut_rag",  "AMBRE")
    alm  = result_alm  or {}
    reg1 = result_reg1 or {}

    # ── Onglet 1 : SCR / MCR / QRT S.14.01 ───────────────────────────────────
    ws = wb.active
    ws.title = "SCR MCR QRT S14"
    _title_block(ws, "SCR/MCR Prévoyance", "Agent P4 Valentin — Art. 145 & 252 RD 2015/35 · QRT S.14.01", rag)
    _col_widths(ws, [36, 18, 18, 28])

    r = 5
    r = _section_title(ws, r, "🔢 QRT S.14.01 — PRÉVOYANCE COLLECTIVE")
    _header_row(ws, r, ["Composante", "Montant (€)", "Ratio/Info", "Référence réglementaire"])
    r += 1

    be   = float(result_p4.get("be_prevoyance",   0) or 0)
    ra   = float(result_p4.get("risk_adjustment", 0) or 0)
    tp   = float(result_p4.get("tp_prevoyance",   0) or 0)
    pm_r = float(result_p4.get("pm_rentes_ip",    0) or 0)
    psap = float(result_p4.get("psap_total",      0) or 0)
    scr_m= float(result_p4.get("scr_morbidite",   0) or 0)
    scr_c= float(result_p4.get("scr_cessation",   0) or 0)
    scr_l= float(result_p4.get("scr_longevite",   0) or 0)
    scr  = float(result_p4.get("scr_invalidite",  0) or 0)
    mcr  = float(result_p4.get("mcr",             0) or 0)
    fpp  = float(result_p4.get("fonds_propres",   0) or 0)
    r_scr= float(result_p4.get("ratio_scr_pct",   0) or 0)
    r_mcr= float(result_p4.get("ratio_mcr_pct",   0) or 0)

    scr_be = scr/be*100 if be else 0
    qrt_rows = [
        ("PM Rentes IP (long terme)",            _f(pm_r),  "",                "Actuariat vie — TH0002"),
        ("PSAP ITT (court terme)",               _f(psap),  "",                "Sinistres déclarés"),
        ("Best Estimate Prévoyance",             _f(be),    "100%",            "Art. 77 §1 S2"),
        ("Risk Adjustment (CoC 8%)",             _f(ra),    _pct(ra/be*100 if be else 0), "IFRS 17 §B119"),
        ("TP Prévoyance (BE + RA)",              _f(tp),    "",                "Art. 77 §1 S2"),
        ("SCR Morbidité (+35% incidence)",       _f(scr_m), "Choc dominant",   "Art. 145 §2(a) RD 2015/35"),
        ("SCR Cessation (-20% guérison)",        _f(scr_c), "Corrélation 0.25","Art. 145 §2(b) RD 2015/35"),
        ("SCR Longévité (-20% mortalité IP)",    _f(scr_l), "Corrélation 0.25","Art. 145 §2(c) RD 2015/35"),
        ("SCR Invalidité (agrégé EIOPA)",        _f(scr),   _pct(scr_be),      "EIOPA Annexe IV"),
        ("MCR Prévoyance",                       _f(mcr),   "Plancher 3,7 M€", "Art. 252 RD 2015/35"),
        ("Fonds Propres",                        _f(fpp),   "",                "Art. 87 S2"),
        ("Ratio SCR (%)",                        _pct(r_scr),"≥ 100% requis",  "Art. 138 S2"),
        ("Ratio MCR (%)",                        _pct(r_mcr),"≥ 100% requis",  "Art. 139 S2"),
    ]
    for i, row_q in enumerate(qrt_rows):
        gold = i >= 10
        _data_row(ws, r, list(row_q), even=(i % 2 == 0), gold=gold, bold=gold)
        r += 1

    # ── Onglet 2 : ALM ────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("ALM")
    _title_block(ws2, "Gestion Actif-Passif", "Duration · LCR · Immunisation Redington", rag)
    _col_widths(ws2, [30, 16, 16, 28])

    r2 = 5
    r2 = _section_title(ws2, r2, "📐 INDICATEURS ALM")
    _header_row(ws2, r2, ["Indicateur", "Valeur", "Norme/Cible", "Référence"])
    r2 += 1

    dur_actif  = float(alm.get("duration_actif",  0) or 0)
    dur_passif = float(alm.get("duration_passif", 0) or 0)
    gap_dur    = dur_actif - dur_passif
    lcr        = float(alm.get("lcr",             0) or 0)
    gap_ok = abs(gap_dur) <= 2
    lcr_ok = lcr >= 1.0

    alm_rows = [
        ("Duration actif",          f"{dur_actif:.2f} ans",  "—",              "Portefeuille obligataire"),
        ("Duration passif",         f"{dur_passif:.2f} ans", "—",              "PM + PSAP actualisés"),
        ("Gap de duration",         f"{gap_dur:+.2f} ans",   "≤ 2 ans idéal",  "Alerte si > 3 ans"),
        ("LCR (Liquidité)",         _pct(lcr*100),           "≥ 100%",         "Ratio couverture liquidité"),
        ("Immunisation Redington",  alm.get("redington_ok","—"), "Convexité ≥ 0","Condition 2e ordre"),
        ("BV01 stress +100bp",      _f(alm.get("bv01_stress_100", 0)), "€ impact", "Sensibilité taux"),
        ("BV01 stress +200bp",      _f(alm.get("bv01_stress_200", 0)), "€ impact", "Sensibilité taux"),
    ]
    for i, row_a in enumerate(alm_rows):
        _data_row(ws2, r2, list(row_a), even=(i % 2 == 0))
        r2 += 1

    r2 += 1
    r2 = _section_title(ws2, r2, "⚡ STATUT ALM")
    _header_row(ws2, r2, ["Contrôle", "Résultat", "Seuil", "Statut"])
    r2 += 1
    alm_checks = [
        ("Gap duration",     f"{abs(gap_dur):.2f} ans",  "≤ 2 ans", "✅" if gap_ok else "⚠️"),
        ("LCR",              _pct(lcr*100),              "≥ 100%",  "✅" if lcr_ok else "❌"),
    ]
    for i, row_c in enumerate(alm_checks):
        _data_row(ws2, r2, list(row_c), even=(i % 2 == 0))
        r2 += 1

    # ── Onglet 3 : Rapport S2 narratif (REG1) ────────────────────────────────
    ws3 = wb.create_sheet("Rapport S2 REG1")
    _title_block(ws3, "Rapport S2 Narratif", "SP-REG1 — Solvabilité 2 Santé-Prévoyance", rag)
    _col_widths(ws3, [24, 16, 16, 16, 24])

    r3 = 5
    r3 = _section_title(ws3, r3, "📋 SCR CONSOLIDÉ SP (Art. 148 + 145 S2)")
    _header_row(ws3, r3, ["Composante", "Montant (€)", "σ EIOPA", "ρ corr.", "Référence"])
    r3 += 1

    scr_sante = float(reg1.get("scr_sante",       0) or 0)
    scr_prev2 = float(reg1.get("scr_prevoyance",  0) or 0)
    scr_conso = float(reg1.get("scr_consolide",   0) or 0)

    consol_rows = [
        ("SCR Santé NSLT",           _f(scr_sante),  "0.10",  "—",    "Art. 148 RD 2015/35"),
        ("SCR Prévoyance SLT",       _f(scr_prev2),  "—",     "0.25", "Art. 145 RD 2015/35"),
        ("SCR Consolidé SP",         _f(scr_conso),  "—",     "0.25", "EIOPA Annexe IV — ρ=0.25"),
        ("Ratio SCR global",         _pct(reg1.get("ratio_scr_pct", 0)), "", "", "≥ 100% réglementaire"),
    ]
    for i, row_r in enumerate(consol_rows):
        gold = i == 2
        _data_row(ws3, r3, list(row_r), even=(i % 2 == 0), gold=gold, bold=gold)
        r3 += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
