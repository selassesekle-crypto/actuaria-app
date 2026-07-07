"""
Palette et helpers communs aux modules de rapport SP.
Reproduit fidèlement le style Navy/Gold de n5_rapport.py.
"""
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# Palette Navy/Gold ActuarIA
NAVY   = "0F2E52"
NAVY_L = "1B3A5C"
GOLD   = "C9A84C"
GOLD_L = "E2C97E"
BLANC  = "F0F4F8"
GRIS   = "8A9BB0"
VERT   = "1E8449"
ROUGE  = "C0392B"
AMBRE  = "E67E22"
BLEU   = "3498DB"

def _font(bold=False, size=10, color="1C2B3A", italic=False):
    return Font(name="Inter", bold=bold, size=size, color=color, italic=italic)

def _fill(hex6):
    return PatternFill("solid", start_color=hex6, end_color=hex6)

def _border_thin():
    s = Side(style="thin", color="DDE4EE")
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _header_row(ws, row, cols, fill_hex=NAVY):
    """Ligne d'en-tête Navy avec texte blanc gras."""
    for col, txt in enumerate(cols, 1):
        c = ws.cell(row=row, column=col, value=txt)
        c.font      = _font(bold=True, size=9, color="FFFFFF")
        c.fill      = _fill(fill_hex)
        c.alignment = _align("center")
        c.border    = _border_thin()

def _data_row(ws, row, values, even=True, bold=False, gold=False):
    """Ligne de données alternée."""
    fill_hex = "F5F7FA" if even else "FFFFFF"
    if gold:
        fill_hex = "F9F4E8"
    for col, v in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=v)
        c.font      = _font(bold=bold, size=9,
                            color=GOLD if gold else "1C2B3A")
        c.fill      = _fill(fill_hex)
        c.alignment = _align()
        c.border    = _border_thin()

def _title_block(ws, titre, sous_titre, rag="VERT"):
    """Bloc titre Navy en haut de la feuille."""
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value     = "ActuarIA — " + titre
    c.font      = _font(bold=True, size=14, color="FFFFFF")
    c.fill      = _fill(NAVY)
    c.alignment = _align("center")

    ws.merge_cells("A2:H2")
    c = ws["A2"]
    c.value     = sous_titre
    c.font      = _font(size=9, color=GOLD_L)
    c.fill      = _fill(NAVY_L)
    c.alignment = _align("center")

    # Statut RAG
    rag_colors = {"VERT": VERT, "AMBRE": AMBRE, "ROUGE": ROUGE}
    ws.merge_cells("A3:H3")
    c = ws["A3"]
    rag_label = {"VERT": "✅ CONFORME", "AMBRE": "⚠️ VIGILANCE", "ROUGE": "❌ ALERTE"}
    c.value     = "Statut RAG : " + rag_label.get(rag, rag)
    c.font      = _font(bold=True, size=9, color=rag_colors.get(rag, GRIS))
    c.fill      = _fill("FAFBFC")
    c.alignment = _align("center")

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 16

def _col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def _section_title(ws, row, titre, n_cols=8):
    ws.merge_cells(f"A{row}:{get_column_letter(n_cols)}{row}")
    c = ws.cell(row=row, column=1, value=titre)
    c.font      = _font(bold=True, size=10, color=GOLD)
    c.fill      = _fill(NAVY)
    c.alignment = _align("left")
    ws.row_dimensions[row].height = 20
    return row + 1

def _f(v, dec=0):
    if v is None: return "—"
    try:
        fv = float(v)
        if dec == 0:
            return f"{fv:,.0f} €".replace(",", " ")
        return f"{fv:,.{dec}f}".replace(",", " ")
    except: return "—"

def _pct(v, dec=1):
    if v is None: return "—"
    try: return f"{float(v):.{dec}f} %"
    except: return "—"


from openpyxl import Workbook
import io

def generer(result_s1: dict) -> bytes:
    """
    Génère le rapport Excel de tarification santé depuis le résultat de S1 Léonie.
    3 onglets : Synthèse · Tarification par poste · Hypothèses
    """
    wb = Workbook()

    # ── Onglet 1 : Synthèse ───────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Synthèse Tarification"

    rag = result_s1.get("statut_rag", "AMBRE")
    _title_block(ws, "Rapport Tarification Santé", "Agent S1 Léonie — DREES 2023 · ANI 2013", rag)
    _col_widths(ws, [22, 18, 18, 18, 18, 18, 18, 18])

    r = 5
    r = _section_title(ws, r, "📊 INDICATEURS CLÉS")
    _header_row(ws, r, ["Indicateur", "Valeur", "Référence"])
    r += 1

    kpis = [
        ("Prime pure unitaire",       _f(result_s1.get("prime_pure")),        "Tables DREES 2023"),
        ("Prime commerciale",          _f(result_s1.get("prime_commerciale")),  "Chargement inclus"),
        ("Prime mensuelle",            _f(result_s1.get("prime_mensuelle")),    "÷ 12"),
        ("Primes acquises portefeuille",_f(result_s1.get("primes_acquises")),   "Prime comm. × assurés"),
        ("Ratio S/P attendu",          _pct(result_s1.get("ratio_sp_attendu", 0) * 100), "Norme [65%-85%]"),
        ("Nb assurés",                 result_s1.get("nb_assures", "—"),       ""),
        ("Âge moyen",                  result_s1.get("age_moyen", "—"),        ""),
        ("Niveau garantie",            result_s1.get("garantie_niveau", "—"),  "eco/confort/premium/luxe"),
        ("Contrat",                    result_s1.get("contrat", "—"),          "individuel/collectif"),
        ("Source données",             result_s1.get("source_donnees", "—"),   ""),
    ]
    for i, (k, v, ref) in enumerate(kpis):
        _data_row(ws, r, [k, v, ref], even=(i % 2 == 0))
        r += 1

    r += 1
    r = _section_title(ws, r, "✅ CONFORMITÉ ANI 2013")
    ani = result_s1.get("ani_detail", {})
    note = ani.get("note_globale", "—")
    ws.merge_cells(f"A{r}:H{r}")
    c = ws.cell(row=r, column=1, value=note)
    c.font      = _font(size=9, color=VERT if "conforme" in note.lower() else ROUGE)
    c.fill      = _fill("FAFBFC")
    c.alignment = _align("left")
    r += 1

    detail = ani.get("detail", {})
    _header_row(ws, r, ["Poste ANI", "Seuil min (€)", "Charge mutuelle (€)", "Conforme", "Note"])
    r += 1
    for i, (poste, d) in enumerate(detail.items()):
        ok   = d.get("ok", True)
        note_p = d.get("note", "N/A")
        _data_row(ws, r, [
            poste,
            _f(d.get("seuil", 0)),
            _f(d.get("charge", 0)),
            "✅" if ok else "❌",
            note_p
        ], even=(i % 2 == 0))
        r += 1

    # ── Onglet 2 : Tarification par poste ─────────────────────────────────────
    ws2 = wb.create_sheet("Postes DREES 2023")
    _title_block(ws2, "Tarification par Poste", "Source : DREES Comptes de la Santé 2023", rag)
    _col_widths(ws2, [22, 12, 14, 14, 16, 16, 14])

    r2 = 5
    r2 = _section_title(ws2, r2, "🔢 SINISTRALITÉ PAR POSTE")
    _header_row(ws2, r2, [
        "Poste", "Fréq./an", "Coût moyen (€)", "Remb. SS (€)",
        "Charge mutuelle (€)", "Sinistre annuel (€)", "Source"
    ])
    r2 += 1

    postes = result_s1.get("postes", {})
    total_sin = 0.0
    for i, (poste, v) in enumerate(postes.items()):
        sin = v.get("sinistre_annuel", 0)
        total_sin += float(sin) if sin else 0
        _data_row(ws2, r2, [
            poste,
            v.get("frequence_an", "—"),
            _f(v.get("cout_moyen")),
            _f(v.get("remb_ss")),
            _f(v.get("charge_mutuelle")),
            _f(sin),
            v.get("source", "—"),
        ], even=(i % 2 == 0))
        r2 += 1

    # Ligne total
    _data_row(ws2, r2, ["TOTAL", "", "", "", "", _f(total_sin), ""], gold=True, bold=True)
    r2 += 2

    # LR par niveau FNMF
    r2 = _section_title(ws2, r2, "📈 LR MARCHÉ FNMF 2023 — RÉFÉRENCE PAR NIVEAU")
    _header_row(ws2, r2, ["Niveau garantie", "LR marché min", "LR marché max", "Référence"])
    r2 += 1
    fnmf = [
        ("eco",     "55%", "65%", "FNMF Rapport sinistralité 2023"),
        ("confort", "65%", "75%", "FNMF Rapport sinistralité 2023"),
        ("premium", "75%", "85%", "FNMF Rapport sinistralité 2023"),
        ("luxe",    "75%", "85%", "FNMF Rapport sinistralité 2023"),
    ]
    for i, row_f in enumerate(fnmf):
        _data_row(ws2, r2, list(row_f), even=(i % 2 == 0))
        r2 += 1

    # ── Onglet 3 : Hypothèses ─────────────────────────────────────────────────
    ws3 = wb.create_sheet("Hypothèses")
    _title_block(ws3, "Hypothèses Actuarielles S1", "Validation standard ActuarIA", rag)
    _col_widths(ws3, [8, 40, 50, 16, 12])

    r3 = 5
    r3 = _section_title(ws3, r3, "📋 HYPOTHÈSES H1-H5")
    _header_row(ws3, r3, ["ID", "Hypothèse", "Valeur", "Statut", "Critique"])
    r3 += 1

    hyp_colors = {"VALIDÉE": VERT, "À JUSTIFIER": AMBRE, "NON VALIDÉE": ROUGE}
    for i, h in enumerate(result_s1.get("hypotheses", [])):
        c_row = r3
        _data_row(ws3, r3, [
            h.get("id", ""),
            h.get("hypothese", ""),
            h.get("valeur", ""),
            h.get("statut", ""),
            "Oui" if h.get("critique") else "Non",
        ], even=(i % 2 == 0))
        # Colorier la cellule statut
        cell_statut = ws3.cell(row=c_row, column=4)
        statut_v = h.get("statut", "")
        cell_statut.font = _font(bold=True, size=9, color=hyp_colors.get(statut_v, GRIS))
        r3 += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
