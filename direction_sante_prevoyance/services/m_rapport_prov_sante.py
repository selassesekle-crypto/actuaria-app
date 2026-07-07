"""
Palette et helpers communs aux modules de rapport SP.
Reproduit fidèlement le style Navy/Gold de n5_rapport.py.
"""
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# Palette Navy/Gold ActuarIA
NAVY   = "0F2E52"
NAVY_L = "1B3A5C"
GOLD   = "C9A84C"
GOLD_L = "E2C97E"
BLANC  = "F0F4F8"
GRIS   = "8A9BB0"
VERT   = "1E8449"
ROUGE  = "C0392B"
AMBRE  = "E67E22"
BLEU   = "3498DB"

def _font(bold=False, size=10, color="1C2B3A", italic=False):
    return Font(name="Inter", bold=bold, size=size, color=color, italic=italic)

def _fill(hex6):
    return PatternFill("solid", start_color=hex6, end_color=hex6)

def _border_thin():
    s = Side(style="thin", color="DDE4EE")
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _header_row(ws, row, cols, fill_hex=NAVY):
    """Ligne d'en-tête Navy avec texte blanc gras."""
    for col, txt in enumerate(cols, 1):
        c = ws.cell(row=row, column=col, value=txt)
        c.font      = _font(bold=True, size=9, color="FFFFFF")
        c.fill      = _fill(fill_hex)
        c.alignment = _align("center")
        c.border    = _border_thin()

def _data_row(ws, row, values, even=True, bold=False, gold=False):
    """Ligne de données alternée."""
    fill_hex = "F5F7FA" if even else "FFFFFF"
    if gold:
        fill_hex = "F9F4E8"
    for col, v in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=v)
        c.font      = _font(bold=bold, size=9,
                            color=GOLD if gold else "1C2B3A")
        c.fill      = _fill(fill_hex)
        c.alignment = _align()
        c.border    = _border_thin()

def _title_block(ws, titre, sous_titre, rag="VERT"):
    """Bloc titre Navy en haut de la feuille."""
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value     = "ActuarIA — " + titre
    c.font      = _font(bold=True, size=14, color="FFFFFF")
    c.fill      = _fill(NAVY)
    c.alignment = _align("center")

    ws.merge_cells("A2:H2")
    c = ws["A2"]
    c.value     = sous_titre
    c.font      = _font(size=9, color=GOLD_L)
    c.fill      = _fill(NAVY_L)
    c.alignment = _align("center")

    # Statut RAG
    rag_colors = {"VERT": VERT, "AMBRE": AMBRE, "ROUGE": ROUGE}
    ws.merge_cells("A3:H3")
    c = ws["A3"]
    rag_label = {"VERT": "✅ CONFORME", "AMBRE": "⚠️ VIGILANCE", "ROUGE": "❌ ALERTE"}
    c.value     = "Statut RAG : " + rag_label.get(rag, rag)
    c.font      = _font(bold=True, size=9, color=rag_colors.get(rag, GRIS))
    c.fill      = _fill("FAFBFC")
    c.alignment = _align("center")

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 16

def _col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def _section_title(ws, row, titre, n_cols=8):
    ws.merge_cells(f"A{row}:{get_column_letter(n_cols)}{row}")
    c = ws.cell(row=row, column=1, value=titre)
    c.font      = _font(bold=True, size=10, color=GOLD)
    c.fill      = _fill(NAVY)
    c.alignment = _align("left")
    ws.row_dimensions[row].height = 20
    return row + 1

def _f(v, dec=0):
    if v is None: return "—"
    try:
        fv = float(v)
        if dec == 0:
            return f"{fv:,.0f} €".replace(",", " ")
        return f"{fv:,.{dec}f}".replace(",", " ")
    except: return "—"

def _pct(v, dec=1):
    if v is None: return "—"
    try: return f"{float(v):.{dec}f} %"
    except: return "—"


from openpyxl import Workbook
import io

def generer(result_s2: dict) -> bytes:
    """
    Génère le rapport Excel de provisionnement santé depuis S2 Selma.
    3 onglets : Synthèse · Provisions par poste · Hypothèses
    """
    wb = Workbook()
    rag = result_s2.get("statut_rag", "AMBRE")

    # ── Onglet 1 : Synthèse ───────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Synthèse Provisionnement"
    _title_block(ws, "Rapport Provisionnement Santé", "Agent S2 Selma — PSAP · IBNR · Cadence", rag)
    _col_widths(ws, [28, 20, 20, 20, 20])

    r = 5
    r = _section_title(ws, r, "📊 PROVISIONS TECHNIQUES")
    _header_row(ws, r, ["Composante", "Montant (€)", "% des Primes", "Référence réglementaire"])
    r += 1

    pa  = float(result_s2.get("primes_acquises",    0) or 0)
    psa = float(result_s2.get("psap_dossiers",       0) or 0)
    ibnr= float(result_s2.get("psap_ibnr",           0) or 0)
    prec= float(result_s2.get("prec",                0) or 0)
    be  = float(result_s2.get("be_sante",            0) or 0)
    ra  = float(result_s2.get("risk_adjustment",     0) or 0)
    tp  = float(result_s2.get("tp_sante",            0) or 0)
    lr  = float(result_s2.get("loss_ratio",          0) or 0)

    prov_rows = [
        ("PSAP Dossiers (sinistres connus)",    _f(psa),  _pct(psa/pa*100 if pa else 0), "Sinistres déclarés"),
        ("IBNR santé (5-25% — liquidation <3m)",_f(ibnr), _pct(ibnr/pa*100 if pa else 0),"Art. 77 S2 — cadence rapide"),
        ("PREC (provision risques en cours)",    _f(prec), _pct(prec/pa*100 if pa else 0),"Art. 78 S2"),
        ("Best Estimate Santé",                  _f(be),   _pct(be/pa*100 if pa else 0),  "Art. 77 §1 S2"),
        ("Risk Adjustment (CoC 8%)",             _f(ra),   _pct(ra/pa*100 if pa else 0),  "IFRS 17 §B119"),
        ("TP Santé (BE + RA)",                   _f(tp),   _pct(tp/pa*100 if pa else 0),  "Art. 77 §1 S2"),
    ]
    for i, row_p in enumerate(prov_rows):
        bold = (i == len(prov_rows) - 1)
        gold = (i == len(prov_rows) - 1)
        _data_row(ws, r, list(row_p), even=(i % 2 == 0), bold=bold, gold=gold)
        r += 1

    r += 1
    r = _section_title(ws, r, "📈 LOSS RATIO ET SINISTRALITÉ")
    _header_row(ws, r, ["Indicateur", "Valeur", "Norme", "Statut"])
    r += 1
    lr_statut = "✅ CONFORME" if lr <= 0.85 else ("⚠️ VIGILANCE" if lr <= 0.95 else "❌ ALERTE")
    _data_row(ws, r, ["Loss Ratio observé", _pct(lr * 100), "≤ 85% (mutuelles santé)", lr_statut])
    r += 1
    _data_row(ws, r, ["Primes acquises", _f(pa), "", "Base de calcul"], even=False)
    r += 1

    # ── Onglet 2 : Provisions par poste ───────────────────────────────────────
    ws2 = wb.create_sheet("Provisions par Poste")
    _title_block(ws2, "PSAP & IBNR par Poste", "Cadence santé : 97% liquidé en 3 mois", rag)
    _col_widths(ws2, [22, 16, 16, 14, 14])

    r2 = 5
    r2 = _section_title(ws2, r2, "🔢 PSAP PAR POSTE")
    _header_row(ws2, r2, ["Poste", "PSAP (€)", "IBNR (€)", "Taux IBNR", "Total (€)"])
    r2 += 1

    psap_p = result_s2.get("psap_postes", {})
    ibnr_p = result_s2.get("ibnr_postes", {})
    total_psap = 0.0
    total_ibnr = 0.0
    for i, poste in enumerate(["medecine", "pharmacie", "hospitalisation", "dentaire", "optique"]):
        ps = float(psap_p.get(poste, 0) or 0)
        ib = float(ibnr_p.get(poste, 0) or 0)
        total_psap += ps
        total_ibnr += ib
        taux_ibnr = ib / ps * 100 if ps > 0 else 0
        _data_row(ws2, r2, [poste, _f(ps), _f(ib), _pct(taux_ibnr), _f(ps + ib)],
                  even=(i % 2 == 0))
        r2 += 1

    _data_row(ws2, r2, ["TOTAL", _f(total_psap), _f(total_ibnr),
                         _pct(total_ibnr/total_psap*100 if total_psap else 0),
                         _f(total_psap + total_ibnr)], gold=True, bold=True)
    r2 += 2

    # Cadence de règlement
    r2 = _section_title(ws2, r2, "⏱️ CADENCE DE RÈGLEMENT SANTÉ")
    _header_row(ws2, r2, ["Horizon", "% Réglé (santé)", "% Réglé (IARD)", "% Réglé (Prévoyance)"])
    r2 += 1
    cadence = [
        ("J0  (déclaration)", "0%",   "0%",   "0%"),
        ("M1  (1 mois)",      "60%",  "10%",  "2%"),
        ("M2  (2 mois)",      "85%",  "25%",  "5%"),
        ("M3  (3 mois)",      "97%",  "40%",  "8%"),
        ("M6  (6 mois)",      "99%",  "60%",  "20%"),
        ("M12 (12 mois)",     "≈100%","80%",  "40%"),
        ("M36 (36 mois)",     "100%", "≈100%","80%"),
        ("M60 (60 mois)",     "100%", "100%", "≈100%"),
    ]
    for i, row_c in enumerate(cadence):
        _data_row(ws2, r2, list(row_c), even=(i % 2 == 0))
        r2 += 1

    # ── Onglet 3 : Hypothèses ─────────────────────────────────────────────────
    ws3 = wb.create_sheet("Hypothèses")
    _title_block(ws3, "Hypothèses Actuarielles S2", "Validation standard ActuarIA", rag)
    _col_widths(ws3, [8, 40, 50, 16, 12])

    r3 = 5
    r3 = _section_title(ws3, r3, "📋 HYPOTHÈSES H1-H4")
    _header_row(ws3, r3, ["ID", "Hypothèse", "Valeur", "Statut", "Critique"])
    r3 += 1

    hyp_colors = {"VALIDÉE": VERT, "À JUSTIFIER": AMBRE, "NON VALIDÉE": ROUGE}
    for i, h in enumerate(result_s2.get("hypotheses", [])):
        c_row = r3
        _data_row(ws3, r3, [
            h.get("id", ""), h.get("hypothese", ""),
            h.get("valeur", ""), h.get("statut", ""),
            "Oui" if h.get("critique") else "Non",
        ], even=(i % 2 == 0))
        ws3.cell(row=c_row, column=4).font = _font(
            bold=True, size=9, color=hyp_colors.get(h.get("statut", ""), GRIS))
        r3 += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
