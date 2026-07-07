"""
Palette et helpers communs aux modules de rapport SP.
Reproduit fidèlement le style Navy/Gold de n5_rapport.py.
"""
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# Palette Navy/Gold ActuarIA
NAVY   = "0F2E52"
NAVY_L = "1B3A5C"
GOLD   = "C9A84C"
GOLD_L = "E2C97E"
BLANC  = "F0F4F8"
GRIS   = "8A9BB0"
VERT   = "1E8449"
ROUGE  = "C0392B"
AMBRE  = "E67E22"
BLEU   = "3498DB"

def _font(bold=False, size=10, color="1C2B3A", italic=False):
    return Font(name="Inter", bold=bold, size=size, color=color, italic=italic)

def _fill(hex6):
    return PatternFill("solid", start_color=hex6, end_color=hex6)

def _border_thin():
    s = Side(style="thin", color="DDE4EE")
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _header_row(ws, row, cols, fill_hex=NAVY):
    """Ligne d'en-tête Navy avec texte blanc gras."""
    for col, txt in enumerate(cols, 1):
        c = ws.cell(row=row, column=col, value=txt)
        c.font      = _font(bold=True, size=9, color="FFFFFF")
        c.fill      = _fill(fill_hex)
        c.alignment = _align("center")
        c.border    = _border_thin()

def _data_row(ws, row, values, even=True, bold=False, gold=False):
    """Ligne de données alternée."""
    fill_hex = "F5F7FA" if even else "FFFFFF"
    if gold:
        fill_hex = "F9F4E8"
    for col, v in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=v)
        c.font      = _font(bold=bold, size=9,
                            color=GOLD if gold else "1C2B3A")
        c.fill      = _fill(fill_hex)
        c.alignment = _align()
        c.border    = _border_thin()

def _title_block(ws, titre, sous_titre, rag="VERT"):
    """Bloc titre Navy en haut de la feuille."""
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value     = "ActuarIA — " + titre
    c.font      = _font(bold=True, size=14, color="FFFFFF")
    c.fill      = _fill(NAVY)
    c.alignment = _align("center")

    ws.merge_cells("A2:H2")
    c = ws["A2"]
    c.value     = sous_titre
    c.font      = _font(size=9, color=GOLD_L)
    c.fill      = _fill(NAVY_L)
    c.alignment = _align("center")

    # Statut RAG
    rag_colors = {"VERT": VERT, "AMBRE": AMBRE, "ROUGE": ROUGE}
    ws.merge_cells("A3:H3")
    c = ws["A3"]
    rag_label = {"VERT": "✅ CONFORME", "AMBRE": "⚠️ VIGILANCE", "ROUGE": "❌ ALERTE"}
    c.value     = "Statut RAG : " + rag_label.get(rag, rag)
    c.font      = _font(bold=True, size=9, color=rag_colors.get(rag, GRIS))
    c.fill      = _fill("FAFBFC")
    c.alignment = _align("center")

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 16

def _col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def _section_title(ws, row, titre, n_cols=8):
    ws.merge_cells(f"A{row}:{get_column_letter(n_cols)}{row}")
    c = ws.cell(row=row, column=1, value=titre)
    c.font      = _font(bold=True, size=10, color=GOLD)
    c.fill      = _fill(NAVY)
    c.alignment = _align("left")
    ws.row_dimensions[row].height = 20
    return row + 1

def _f(v, dec=0):
    if v is None: return "—"
    try:
        fv = float(v)
        if dec == 0:
            return f"{fv:,.0f} €".replace(",", " ")
        return f"{fv:,.{dec}f}".replace(",", " ")
    except: return "—"

def _pct(v, dec=1):
    if v is None: return "—"
    try: return f"{float(v):.{dec}f} %"
    except: return "—"


from openpyxl import Workbook
import io

def generer(result_s3: dict, result_reg2: dict = None, result_reg3: dict = None) -> bytes:
    """
    Rapport réglementation santé : S3 (SCR/MCR NSLT) + REG2 (IFRS17) + REG3 (ANI/100%Santé).
    3 onglets : SCR/MCR · IFRS 17 · ANI & 100% Santé
    """
    wb = Workbook()
    rag = result_s3.get("statut_rag", "AMBRE")
    reg2 = result_reg2 or {}
    reg3 = result_reg3 or {}

    # ── Onglet 1 : SCR / MCR NSLT ────────────────────────────────────────────
    ws = wb.active
    ws.title = "SCR MCR NSLT"
    _title_block(ws, "SCR/MCR Santé NSLT", "Agent S3 — Art. 148-162 RD 2015/35 · QRT S.13.01", rag)
    _col_widths(ws, [32, 18, 18, 30])

    r = 5
    r = _section_title(ws, r, "🔢 QRT S.13.01 — SANTÉ NON-SLT")
    _header_row(ws, r, ["Composante", "Montant (€)", "% BE", "Référence réglementaire"])
    r += 1

    be   = float(result_s3.get("be_sante",       0) or 0)
    scr  = float(result_s3.get("scr_sante",      0) or 0)
    mcr  = float(result_s3.get("mcr",            0) or 0)
    fpp  = float(result_s3.get("fonds_propres",  0) or 0)
    r_scr= float(result_s3.get("ratio_scr_pct",  0) or 0)
    r_mcr= float(result_s3.get("ratio_mcr_pct",  0) or 0)

    rows_scr = [
        ("Best Estimate Santé",        _f(be),   "100%",              "Art. 77 §1 S2"),
        ("SCR Santé NSLT",             _f(scr),  _pct(scr/be*100 if be else 0), "Art. 148-162 S2"),
        ("MCR Santé",                  _f(mcr),  "",                  "Art. 252 RD 2015/35"),
        ("Fonds Propres",              _f(fpp),  "",                  "Art. 87 S2"),
        ("Ratio SCR (%)",              _pct(r_scr), "",               "≥ 100% réglementaire"),
        ("Ratio MCR (%)",              _pct(r_mcr), "",               "≥ 100% réglementaire"),
    ]
    for i, row_s in enumerate(rows_scr):
        gold = i >= 4
        _data_row(ws, r, list(row_s), even=(i % 2 == 0), gold=gold)
        r += 1

    # ── Onglet 2 : IFRS 17 ────────────────────────────────────────────────────
    ws2 = wb.create_sheet("IFRS 17")
    _title_block(ws2, "IFRS 17 — Santé", "Risk Adjustment CoC 6% · PAA/GMM", rag)
    _col_widths(ws2, [32, 18, 18, 30])

    r2 = 5
    r2 = _section_title(ws2, r2, "📐 COMPOSANTES IFRS 17")
    _header_row(ws2, r2, ["Composante", "Montant (€)", "Méthode", "Référence"])
    r2 += 1

    be17  = float(reg2.get("be_sante",        be)  or be)
    ra17  = float(reg2.get("risk_adjustment",  0)   or 0)
    csm17 = float(reg2.get("csm",             0)   or 0)
    tp17  = be17 + ra17

    rows_ifrs = [
        ("Best Estimate",           _f(be17),  "Flux futurs actualisés",     "IFRS 17 §33"),
        ("Risk Adjustment (8% BE)", _f(ra17),  "CoC 6% — floor 3%",          "IFRS 17 §B119"),
        ("TP Santé (BE + RA)",      _f(tp17),  "",                            "IFRS 17 §32"),
        ("CSM (si GMM)",            _f(csm17), "Profit reporté",              "IFRS 17 §38"),
    ]
    for i, row_i in enumerate(rows_ifrs):
        _data_row(ws2, r2, list(row_i), even=(i % 2 == 0))
        r2 += 1

    # ── Onglet 3 : ANI & 100% Santé ───────────────────────────────────────────
    ws3 = wb.create_sheet("ANI 100pct Santé")
    _title_block(ws3, "ANI 2013 & 100% Santé", "Art. L911-7 CSS · Réforme 2019", rag)
    _col_widths(ws3, [24, 16, 16, 14, 24])

    r3 = 5
    r3 = _section_title(ws3, r3, "✅ PANIER ANI 2013 — COLLECTIF OBLIGATOIRE")
    _header_row(ws3, r3, ["Poste", "Seuil ANI (€)", "Couverture (€)", "Conforme", "Note"])
    r3 += 1

    ani_ref = {"medecine": 30, "hospitalisation": 100, "dentaire": 75, "optique": 100}
    ani_detail = reg3.get("ani_detail", {}).get("detail", {})
    for i, (poste, seuil) in enumerate(ani_ref.items()):
        d = ani_detail.get(poste, {})
        charge = d.get("charge", 0)
        ok = charge >= seuil if charge else False
        _data_row(ws3, r3, [
            poste, _f(seuil), _f(charge),
            "✅" if ok else "❌",
            d.get("note", "—")
        ], even=(i % 2 == 0))
        r3 += 1

    r3 += 1
    r3 = _section_title(ws3, r3, "🏥 100% SANTÉ — PANIERS RÉFORME 2019")
    _header_row(ws3, r3, ["Catégorie", "Panier A (RAZ)", "Panier B (encadré)", "Panier C (libre)"])
    r3 += 1
    paniers = [
        ("Optique",     "Montures+verres RAZ",  "Plafonds UNOCAM",   "Libre"),
        ("Dentaire",    "Soins RAZ + prothèses","Plafonds UNOCAM",   "Libre"),
        ("Audiologie",  "Aides auditives RAZ",  "Plafonds UNOCAM",   "Libre"),
    ]
    for i, row_p in enumerate(paniers):
        _data_row(ws3, r3, list(row_p), even=(i % 2 == 0))
        r3 += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
