"""
Palette et helpers communs aux modules de rapport SP.
Reproduit fidèlement le style Navy/Gold de n5_rapport.py.
"""
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# Palette Navy/Gold ActuarIA
NAVY   = "0F2E52"
NAVY_L = "1B3A5C"
GOLD   = "C9A84C"
GOLD_L = "E2C97E"
GRIS   = "8A9BB0"
VERT   = "1E8449"
ROUGE  = "C0392B"
AMBRE  = "E67E22"

def _font(bold=False, size=10, color="1C2B3A", italic=False):
    return Font(name="Inter", bold=bold, size=size, color=color, italic=italic)

def _fill(hex6):
    return PatternFill("solid", start_color=hex6, end_color=hex6)

def _border_thin():
    s = Side(style="thin", color="DDE4EE")
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _header_row(ws, row, cols, fill_hex=NAVY):
    """Ligne d'en-tête Navy avec texte blanc gras."""
    for col, txt in enumerate(cols, 1):
        c = ws.cell(row=row, column=col, value=txt)
        c.font      = _font(bold=True, size=9, color="FFFFFF")
        c.fill      = _fill(fill_hex)
        c.alignment = _align("center")
        c.border    = _border_thin()

def _data_row(ws, row, values, even=True, bold=False, gold=False):
    """Ligne de données alternée."""
    fill_hex = "F5F7FA" if even else "FFFFFF"
    if gold:
        fill_hex = "F9F4E8"
    for col, v in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=v)
        c.font      = _font(bold=bold, size=9,
                            color=GOLD if gold else "1C2B3A")
        c.fill      = _fill(fill_hex)
        c.alignment = _align()
        c.border    = _border_thin()

def _title_block(ws, titre, sous_titre, rag="VERT"):
    """Bloc titre Navy en haut de la feuille."""
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value     = "ActuarIA — " + titre
    c.font      = _font(bold=True, size=14, color="FFFFFF")
    c.fill      = _fill(NAVY)
    c.alignment = _align("center")

    ws.merge_cells("A2:H2")
    c = ws["A2"]
    c.value     = sous_titre
    c.font      = _font(size=9, color=GOLD_L)
    c.fill      = _fill(NAVY_L)
    c.alignment = _align("center")

    # Statut RAG
    rag_colors = {"VERT": VERT, "AMBRE": AMBRE, "ROUGE": ROUGE}
    ws.merge_cells("A3:H3")
    c = ws["A3"]
    rag_label = {"VERT": "✅ CONFORME", "AMBRE": "⚠️ VIGILANCE", "ROUGE": "❌ ALERTE"}
    c.value     = "Statut RAG : " + rag_label.get(rag, rag)
    c.font      = _font(bold=True, size=9, color=rag_colors.get(rag, GRIS))
    c.fill      = _fill("FAFBFC")
    c.alignment = _align("center")

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 16

def _col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def _section_title(ws, row, titre, n_cols=8):
    ws.merge_cells(f"A{row}:{get_column_letter(n_cols)}{row}")
    c = ws.cell(row=row, column=1, value=titre)
    c.font      = _font(bold=True, size=10, color=GOLD)
    c.fill      = _fill(NAVY)
    c.alignment = _align("left")
    ws.row_dimensions[row].height = 20
    return row + 1

def _f(v, dec=0):
    if v is None: return "—"
    try:
        fv = float(v)
        if dec == 0:
            return f"{fv:,.0f} €".replace(",", " ")
        return f"{fv:,.{dec}f}".replace(",", " ")
    except: return "—"

def _pct(v, dec=1):
    if v is None: return "—"
    try: return f"{float(v):.{dec}f} %"
    except: return "—"


from openpyxl import Workbook
import io

def generer(result_p1: dict, result_p2: dict = None) -> bytes:
    """
    Rapport tarification prévoyance : P1 (primes ITT/IP) + P2 (tables BCAC/Markov).
    3 onglets : Tarification · Tables Morbidité · Markov
    """
    wb = Workbook()
    rag = result_p1.get("statut_rag", "AMBRE")
    p2  = result_p2 or {}

    # ── Onglet 1 : Tarification P1 ────────────────────────────────────────────
    ws = wb.active
    ws.title = "Tarification Prévoyance"
    _title_block(ws, "Tarification Prévoyance", "Agent P1 Axel — BCAC 2019 · CTIP", rag)
    _col_widths(ws, [30, 18, 18, 24])

    r = 5
    r = _section_title(ws, r, "📊 PRIMES ET TAUX")
    _header_row(ws, r, ["Indicateur", "Valeur", "Unité", "Référence"])
    r += 1

    kpis = [
        ("Prime commerciale",          _f(result_p1.get("prime_commerciale")),     "€/an", "Chargement inclus"),
        ("Taux de cotisation",         _pct(result_p1.get("taux_cotisation_pct")), "%",    "% salaire brut"),
        ("Salaire brut",               _f(result_p1.get("salaire_brut")),          "€/an", "Base de tarification"),
        ("Taux rente IPP",             _pct(result_p1.get("taux_rente_ipp", 0)*100), "%",  "% salaire"),
        ("Franchise ITT",              str(result_p1.get("franchise_jours", 90)) + " jours", "", "Délai de carence"),
        ("Durée contrat",              str(result_p1.get("duree_contrat", 20)) + " ans", "", ""),
        ("Taux ITT (BCAC)",            _pct(result_p1.get("taux_itt", 0)*100),     "%/an", "BCAC 2019"),
        ("Taux IP annuel",             _pct(result_p1.get("taux_ip", 0)*100),      "%/an", "BCAC 2019"),
        ("Âge assuré",                 result_p1.get("age", "—"),                  "ans",  ""),
        ("CSP",                        result_p1.get("categorie", "—"),             "",     "ouvrier/employé/cadre"),
    ]
    for i, (k, v, u, ref) in enumerate(kpis):
        _data_row(ws, r, [k, v, u, ref], even=(i % 2 == 0))
        r += 1

    # ── Onglet 2 : Tables Morbidité BCAC 2019 ─────────────────────────────────
    ws2 = wb.create_sheet("Tables Morbidité BCAC")
    _title_block(ws2, "Tables BCAC 2019", "Probabilités de transition — Chaîne de Markov", rag)
    _col_widths(ws2, [30, 16, 16, 24])

    r2 = 5
    r2 = _section_title(ws2, r2, "🔢 PROBABILITÉS ANNUELLES DE TRANSITION")
    _header_row(ws2, r2, ["Transition", "Probabilité", "Facteur CSP", "Source"])
    r2 += 1

    trans = p2.get("transitions", {})
    trans_rows = [
        ("q_AI : Actif → ITT (incidence)",    _pct(trans.get("q_AI", 0)*100, 3),   _pct(trans.get("fact_csp", 1)*100, 0), "BCAC 2019"),
        ("q_IA : ITT → Actif (guérison)",     _pct(trans.get("q_IA", 0)*100),      "—",                                    "BCAC 2019"),
        ("q_IP|ITT : ITT → IP (consolidation)",_pct(trans.get("q_IP_cond", 0)*100),"—",                                    "BCAC 2019"),
        ("q_IP annuel : Actif → IP",           _pct(trans.get("q_IP_annuel", 0)*100, 3), "—",                               "q_AI × q_IP|ITT"),
        ("q_PD : IP → Décès",                  _pct(trans.get("q_PD", 0)*100),      "—",                                    "BCAC 2019 / TH0002"),
        ("q_AD : Actif → Décès",               _pct(trans.get("q_AD", 0)*100, 3),   "—",                                    "TH0002 mortalité"),
    ]
    for i, row_t in enumerate(trans_rows):
        _data_row(ws2, r2, list(row_t), even=(i % 2 == 0))
        r2 += 1

    r2 += 1
    r2 = _section_title(ws2, r2, "⏱️ ESPÉRANCES DE DURÉE")
    _header_row(ws2, r2, ["Indicateur", "Valeur", "Unité", "Source"])
    r2 += 1

    esp = p2.get("esperances", {})
    esp_rows = [
        ("Durée moyenne ITT (>franchise)", esp.get("duree_moy_itt_mois", "—"), "mois", "BCAC 2019"),
        ("Espérance durée IP",             esp.get("esperance_duree_ip_ans", "—"), "ans", "Jusqu'à 65 ans"),
        ("P(guérison totale)",             _pct(esp.get("prob_guerison_tt", 0)*100), "%", "BCAC 2019"),
        ("P(passage IP vie active)",       _pct(esp.get("prob_passage_ip_vie", 0)*100), "%", "q_IP × (65-âge)"),
    ]
    for i, row_e in enumerate(esp_rows):
        _data_row(ws2, r2, list(row_e), even=(i % 2 == 0))
        r2 += 1

    r2 += 1
    r2 = _section_title(ws2, r2, "📈 PROBABILITÉS DE MAINTIEN EN ITT")
    _header_row(ws2, r2, ["Horizon", "Probabilité maintien", "Norme BCAC", "Statut"])
    r2 += 1

    maint = p2.get("prob_maintien", {})
    maint_data = [
        ("6 mois",  maint.get("mois_6",  0), "5%-50%"),
        ("12 mois", maint.get("mois_12", 0), "2%-25%"),
        ("24 mois", maint.get("mois_24", 0), "1%-15%"),
    ]
    for i, (h, v, norme) in enumerate(maint_data):
        pct_v = float(v) * 100 if v else 0
        ok = 0.05 <= float(v) <= 0.50 if v and h == "6 mois" else True
        _data_row(ws2, r2, [h, _pct(pct_v), norme, "✅" if ok else "⚠️"],
                  even=(i % 2 == 0))
        r2 += 1

    # ── Onglet 3 : Matrice Markov ──────────────────────────────────────────────
    ws3 = wb.create_sheet("Matrice Markov")
    _title_block(ws3, "Matrice de Transition Markov", "4 états : Actif / ITT / IP / Décès", rag)
    _col_widths(ws3, [18, 14, 14, 14, 14])

    r3 = 5
    r3 = _section_title(ws3, r3, "🔢 MATRICE P (ANNUELLE) — PROBABILITÉS %")
    etats = ["Actif", "ITT", "IP", "Décès"]
    _header_row(ws3, r3, ["État \ Vers →"] + etats)
    r3 += 1

    mat = p2.get("matrice_P", [[0]*4]*4)
    for i, etat in enumerate(etats):
        row_m = [etat] + [_pct(mat[i][j]*100 if mat and len(mat) > i else 0)
                          for j in range(4)]
        _data_row(ws3, r3, row_m, even=(i % 2 == 0))
        r3 += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
