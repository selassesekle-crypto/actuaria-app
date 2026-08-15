# -*- coding: utf-8 -*-
"""
=============================================================================
  ActuarIA — IFRS 17 : LE RENDU, QUI NE DEMANDE RIEN ET N'OMET RIEN
=============================================================================

⚠️⚠️ CE MODULE A ÉTÉ MIS EN DERNIER DÉLIBÉRÉMENT, ET LA RAISON DOIT SE LIRE
AVANT SON CODE. Un HTML titré « États financiers IFRS 17 » portant le §78 et
le §80 sans les §99-109 serait PLAUSIBLE, IMPRIMABLE, JOIGNABLE — ET
INCOMPLET. Aujourd'hui l'absence est visible ; un document la rendrait
invisible.

⚠️ D'OÙ LE VERDICT QUI GOUVERNE TOUT CE QUI SUIT, ET QUI DESCEND DANS LE
DOCUMENT LUI-MÊME :

    UN BEAU DOCUMENT NE COMBLE AUCUN DE CES TROUS — IL LES AFFICHE.

⚠️ LE TITRE N'EST PAS « ÉTATS FINANCIERS », ET CE N'EST PAS UNE PRUDENCE DE
FORME. Un jeu d'états au sens d'IAS 1 comprend un état des flux de
trésorerie, un état de variation des capitaux propres et un comparatif : la
plateforme n'en produit AUCUN, et §93 nomme lui-même le premier comme
extérieur. S'y ajoutent cinq pièces sur huit légitimement absentes, aucun
tableau du §99 a), et le cédé séparé nulle part (§78 c) et d), §82).
⚠️ ET LE PRÉCÉDENT EST DANS CE DÉPÔT : le préambule du périmètre publié a
affirmé « la plateforme couvre l'évaluation, la présentation et la clôture »
pendant que quatre pans étaient déclarés non bâtis. Le titre est la seule
ligne dont on soit certain qu'elle sera lue — c'est là qu'une
sur-affirmation coûte le plus cher.

⚠️⚠️ TROIS PROPRIÉTÉS SONT VÉRIFIÉES SUR LE DOCUMENT PRODUIT, PAS PROMISES :

  1. AUCUNE PIÈCE N'EST OMISE. Le document est REFUSÉ si une pièce de
     l'état n'y figure pas. Un rendu capable de sauter une section en
     silence rouvrirait, un cran plus bas, le trou que l'assemblage vient de
     fermer — là où personne ne regarde.
  2. AUCUNE REQUÊTE EXTERNE. ⚠️ La chaîne de rapport A7 n'est PAS AUTONOME :
     mesuré, elle importe ses polices de `fonts.googleapis.com` et un SCRIPT
     de `cdn.jsdelivr.net`. Un document d'archive qui va chercher une
     feuille de style et un script chez deux tiers AU MOMENT OÙ ON L'OUVRE
     dépend d'eux à la lecture, émet deux requêtes quand un commissaire
     l'ouvre, et ne rend pas la même chose hors ligne ou dans cinq ans.
     Ce n'est pas une question d'esthétique, c'est une question
     d'OPPOSABILITÉ. ⚠️ C'est aussi pourquoi la reprise d'A7 vaut ~25 % et
     non ~55 % comme estimé d'abord : l'essentiel était à réécrire.
  3. AUCUN CHAMP LIBRE. Ce module n'accepte QUE l'état daté. ⚠️ Le verrou
     qui interdit à `normes/` d'atteindre le modèle de langage regarde les
     IMPORTS, pas les chaînes : un document qui imprimerait un
     « commentaire » fourni par l'appelant laisserait entrer du texte généré
     SANS QU'UNE SEULE LIGNE DE CODE CHANGE. Le rendu n'imprime donc que ce
     que l'état porte — `motif`, `raison`, `reserve` — tous produits par du
     code et testés.

⚠️ ET CE MODULE VIT SOUS `normes/`, CE QUI N'EST PAS INDIFFÉRENT : c'est là
que porte le contrôle AST qui interdit d'atteindre `appeler` ou `anthropic`.
Ailleurs, il pourrait appeler la frontière sans rien violer.

RÉFÉRENCES — IFRS 17, annexe au règlement (UE) 2023/1803, JO L 237 du
26.9.2023, §78, §80, §93, §96, §99 a).
=============================================================================
"""

import html as _html
import re
from pathlib import Path

from normes.ifrs17.etats.assemblage import (
    NON_ETABLIE,
    PIECES,
    EtatDate,
)
from normes.ifrs17.mesure.lrc_paa import RefusMesure

#: ⚠️ TROIS CHOSES VRAIES, ET PAS UNE DE PLUS : c'est une MESURE, elle est
#: DATÉE, et c'est un INTRANT — pas un livrable final.
TITRE = 'ÉTAT DE LA MESURE IFRS 17'
SOUS_TITRE = ("pièces destinées à être intégrées aux états financiers "
              "de l'entité")

#: ⚠️ LA PHRASE QUI GOUVERNE CE MODULE, ET ELLE DESCEND DANS LE DOCUMENT.
VERDICT_DU_DOCUMENT = (
    "UN BEAU DOCUMENT NE COMBLE AUCUN DE CES TROUS — IL LES AFFICHE. Ce "
    "document rend visible ce qui manque ; il ne le remplace pas. Cinq "
    "pièces sur huit peuvent en être légitimement absentes, et une absence "
    "y occupe la même place qu'une présence.")

#: ⚠️ CE QUE CE DOCUMENT N'EST PAS — EN TÊTE, JAMAIS EN FIN. Le périmètre
#: publié a payé exactement l'inverse : sa sur-affirmation vivait dans le
#: préambule, là où l'on lit.
CE_QUE_CE_DOCUMENT_N_EST_PAS = (
    "Ce document N'EST PAS un jeu d'états financiers. Un jeu au sens d'IAS 1 "
    "comprend un état des flux de trésorerie, un état de variation des "
    "capitaux propres et un comparatif : aucun n'est produit ici, et le §93 "
    "d'IFRS 17 nomme lui-même le premier comme extérieur à la norme. Les "
    "rapprochements ne sont pas présentés sous forme de tableau, ce que le "
    "§99 a) exige. La réassurance détenue n'est séparée ni au bilan (§78 c) "
    "et d)) ni au compte de résultat (§82). Les pièces ci-dessous sont "
    "destinées à être INTÉGRÉES par l'entité dans ses propres états.")

MOTIF_PIECE_OMISE = 'piece_absente_du_document_rendu'
MOTIF_REQUETE_EXTERNE = 'document_non_autonome'

#: ⚠️ CE QU'UNE REQUÊTE EXTERNE RECOUVRE. Le balayage est STRICT parce que ce
#: module n'émet ni image, ni graphique, ni SVG : aucun `http` légitime ne
#: peut donc apparaître, pas même un espace de noms XML. ⚠️ Le jour où un
#: graphique entrera, cette strictesse devra être RENÉGOCIÉE explicitement —
#: et non contournée par une exception discrète.
FORMES_EXTERNES = ('http://', 'https://', '@import', '<link', '<script',
                   'src=', 'url(')


def _exiger_aucune_piece_omise(document: str, etat: EtatDate) -> None:
    """⚠️⚠️ LE CŒUR DU LOT : LE DOCUMENT EST REFUSÉ S'IL OMET UNE PIÈCE.

    L'assemblage a fermé l'oubli au niveau de l'état ; un rendu capable de
    sauter une section le rouvrirait un cran plus bas, là où personne ne
    regarde. La vérification porte sur le DOCUMENT PRODUIT, pas sur
    l'intention du gabarit.
    """
    omises = sorted(p.nom for p in etat.pieces if p.nom not in document)
    if omises:
        raise RefusMesure(
            MOTIF_PIECE_OMISE,
            f"{len(omises)} pièce(s) de l'état n'apparaissent pas dans le "
            f"document rendu : {omises}. ⚠️ UNE PIÈCE ABSENTE DU DOCUMENT "
            f"EST INDISCERNABLE D'UNE PIÈCE QUI N'EXISTE PAS. L'assemblage "
            f"interdit d'oublier une pièce ; le rendu doit l'interdire "
            f"aussi, sans quoi le contrôle s'arrête juste avant l'endroit où "
            f"le lecteur regarde.")


def _exiger_autonomie(document: str) -> None:
    """⚠️ AUCUNE REQUÊTE EXTERNE, ET C'EST VÉRIFIÉ SUR LE DOCUMENT PRODUIT.

    Un document d'archive doit rendre la même chose dans cinq ans, hors
    ligne, sur un poste d'audit isolé — et ne rien émettre quand on l'ouvre.
    """
    trouvees = sorted({f for f in FORMES_EXTERNES if f in document})
    if trouvees:
        raise RefusMesure(
            MOTIF_REQUETE_EXTERNE,
            f"le document rendu porte {len(trouvees)} forme(s) de dépendance "
            f"externe : {trouvees}. ⚠️ UN DOCUMENT QUI VA CHERCHER QUELQUE "
            f"CHOSE AU MOMENT OÙ ON L'OUVRE DÉPEND D'UN TIERS À LA LECTURE : "
            f"il émet une requête quand un commissaire aux comptes l'ouvre, "
            f"et ne rend pas la même chose hors ligne ou dans cinq ans. "
            f"C'est une question d'opposabilité, pas d'esthétique — et c'est "
            f"exactement ce qui disqualifie la chaîne de rapport existante, "
            f"qui importe ses polices et un script chez deux tiers.")


def _style() -> str:
    """⚠️ EMBARQUÉ, SANS UNE SEULE POLICE DISTANTE. Les familles sont
    génériques : elles existent partout et n'appellent personne."""
    return (
        '<style>'
        'body{font-family:Georgia,"Times New Roman",serif;color:#1C2B3A;'
        'background:#FFFFFF;margin:2.5em auto;max-width:52em;line-height:1.5}'
        'h1{font-size:1.5em;margin-bottom:.1em}'
        'h2{font-size:1em;text-transform:uppercase;letter-spacing:.08em;'
        'border-bottom:1px solid #1C2B3A;padding-bottom:.2em;margin-top:2em}'
        '.sous-titre{font-style:italic;color:#55606B;margin-top:0}'
        '.reserve{border:2px solid #1C2B3A;padding:1em;margin:1.5em 0}'
        '.reserve h2{margin-top:0;border:none}'
        '.absente{border-left:4px solid #1C2B3A;padding-left:.8em;'
        'background:#F2F2F2}'
        '.etiquette{font-family:monospace;font-weight:bold}'
        'table{border-collapse:collapse;width:100%}'
        'td,th{border:1px solid #99A2AB;padding:.4em .6em;'
        'text-align:left;vertical-align:top}'
        '@media print{body{margin:0;max-width:none}'
        '*{-webkit-print-color-adjust:exact;print-color-adjust:exact}}'
        '</style>')


def _e(texte) -> str:
    """Échappe pour un NŒUD DE TEXTE — sans toucher aux apostrophes.

    ⚠️⚠️ `html.escape` ÉCHAPPE `'` PAR DÉFAUT, ET C'EST UN COÛT SANS
    CONTREPARTIE ICI. Ce module n'écrit dans aucun attribut : `<` `>` et `&`
    suffisent à la bonne formation. Échapper l'apostrophe rendrait
    « l'entité » en « l&#x27;entité » — illisible dans la source, et surtout
    INTROUVABLE : un lecteur qui cherche une phrase du document dans son
    source ne la trouverait pas.
    ⚠️ C'est un test qui l'a révélé, en échouant à retrouver dans le document
    une constante que le module y écrit pourtant. La leçon vaut plus que le
    correctif : un document destiné à être VÉRIFIÉ doit contenir ses phrases
    telles qu'on les cherchera.
    """
    return _html.escape(str(texte), quote=False)


def _section_piece(piece) -> str:
    """Une section par pièce — présente ou non, à sa place dans le flux.

    ⚠️ UNE ABSENCE N'EST PAS UNE NOTE DE BAS DE PAGE. Elle occupe la même
    surface qu'une présence : c'est ce qui la rend visible plutôt que muette.
    """
    e = _e
    reference = e(PIECES[piece.nom])
    if piece.presente:
        return (f'<h2><span class="etiquette">{e(piece.nom)}</span> — '
                f'{reference}</h2>\n<p>Pièce FOURNIE.</p>')
    return (f'<h2><span class="etiquette">{e(piece.nom)}</span> — '
            f'{reference}</h2>\n'
            f'<div class="absente"><p><strong>PIÈCE ABSENTE.</strong> '
            f'{e(piece.motif)}</p>'
            f'<p>Cette absence est DÉCLARÉE, avec son motif. Elle n\'est ni '
            f'un oubli ni un zéro.</p></div>')


def rendre_html(etat: EtatDate) -> str:
    """L'état daté, en un document autonome — ou un REFUS.

    ⚠️ UN SEUL PARAMÈTRE, ET C'EST LE VERROU LE PLUS DISCRET DE CE MODULE.
    Aucun champ libre n'entre ici : pas de « commentaire », pas de titre
    personnalisé, pas de note. Le document n'imprime que ce que l'état
    porte, et l'état ne porte que ce que du code testé a produit.
    """
    e = _e
    lignes = [
        '<!DOCTYPE html>', '<html lang="fr"><head>',
        '<meta charset="utf-8">',
        f'<title>{e(TITRE)} au {e(etat.arrete)}</title>',
        _style(), '</head><body>',
        f'<h1>{e(TITRE)} AU {e(etat.arrete)} — {e(etat.entite)}</h1>',
        f'<p class="sous-titre">{e(SOUS_TITRE)}</p>',
        '<div class="reserve">',
        '<h2>Ce que ce document n\'est pas</h2>',
        f'<p>{e(CE_QUE_CE_DOCUMENT_N_EST_PAS)}</p>',
        f'<p><strong>{e(VERDICT_DU_DOCUMENT)}</strong></p>',
        '</div>',
        '<h2>Articulations vérifiées</h2>', '<table>',
        '<tr><th>contrôle</th><th>verdict</th><th>ce qu\'il établit</th></tr>',
    ]
    for a in etat.articulations:
        marque = ('NON ÉTABLIE — le contrôle n\'a pas pu tourner'
                  if a.verdict == NON_ETABLIE else 'ÉTABLIE')
        lignes.append(f'<tr><td>{e(a.nom)}</td><td>{e(marque)}</td>'
                      f'<td>{e(a.motif)}</td></tr>')
    lignes.append('</table>')
    lignes.append('<p>« NON ÉTABLIE » n\'est pas « concordante » : le '
                  'contrôle n\'a pas pu tourner, il n\'a donc rien '
                  'constaté.</p>')
    for piece in etat.pieces:
        lignes.append(_section_piece(piece))
    lignes += ['<h2>Portée de cet état</h2>', f'<p>{e(etat.motif)}</p>',
               '</body></html>']

    document = '\n'.join(lignes)
    _exiger_aucune_piece_omise(document, etat)
    _exiger_autonomie(document)
    return document


def rendre_classeur(etat: EtatDate, chemin) -> Path:
    """Le LIVRABLE RÉEL : la matière que l'entité réintègre chez elle.

    ⚠️⚠️ AUCUN DE NOS FORMATS N'EST « JOINT À DES COMPTES ». Les annexes d'un
    rapport annuel sont produites par l'entité, dans son outillage. Ce que la
    plateforme livre est la MATIÈRE — et une matière se réutilise mieux dans
    un classeur que dans une page.

    ⚠️ ET LE WORD EST ÉCARTÉ, PAS OUBLIÉ : c'est le format qui invite le plus
    à être collé tel quel dans un rapport annuel. Tant que les tableaux du
    §99 a) et la séparation du cédé manquent, en produire reviendrait à
    fabriquer l'objet même que ce lot a été retardé pour éviter.
    """
    # ⚠️ IMPORT TARDIF, ET IL EST DÉLIBÉRÉ : `openpyxl` n'est utile qu'ici.
    # L'importer en tête ferait échouer TOUT le module — dont le rendu HTML,
    # qui n'en a aucun besoin — sur un poste où le paquet manque.
    from openpyxl import Workbook
    chemin = Path(chemin)
    chemin.parent.mkdir(parents=True, exist_ok=True)
    livre = Workbook()

    tete = livre.active
    tete.title = 'En-tete'
    for ligne in (('Titre', TITRE), ('Sous-titre', SOUS_TITRE),
                  ('Arrete', etat.arrete), ('Entite', etat.entite),
                  ("Ce que ce document n'est pas",
                   CE_QUE_CE_DOCUMENT_N_EST_PAS),
                  ('Verdict', VERDICT_DU_DOCUMENT),
                  ('Portee', etat.motif)):
        tete.append(list(ligne))

    feuille = livre.create_sheet('Pieces')
    feuille.append(['piece', 'reference', 'etat', 'motif'])
    for p in etat.pieces:
        feuille.append([p.nom, PIECES[p.nom],
                        'FOURNIE' if p.presente else 'ABSENTE', p.motif])

    art = livre.create_sheet('Articulations')
    art.append(['controle', 'verdict', 'motif'])
    for a in etat.articulations:
        art.append([a.nom, a.verdict, a.motif])

    livre.save(chemin)
    return chemin


def pieces_du_document(document: str) -> tuple:
    """Les pièces effectivement écrites dans un document rendu.

    ⚠️ ELLE SE LIT SUR LE DOCUMENT, PAS SUR L'ÉTAT — c'est ce qui permet de
    vérifier le rendu plutôt que de le croire sur parole.
    """
    return tuple(sorted(n for n in PIECES
                        if re.search(r'\b' + re.escape(n) + r'\b', document)))
