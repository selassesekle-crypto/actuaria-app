# -*- coding: utf-8 -*-
"""Tests R1-R3 — le document qui n'omet rien, n'appelle personne, et le dit.

⚠️ GATE : `py -m unittest discover -s normes -t .`

⚠️ LES TROIS TESTS QUI PORTENT CE LOT sont ceux qui vérifient le DOCUMENT
PRODUIT : aucune pièce omise, aucune requête externe, aucun champ libre. Les
trois se lisent sur la sortie, jamais sur l'intention du gabarit.
"""
import ast
import inspect
import pathlib
import shutil
import tempfile
import unittest

from normes.ifrs17.etats.assemblage import PIECES, assembler
from normes.ifrs17.etats.rendu import (
    CE_QUE_CE_DOCUMENT_N_EST_PAS,
    FORMES_EXTERNES,
    MOTIF_PIECE_OMISE,
    MOTIF_REQUETE_EXTERNE,
    SOUS_TITRE,
    TITRE,
    VERDICT_DU_DOCUMENT,
    _exiger_autonomie,
    pieces_du_document,
    rendre_classeur,
    rendre_html,
)
from normes.ifrs17.mesure.bilan import SoldeGroupe, etat_situation_financiere
from normes.ifrs17.mesure.declaration import ContexteEvaluation
from normes.ifrs17.mesure.lrc_paa import RefusMesure

CONTEXTE = ContexteEvaluation(arrete='2026-12-31', portefeuilles=('DO', 'MRH'))
BILAN = etat_situation_financiere(
    [SoldeGroupe('DO', 'DO|AUTRES|2026', 1000.0),
     SoldeGroupe('MRH', 'MRH|AUTRES|2026', -300.0)], CONTEXTE)


class _Resultat:
    insurance_revenue = 800.0
    insurance_service_expenses = -500.0
    charges_financieres = -60.0


def _etat():
    return assembler(
        arrete='2026-12-31', entite='MutuelleTest',
        pieces={'PERIMETRE_PUBLIE': 'PÉRIMÈTRE IFRS 17 — …',
                'BILAN_78': BILAN, 'RESULTAT_80': _Resultat()},
        absences={
            'RAPPROCHEMENTS_100': 'premier exercice, aucune ouverture signee',
            'NOTES_97_119_120': 'les cinq declarations ne sont pas signees',
            'REGISTRE_DES_GROUPES': 'premier exercice, aucun registre',
            'DEVELOPPEMENT_130': 'aucune donnee de sinistres remise',
            'FINANCEMENT_56': 'aucun groupe a composante financement'})


class R1_AucunePieceOmise(unittest.TestCase):
    """⚠️⚠️ LE CŒUR DU LOT. L'assemblage a fermé l'oubli au niveau de l'état ;
    un rendu capable de sauter une section le rouvrirait un cran plus bas,
    là où personne ne regarde."""

    def test_les_HUIT_pieces_figurent_au_document(self):
        doc = rendre_html(_etat())
        self.assertEqual(pieces_du_document(doc), tuple(sorted(PIECES)))
        print(f"    OK R1 : les {len(PIECES)} pieces figurent au document")

    def test_une_ABSENCE_occupe_la_meme_surface_qu_une_presence(self):
        """⚠️ Une absence n'est pas une note de bas de page : elle a sa
        section, à sa place dans le flux, avec son motif."""
        doc = rendre_html(_etat())
        self.assertEqual(doc.count('<h2><span class="etiquette">'),
                         len(PIECES))
        self.assertIn('PIÈCE ABSENTE.', doc)
        self.assertIn('aucune donnee de sinistres remise', doc)
        self.assertIn("ni un oubli ni un zéro", doc)

    def test_UN_DOCUMENT_QUI_OMET_UNE_PIECE_EST_REFUSE(self):
        """⚠️ La vérification porte sur le DOCUMENT PRODUIT, pas sur
        l'intention du gabarit — c'est ce qui la rend vérifiable."""
        etat = _etat()
        ampute = rendre_html(etat).replace('DEVELOPPEMENT_130', 'X')
        from normes.ifrs17.etats.rendu import _exiger_aucune_piece_omise
        with self.assertRaises(RefusMesure) as e:
            _exiger_aucune_piece_omise(ampute, etat)
        self.assertEqual(e.exception.motif, MOTIF_PIECE_OMISE)
        self.assertIn('INDISCERNABLE', str(e.exception))
        print("    OK R1b : un document ampute d'une piece -> REFUSE")

    def test_les_articulations_NON_ETABLIES_sont_ECRITES_pas_omises(self):
        """⚠️ « NON ÉTABLIE » n'est pas « concordante » — et un tiret ne le
        dirait pas."""
        doc = rendre_html(_etat())
        self.assertIn("NON ÉTABLIE — le contrôle n'a pas pu tourner", doc)
        self.assertIn("n'a donc rien constaté", doc)


class R1b_LeTitreEtLaReserveEnTETE(unittest.TestCase):
    """⚠️ LE TITRE EST LA SEULE LIGNE DONT ON SOIT CERTAIN QU'ELLE SERA LUE.
    Le périmètre publié a déjà payé une sur-affirmation en tête."""

    def test_le_document_ne_s_appelle_PAS_etats_financiers(self):
        doc = rendre_html(_etat())
        self.assertIn(TITRE, doc)
        self.assertEqual(TITRE, 'ÉTAT DE LA MESURE IFRS 17')
        self.assertNotIn('États financiers IFRS 17', doc)
        self.assertIn(SOUS_TITRE, doc)
        print(f"    OK R1c : titre « {TITRE} » -- une mesure, datee, un "
              "intrant")

    def test_la_reserve_est_AVANT_les_pieces_pas_apres(self):
        """⚠️ Le périmètre publié portait sa sur-affirmation dans le
        préambule : c'est là qu'on lit."""
        doc = rendre_html(_etat())
        self.assertLess(doc.index("Ce que ce document n'est pas"),
                        doc.index('BILAN_78'))
        for morceau in ('IAS 1', 'flux de trésorerie', '§99 a)',
                        '§78 c)', '§82', 'INTÉGRÉES'):
            self.assertIn(morceau, CE_QUE_CE_DOCUMENT_N_EST_PAS, morceau)

    def test_le_verdict_du_chantier_DESCEND_dans_le_document(self):
        """⚠️ « Un beau document ne comble aucun de ces trous — il les
        affiche. »"""
        self.assertIn(VERDICT_DU_DOCUMENT, rendre_html(_etat()))
        self.assertIn('NE COMBLE AUCUN DE CES TROUS', VERDICT_DU_DOCUMENT)


class R2_AucuneRequeteExterne(unittest.TestCase):
    """⚠️⚠️ LA CHAÎNE A7 N'EST PAS AUTONOME — mesuré : polices chez Google,
    script chez jsDelivr. Un document d'archive qui dépend d'un tiers À LA
    LECTURE n'est pas opposable."""

    def test_le_document_produit_n_appelle_PERSONNE(self):
        doc = rendre_html(_etat())
        for forme in ('http://', 'https://', '@import', '<link', '<script',
                      'src=', 'url('):
            self.assertNotIn(forme, doc, forme)
        print(f"    OK R2 : {len(doc)} octets, zero requete externe")

    def test_LE_BALAYAGE_ATTRAPE_CHAQUE_FORME_UNE_PAR_UNE(self):
        """⚠️⚠️ UN CONTRÔLE QUI N'A JAMAIS ÉCHOUÉ NE PROUVE RIEN — et un
        exemple qui déclenche TROIS formes à la fois n'en éprouve aucune.

        ⚠️ La première rédaction annonçait « les 6 formes » alors qu'il y en
        a SEPT, et son exemple pour « https:// » était
        `@import url('https://…')` : il portait `https://`, `@import` ET
        `url(` ensemble. Le refus venait donc peut-être d'une autre forme, et
        `@import` seul n'était jamais éprouvé. Chaque exemple ci-dessous ne
        porte QU'UNE forme, et un test le vérifie avant de conclure.
        """
        exemples = {
            'http://': '<a href=[http://x]>',
            'https://': '<a href=[https://x]>',
            '@import': '<style>@import "fond";</style>',
            '<link': '<link rel=[stylesheet]>',
            '<script': '<script>1</script>',
            'src=': '<img src="logo.png">',
            'url(': '<style>background:url(fond.png)</style>',
        }
        self.assertEqual(sorted(exemples), sorted(FORMES_EXTERNES))
        for forme, exemple in exemples.items():
            portees = [f for f in FORMES_EXTERNES if f in exemple]
            self.assertEqual(
                portees, [forme],
                f"l'exemple de « {forme} » porte aussi {portees} : il "
                f"n'éprouve pas ce qu'il prétend")
            with self.assertRaises(RefusMesure, msg=forme) as e:
                _exiger_autonomie('<html>' + exemple + '</html>')
            self.assertEqual(e.exception.motif, MOTIF_REQUETE_EXTERNE)
        print(f"    OK R2b : les {len(FORMES_EXTERNES)} formes eprouvees "
              "SEULES, chacune verifiee comme telle")

    def test_le_refus_dit_que_c_est_une_question_D_OPPOSABILITE(self):
        with self.assertRaises(RefusMesure) as e:
            _exiger_autonomie('<html><script src="x"></script></html>')
        self.assertIn("opposabilité, pas d'esthétique", str(e.exception))
        self.assertIn('cinq ans', str(e.exception))

    def test_le_style_est_EMBARQUE_et_ses_polices_sont_GENERIQUES(self):
        doc = rendre_html(_etat())
        self.assertIn('<style>', doc)
        self.assertIn('Georgia', doc)
        self.assertIn('print-color-adjust', doc)


class R3_LeClasseurEstLeLivrableReel(unittest.TestCase):
    """⚠️ AUCUN DE NOS FORMATS N'EST « JOINT À DES COMPTES » : les annexes
    sont produites par l'entité. Ce que la plateforme livre est la MATIÈRE,
    et une matière se réutilise mieux dans un classeur."""

    def _fichier(self):
        d = tempfile.mkdtemp(prefix='rendu_')
        self.addCleanup(shutil.rmtree, d, ignore_errors=True)
        return pathlib.Path(d) / 'etat.xlsx'

    def test_le_classeur_porte_les_trois_feuilles_et_toutes_les_pieces(self):
        from openpyxl import load_workbook
        livre = load_workbook(rendre_classeur(_etat(), self._fichier()))
        self.assertEqual(livre.sheetnames,
                         ['En-tete', 'Pieces', 'Articulations'])
        lues = {r[0] for r in livre['Pieces'].iter_rows(min_row=2,
                                                        values_only=True)}
        self.assertEqual(lues, set(PIECES))
        print(f"    OK R3 : 3 feuilles, {len(lues)} pieces au classeur")

    def test_le_classeur_porte_les_ABSENCES_avec_leur_motif(self):
        from openpyxl import load_workbook
        livre = load_workbook(rendre_classeur(_etat(), self._fichier()))
        absentes = [(r[0], r[3]) for r in
                    livre['Pieces'].iter_rows(min_row=2, values_only=True)
                    if r[2] == 'ABSENTE']
        self.assertEqual(len(absentes), 5)
        for _, motif in absentes:
            self.assertTrue((motif or '').strip())

    def test_le_classeur_porte_la_reserve_en_tete(self):
        from openpyxl import load_workbook
        livre = load_workbook(rendre_classeur(_etat(), self._fichier()))
        tete = {r[0]: r[1] for r in
                livre['En-tete'].iter_rows(values_only=True)}
        self.assertEqual(tete['Titre'], TITRE)
        self.assertIn('IAS 1', tete["Ce que ce document n'est pas"])
        self.assertIn('NE COMBLE AUCUN', tete['Verdict'])


class Z_AucunCHAMP_LIBRE(unittest.TestCase):
    """⚠️⚠️ LE BIAIS QUE LE VERROU AST NE VERRAIT PAS. Il regarde les
    IMPORTS, pas les chaînes : un document qui imprimerait un
    « commentaire » fourni par l'appelant laisserait entrer du texte généré
    SANS QU'UNE SEULE LIGNE DE CODE CHANGE."""

    def test_le_rendu_n_accepte_QUE_l_etat(self):
        params = list(inspect.signature(rendre_html).parameters)
        self.assertEqual(params, ['etat'])
        self.assertEqual(list(inspect.signature(rendre_classeur).parameters),
                         ['etat', 'chemin'])
        print("    OK Rz : rendre_html n'accepte QUE l'etat -- aucun champ "
              "libre, aucun commentaire, aucun titre personnalise")

    def test_aucune_fonction_publique_n_accepte_de_texte_libre(self):
        from normes.ifrs17.etats import rendu
        interdits = {'commentaire', 'note', 'texte', 'narration', 'titre',
                     'libelle_personnalise', 'entete'}
        for nom, f in vars(rendu).items():
            if nom.startswith('_') or not inspect.isfunction(f):
                continue
            fuite = set(inspect.signature(f).parameters) & interdits
            self.assertEqual(fuite, set(), f'{nom} accepte {fuite}')

    def test_ce_module_n_atteint_PAS_le_modele_de_langage(self):
        """⚠️ Le verrou global le couvre parce que ce module vit sous
        `normes/`. Ailleurs, il pourrait appeler la frontière sans rien
        violer — d'où sa place."""
        from normes.ifrs17.etats import rendu
        arbre = ast.parse(pathlib.Path(rendu.__file__).read_text(
            encoding='utf-8'))
        noms = {n.id for n in ast.walk(arbre) if isinstance(n, ast.Name)}
        noms |= {n.attr for n in ast.walk(arbre)
                 if isinstance(n, ast.Attribute)}
        self.assertEqual(noms & {'appeler', 'anthropic'}, set())
        self.assertIn('normes' + '/ifrs17', pathlib.Path(
            rendu.__file__).as_posix())


if __name__ == '__main__':
    unittest.main(verbosity=2)
